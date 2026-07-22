"""
Create bundle - tool สร้าง bundle อัตโนมัติบนเว็บ aztek-tools (combo-interactive)
- login ด้วย persistent chrome profile
- เลือกเกม/เซิร์ฟ -> Shop V2 -> Bundles
- สร้าง bundle: ชื่อ/Type/ส่งของทันที + เพิ่มไอเทม + reward (Credit/Debit/Mileage/PlayerExp)
- ตั้งจำนวน/Tier, Save, ดึงเลข bundle + รีเชค
- import .xlsx/.csv (หลาย bundle), Fetch ตัวเลือก reward, สร้างทุก bundle อัตโนมัติ
รัน:  python new_tool.py
"""

import os
import re
import sys
import csv
import json
import asyncio
import threading
import traceback
import tkinter as tk
from tkinter import ttk, messagebox

try:
    from playwright.async_api import async_playwright
    PW_OK = True
except ImportError:
    PW_OK = False

# ธีม/ฟอนต์/ไอคอน จากของกลาง
from ui_common import C, FM, FB, _find_icon, _set_window_icon, allow_wide_popdown
import aztek_core as core

# เกม/เซิร์ฟ + Chrome + prefs ใช้จาก aztek_core (แหล่งเดียว — เพิ่มเซิร์ฟใหม่แก้ที่ core ที่เดียว)
GAME_BASES = core.GAME_BASES
GAME_NAMES = core.GAME_NAMES
BUNDLE_TYPES = ['CHOICE', 'FIXED', 'GACHAPON', 'GACHAPON_LIMIT', 'RANDOM']
DEFAULT_BUNDLE_TYPE = 'FIXED'          # ประเภทเริ่มต้นของทุกบันเดิล
TIERS = ['Common', 'Rare', 'Epic', 'Mystic', 'Legend']
DEFAULT_TIER = 'Common'
REWARD_KINDS = ['CREDIT', 'DEBIT', 'MILEAGE', 'PLAYER_EXP']
REWARD_COLS = [('credit_value', 'CREDIT', 'Credit'),
               ('debit_value', 'DEBIT', 'Debit'),
               ('mileage_value', 'MILEAGE', 'Mileage'),
               ('playerexp_value', 'PLAYER_EXP', 'Player Experience')]
REWARD_LABELS = {kind: label for _, kind, label in REWARD_COLS}

SEL = {
    'add_bundle_btn':   "เพิ่ม Bundle",
    'name_label':       "ชื่อบันเดิล",
    'type_placeholder': "เลือกหมวดหมู่",
    'deliver_label':    "ส่งของทันที",
    'item_label':       "ค้นหาไอเท็ม",
    'search_btn':       "Search",
    'add_to_bundle':    "เพิ่มเข้าบันเดิล",
    'expand_all':       "Expand All",
    'qty_label':        "จำนวน",
    'tier_label':       "Tier",
    'save_btn':         "ยืนยันการสร้างบันเดิล",
    'bundle_search':    "Aztek Bundle Id",
    'currency_ph':      "เลือกสกุลเงิน",
    'rank_ph':          "เลือกยศ",
    'rate_label':       "เรทสุ่ม",
}


def game_url(game, page='bundles'):
    return core.game_url(game, page)


_get_app_dir = core._get_app_dir
find_chrome_exe = core.find_chrome_exe
# login ร่วมกับ tool อื่น: Chrome profile + prefs 'game' ใช้ไฟล์เดียวกับ core
CHROME_PROFILE = core.CHROME_PROFILE
load_prefs = core.load_prefs
save_prefs = core.save_prefs
# options (รายการ CREDIT/DEBIT/... ) เป็นข้อมูลเฉพาะ tool นี้ -> เก็บไฟล์แยกของตัวเอง
OPTIONS_FILE = os.path.join(core.APP_DIR, '.new_tool_options.json')


def _read_options_file():
    try:
        with open(OPTIONS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f) or {}
    except Exception:
        return {}


def load_options(game=None):
    """ตัวเลือก reward (สกุลเงิน/ยศ) ของเกม/เซิร์ฟที่ระบุ -> {KIND: [ชื่อ...]}
    ไฟล์เก็บแยกตามเกม: {game: {KIND: [...]}} — คนละเซิร์ฟสกุลเงินไม่เหมือนกัน
    ยังอ่านไฟล์แบบเก่า ({KIND: [...]} ไม่มีชั้นเกม) ได้ ใช้เป็นค่าสำรองให้ทุกเกม"""
    d = _read_options_file()
    legacy = {k: list(d.get(k, [])) for k in REWARD_KINDS} if any(k in d for k in REWARD_KINDS) else None
    if game and isinstance(d.get(game), dict):
        return {k: list(d[game].get(k, [])) for k in REWARD_KINDS}
    if game:
        return legacy or {k: [] for k in REWARD_KINDS}   # เกมนี้ยังไม่เคย fetch
    return legacy or {k: [] for k in REWARD_KINDS}


def save_options(game, opts):
    """เก็บตัวเลือกของเกมนั้น โดยไม่ทับของเกมอื่น"""
    d = _read_options_file()
    if any(k in d for k in REWARD_KINDS):        # ไฟล์แบบเก่า -> ย้ายขึ้นชั้นเกมก่อน
        d = {}
    d[game] = {k: list(opts.get(k, [])) for k in REWARD_KINDS}
    with open(OPTIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(d, f, ensure_ascii=False, indent=2)


# ---------- template ----------
TEMPLATE_HEADERS = ['bundle_name', 'bundle_type', 'deliver_now', 'item_id', 'qty', 'tier',
                    'credit_value', 'debit_value', 'mileage_value', 'playerexp_value',
                    'random_rate']


def _template_rows(opts):
    c = (opts.get('CREDIT') or ['ใส่ค่า Credit'])[0]
    p = (opts.get('PLAYER_EXP') or ['ใส่ยศ'])[0]
    # คอลัมน์สุดท้าย random_rate: ใส่เฉพาะ bundle_type = RANDOM (0.000 - 100.000)
    return [
        ['Bundle A', 'FIXED',  'TRUE',  221942, 1,  'Common', '', '', '', '', ''],
        ['Bundle A', 'FIXED',  'TRUE',  221943, 5,  'Epic',   '', '', '', '', ''],
        ['Bundle A', 'FIXED',  'TRUE',  '',     100, '',      c,  '', '', '', ''],
        ['Bundle A', 'FIXED',  'TRUE',  '',     50,  '',      '', '', '', p,  ''],
        ['Bundle B', 'CHOICE', 'FALSE', 221950, 1,  'Common', '', '', '', '', ''],
        ['Bundle C', 'RANDOM', 'TRUE',  221960, 1,  'Common', '', '', '', '', '12.500'],
        ['Bundle C', 'RANDOM', 'TRUE',  221961, 1,  'Rare',   '', '', '', '', '5.250'],
    ]


def build_template_csv(path, opts=None):
    opts = opts or load_options()
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(TEMPLATE_HEADERS)
        for r in _template_rows(opts):
            w.writerow(r)


def build_template_xlsx(path, opts=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    opts = opts or load_options()
    rows = _template_rows(opts)
    wb = Workbook()
    ws = wb.active
    ws.title = 'bundles'
    ws.append(TEMPLATE_HEADERS)
    for r in rows:
        ws.append(r)

    hdr_fill = PatternFill('solid', fgColor='1F6FEB')
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 22
    for row in ws.iter_rows(min_row=2, max_row=1 + len(rows), max_col=len(TEMPLATE_HEADERS)):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            if cell.column_letter in ('D', 'E'):
                cell.alignment = Alignment(horizontal='center')
    for col, w in {'A': 16, 'B': 14, 'C': 12, 'D': 11, 'E': 7, 'F': 11,
                   'G': 16, 'H': 16, 'I': 16, 'J': 18, 'K': 13}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'

    wl = wb.create_sheet('lists')
    wl.sheet_state = 'hidden'
    list_cols = {}
    for ci, kind in enumerate(REWARD_KINDS):
        col_letter = chr(ord('A') + ci)
        vals = list(dict.fromkeys(opts.get(kind, [])))
        for ri, v in enumerate(vals, 1):
            wl.cell(row=ri, column=ci + 1, value=v)
        list_cols[kind] = (col_letter, len(vals))

    last = 500
    dv_type = DataValidation(type='list', formula1='"%s"' % ','.join(BUNDLE_TYPES), allow_blank=True)
    dv_dlv = DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=True)
    dv_tier = DataValidation(type='list', formula1='"%s"' % ','.join(TIERS), allow_blank=True)
    ws.add_data_validation(dv_type); dv_type.add('B2:B%d' % last)
    ws.add_data_validation(dv_dlv); dv_dlv.add('C2:C%d' % last)
    ws.add_data_validation(dv_tier); dv_tier.add('F2:F%d' % last)
    reward_target = {'CREDIT': 'G', 'DEBIT': 'H', 'MILEAGE': 'I', 'PLAYER_EXP': 'J'}
    for kind in REWARD_KINDS:
        sheet_col, n = list_cols[kind]
        if n <= 0:
            continue
        tgt = reward_target[kind]
        dv = DataValidation(type='list',
                            formula1='lists!$%s$1:$%s$%d' % (sheet_col, sheet_col, n),
                            allow_blank=True)
        ws.add_data_validation(dv); dv.add('%s2:%s%d' % (tgt, tgt, last))

    ws2 = wb.create_sheet('คำอธิบาย')
    ws2.column_dimensions['A'].width = 100
    info = [
        ('วิธีใช้ template นำเข้า Bundle', True),
        ('', False),
        ('1 แถว = 1 รายการในบันเดิล (ไอเท็ม หรือ reward)', False),
        ('แถวที่มี bundle_name เดียวกัน = อยู่ bundle เดียวกัน (เรียงต่อกัน)', False),
        ('bundle_type / deliver_now ใช้ค่าจากแถวแรกของแต่ละ bundle', False),
        ('', False),
        ('แต่ละแถวกรอก "ช่องเดียว" ว่าจะเป็นอะไร + qty:', True),
        ('  - ไอเท็ม: item_id (+tier)', False),
        ('  - Credit: credit_value', False),
        ('  - Debit: debit_value', False),
        ('  - Mileage: mileage_value', False),
        ('  - Player Experience: playerexp_value', False),
        ('', False),
        ('qty = จำนวน (ไม่ใส่ = 1) | tier = Common/Rare/Epic/Mystic/Legend (เฉพาะไอเท็ม)', False),
        ('random_rate = เรทสุ่ม 0.000-100.000 (ใส่เฉพาะ bundle_type = RANDOM, ต่อไอเท็ม)', False),
        ('* dropdown reward มาจากปุ่ม "Fetch ตัวเลือก" ในตัว tool', False),
    ]
    for i, (txt, bold) in enumerate(info, 1):
        cell = ws2.cell(row=i, column=1, value=txt)
        cell.font = Font(name='Arial', size=12 if bold else 10, bold=bold,
                         color='1F6FEB' if bold else '000000')
    wb.save(path)


def extract_bundle_id(data):
    prefer = ['bundleid', 'bundle_id', 'bundleno', 'bundle_no', 'id', 'no', 'number']
    found = {}
    queue = [data]
    while queue:
        o = queue.pop(0)
        if isinstance(o, dict):
            for k, v in o.items():
                kl = str(k).lower().replace(' ', '')
                if (isinstance(v, int) and not isinstance(v, bool)) or (isinstance(v, str) and v.isdigit()):
                    found.setdefault(kl, str(v))
                if isinstance(v, (dict, list)):
                    queue.append(v)
        elif isinstance(o, list):
            for v in o:
                if isinstance(v, (dict, list)):
                    queue.append(v)
    for k in prefer:
        if k in found:
            return found[k]
    for k, v in found.items():
        if k.endswith('id'):
            return v
    return None


def extract_id_by_name(data, name):
    target = str(name).strip().lower()
    prefer = ('bundleid', 'bundle_id', 'bundleno', 'bundle_no', 'id', 'no', 'number')
    result = []

    def walk(o):
        if result:
            return
        if isinstance(o, dict):
            has_name = any(isinstance(v, str) and v.strip().lower() == target for v in o.values())
            if has_name:
                lk = {str(k).lower().replace(' ', ''): v for k, v in o.items()}
                for k in prefer:
                    v = lk.get(k)
                    if isinstance(v, int) and not isinstance(v, bool):
                        result.append(str(v)); return
                    if isinstance(v, str) and v.isdigit():
                        result.append(v); return
                for v in o.values():
                    if isinstance(v, int) and not isinstance(v, bool):
                        result.append(str(v)); return
            for v in o.values():
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)
    walk(data)
    return result[0] if result else None


def extract_ids_by_name(data, name):
    """เก็บ 'ทุก' เลข bundle ที่ชื่อตรงกับ name (ไม่หยุดที่ตัวแรก) -> list สตริงเลข (ไม่ซ้ำ คงลำดับ)
    ใช้ตอนรีเชค: ชื่อ bundle ซ้ำกันได้ (ของเก่าในเว็บชื่อเดียวกัน) การหยิบตัวแรกอาจได้ผิดตัว
    -> เก็บหมดแล้วค่อยเลือก (ตัวที่ตรงกับเลขตอนสร้าง หรือเลขล่าสุด)"""
    target = str(name).strip().lower()
    prefer = ('bundleid', 'bundle_id', 'bundleno', 'bundle_no', 'id', 'no', 'number')
    out = []

    def walk(o):
        if isinstance(o, dict):
            if any(isinstance(v, str) and v.strip().lower() == target for v in o.values()):
                lk = {str(k).lower().replace(' ', ''): v for k, v in o.items()}
                picked = None
                for k in prefer:
                    v = lk.get(k)
                    if isinstance(v, bool):
                        continue
                    if isinstance(v, int):
                        picked = str(v); break
                    if isinstance(v, str) and v.isdigit():
                        picked = v; break
                if picked is None:
                    for v in o.values():
                        if isinstance(v, int) and not isinstance(v, bool):
                            picked = str(v); break
                if picked is not None:
                    out.append(picked)
            for v in o.values():
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)
    walk(data)
    seen, uniq = set(), []
    for x in out:
        if x not in seen:
            seen.add(x); uniq.append(x)
    return uniq


def pick_recheck_id(ids, created_id):
    """เลือกเลขรีเชคจากผลค้นหาชื่อ (ที่อาจซ้ำ):
    1) ถ้าเลขตอนสร้างอยู่ในผล -> ใช้เลขนั้น (ยืนยันตรง กันชื่อซ้ำหลอกให้ mismatch)
    2) ไม่งั้น -> เลขมากสุด (bundle ที่เพิ่งสร้าง = id สูงสุด)
    3) ไม่มีเลย -> None"""
    if created_id and str(created_id) in ids:
        return str(created_id)
    if ids:
        return max(ids, key=lambda s: int(s))
    return None


# C/FM/FB import จาก ui_common ข้างบนแล้ว


async def first_visible(page, selectors, timeout_each=2500):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            await loc.wait_for(state='visible', timeout=timeout_each)
            return loc
        except Exception:
            continue
    return None


class _BundleEditor:
    """ไส้ในฟอร์ม 1 bundle (ชื่อ/Type/ส่งของ/คอลัมน์ไอเทม/รางวัล) ที่ render ลง frame ใดก็ได้
    ใช้ร่วมกัน: open_bundle_dialog (Toplevel) กับ open_bundle_review (ฟอร์มขวาใน master-detail)
    app = instance ของ App (ใช้ helper: _parse_columns/set_combo_values/_fetch_options/log/_item_names/vgame)"""

    def __init__(self, app, parent, item_mode='columns'):
        # item_mode='columns' = วาง Excel ทีละคอลัมน์ (ฟอร์มเดิม) | 'rows' = รายไอเทม + ปุ่มตัด/เรียง (หน้ารีวิว)
        self.app = app
        self.mode = item_mode
        self._win = parent.winfo_toplevel()   # ไว้เป็น parent ของ messagebox
        self.rewards = []                      # [{'type','value','qty'}]
        self._items = []                       # rows mode: รายไอเทมตามลำดับ [{'id','name','qty','tier','rate'?}]
        self._build(parent)

    def _build(self, parent):
        app = self.app
        # --- หัว bundle: ชื่อ / Type / ส่งของทันที ---
        top = tk.LabelFrame(parent, text='ข้อมูล Bundle', bg=C['bg_med'], fg=C['muted'],
                            font=FM, bd=1, relief='solid')
        top.pack(fill='x', padx=10, pady=(10, 6))
        r1 = tk.Frame(top, bg=C['bg_med'])
        r1.pack(fill='x', padx=10, pady=(8, 4))
        tk.Label(r1, text='ชื่อบันเดิล:', bg=C['bg_med'], fg=C['text'], font=FM,
                 width=10, anchor='w').pack(side='left')
        self.v_name = tk.StringVar()
        self.e_name = tk.Entry(r1, textvariable=self.v_name, bg=C['bg_inp'], fg=C['text'],
                               insertbackground=C['text'], font=FM, relief='flat')
        self.e_name.pack(side='left', fill='x', expand=True, ipady=3)
        r2 = tk.Frame(top, bg=C['bg_med'])
        r2.pack(fill='x', padx=10, pady=(4, 8))
        tk.Label(r2, text='Type:', bg=C['bg_med'], fg=C['text'], font=FM,
                 width=10, anchor='w').pack(side='left')
        self.v_type = tk.StringVar(value=app.vtype.get() or DEFAULT_BUNDLE_TYPE)
        ttk.Combobox(r2, textvariable=self.v_type, values=BUNDLE_TYPES, state='readonly',
                     width=16, font=FM).pack(side='left')
        self.v_dlv = tk.BooleanVar(value=app.vdeliver.get())
        tk.Checkbutton(r2, text='ส่งของทันที', variable=self.v_dlv, bg=C['bg_med'], fg=C['text'],
                       font=FM, selectcolor=C['bg_inp'], activebackground=C['bg_med'],
                       activeforeground=C['text']).pack(side='left', padx=12)

        # --- พื้นที่ไอเทม: คอลัมน์ (วาง Excel) หรือ รายไอเทม (ตัด/เรียง) ตามโหมด ---
        if self.mode == 'rows':
            self._build_item_rows(parent)
        else:
            self._build_item_columns(parent)

        # --- รางวัล: Credit / Debit / Mileage / Player Experience ---
        rw = tk.LabelFrame(parent, text='Credit / Debit / Mileage / Player Experience (ใส่กี่รายการก็ได้)',
                           bg=C['bg_med'], fg=C['muted'], font=FM, bd=1, relief='solid')
        rw.pack(fill='x', padx=10, pady=(0, 6))
        rline = tk.Frame(rw, bg=C['bg_med'])
        rline.pack(fill='x', padx=8, pady=(6, 2))
        self.v_kind = tk.StringVar(value=REWARD_COLS[0][2])
        self._kind_by_label = {label: kind for _c, kind, label in REWARD_COLS}
        ttk.Combobox(rline, textvariable=self.v_kind, values=[l for _c, _k, l in REWARD_COLS],
                     state='readonly', width=17, font=FM).pack(side='left')
        self.v_val = tk.StringVar()
        # readonly = คลิกที่ไหนของช่องก็เปิด dropdown ได้ (ไม่ใช่แค่ลูกศร) + กันพิมพ์มั่ว
        self.cb_val = ttk.Combobox(rline, textvariable=self.v_val, width=22, font=FM, state='readonly')
        self.cb_val.pack(side='left', padx=6)
        tk.Label(rline, text='จำนวน', bg=C['bg_med'], fg=C['text'], font=FM).pack(side='left')
        self.v_qty = tk.StringVar()
        tk.Entry(rline, textvariable=self.v_qty, width=8, bg=C['bg_inp'], fg=C['text'],
                 insertbackground=C['text'], font=FM, relief='flat').pack(side='left', padx=6, ipady=2)

        def _sync_vals(*_):
            kind = self._kind_by_label.get(self.v_kind.get(), REWARD_COLS[0][1])
            vals = load_options(app.vgame.get()).get(kind, [])
            app.set_combo_values(self.cb_val, vals, state='readonly')
            if self.v_val.get() not in vals:
                self.v_val.set(vals[0] if vals else '')
        self.v_kind.trace_add('write', _sync_vals)
        _sync_vals()
        self._sync_vals = _sync_vals

        def _redraw():
            self.lb_rw.delete(0, tk.END)
            for r in self.rewards:
                self.lb_rw.insert(tk.END, '%-18s %-24s x %s'
                                  % (REWARD_LABELS.get(r['type'], r['type']), r['value'], r['qty']))
        self._redraw = _redraw

        def _add_rw():
            val = self.v_val.get().strip()
            qty = self.v_qty.get().strip()
            if not val:
                messagebox.showwarning('ยังไม่ได้เลือก', 'เลือก/พิมพ์สกุลเงินหรือยศก่อน', parent=self._win)
                return
            if not re.fullmatch(r'\d+', qty or ''):
                messagebox.showwarning('จำนวนไม่ถูกต้อง', 'จำนวนต้องเป็นตัวเลข', parent=self._win)
                return
            self.rewards.append({'type': self._kind_by_label.get(self.v_kind.get(), REWARD_COLS[0][1]),
                                 'value': val, 'qty': qty})
            self.v_qty.set('')
            _redraw()
        self._add_rw = _add_rw

        def _del_rw():
            for i in reversed(self.lb_rw.curselection()):
                del self.rewards[i]
            _redraw()
        self._del_rw = _del_rw

        def _fetch_now():
            app.log('ดึงตัวเลือกของ [%s] ...' % app.vgame.get(), 'STEP')
            app._fetch_options(on_done=_sync_vals)

        tk.Button(rline, text='＋ เพิ่ม', command=_add_rw, bg=C['accent2'], fg='#fff',
                  font=FM, relief='flat', padx=10, cursor='hand2').pack(side='left')
        tk.Button(rline, text='🔄 ดึงตัวเลือก', command=_fetch_now, bg=C['bg_card'], fg=C['text'],
                  font=FM, relief='flat', padx=8, cursor='hand2').pack(side='left', padx=6)
        tk.Button(rline, text='ลบที่เลือก', command=_del_rw, bg=C['bg_card'], fg=C['muted'],
                  font=FM, relief='flat', padx=8, cursor='hand2').pack(side='right')
        self.lb_rw = tk.Listbox(rw, height=4, bg=C['bg_inp'], fg=C['text'], font=('Consolas', 10),
                                relief='flat', selectmode='extended', activestyle='none')
        self.lb_rw.pack(fill='x', padx=8, pady=(2, 8))

    def _build_item_columns(self, parent):
        """โหมด columns: วางจาก Excel ทีละคอลัมน์ (ID/ชื่อ/qty/Tier/เรทสุ่ม) — ใช้ในฟอร์ม 'เพิ่ม Bundle เข้าคิว'"""
        tk.Label(parent, text='วางจาก Excel ได้ทีละคอลัมน์  •  1 บรรทัด = 1 ไอเท็ม  '
                             '(แถวของทุกช่องต้องตรงกัน)  •  ตัดไอเทม = ลบบรรทัด  •  '
                             '"เรทสุ่ม" โผล่เมื่อ Type = RANDOM:',
                 bg=C['bg_dark'], fg=C['muted'], font=('Segoe UI', 9),
                 anchor='w', justify='left').pack(fill='x', padx=10)
        cols = tk.Frame(parent, bg=C['bg_dark'])
        cols.pack(fill='both', expand=True, padx=10, pady=(2, 6))
        self._texts = []

        def _col(title, w):
            f = tk.Frame(cols, bg=C['bg_dark'])
            tk.Label(f, text=title, bg=C['bg_dark'], fg=C['text'], font=FB).pack(anchor='w')
            t = tk.Text(f, bg=C['bg_inp'], fg=C['text'], insertbackground=C['text'],
                        font=('Consolas', 10), relief='flat', wrap='none', width=w, height=12)
            t.pack(fill='both', expand=True)
            self._texts.append(t)
            return f, t

        f_id, self.t_id = _col('Item ID', 10)
        f_name, self.t_name = _col('ชื่อไอเท็ม (แสดงเฉย ๆ)', 20)
        self.t_name.config(fg=C['muted'])
        f_qty, self.t_qty = _col('จำนวน (qty)', 8)
        f_tier, self.t_tier = _col('Tier', 10)
        self._f_rate, self.t_rate = _col('เรทสุ่ม*', 10)
        f_id.pack(side='left', fill='y', padx=(0, 6))
        f_name.pack(side='left', fill='both', expand=True, padx=(0, 6))
        f_qty.pack(side='left', fill='y', padx=(0, 6))
        f_tier.pack(side='left', fill='y', padx=(0, 6))

        def _sync_rate_col(*_):
            if self.v_type.get().upper() == 'RANDOM':
                self._f_rate.pack(side='left', fill='y')
            else:
                self._f_rate.pack_forget()
        self.v_type.trace_add('write', _sync_rate_col)
        _sync_rate_col()

        def _on_wheel(e):                       # เลื่อนทุกช่องพร้อมกัน แถวจะตรงเสมอ
            step = -1 if e.delta > 0 else 1
            for t in self._texts:
                t.yview_scroll(step, 'units')
            return 'break'
        for t in self._texts:
            t.bind('<MouseWheel>', _on_wheel)

    def _build_item_rows(self, parent):
        """โหมด rows: รายไอเทมเรียงลำดับ (บน→ล่าง) + ปุ่มตัดออกรายตัว + ▲▼ ย้ายลำดับ — ใช้ในหน้ารีวิว
        ลำดับใน self._items = ลำดับที่จะขึ้นบนเว็บ (สำคัญ)"""
        tk.Label(parent, text='ไอเทมในบันเดิล  •  เรียงตามลำดับ import อัตโนมัติ (บน→ล่าง = ลำดับบนเว็บ)  •  '
                             'เลือกแถวแล้ว ✕ ตัดออก / ▲▼ ย้ายเอง',
                 bg=C['bg_dark'], fg=C['muted'], font=('Segoe UI', 9),
                 anchor='w', justify='left').pack(fill='x', padx=10)
        box = tk.Frame(parent, bg=C['bg_dark'])
        box.pack(fill='both', expand=True, padx=10, pady=(2, 4))
        tv = ttk.Treeview(box, columns=('no', 'id', 'name', 'qty', 'tier', 'rate'), show='headings',
                          selectmode='browse', height=8)
        # ความกว้างคุมให้รวมพอดี (~492px) เพื่อให้คอลัมน์ 'เรทสุ่ม' โผล่ครบไม่โดนตัดขอบ
        # แม้แพเนลจะแคบ + มีสกอลล์แนวนอนกันเหนียวถ้าย่อจนแคบมาก (minwidth กันคอลัมน์ถูกบีบหาย)
        for c, txt, w, anc, st in (('no', '#', 36, 'center', False), ('id', 'Item ID', 80, 'w', False),
                                   ('name', 'ชื่อไอเท็ม', 170, 'w', True), ('qty', 'จำนวน', 54, 'center', False),
                                   ('tier', 'Tier', 80, 'w', False), ('rate', 'เรทสุ่ม', 72, 'center', False)):
            tv.heading(c, text=txt)
            tv.column(c, width=w, minwidth=w, anchor=anc, stretch=st)
        # คอลัมน์ 'rate' โผล่เฉพาะ Type=RANDOM (คุมด้วย displaycolumns ใน _sync_rate_row)
        tv.configure(displaycolumns=('no', 'id', 'name', 'qty', 'tier'))
        vsb = ttk.Scrollbar(box, orient='vertical', command=tv.yview)
        hsb = ttk.Scrollbar(box, orient='horizontal', command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        box.rowconfigure(0, weight=1)
        box.columnconfigure(0, weight=1)
        self.item_tree = tv

        ctl = tk.Frame(parent, bg=C['bg_dark'])
        ctl.pack(fill='x', padx=10, pady=(0, 4))
        tk.Button(ctl, text='✕ ตัดออก', command=self._row_remove, bg=C['bg_card'], fg=C['danger'],
                  font=FM, relief='flat', padx=10, cursor='hand2').pack(side='left')
        tk.Button(ctl, text='▲', command=lambda: self._row_move(-1), bg=C['bg_card'], fg=C['text'],
                  font=FB, relief='flat', padx=10, cursor='hand2').pack(side='left', padx=(6, 0))
        tk.Button(ctl, text='▼', command=lambda: self._row_move(1), bg=C['bg_card'], fg=C['text'],
                  font=FB, relief='flat', padx=10, cursor='hand2').pack(side='left', padx=(2, 0))
        tk.Label(ctl, text='   แก้ที่เลือก → จำนวน', bg=C['bg_dark'], fg=C['muted'],
                 font=('Segoe UI', 9)).pack(side='left')
        self._row_qty = tk.StringVar()
        tk.Entry(ctl, textvariable=self._row_qty, width=6, bg=C['bg_inp'], fg=C['text'],
                 insertbackground=C['text'], font=FM, relief='flat').pack(side='left', padx=4, ipady=2)
        tk.Label(ctl, text='Tier', bg=C['bg_dark'], fg=C['muted'], font=('Segoe UI', 9)).pack(side='left')
        self._row_tier = tk.StringVar()
        ttk.Combobox(ctl, textvariable=self._row_tier, values=TIERS, state='readonly',
                     width=9, font=FM).pack(side='left', padx=4)
        tk.Button(ctl, text='ตั้งค่า', command=self._row_apply, bg=C['bg_card'], fg=C['text'],
                  font=FM, relief='flat', padx=8, cursor='hand2').pack(side='left', padx=(4, 0))

        # วางจำนวน/Tier ทีละคอลัมน์ (ก็อปจาก Excel) -> ใช้กับไอเทมตามลำดับบน→ล่าง
        paste = tk.LabelFrame(parent, text='วางจำนวน / Tier ทีละคอลัมน์ (ก็อปจาก Excel) — บรรทัดที่ i = ไอเทมลำดับที่ i',
                              bg=C['bg_med'], fg=C['muted'], font=('Segoe UI', 9), bd=1, relief='solid')
        paste.pack(fill='x', padx=10, pady=(2, 6))
        prow = tk.Frame(paste, bg=C['bg_med'])
        prow.pack(fill='x', padx=8, pady=6)
        tk.Label(prow, text='จำนวน', bg=C['bg_med'], fg=C['text'], font=('Segoe UI', 9)).grid(row=0, column=0, sticky='w')
        tk.Label(prow, text='Tier', bg=C['bg_med'], fg=C['text'], font=('Segoe UI', 9)).grid(row=0, column=1, sticky='w', padx=(8, 0))
        self._paste_qty = tk.Text(prow, width=9, height=3, bg=C['bg_inp'], fg=C['text'],
                                  insertbackground=C['text'], font=('Consolas', 10), relief='flat', wrap='none')
        self._paste_qty.grid(row=1, column=0, sticky='w')
        self._paste_tier = tk.Text(prow, width=12, height=3, bg=C['bg_inp'], fg=C['text'],
                                   insertbackground=C['text'], font=('Consolas', 10), relief='flat', wrap='none')
        self._paste_tier.grid(row=1, column=1, sticky='w', padx=(8, 0))
        # คอลัมน์วาง 'เรทสุ่ม' (โผล่เฉพาะ Type=RANDOM) — ฉลาก+ช่องเก็บไว้ toggle พร้อมกัน
        self._paste_rate_lbl = tk.Label(prow, text='เรทสุ่ม', bg=C['bg_med'], fg=C['warn'], font=('Segoe UI', 9))
        self._paste_rate = tk.Text(prow, width=10, height=3, bg=C['bg_inp'], fg=C['warn'],
                                   insertbackground=C['text'], font=('Consolas', 10), relief='flat', wrap='none')
        pbtn = tk.Frame(prow, bg=C['bg_med'])
        pbtn.grid(row=1, column=3, sticky='n', padx=8)
        tk.Button(pbtn, text='📋 ใช้กับไอเทม (ตามลำดับ)', command=self._row_paste_apply,
                  bg=C['accent2'], fg='#fff', font=FM, relief='flat', padx=10, cursor='hand2').pack(anchor='w')
        tk.Button(pbtn, text='↧ เติมจากค่าปัจจุบัน (แก้แล้ววางกลับ)', command=self._row_paste_fill,
                  bg=C['bg_card'], fg=C['muted'], font=('Segoe UI', 9), relief='flat', padx=8,
                  cursor='hand2').pack(anchor='w', pady=(4, 0))

        # เรทสุ่ม (เฉพาะ RANDOM) — แก้ของตัวที่เลือก
        self._row_rate = tk.StringVar()
        self._rate_ctl = tk.Frame(parent, bg=C['bg_dark'])
        tk.Label(self._rate_ctl, text='เรทสุ่มของตัวที่เลือก:', bg=C['bg_dark'], fg=C['muted'],
                 font=('Segoe UI', 9)).pack(side='left', padx=(10, 4))
        tk.Entry(self._rate_ctl, textvariable=self._row_rate, width=10, bg=C['bg_inp'], fg=C['text'],
                 insertbackground=C['text'], font=FM, relief='flat').pack(side='left', ipady=2)
        tk.Button(self._rate_ctl, text='ตั้งเรทสุ่ม', command=self._row_apply, bg=C['bg_card'],
                  fg=C['text'], font=FM, relief='flat', padx=8, cursor='hand2').pack(side='left', padx=4)

        def _sync_rate_row(*_):
            if self.v_type.get().upper() == 'RANDOM':
                self.item_tree.configure(displaycolumns=('no', 'id', 'name', 'qty', 'tier', 'rate'))
                self._paste_rate_lbl.grid(row=0, column=2, sticky='w', padx=(8, 0))
                self._paste_rate.grid(row=1, column=2, sticky='w', padx=(8, 0))
                self._rate_ctl.pack(fill='x', pady=(0, 4))
            else:
                self.item_tree.configure(displaycolumns=('no', 'id', 'name', 'qty', 'tier'))
                self._paste_rate_lbl.grid_forget()
                self._paste_rate.grid_forget()
                self._rate_ctl.pack_forget()
        self.v_type.trace_add('write', _sync_rate_row)
        _sync_rate_row()

        tv.bind('<<TreeviewSelect>>', self._row_on_select)

    # ---- rows mode: ไอเทมรายตัว (ลำดับสำคัญ) ----
    def _row_name(self, it):
        return str(self.app._item_names.get(it.get('id', ''), it.get('name', '')) or '')

    def _selected_row_index(self):
        s = self.item_tree.selection()
        if not s:
            return None
        try:
            i = int(s[0])
        except (ValueError, TypeError):
            return None
        return i if 0 <= i < len(self._items) else None

    def _row_load_edit(self, idx):
        it = self._items[idx]
        self._row_qty.set(str(it.get('qty', '') or '1'))
        self._row_tier.set(str(it.get('tier', '') or DEFAULT_TIER))
        self._row_rate.set(str(it.get('rate', '') or ''))

    def _row_on_select(self, _evt=None):
        idx = self._selected_row_index()
        if idx is not None:
            self._row_load_edit(idx)

    def _row_commit_edit(self):
        """เขียนค่าในช่องแก้ (qty/tier/rate) กลับเข้าไอเทมที่เลือก — เรียกก่อน remove/move/to_bundle"""
        idx = self._selected_row_index()
        if idx is None:
            return
        it = self._items[idx]
        q = self._row_qty.get().strip()
        if q:
            it['qty'] = q
        t = self._row_tier.get().strip()
        if t:
            it['tier'] = t
        if self.v_type.get().upper() == 'RANDOM':
            r = self._row_rate.get().strip()
            if r:
                it['rate'] = r

    def _render_rows(self, keep=None):
        tv = self.item_tree
        if keep is None:
            keep = self._selected_row_index()
        tv.delete(*tv.get_children())
        for i, it in enumerate(self._items):
            tv.insert('', 'end', iid=str(i),
                      values=(i + 1, it.get('id', ''), self._row_name(it),
                              it.get('qty', '') or '1', it.get('tier', '') or DEFAULT_TIER,
                              it.get('rate', '') or ''))
        if self._items:
            j = min(keep if keep is not None else 0, len(self._items) - 1)
            tv.selection_set(str(j))
            tv.see(str(j))
            self._row_load_edit(j)
        else:
            self._row_qty.set('')
            self._row_tier.set('')
            self._row_rate.set('')

    def _row_apply(self):
        if self._selected_row_index() is None:
            return
        self._row_commit_edit()
        self._render_rows()

    def _row_remove(self):
        idx = self._selected_row_index()
        if idx is None:
            return
        del self._items[idx]
        self._render_rows(keep=idx)

    def _row_move(self, delta):
        idx = self._selected_row_index()
        if idx is None:
            return
        self._row_commit_edit()
        j = idx + delta
        if not (0 <= j < len(self._items)):
            return
        self._items[idx], self._items[j] = self._items[j], self._items[idx]
        self._render_rows(keep=j)

    def _row_paste_apply(self):
        """เอาค่าที่วางในกล่อง จำนวน/Tier ไปใส่ไอเทมตามลำดับบรรทัด (บน→ล่าง) — เว้นบรรทัด = ไม่แตะตัวนั้น"""
        self._row_commit_edit()
        qlines = self._paste_qty.get('1.0', 'end').splitlines()
        tlines = self._paste_tier.get('1.0', 'end').splitlines()
        rlines = self._paste_rate.get('1.0', 'end').splitlines()
        is_rand = self.v_type.get().upper() == 'RANDOM'
        tier_map = {t.lower(): t for t in TIERS}
        for i, it in enumerate(self._items):
            if i < len(qlines):
                m = re.search(r'\d+', qlines[i])
                if m:
                    it['qty'] = m.group()
            if i < len(tlines) and tlines[i].strip():
                it['tier'] = tier_map.get(tlines[i].strip().lower(), it.get('tier') or DEFAULT_TIER)
            if is_rand and i < len(rlines):        # เรทสุ่ม = ทศนิยม/จำนวนเต็ม (0.000-100.000)
                m = re.search(r'\d+(\.\d+)?', rlines[i])
                if m:
                    it['rate'] = m.group()
        self._render_rows()

    def _row_paste_fill(self):
        """เติมกล่องวางด้วยค่าปัจจุบันของไอเทม (จะได้แก้ทีละคอลัมน์แล้ววางกลับ)"""
        self._paste_qty.delete('1.0', 'end')
        self._paste_tier.delete('1.0', 'end')
        self._paste_rate.delete('1.0', 'end')
        if self._items:
            self._paste_qty.insert('1.0', '\n'.join(str(it.get('qty', '') or '1') for it in self._items))
            self._paste_tier.insert('1.0', '\n'.join(str(it.get('tier', '') or DEFAULT_TIER) for it in self._items))
            if self.v_type.get().upper() == 'RANDOM':
                self._paste_rate.insert('1.0', '\n'.join(str(it.get('rate', '') or '') for it in self._items))

    # ---- โหลด/ดึงข้อมูล ----
    @staticmethod
    def _set_col(t, lines):
        t.delete('1.0', 'end')
        if any(str(x).strip() for x in lines):
            t.insert('1.0', '\n'.join(str(x) for x in lines) + '\n')

    def load(self, b):
        """เติมฟอร์มจาก bundle dict (ชื่อ/type/deliver/items/rewards) — คงลำดับไอเทมตามที่รับมา"""
        self.v_name.set(b.get('name', '') or '')
        if b.get('type') in BUNDLE_TYPES:
            self.v_type.set(b['type'])
        self.v_dlv.set(bool(b.get('deliver')))
        if self.mode == 'rows':
            self._items = []
            for it in (b.get('items') or []):        # เก็บตามลำดับเป๊ะ (index = ลำดับบนเว็บ)
                iid = str(it.get('id', '') or '')
                row = {'id': iid, 'name': str(self.app._item_names.get(iid, it.get('name', '')) or ''),
                       'qty': str(it.get('qty', '') or '1'), 'tier': str(it.get('tier', '') or DEFAULT_TIER)}
                if str(it.get('rate', '') or '').strip():
                    row['rate'] = str(it.get('rate'))
                self._items.append(row)
            self._render_rows(keep=0)
        else:
            ids, names, qtys, tiers, rates = [], [], [], [], []
            for it in (b.get('items') or []):
                iid = str(it.get('id', '') or '')
                ids.append(iid)
                names.append(str(self.app._item_names.get(iid, it.get('name', '')) or ''))
                qtys.append(str(it.get('qty', '') or ''))
                tiers.append(str(it.get('tier', '') or ''))
                rates.append(str(it.get('rate', '') or ''))
            self._set_col(self.t_id, ids)
            self._set_col(self.t_name, names)
            self._set_col(self.t_qty, qtys)
            self._set_col(self.t_tier, tiers)
            self._set_col(self.t_rate, rates)
        self.rewards[:] = [dict(r) for r in (b.get('rewards') or [])]
        self._redraw()

    def set_prefill(self, ids, names):
        """เติม id/ชื่อไอเท็มล่วงหน้า (จาก Item Finder / handoff)"""
        self._set_col(self.t_id, [str(x) for x in (ids or [])])
        if names:
            self._set_col(self.t_name, [str(x) for x in names])
        for _id, _nm in zip(ids or [], names or []):
            _id, _nm = str(_id).strip(), str(_nm).strip()
            if _id and _nm:
                self.app._item_names[_id] = _nm

    def to_bundle(self, group=''):
        """อ่านค่าจากฟอร์ม -> bundle dict (ยังไม่กันชื่อซ้ำ — ให้ผู้เรียกจัดการ)
        rows mode: ไอเทมเรียงตามลำดับใน list เป๊ะ (index = ลำดับบนเว็บ)"""
        if self.mode == 'rows':
            self._row_commit_edit()          # เก็บค่าที่ค้างในช่องแก้ก่อน
            items = []
            for it in self._items:
                d = {'id': str(it.get('id', '') or ''), 'qty': str(it.get('qty', '') or '1'),
                     'tier': it.get('tier') or DEFAULT_TIER}
                if str(it.get('rate', '') or '').strip():
                    d['rate'] = str(it['rate'])
                items.append(d)
        else:
            items = self.app._parse_columns(self.t_id.get('1.0', 'end'),
                                            self.t_qty.get('1.0', 'end'),
                                            self.t_tier.get('1.0', 'end'),
                                            self.t_rate.get('1.0', 'end'))
        return {'name': self.v_name.get().strip(), 'type': self.v_type.get(),
                'deliver': bool(self.v_dlv.get()), 'items': items,
                'rewards': [dict(r) for r in self.rewards], 'group': group}

    def focus_name(self):
        try:
            self.e_name.focus_set()
        except Exception:
            pass


class App:
    def __init__(self, root, container=None, game_var=None, on_bundle_created=None, on_go_next=None):
        self.root = root
        # callback(target_key) ให้ launcher พาไปสร้าง Item Code / Event ต่อ (+ ดึงข้อมูลเข้าให้)
        self._on_go_next = on_go_next
        # container = frame ที่จะฝัง UI ลงไป (ถ้า None = สร้างลง root โดยตรง = โหมดรันเดี่ยว)
        self.container = container if container is not None else root
        self._embedded = container is not None    # เปิดผ่าน launcher -> ไม่มีแท็บ ① (เกม/Login อยู่แถบบน)
        # game_var = StringVar เกม/เซิร์ฟที่แชร์ร่วมกับ tool อื่น (เลือกที่ไหนก็เปลี่ยนพร้อมกัน)
        self._game_var = game_var
        self._shared_game = game_var is not None
        # callback(bid) เรียกเมื่อสร้าง bundle เสร็จ -> launcher ผูก id เข้าคิว Event (Item Code)
        self._on_bundle_created = on_bundle_created
        if container is None:
            self.root.title('Create bundle')
            self.root.configure(bg=C['bg_dark'])
            self.root.minsize(600, 660)
            _set_window_icon(self.root)
        self._busy = False
        self._cancel = False
        self._bundles = []
        self._results = []
        self._last_bundle_id = None
        self._current_group = ''         # กลุ่ม (Code) ของ bundle ที่กำลังสร้าง -> ส่ง id กลับให้ถูกกลุ่ม
        self._item_names = {}            # id -> ชื่อไอเทม (จาก Item Finder) โชว์กำกับ ID (ไม่ยุ่งกับการสร้าง)
        self._current_rewards = []
        self._current_rewards_name = None
        self._sel_bundle = None          # ชื่อบันเดิลในคิวที่กำลังโหลดอยู่ในฟอร์ม (ไว้เซฟกลับตอนสลับ)
        self.vgame = self._game_var if self._game_var is not None else tk.StringVar(value=GAME_NAMES[0])
        self.vname = tk.StringVar()
        self.vtype = tk.StringVar(value=DEFAULT_BUNDLE_TYPE)
        self.vdeliver = tk.BooleanVar(value=True)
        self.vsave = tk.BooleanVar(value=False)
        self.vbundle = tk.StringVar()
        self._build_ui()
        self._load_prefs()

    # ---------------- UI ----------------
    def _build_ui(self):
        hdr = tk.Frame(self.container, bg=C['bg_med'], height=48)
        hdr.pack(fill='x')
        tk.Label(hdr, text='Create bundle', bg=C['bg_med'], fg=C['text'],
                 font=('Segoe UI', 13, 'bold')).pack(side='left', padx=14, pady=10)
        tk.Label(hdr, text='combo-interactive', bg=C['bg_med'], fg=C['teal'],
                 font=FM).pack(side='left', pady=10)

        s = ttk.Style()
        s.theme_use('clam')
        s.configure('TCombobox', fieldbackground=C['bg_inp'], background=C['bg_card'],
                    foreground=C['text'], arrowcolor=C['text'],
                    selectbackground=C['bg_inp'], selectforeground=C['text'])
        s.map('TCombobox',
              fieldbackground=[('readonly', C['bg_inp']), ('disabled', C['bg_inp'])],
              foreground=[('readonly', C['text']), ('disabled', C['muted'])],
              selectbackground=[('readonly', C['bg_inp'])],
              selectforeground=[('readonly', C['text'])])
        self.root.option_add('*TCombobox*Listbox.background', C['bg_card'])
        self.root.option_add('*TCombobox*Listbox.foreground', C['text'])
        self.root.option_add('*TCombobox*Listbox.selectBackground', C['accent'])
        self.root.option_add('*TCombobox*Listbox.selectForeground', '#ffffff')
        s.configure('TNotebook', background=C['bg_dark'], borderwidth=0)
        s.configure('TNotebook.Tab', background=C['bg_card'], foreground=C['muted'], padding=[14, 6])
        s.map('TNotebook.Tab', background=[('selected', C['bg_med'])], foreground=[('selected', C['text'])])

        self.nb = ttk.Notebook(self.container)
        self.nb.pack(fill='both', expand=True, padx=8, pady=8)
        tab_main = tk.Frame(self.nb, bg=C['bg_dark'])
        tab_detail = tk.Frame(self.nb, bg=C['bg_dark'])
        tab_result = tk.Frame(self.nb, bg=C['bg_dark'])
        tab_log = tk.Frame(self.nb, bg=C['bg_dark'])
        self._tab_detail = tab_detail
        self._tab_result = tab_result
        if not self._embedded:
            # รันเดี่ยว -> มีแท็บ ① ตั้งค่า & Login
            self.nb.add(tab_main, text='  ① ตั้งค่า & Login  ')
            self.nb.add(tab_detail, text='  ② สร้าง Bundle  ')
        else:
            # เปิดผ่าน launcher -> เกม/Login อยู่แถบบน ข้ามแท็บ ① (widget ในนั้นสร้างไว้แต่ไม่โชว์)
            self.nb.add(tab_detail, text='  สร้าง Bundle  ')
        self.nb.add(tab_result, text='  เลข Bundle  ')
        self.nb.add(tab_log, text='  Log  ')

        # ----- tab ① ตั้งค่า & Login (โชว์เฉพาะรันเดี่ยว; embedded สร้างไว้แต่ไม่ add) -----
        body = tk.Frame(tab_main, bg=C['bg_dark'])
        body.pack(fill='both', expand=True, padx=10, pady=10)

        lf = tk.LabelFrame(body, text='เลือกเกม / เซิร์ฟเวอร์', bg=C['bg_med'],
                           fg=C['muted'], font=FM, bd=1, relief='solid')
        lf.pack(fill='x', pady=(0, 8))
        row = tk.Frame(lf, bg=C['bg_med'])
        row.pack(fill='x', padx=10, pady=8)
        tk.Label(row, text='เกม:', bg=C['bg_med'], fg=C['text'], font=FM).pack(side='left')
        self.game_cb = ttk.Combobox(row, textvariable=self.vgame, values=GAME_NAMES,
                                     state='readonly', width=18, font=FM)
        self.game_cb.pack(side='left', padx=8)
        self.game_cb.bind('<<ComboboxSelected>>', lambda e: self._save_prefs())
        self.url_lbl = tk.Label(lf, text='', bg=C['bg_med'], fg=C['muted'],
                                font=('Consolas', 9), anchor='w', justify='left')
        self.url_lbl.pack(fill='x', padx=10, pady=(0, 8))
        self.vgame.trace_add('write', lambda *a: self._refresh_url())

        lf2 = tk.LabelFrame(body, text='เข้าสู่ระบบ (เก็บค่า login)', bg=C['bg_med'],
                            fg=C['muted'], font=FM, bd=1, relief='solid')
        lf2.pack(fill='x', pady=(0, 8))
        btnrow = tk.Frame(lf2, bg=C['bg_med'])
        btnrow.pack(fill='x', padx=10, pady=8)
        tk.Button(btnrow, text='เปิดหน้า Login', command=self._open_login,
                  bg=C['accent'], fg='#fff', font=FB, relief='flat', padx=14, pady=5).pack(side='left')
        tk.Button(btnrow, text='ล้าง session', command=self._clear_profile,
                  bg=C['bg_card'], fg=C['muted'], font=FM, relief='flat', padx=10, pady=5).pack(side='left', padx=8)
        tk.Button(btnrow, text='ทดสอบเข้าหน้า Bundles', command=self._test_nav,
                  bg=C['warn'], fg='#000', font=FM, relief='flat', padx=10, pady=5).pack(side='left')

        # บอกทางไปแท็บถัดไป (ชื่อ/Type/ไอเทม/ปุ่มสร้าง อยู่แท็บ '② สร้าง Bundle' ทั้งหมด)
        hint = tk.LabelFrame(body, text='ขั้นตอนถัดไป', bg=C['bg_med'],
                             fg=C['muted'], font=FM, bd=1, relief='solid')
        hint.pack(fill='x', pady=(0, 8))
        tk.Label(hint,
                 text='เลือกเกม + กด "เปิดหน้า Login" (ทำครั้งเดียว จำ session ไว้)\n'
                      'จากนั้นไปแท็บ  ②  สร้าง Bundle  เพื่อกรอกชื่อ/ไอเทม แล้วกดสร้างได้เลย',
                 bg=C['bg_med'], fg=C['text'], font=FM, anchor='w', justify='left').pack(
                     fill='x', padx=12, pady=10)

        # ----- tab ② สร้าง Bundle (ชื่อ/Type/ไอเทม/สร้าง — จบในหน้าเดียว) -----
        # ข้อมูลหัว bundle (ชื่อ/Type/ส่งของ) — อยู่ที่นี่ที่เดียว (แท็บ ① ไม่มีช่องนี้แล้ว)
        meta = tk.LabelFrame(tab_detail, text='ข้อมูล Bundle (ชื่อ / ประเภท / ส่งของ)', bg=C['bg_med'],
                             fg=C['muted'], font=FM, bd=1, relief='solid')
        meta.pack(fill='x', padx=8, pady=(8, 2))
        mrow = tk.Frame(meta, bg=C['bg_med'])
        mrow.pack(fill='x', padx=10, pady=6)
        tk.Label(mrow, text='ชื่อบันเดิล:', bg=C['bg_med'], fg=C['text'], font=FM).pack(side='left')
        self.detail_name_entry = tk.Entry(mrow, textvariable=self.vname, bg=C['bg_inp'],
                                          fg=C['text'], insertbackground=C['text'], font=FM,
                                          relief='flat', width=22)
        self.detail_name_entry.pack(side='left', padx=(6, 12), ipady=2)
        tk.Label(mrow, text='Type:', bg=C['bg_med'], fg=C['text'], font=FM).pack(side='left')
        ttk.Combobox(mrow, textvariable=self.vtype, values=BUNDLE_TYPES,
                     state='readonly', width=14, font=FM).pack(side='left', padx=(6, 12))
        tk.Checkbutton(mrow, text='ส่งของทันที', variable=self.vdeliver,
                       bg=C['bg_med'], fg=C['text'], font=FM, selectcolor=C['bg_inp'],
                       activebackground=C['bg_med'], activeforeground=C['text']).pack(side='left')

        dtop = tk.Frame(tab_detail, bg=C['bg_dark'])
        dtop.pack(fill='x', padx=8, pady=(8, 2))
        tk.Label(dtop, text='ไฟล์ template', bg=C['bg_dark'], fg=C['text'], font=FB).pack(side='left')
        tk.Button(dtop, text='➕ เพิ่ม Bundle เข้าคิว', command=self.open_bundle_dialog,
                  bg=C['accent2'], fg='#fff', font=('Segoe UI', 9, 'bold'), relief='flat',
                  padx=10, pady=3).pack(side='left', padx=(10, 0))
        tk.Button(dtop, text='Import .xlsx/.csv', command=self._import_ids,
                  bg=C['bg_card'], fg=C['muted'], font=('Segoe UI', 9), relief='flat',
                  padx=8, pady=3).pack(side='right')
        tk.Button(dtop, text='ดาวน์โหลด template', command=self._save_template,
                  bg=C['bg_card'], fg=C['teal'], font=('Segoe UI', 9), relief='flat',
                  padx=8, pady=3).pack(side='right', padx=(0, 6))
        tk.Button(dtop, text='Fetch ตัวเลือก', command=self._fetch_options,
                  bg=C['bg_card'], fg=C['accent'], font=('Segoe UI', 9), relief='flat',
                  padx=8, pady=3).pack(side='right', padx=(0, 6))

        drow = tk.Frame(tab_detail, bg=C['bg_dark'])
        drow.pack(fill='x', padx=8, pady=2)
        tk.Label(drow, text='เลือก bundle จากไฟล์:', bg=C['bg_dark'], fg=C['muted'],
                 font=('Segoe UI', 9)).pack(side='left')
        tk.Button(drow, text='สร้างทุก bundle อัตโนมัติ', command=self._create_all,
                  bg=C['accent'], fg='#fff', font=('Segoe UI', 9, 'bold'), relief='flat',
                  padx=10, pady=3).pack(side='right')
        # ยืดเต็มแถว + รายการที่กางลงมากว้างตามชื่อจริง (ชื่อ bundle ยาว ๆ จะได้อ่านออก)
        self.bundle_picker = ttk.Combobox(drow, textvariable=self.vbundle, values=[],
                                          state='disabled', width=32, font=FM)
        self.bundle_picker.pack(side='left', padx=8, fill='x', expand=True)
        self.bundle_picker.bind('<<ComboboxSelected>>', lambda e: self._on_pick_bundle())
        # ลบบันเดิลที่เลือกออกจากคิว
        tk.Button(drow, text='🗑 ลบบันเดิลที่เลือก', command=self._del_selected_bundle,
                  bg=C['bg_card'], fg=C['danger'], font=('Segoe UI', 9), relief='flat',
                  padx=10, pady=3, cursor='hand2').pack(side='left', padx=(0, 6))

        tk.Label(tab_detail,
                 text='รายการไอเท็ม — 1 บรรทัด = 1 ไอเท็ม:  ID  จำนวน   (คัดลอกจาก Excel มาวางได้)  •  '
                      'ช่อง "ชื่อไอเท็ม" กำกับตาม ID อัตโนมัติ  •  ช่อง "Tier" เรียงตามบรรทัด  •  '
                      '"เรทสุ่ม" โผล่เมื่อ Type=RANDOM:',
                 bg=C['bg_dark'], fg=C['text'], font=FM, anchor='w', justify='left'
                 ).pack(fill='x', padx=8, pady=(6, 0))
        # หัวคอลัมน์ (จัดให้ตรงกับช่องด้านล่าง)
        hcol = tk.Frame(tab_detail, bg=C['bg_dark'])
        hcol.pack(fill='x', padx=8)
        tk.Label(hcol, text='ไอเท็ม (ID  จำนวน)', bg=C['bg_dark'], fg=C['muted'],
                 font=('Segoe UI', 8), width=15, anchor='w').pack(side='left')
        tk.Label(hcol, text='ชื่อไอเท็ม (อัตโนมัติ)', bg=C['bg_dark'], fg=C['muted'],
                 font=('Segoe UI', 8), anchor='w').pack(side='left', fill='x', expand=True)
        self._tier_hdr = tk.Label(hcol, text='Tier', bg=C['bg_dark'], fg=C['teal'],
                                  font=('Segoe UI', 8, 'bold'), width=10, anchor='w')
        self._tier_hdr.pack(side='left')
        self._rate_hdr = tk.Label(hcol, text='เรทสุ่ม', bg=C['bg_dark'], fg=C['warn'],
                                  font=('Segoe UI', 8, 'bold'), width=11, anchor='w')

        itembox = tk.Frame(tab_detail, bg=C['bg_dark'])
        itembox.pack(fill='x', padx=8, pady=(0, 2))
        self.ids_text = tk.Text(itembox, height=8, width=15, bg=C['bg_inp'], fg=C['text'],
                                insertbackground=C['text'], font=('Consolas', 10),
                                relief='flat', wrap='none')
        self.ids_text.pack(side='left', fill='y')
        # ช่องชื่อไอเท็ม (อ่านอย่างเดียว) — กำกับชื่อตาม ID ในแต่ละบรรทัดให้อัตโนมัติ
        # ช่องกว้าง + ขยายเต็มพื้นที่ที่เหลือ ชื่อยาว ๆ จะได้อ่านครบ
        self._name_text = tk.Text(itembox, height=8, width=28, bg=C['bg_med'], fg=C['muted'],
                                  font=('Segoe UI', 9), relief='flat', wrap='none',
                                  state='disabled', cursor='arrow')
        self._name_text.pack(side='left', fill='both', expand=True, padx=(6, 0))
        # ช่อง Tier (โชว์ตลอด — ใช้ได้ทุก Type) เรียงบรรทัดตรงกับรายการไอเทม
        self._tier_text = tk.Text(itembox, height=8, width=10, bg=C['bg_inp'], fg=C['teal'],
                                  insertbackground=C['text'], font=('Consolas', 10),
                                  relief='flat', wrap='none')
        self._tier_text.pack(side='left', fill='y', padx=(6, 0))
        # ช่องเรทสุ่ม (โผล่เมื่อ Type=RANDOM) เรียงบรรทัดตรงกับรายการไอเทม
        self._rate_text = tk.Text(itembox, height=8, width=11, bg=C['bg_inp'], fg=C['warn'],
                                  insertbackground=C['text'], font=('Consolas', 10),
                                  relief='flat', wrap='none')
        # เลื่อนลูกกลิ้ง -> ทุกช่องเลื่อนพร้อมกัน แถวจะตรงกันเสมอ
        def _items_wheel(e):
            step = -1 if e.delta > 0 else 1
            for w in (self.ids_text, self._name_text, self._tier_text, self._rate_text):
                w.yview_scroll(step, 'units')
            return 'break'
        for w in (self.ids_text, self._name_text, self._tier_text, self._rate_text):
            w.bind('<MouseWheel>', _items_wheel)
        # พิมพ์/วาง/ลบ ID -> อัปเดตชื่อกำกับ (debounce กันหน่วงตอนวางเยอะ)
        self.ids_text.bind('<KeyRelease>', lambda e: self._schedule_name_sync())
        self.vtype.trace_add('write', self._sync_detail_rate)
        self._sync_detail_rate()
        self.reward_lbl = tk.Label(tab_detail, text='', bg=C['bg_dark'], fg=C['teal'],
                                   font=('Segoe UI', 9), anchor='w', justify='left', wraplength=580)
        self.reward_lbl.pack(fill='x', padx=8, pady=(0, 4))

        r4 = tk.Frame(tab_detail, bg=C['bg_dark'])
        r4.pack(fill='x', padx=8, pady=(2, 6))
        self.btn_create = tk.Button(r4, text='เริ่ม (bundle ที่เลือก)', command=self._create_bundle_fields,
                                    bg=C['accent2'], fg='#fff', font=FB, relief='flat', padx=14, pady=5)
        self.btn_create.pack(side='left')
        self.btn_stop = tk.Button(r4, text='หยุด', command=self._cancel_run,
                                  bg=C['danger'], fg='#fff', font=FB, relief='flat',
                                  padx=16, pady=5, state='disabled')
        self.btn_stop.pack(side='left', padx=8)
        tk.Checkbutton(r4, text='กดบันทึก (Save) หลังกรอกเสร็จ', variable=self.vsave,
                       bg=C['bg_dark'], fg=C['warn'], font=FM, selectcolor=C['bg_inp'],
                       activebackground=C['bg_dark'], activeforeground=C['warn']).pack(side='left', padx=10)

        btop = tk.Frame(tab_detail, bg=C['bg_dark'])
        btop.pack(fill='x', padx=8, pady=(4, 0))
        tk.Label(btop, text='สรุปทั้งหมดที่จะใส่ในบันเดิลนี้ (ไอเท็ม + reward):',
                 bg=C['bg_dark'], fg=C['text'], font=FB).pack(side='left')
        tk.Button(btop, text='รีเฟรช', command=self._refresh_detail,
                  bg=C['bg_card'], fg=C['muted'], font=('Segoe UI', 9), relief='flat',
                  padx=10, pady=3).pack(side='right')
        dwrap = tk.Frame(tab_detail, bg=C['bg_dark'])
        dwrap.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        dvsb = tk.Scrollbar(dwrap)
        dvsb.pack(side='right', fill='y')
        self.detail_list = tk.Listbox(dwrap, bg=C['bg_dark'], fg=C['text'],
                                      font=('Consolas', 10), relief='flat',
                                      activestyle='none', yscrollcommand=dvsb.set)
        self.detail_list.pack(side='left', fill='both', expand=True)
        dvsb.config(command=self.detail_list.yview)

        # ----- ขั้นต่อไป: ไปสร้าง Item Code / Event (เฉพาะเปิดผ่าน launcher) -----
        if self._on_go_next is not None:
            nxt = tk.Frame(tab_detail, bg=C['bg_med'])
            nxt.pack(fill='x', padx=8, pady=(0, 8))
            tk.Label(nxt, text='ขั้นต่อไป (หลังสร้าง bundle):', bg=C['bg_med'], fg=C['text'],
                     font=FM).pack(side='left', padx=(10, 8), pady=6)
            tk.Button(nxt, text='🎟️ ไปสร้าง Item Code', command=lambda: self._on_go_next('itemcode'),
                      bg=C['warn'], fg='#000', font=('Segoe UI', 9, 'bold'), relief='flat',
                      padx=10, pady=3, cursor='hand2').pack(side='left', padx=(0, 6), pady=6)
            tk.Button(nxt, text='🎉 ไปสร้าง Event', command=lambda: self._on_go_next('event'),
                      bg=C['danger'], fg='#fff', font=('Segoe UI', 9, 'bold'), relief='flat',
                      padx=10, pady=3, cursor='hand2').pack(side='left', pady=6)

        # ----- tab เลข Bundle -----
        rtop = tk.Frame(tab_result, bg=C['bg_dark'])
        rtop.pack(fill='x', padx=8, pady=(8, 0))
        tk.Label(rtop, text='เลข Bundle ที่สร้าง (+ ผลรีเชค)', bg=C['bg_dark'],
                 fg=C['text'], font=FB).pack(side='left')
        tk.Button(rtop, text='บันทึก .csv', command=self._export_results,
                  bg=C['bg_card'], fg=C['teal'], font=('Segoe UI', 9), relief='flat',
                  padx=10, pady=3).pack(side='right')
        tk.Button(rtop, text='คัดลอกเลข', command=self._copy_ids,
                  bg=C['accent'], fg='#fff', font=('Segoe UI', 9), relief='flat',
                  padx=10, pady=3).pack(side='right', padx=6)
        tk.Button(rtop, text='ล้าง', command=self._clear_results,
                  bg=C['bg_card'], fg=C['muted'], font=('Segoe UI', 9), relief='flat',
                  padx=10, pady=3).pack(side='right', padx=6)
        tk.Label(tab_result, text='เลือกแถวที่ต้องการแล้วกด "คัดลอกเลข" (ไม่เลือก = คัดลอกทั้งหมด)',
                 bg=C['bg_dark'], fg=C['muted'], font=('Segoe UI', 9)).pack(fill='x', padx=8, pady=(0, 2))
        lwrap = tk.Frame(tab_result, bg=C['bg_dark'])
        lwrap.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        vsb = tk.Scrollbar(lwrap)
        vsb.pack(side='right', fill='y')
        hsb = tk.Scrollbar(lwrap, orient='horizontal')     # ชื่อยาว ๆ เลื่อนดูจนจบได้
        hsb.pack(side='bottom', fill='x')
        self.result_list = tk.Listbox(lwrap, bg=C['bg_dark'], fg=C['text'],
                                      font=('Consolas', 10), relief='flat',
                                      selectmode='extended', selectbackground=C['accent'],
                                      selectforeground='#ffffff', activestyle='none',
                                      yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.result_list.pack(side='left', fill='both', expand=True)
        vsb.config(command=self.result_list.yview)
        hsb.config(command=self.result_list.xview)
        self.result_list.bind('<Control-c>', lambda e: self._copy_ids())

        # ----- tab Log -----
        logtop = tk.Frame(tab_log, bg=C['bg_dark'])
        logtop.pack(fill='x', padx=8, pady=(8, 0))
        tk.Label(logtop, text='Log', bg=C['bg_dark'], fg=C['text'], font=FB).pack(side='left')
        tk.Button(logtop, text='ล้าง Log', command=self._clear_log,
                  bg=C['bg_card'], fg=C['muted'], font=('Segoe UI', 9), relief='flat',
                  padx=10, pady=3).pack(side='right')
        self.log_area = tk.Text(tab_log, bg=C['bg_dark'], fg=C['text'],
                                font=('Consolas', 10), relief='flat', state='disabled', wrap='word')
        self.log_area.pack(fill='both', expand=True, padx=8, pady=8)
        for lvl, col in (('SUCCESS', C['accent2']), ('WARNING', C['warn']),
                         ('ERROR', C['danger']), ('INFO', C['text']), ('STEP', C['accent'])):
            self.log_area.tag_config(lvl, foreground=col)

        self._refresh_url()
        self._refresh_detail()
        if not PW_OK:
            self.log('ไม่พบ playwright - รัน: pip install playwright แล้ว playwright install chromium', 'WARNING')

    def _refresh_url(self):
        self.url_lbl.config(text=game_url(self.vgame.get(), 'bundles'))

    def _load_prefs(self):
        if self._shared_game:
            return
        p = load_prefs()
        if p.get('game') in GAME_BASES:
            self.vgame.set(p['game'])

    def _save_prefs(self):
        save_prefs({'game': self.vgame.get()})

    def log(self, msg, lvl='INFO'):
        self.root.after(0, self._log_main, msg, lvl)

    def _log_main(self, msg, lvl):
        self.log_area.config(state='normal')
        self.log_area.insert('end', msg + '\n', lvl)
        self.log_area.see('end')
        self.log_area.config(state='disabled')

    def _set_busy(self, busy):
        self._busy = busy
        st = 'disabled' if busy else 'normal'
        sst = 'normal' if busy else 'disabled'
        self.root.after(0, lambda: (self.btn_create.config(state=st), self.btn_stop.config(state=sst)))

    def _cancel_run(self):
        self._cancel = True
        self.log('>> กำลังหยุด... (จะหยุดเมื่อจบสเต็ปปัจจุบัน)', 'WARNING')

    def _clear_log(self):
        self.log_area.config(state='normal')
        self.log_area.delete('1.0', tk.END)
        self.log_area.config(state='disabled')

    async def _idle_until_closed(self, browser, page):
        """ปิด browser ให้เรียบร้อยเสมอ — ถ้าปล่อยให้ Chrome โดนฆ่าตอน teardown
        cookie/session จะยังไม่ถูกเขียนลง profile แล้วรอบหน้าจะกลายเป็น 'login หลุด'"""
        try:
            while not self._cancel:
                if not browser.pages:
                    break
                await page.wait_for_timeout(500)
        except Exception:
            pass
        finally:
            try:
                await browser.close()      # flush cookie/session ลง profile
            except Exception:
                pass

    @staticmethod
    def set_combo_values(cb, values, state='readonly', maxw=110):
        """ใส่ค่าลง combobox แล้วขยาย 'รายการที่กางลงมา' ให้กว้างพอเห็นข้อความเต็ม
        (ช่องด้านหน้าจะแคบกว่าก็ได้ ไม่ต้องยืดหน้าต่าง) — ชื่อ bundle/slug ยาวจะได้ไม่โดนตัด"""
        values = [str(v) for v in values]
        cb.config(values=values, state=state)
        longest = max((len(v) for v in values), default=0)
        allow_wide_popdown(cb)        # ปลดล็อกก่อน ไม่งั้น Tk บังคับกว้างเท่าช่องเสมอ
        try:
            w = max(int(cb.cget('width') or 0), min(longest + 2, maxw))
            pop = cb.tk.call('ttk::combobox::PopdownWindow', cb)
            cb.tk.call('%s.f.l' % pop, 'configure', '-width', w)
        except Exception:
            pass          # ธีม/แพลตฟอร์มอื่นไม่มี popdown ภายใน -> ปล่อยตามเดิม

    def _launch_kwargs(self):
        return core.launch_kwargs()

    # ---------------- รายการ ----------------
    @staticmethod
    def _dupes(ids):
        seen, dup = set(), []
        for i in ids:
            if i in seen and i not in dup:
                dup.append(i)
            seen.add(i)
        return dup

    @staticmethod
    def _parse_item_line(line):
        """แยก 1 บรรทัดเป็น (id, qty, tier, rate)
        รองรับ: เว้นวรรค / คอมมา / TAB (คัดลอกจาก Excel) และมีคอลัมน์ชื่อปนได้
        - จำนวน = เลขจำนวนเต็มตัวถัดจาก ID
        - เรทสุ่ม (rate) = เลขทศนิยม เช่น 12.500 (เฉพาะ bundle RANDOM)
        คืน None ถ้าไม่พบเลข ID ในบรรทัด"""
        tier_map = {t.lower(): t for t in TIERS}
        if '\t' in line:
            cols = [c.strip() for c in line.split('\t')]
        else:
            cols = [c for c in re.split(r'[\s,]+', line.strip()) if c]
        cols = [c for c in cols if c != '']
        # ID = เลขล้วนตัวแรก
        iid, id_pos = None, -1
        for i, c in enumerate(cols):
            if c.isdigit():
                iid, id_pos = c, i
                break
        if iid is None:
            return None
        # จำนวน = จำนวนเต็มตัวแรกถัดจาก ID | เรทสุ่ม = ทศนิยมตัวแรก (มีจุด)
        qty, rate = None, ''
        for c in cols[id_pos + 1:]:
            if c.lower() in tier_map:
                continue
            if re.fullmatch(r'\d+\.\d+', c):        # ทศนิยม = เรทสุ่ม
                if not rate:
                    rate = c
                continue
            if qty is None:
                m = re.search(r'\d+', c)
                if m:
                    qty = m.group()
        # Tier = token ที่ตรงชื่อ tier (Common/Rare/Epic/Mystic/Legend) ถ้ามีในบรรทัด (ไม่มี = None -> ใช้ช่อง Tier)
        tier = None
        for c in cols:
            if c.lower() in tier_map:
                tier = tier_map[c.lower()]
                break
        return iid, (qty or '1'), tier, rate

    @staticmethod
    def _match_tier(text):
        """แปลงข้อความในช่อง Tier เป็นชื่อ tier มาตรฐาน (รองรับพิมพ์ย่อ เช่น 'ep' -> Epic), None ถ้าไม่ตรง"""
        t = (text or '').strip().lower()
        if not t:
            return None
        tier_map = {x.lower(): x for x in TIERS}
        if t in tier_map:
            return tier_map[t]
        for k, v in tier_map.items():        # พิมพ์ย่อ = prefix ของชื่อ tier
            if k.startswith(t):
                return v
        return None

    def _raw_ids(self):
        out = []
        for line in self.ids_text.get('1.0', 'end').splitlines():
            if not line.strip():
                continue
            parsed = self._parse_item_line(line)
            if parsed:
                out.append(parsed[0])
        return out

    def _sync_detail_rate(self, *_):
        """โชว์/ซ่อนช่องเรทสุ่ม (+หัวคอลัมน์) บนแท็บหลักตาม Type (RANDOM = โชว์)"""
        rt = getattr(self, '_rate_text', None)
        if rt is None:
            return
        hdr = getattr(self, '_rate_hdr', None)
        if self.vtype.get().upper() == 'RANDOM':
            rt.pack(side='left', fill='y', padx=(6, 0))
            if hdr is not None:
                hdr.pack(side='left')
        else:
            rt.pack_forget()
            if hdr is not None:
                hdr.pack_forget()

    def _schedule_name_sync(self):
        # debounce: พิมพ์/วางรัว ๆ ไม่รีเฟรชทุกคีย์ (ตั้งใหม่ 150ms)
        job = getattr(self, '_name_sync_job', None)
        if job:
            try:
                self.root.after_cancel(job)
            except Exception:
                pass
        self._name_sync_job = self.root.after(150, self._sync_name_col)

    def _sync_name_col(self):
        """เติมช่อง "ชื่อไอเท็ม" (อ่านอย่างเดียว) ให้ตรงทุกบรรทัดของช่อง ID
        ชื่อดึงจาก self._item_names ตาม ID จึงเกาะไปกับไอเทมเสมอ แม้จะสลับ/แก้บรรทัด"""
        self._name_sync_job = None
        nt = getattr(self, '_name_text', None)
        if nt is None:
            return
        names = []
        for line in self.ids_text.get('1.0', 'end').splitlines():
            parsed = self._parse_item_line(line) if line.strip() else None
            iid = parsed[0] if parsed else ''
            nm = self._item_names.get(iid, '') if iid else ''
            names.append((nm[:39] + '…') if len(nm) > 40 else nm)
        nt.config(state='normal')
        nt.delete('1.0', 'end')
        if names:
            nt.insert('1.0', '\n'.join(names) + '\n')
        nt.config(state='disabled')
        try:
            nt.yview_moveto(self.ids_text.yview()[0])   # เลื่อนแนวตั้งให้ตรงกับช่อง ID
        except Exception:
            pass

    def _get_items(self):
        id_lines = self.ids_text.get('1.0', 'end').splitlines()
        rt = getattr(self, '_rate_text', None)
        rate_lines = rt.get('1.0', 'end').splitlines() if rt is not None else []
        tt = getattr(self, '_tier_text', None)
        tier_lines = tt.get('1.0', 'end').splitlines() if tt is not None else []
        seen, out = set(), []
        for i, line in enumerate(id_lines):
            if not line.strip():
                continue
            parsed = self._parse_item_line(line)
            if not parsed:
                continue
            iid, qty, tier, rate = parsed
            if iid in seen:
                continue
            seen.add(iid)
            # Tier: ในบรรทัดไอเทม (inline) -> ช่อง "Tier" (เรียงตามบรรทัด) -> ค่า default
            if not tier and i < len(tier_lines):
                tier = self._match_tier(tier_lines[i])
            tier = tier or DEFAULT_TIER
            # เติมเรทจากช่อง "เรทสุ่ม" (เรียงตามบรรทัด) ถ้าในบรรทัดไอเทมยังไม่มี
            if not rate and i < len(rate_lines):
                rm = re.search(r'\d+(\.\d+)?', rate_lines[i])
                if rm:
                    rate = rm.group()
            item = {'id': iid, 'qty': qty, 'tier': tier}
            if rate:
                item['rate'] = rate
            out.append(item)
        return out

    def _refresh_detail(self):
        if not hasattr(self, 'detail_list'):
            return
        self.detail_list.delete(0, tk.END)
        for it in self._get_items():
            nm = self._item_names.get(it['id'], '')
            nm_disp = ('%-24s' % ((nm[:23] + '…') if len(nm) > 24 else nm)) if nm else '%-24s' % '-'
            line = 'ITEM | %-7s %s | x%s | %s' % (it['id'], nm_disp, it['qty'], it['tier'])
            if it.get('rate'):
                line += ' | เรท %s' % it['rate']
            self.detail_list.insert('end', line)
            self.detail_list.itemconfig(self.detail_list.size() - 1, foreground=C['text'])
        rewards = (self._current_rewards
                   if self.vname.get().strip() == self._current_rewards_name else [])
        for r in rewards:
            self.detail_list.insert('end', '%-12s | %s | x%s' % (r['type'], r['value'], r['qty']))
            self.detail_list.itemconfig(self.detail_list.size() - 1, foreground=C['teal'])
        if self.detail_list.size() == 0:
            self.detail_list.insert('end', '(ยังไม่มีรายการ - import หรือพิมพ์ในช่องด้านบน)')

    def add_external_items(self, items):
        """รับไอเทมจากเครื่องมืออื่น (เช่น Item Finder) มาเติมต่อท้ายช่องรายการ
        items: list ของ dict ที่มี key 'id' หรือ 'aztek_id' (optional 'qty', 'tier')
        ตัด id ที่ซ้ำกับรายการที่มีอยู่แล้วออก คืนจำนวนที่เพิ่มจริง"""
        existing = {it['id'] for it in self._get_items()}
        lines = []
        for it in items:
            iid = str(it.get('id') or it.get('aztek_id') or '').strip()
            if not iid.isdigit() or iid in existing:
                continue
            existing.add(iid)
            qty = str(it.get('qty') or '1').strip() or '1'
            tier = it.get('tier') or DEFAULT_TIER
            if tier not in TIERS:
                tier = DEFAULT_TIER
            # tier != ค่า default -> ใส่ inline (ตรวจจับได้); ค่า default -> ปล่อยว่าง (ช่อง Tier = Common)
            lines.append('%s %s %s' % (iid, qty, tier) if tier != DEFAULT_TIER else '%s %s' % (iid, qty))
        added = len(lines)
        if lines:
            cur = self.ids_text.get('1.0', 'end').rstrip('\n')
            block = '\n'.join(lines)
            newtext = (cur + '\n' + block if cur.strip() else block) + '\n'
            self.ids_text.delete('1.0', 'end')
            self.ids_text.insert('1.0', newtext)
            self._refresh_detail()
            try:
                self.nb.select(self._tab_detail)
                if not self.vname.get().strip():
                    self.detail_name_entry.focus_set()
            except Exception:
                pass
        skipped = len(items) - added
        msg = 'รับจาก Item Finder: เพิ่ม %d ไอเทม' % added
        if skipped:
            msg += ' (ข้ามซ้ำ/ไม่ถูกต้อง %d)' % skipped
        self.log(msg, 'SUCCESS' if added else 'WARNING')
        return added

    # ---------------- popup: เพิ่ม bundle เข้าคิว ----------------
    @staticmethod
    def _parse_columns(id_block, qty_block, tier_block, rate_block=''):
        """แปลงคอลัมน์ (ID / qty / tier / เรทสุ่ม) ที่วางจาก Excel เป็น list ของ item
        จับคู่ตามลำดับบรรทัด (แถวที่ i ของทุกช่อง = ไอเท็มเดียวกัน)
        ตัด id ซ้ำออก คืน list[{'id','qty','tier'[,'rate']}]"""
        tier_map = {t.lower(): t for t in TIERS}
        id_lines = id_block.splitlines()
        qty_lines = qty_block.splitlines()
        tier_lines = tier_block.splitlines()
        rate_lines = (rate_block or '').splitlines()
        n = max(len(id_lines), len(qty_lines), len(tier_lines), len(rate_lines))
        seen, items = set(), []
        for i in range(n):
            idc = id_lines[i] if i < len(id_lines) else ''
            m = re.search(r'\d+', idc)
            if not m:
                continue
            iid = m.group()
            if iid in seen:
                continue
            seen.add(iid)
            qraw = qty_lines[i] if i < len(qty_lines) else ''
            qm = re.search(r'\d+', qraw)
            qty = qm.group() if qm else '1'
            traw = (tier_lines[i] if i < len(tier_lines) else '').strip()
            tier = tier_map.get(traw.lower(), DEFAULT_TIER)
            item = {'id': iid, 'qty': qty, 'tier': tier}
            rraw = rate_lines[i] if i < len(rate_lines) else ''
            rm = re.search(r'\d+(\.\d+)?', rraw)
            if rm:
                item['rate'] = rm.group()
            items.append(item)
        return items

    def open_bundle_dialog(self, prefill_ids=None, group='', prefill_names=None,
                           name='', on_close=None, seq=None):
        """เปิดหน้าต่างแยกสำหรับกรอกข้อมูล 1 bundle แล้วเก็บเข้าคิว (self._bundles)
        prefill_ids: list ของ id เติมในคอลัมน์ Item ID ล่วงหน้า (เช่น จาก Item Finder)
        prefill_names: list ชื่อไอเทม (เรียงตรงกับ prefill_ids) โชว์ในคอลัมน์ "ชื่อไอเท็ม"
        group: กลุ่ม (Code) ต้นทางจาก handoff — ผูกกับ bundle นี้เพื่อส่ง id กลับให้ถูกกลุ่ม
        name: ชื่อบันเดิลตั้งต้น (เช่น ชื่อกลุ่มจาก Item Finder) แก้ได้
        on_close(reason): เรียกตอนปิดฟอร์ม reason='saved'/'cancel'/'close' (ไว้ต่อคิวฟอร์มถัดไป)
        seq: (i, n) โชว์ความคืบหน้าบนหัวฟอร์มตอนเปิดทีละกลุ่ม"""
        win = tk.Toplevel(self.root)
        win.title('เพิ่ม Bundle เข้าคิว' + (' — กลุ่ม %d/%d' % seq if seq else ''))
        win.configure(bg=C['bg_dark'])
        win.geometry('760x600')
        win.minsize(680, 520)
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass

        _done = {'v': False}

        def _finish(reason):
            if _done['v']:
                return
            _done['v'] = True
            try:
                win.destroy()
            except Exception:
                pass
            if on_close:
                on_close(reason)
        win.protocol('WM_DELETE_WINDOW', lambda: _finish('close'))

        # ไส้ในฟอร์ม (หัว/คอลัมน์ไอเทม/รางวัล) reuse จาก _BundleEditor
        ed = _BundleEditor(self, win)
        if name:
            ed.v_name.set(name)                       # ตั้งชื่อจากกลุ่ม (Item Finder) มาให้ แก้ได้
        if prefill_ids:
            ed.set_prefill(prefill_ids, prefill_names or [])

        # --- ปุ่ม ---
        bar = tk.Frame(win, bg=C['bg_dark'])
        bar.pack(fill='x', padx=10, pady=(0, 10))
        tk.Label(bar, text='ในคิวตอนนี้: %d bundle' % len(self._bundles),
                 bg=C['bg_dark'], fg=C['teal'], font=('Segoe UI', 9)).pack(side='right')

        def _save():
            name_ = ed.v_name.get().strip()
            if not name_:
                messagebox.showwarning('กรอกชื่อก่อน', 'กรุณาใส่ชื่อบันเดิล', parent=win)
                ed.focus_name()
                return
            bundle = ed.to_bundle(group=group)
            items, rewards = bundle['items'], bundle['rewards']
            if not items and not rewards and not messagebox.askyesno(
                    'ไม่มีไอเท็ม', 'ยังไม่มีไอเท็ม/reward ในบันเดิลนี้ จะเก็บเข้าคิวเลยไหม?', parent=win):
                return
            if bundle['type'].upper() == 'RANDOM':
                missing = [it['id'] for it in items if not it.get('rate')]
                if missing and not messagebox.askyesno(
                        'เรทสุ่มไม่ครบ',
                        'RANDOM แต่ไอเท็มบางตัวยังไม่มีเรทสุ่ม:\n%s\nจะเก็บเข้าคิวเลยไหม?'
                        % ', '.join(missing[:15]), parent=win):
                    return
            existing = {b['name'] for b in self._bundles}
            uniq, k = name_, 2
            while uniq in existing:
                uniq = '%s (%d)' % (name_, k)
                k += 1
            bundle['name'] = uniq
            self._bundles.append(bundle)
            self.set_combo_values(self.bundle_picker, [b['name'] for b in self._bundles])
            self.vbundle.set(uniq)
            self._load_bundle(bundle)
            self.log('เก็บ Bundle "%s" เข้าคิว (%d ไอเท็ม%s) — คิวรวม %d bundle'
                     % (uniq, len(items),
                        (' + %d reward' % len(rewards)) if rewards else '',
                        len(self._bundles)), 'SUCCESS')
            _finish('saved')

        save_txt = 'เก็บเข้าคิว → กลุ่มถัดไป' if seq and seq[0] < seq[1] else 'เก็บเข้าคิว'
        tk.Button(bar, text=save_txt, command=_save, bg=C['accent2'], fg='#fff',
                  font=FB, relief='flat', padx=16, pady=6, cursor='hand2').pack(side='left')
        tk.Button(bar, text='ข้ามกลุ่มนี้' if seq else 'ยกเลิก', command=lambda: _finish('cancel'),
                  bg=C['bg_card'], fg=C['muted'], font=FM, relief='flat',
                  padx=14, pady=6, cursor='hand2').pack(side='left', padx=8)

        # handle สำหรับเรียกใช้/ทดสอบภายนอก (ชี้ไปที่ _BundleEditor)
        win.name_var, win.type_var = ed.v_name, ed.v_type
        win.id_text, win.qty_text, win.tier_text, win.rate_text = ed.t_id, ed.t_qty, ed.t_tier, ed.t_rate
        win.item_name_text = ed.t_name
        win.save = _save
        win.rw_kind_var, win.rw_value_var, win.rw_qty_var = ed.v_kind, ed.v_val, ed.v_qty
        win.rw_add, win.rw_del, win.rw_list, win.rw_box, win.rw_refresh = \
            ed._add_rw, ed._del_rw, ed.rewards, ed.lb_rw, ed._sync_vals
        win.finish = _finish
        win.editor = ed

        ed.focus_name()
        return win

    def queue_bundles_from_finder(self, bundles):
        """รับบันเดิลที่ Item Finder รวมมาให้ (รายกลุ่ม) แล้วเก็บเข้าคิวโดยตรง — ไม่เปิดฟอร์มทีละอัน
        bundles: list ของ {'name','group','items':[{'id','name'}]}
          · qty=1, tier=ค่าเริ่มต้น, Type/ส่งของทันที = ตามที่เลือกอยู่บนหน้า Create Bundle
          · ชื่อชนกันในคิว -> เติม (2)(3)… ให้ไม่ซ้ำ (กติกาเดียวกับ _save)
        คืนจำนวนบันเดิลที่เพิ่มจริง"""
        added, last = 0, None
        for b in bundles:
            items, seen = [], set()
            for it in (b.get('items') or []):
                m = re.search(r'\d+', str(it.get('id', '')))
                if not m or m.group() in seen:
                    continue
                iid = m.group(); seen.add(iid)
                items.append({'id': iid, 'qty': '1', 'tier': DEFAULT_TIER})
                nm = str(it.get('name', '')).strip()
                if nm:
                    self._item_names[iid] = nm       # ชื่อเกาะไปกับ ID (โชว์กำกับในคิว/สรุป)
            if not items:
                continue
            name = (b.get('name') or '').strip() or 'Bundle'
            existing = {x['name'] for x in self._bundles}
            uniq, k = name, 2
            while uniq in existing:
                uniq = '%s (%d)' % (name, k); k += 1
            last = {'name': uniq, 'type': self.vtype.get() or DEFAULT_BUNDLE_TYPE,
                    'deliver': bool(self.vdeliver.get()), 'items': items,
                    'rewards': [], 'group': b.get('group', '') or ''}
            self._bundles.append(last); added += 1
        if added:
            self.set_combo_values(self.bundle_picker, [x['name'] for x in self._bundles])
            self.vbundle.set(last['name'])
            self._load_bundle(last)
            self.log('รับ %d บันเดิลจาก Item Finder เข้าคิว — คิวรวม %d bundle'
                     % (added, len(self._bundles)), 'SUCCESS')
        return added

    def open_bundle_review(self, bundles, on_done=None):
        """หน้ารีวิวรวมบันเดิล (master-detail): ลิสต์ซ้าย (เลือก/สลับใช้-ไม่ใช้/ลบ) + ฟอร์มขวา (แก้เต็ม)
        + ปุ่มเดียว 'ส่งเข้าคิวทั้งหมด' -> เก็บทุกบันเดิลที่ติ๊กไว้เข้าคิว self._bundles รวดเดียว
        bundles: list ของ {'name','group','items':[{'id','name'}]} (จาก Item Finder)
        on_done(added): เรียกหลังส่งเข้าคิว (ไว้ให้ launcher โฟกัสหน้าคิว)"""
        # เตรียมข้อมูลเป็น dict ที่แก้ได้ (qty=1, tier=ค่าเริ่มต้น, type/deliver = ตามหน้า Create Bundle)
        data = []
        for b in bundles:
            items, seen = [], set()
            for it in (b.get('items') or []):
                m = re.search(r'\d+', str(it.get('id', '')))
                if not m or m.group() in seen:
                    continue
                iid = m.group(); seen.add(iid)
                items.append({'id': iid, 'qty': '1', 'tier': DEFAULT_TIER})
                nm = str(it.get('name', '')).strip()
                if nm:
                    self._item_names[iid] = nm
            if not items:
                continue
            data.append({'name': (b.get('name') or 'Bundle').strip() or 'Bundle',
                         'type': self.vtype.get() or DEFAULT_BUNDLE_TYPE,
                         'deliver': bool(self.vdeliver.get()),
                         'items': items, 'rewards': [], 'group': b.get('group', '') or '',
                         'include': True})
        if not data:
            messagebox.showinfo('รวมเป็นบันเดิล', 'ไม่มีไอเทมให้รวม')
            return None

        win = tk.Toplevel(self.root)
        win.title('รวมเป็นบันเดิล — รีวิวก่อนส่งเข้าคิว')
        win.configure(bg=C['bg_dark'])
        win.geometry('1060x730')
        win.minsize(920, 640)
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass

        # ล่าง: แถบส่งเข้าคิว (pack ก่อน เพื่อจองขอบล่างไว้)
        bar = tk.Frame(win, bg=C['bg_dark'])
        bar.pack(side='bottom', fill='x', padx=10, pady=(0, 10))
        body = tk.Frame(win, bg=C['bg_dark'])
        body.pack(side='top', fill='both', expand=True)

        # ---- ซ้าย: ลิสต์บันเดิลทั้งหมด ----
        left = tk.Frame(body, bg=C['bg_med'], width=280)
        left.pack(side='left', fill='y')
        left.pack_propagate(False)
        tk.Label(left, text='บันเดิลทั้งหมด (☑ = ส่ง)', bg=C['bg_med'], fg=C['text'],
                 font=FB).pack(anchor='w', padx=10, pady=(10, 2))
        tk.Label(left, text='ดับเบิลคลิก = สลับใช้/ไม่ใช้', bg=C['bg_med'], fg=C['muted'],
                 font=('Segoe UI', 9)).pack(anchor='w', padx=10)
        lb = tk.Listbox(left, bg=C['bg_inp'], fg=C['text'], font=('Consolas', 10), relief='flat',
                        activestyle='none', selectbackground=C['accent'], exportselection=False)
        lb.pack(fill='both', expand=True, padx=10, pady=6)
        lbtn = tk.Frame(left, bg=C['bg_med'])
        lbtn.pack(fill='x', padx=10, pady=(0, 10))

        # ---- ขวา: ฟอร์มแก้บันเดิลที่เลือก (reuse _BundleEditor) ----
        right = tk.Frame(body, bg=C['bg_dark'])
        right.pack(side='left', fill='both', expand=True)
        ed = _BundleEditor(self, right, item_mode='rows')   # รายไอเทม + ปุ่มตัด/เรียงลำดับ

        state = {'cur': None}

        def _commit_current():
            i = state['cur']
            if i is None or not (0 <= i < len(data)):
                return
            nb = ed.to_bundle(group=data[i]['group'])
            d = data[i]
            d['name'] = nb['name'] or d['name']
            d['type'] = nb['type']
            d['deliver'] = nb['deliver']
            d['items'] = nb['items']
            d['rewards'] = nb['rewards']

        def _refresh_list(keep=None):
            lb.delete(0, tk.END)
            for d in data:
                mark = '☑' if d['include'] else '☐'
                lb.insert(tk.END, '%s %s  (%d)' % (mark, d['name'], len(d['items'])))
            if keep is not None and 0 <= keep < len(data):
                lb.selection_clear(0, tk.END)
                lb.selection_set(keep)
                lb.see(keep)

        def _load_index(i):
            if 0 <= i < len(data):
                state['cur'] = i
                ed.load(data[i])

        def _select(i):
            _commit_current()
            _load_index(i)
            _refresh_list(keep=i)

        def _on_lb(evt=None):
            sel = lb.curselection()
            if sel and sel[0] != state['cur']:
                _select(sel[0])

        def _toggle_include(evt=None):
            sel = lb.curselection()
            if not sel:
                return 'break'
            i = sel[0]
            data[i]['include'] = not data[i]['include']
            _refresh_list(keep=i)
            return 'break'

        def _remove():
            i = state['cur']
            if i is None or not (0 <= i < len(data)):
                return
            if not messagebox.askyesno('ลบบันเดิล', 'เอาบันเดิล "%s" ออกจากรายการ?' % data[i]['name'],
                                       parent=win):
                return
            del data[i]
            if not data:
                win.destroy()
                return
            j = min(i, len(data) - 1)
            _load_index(j)
            _refresh_list(keep=j)

        def _send_all():
            _commit_current()
            chosen = [d for d in data if d['include'] and d['items']]
            if not chosen:
                messagebox.showinfo('ส่งเข้าคิว', 'ยังไม่มีบันเดิลที่ติ๊กไว้ (หรือบันเดิลว่างเปล่า)',
                                    parent=win)
                return
            added, last = 0, None
            for d in chosen:
                base = d['name'] or 'Bundle'
                existing = {x['name'] for x in self._bundles}
                uniq, k = base, 2
                while uniq in existing:
                    uniq = '%s (%d)' % (base, k); k += 1
                last = {'name': uniq, 'type': d['type'], 'deliver': d['deliver'],
                        'items': d['items'], 'rewards': d['rewards'], 'group': d['group']}
                self._bundles.append(last); added += 1
            self.set_combo_values(self.bundle_picker, [x['name'] for x in self._bundles])
            if last:
                try:
                    self.bundle_picker.config(state='readonly')
                except Exception:
                    pass
                self.vbundle.set(last['name'])
                self._load_bundle(last)
            self.log('ส่ง %d บันเดิลจากหน้ารีวิวเข้าคิว — คิวรวม %d bundle'
                     % (added, len(self._bundles)), 'SUCCESS')
            win.destroy()
            if on_done:
                on_done(added)

        tk.Button(lbtn, text='สลับใช้/ไม่ใช้', command=_toggle_include, bg=C['bg_card'], fg=C['text'],
                  font=FM, relief='flat', padx=8, cursor='hand2').pack(side='left')
        tk.Button(lbtn, text='ลบออกจากรายการ', command=_remove, bg=C['bg_card'], fg=C['danger'],
                  font=FM, relief='flat', padx=8, cursor='hand2').pack(side='left', padx=6)
        lb.bind('<<ListboxSelect>>', _on_lb)
        lb.bind('<Double-Button-1>', _toggle_include)

        tk.Button(bar, text='✅ ส่งเข้าคิวทั้งหมด → Create Bundle', command=_send_all,
                  bg=C['accent2'], fg='#fff', font=FB, relief='flat', padx=16, pady=7,
                  cursor='hand2').pack(side='left')
        tk.Button(bar, text='ยกเลิก', command=win.destroy, bg=C['bg_card'], fg=C['muted'],
                  font=FM, relief='flat', padx=14, pady=7, cursor='hand2').pack(side='left', padx=8)
        tk.Label(bar, text='%d บันเดิล' % len(data), bg=C['bg_dark'], fg=C['teal'],
                 font=('Segoe UI', 9)).pack(side='right')

        _load_index(0)
        _refresh_list(keep=0)

        # handle สำหรับทดสอบ/ใช้ภายนอก
        win.editor = ed
        win.listbox = lb
        win.review_data = data
        win.select_bundle = _select
        win.toggle_include = _toggle_include
        win.remove_bundle = _remove
        win.send_all = _send_all
        win.commit_current = _commit_current
        ed.focus_name()
        return win

    # ---------------- import / template ----------------
    def _save_template(self):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(title='บันทึก template', defaultextension='.xlsx',
                                            initialfile='bundle_template.xlsx',
                                            filetypes=[('Excel', '*.xlsx'), ('CSV', '*.csv')])
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        opts = load_options(self.vgame.get())        # dropdown ใน template = ของเกมที่เลือกอยู่
        try:
            if ext == '.csv':
                build_template_csv(path, opts)
            else:
                try:
                    build_template_xlsx(path, opts)
                except ImportError:
                    messagebox.showerror('ผิดพลาด', 'สร้าง .xlsx ต้องมี openpyxl (pip install openpyxl) หรือบันทึกเป็น .csv')
                    return
        except Exception as e:
            messagebox.showerror('บันทึกไม่สำเร็จ', str(e))
            return
        self.log('บันทึก template: %s' % os.path.basename(path), 'SUCCESS')

    def _read_rows(self, path, ext):
        if ext == '.csv':
            with open(path, newline='', encoding='utf-8-sig') as f:
                return [r for r in csv.reader(f)]
        if ext in ('.xlsx', '.xlsm'):
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror('ผิดพลาด', 'ต้องติดตั้ง openpyxl ก่อน (pip install openpyxl)')
                return None
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            return [list(r) for r in wb.active.iter_rows(values_only=True)]
        messagebox.showwarning('ไฟล์ไม่รองรับ', 'รองรับเฉพาะ .xlsx / .csv')
        return None

    def _parse_bundles(self, header, data_rows):
        def idx(name):
            return header.index(name) if name in header else -1
        i_name, i_type = idx('bundle_name'), idx('bundle_type')
        i_dlv, i_id = idx('deliver_now'), idx('item_id')
        i_qty, i_tier = idx('qty'), idx('tier')
        i_rate = idx('random_rate')
        reward_idx = [(idx(col), kind) for col, kind, _ in REWARD_COLS]
        tier_map = {t.lower(): t for t in TIERS}

        def cell(row, i):
            if i < 0 or i >= len(row) or row[i] is None:
                return ''
            return str(row[i]).strip()

        order, by_name = [], {}
        for row in data_rows:
            name = cell(row, i_name)
            if not name:
                continue
            qty_m = re.search(r'\d+', cell(row, i_qty))
            qty = qty_m.group() if qty_m else '1'
            iid_m = re.search(r'\d+', cell(row, i_id))
            reward = None
            for ridx, kind in reward_idx:
                val = cell(row, ridx)
                if val:
                    reward = (kind, val)
                    break
            if not iid_m and reward is None:
                continue
            if name not in by_name:
                btype = cell(row, i_type).upper()
                if btype not in BUNDLE_TYPES:
                    btype = DEFAULT_BUNDLE_TYPE
                deliver = cell(row, i_dlv).lower() in ('true', '1', 'yes', 'y', 'ใช่', 'on')
                by_name[name] = {'name': name, 'type': btype, 'deliver': deliver,
                                 'items': [], 'rewards': []}
                order.append(name)
            if iid_m:
                tier = tier_map.get(cell(row, i_tier).lower(), DEFAULT_TIER)
                item = {'id': iid_m.group(), 'qty': qty, 'tier': tier}
                rate_m = re.search(r'\d+(\.\d+)?', cell(row, i_rate))
                if rate_m:
                    item['rate'] = rate_m.group()
                by_name[name]['items'].append(item)
            else:
                by_name[name]['rewards'].append({'type': reward[0], 'value': reward[1], 'qty': qty})
        return [by_name[n] for n in order]

    def _import_ids(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(filetypes=[('Excel/CSV', '*.xlsx *.xlsm *.csv'), ('ทั้งหมด', '*.*')])
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            rows = self._read_rows(path, ext)
        except Exception as e:
            messagebox.showerror('อ่านไฟล์ไม่สำเร็จ', str(e))
            return
        if rows is None:
            return
        if not rows:
            self.log('ไฟล์ว่าง', 'WARNING')
            return
        header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
        if 'bundle_name' in header and 'item_id' in header:
            bundles = self._parse_bundles(header, rows[1:])
            if not bundles:
                self.log('ไม่พบข้อมูล bundle ในไฟล์', 'WARNING')
                return
            self._bundles = bundles
            names = [b['name'] for b in bundles]
            self.set_combo_values(self.bundle_picker, names)
            self.vbundle.set(names[0])
            self._load_bundle(bundles[0])
            self.log('Import %d bundle: %s' % (len(bundles), ', '.join(names)), 'SUCCESS')
            if len(bundles) > 1:
                self.log('  มีหลาย bundle - เลือกจาก dropdown หรือกด "สร้างทุก bundle อัตโนมัติ"', 'INFO')
        else:
            lines = []
            for r in rows:
                vals = [str(c).strip() for c in r if c is not None and str(c).strip() != '']
                if any(re.fullmatch(r'\d+', v) for v in vals):
                    lines.append(' '.join(vals))
            if lines:
                self.ids_text.insert('end', '\n'.join(lines) + '\n')
            self._refresh_detail()
            self.log('Import %d แถว จาก %s' % (len(lines), os.path.basename(path)), 'SUCCESS')

    def _on_pick_bundle(self):
        self._commit_fields_to_queue()      # เซฟที่แก้ของบันเดิลก่อนหน้าไว้ก่อนสลับ
        name = self.vbundle.get()
        for b in self._bundles:
            if b['name'] == name:
                self._load_bundle(b)
                self.log('โหลด bundle: %s (%d ไอเท็ม, %d reward)'
                         % (b['name'], len(b['items']), len(b.get('rewards', []))), 'INFO')
                break

    def _del_selected_bundle(self):
        """ลบบันเดิลที่เลือกอยู่ (ตาม dropdown) ออกจากคิว แล้วเลื่อนไปโหลดตัวถัดไป"""
        name = self.vbundle.get()
        idx = next((i for i, b in enumerate(self._bundles) if b['name'] == name), -1)
        if idx < 0:
            if self._bundles:
                messagebox.showinfo('ลบบันเดิล', 'ยังไม่ได้เลือกบันเดิลจากคิว')
            return
        if not messagebox.askyesno('ลบบันเดิล', 'ลบบันเดิล "%s" ออกจากคิว?' % name):
            return
        del self._bundles[idx]
        names = [b['name'] for b in self._bundles]
        self.set_combo_values(self.bundle_picker, names)
        if names:
            j = min(idx, len(names) - 1)
            self.vbundle.set(names[j])
            self._load_bundle(self._bundles[j])
        else:
            self.vbundle.set('')
            try:
                self.bundle_picker.config(state='disabled')
            except Exception:
                pass
            self._load_bundle({'name': '', 'type': self.vtype.get() or DEFAULT_BUNDLE_TYPE,
                               'deliver': False, 'items': [], 'rewards': []})
        self._refresh_detail()
        self.log('ลบบันเดิล "%s" ออกจากคิว — เหลือ %d bundle' % (name, len(self._bundles)), 'INFO')

    def _commit_fields_to_queue(self):
        """เขียนสิ่งที่แก้ในฟอร์ม (ชื่อ/Type/ส่งของ/ไอเทม) กลับเข้าบันเดิลที่กำลังเลือกในคิว
        -> สลับบันเดิลไปมาแล้วแก้ไม่หาย และ "สร้างทุก bundle" ใช้ค่าล่าสุด (reward/group คงเดิม)"""
        prev = self._sel_bundle
        if not prev:
            return
        idx = next((i for i, b in enumerate(self._bundles) if b['name'] == prev), -1)
        if idx < 0:
            return
        b = self._bundles[idx]
        new_name = self.vname.get().strip() or prev
        if new_name != prev:                         # กันชื่อชนตัวอื่นในคิว
            others = {x['name'] for j, x in enumerate(self._bundles) if j != idx}
            uniq, k = new_name, 2
            while uniq in others:
                uniq = '%s (%d)' % (new_name, k); k += 1
            new_name = uniq
        b['name'] = new_name
        if self.vtype.get() in BUNDLE_TYPES:
            b['type'] = self.vtype.get()
        b['deliver'] = bool(self.vdeliver.get())
        b['items'] = self._get_items()               # id/qty/tier/rate จากฟอร์ม (reward คงเดิม)
        if new_name != prev:
            self._sel_bundle = new_name
            self.set_combo_values(self.bundle_picker, [x['name'] for x in self._bundles])
            if self.vbundle.get() == prev:
                self.vbundle.set(new_name)
            if self._current_rewards_name == prev:
                self._current_rewards_name = new_name

    def _load_bundle(self, b):
        self._sel_bundle = b.get('name')
        self.vname.set(b['name'])
        if b['type'] in BUNDLE_TYPES:
            self.vtype.set(b['type'])
        self.vdeliver.set(bool(b['deliver']))
        self.ids_text.delete('1.0', 'end')
        lines, tiers, rates = [], [], []
        for it in b['items']:
            lines.append('%s %s' % (it['id'], it['qty']))     # tier แยกไปช่อง Tier (เหมือนเรทสุ่ม)
            tiers.append(str(it.get('tier', '') or ''))
            rates.append(str(it.get('rate', '') or ''))
        if lines:
            self.ids_text.insert('1.0', '\n'.join(lines) + '\n')
        # เติมช่อง Tier ให้ตรงบรรทัด
        tt = getattr(self, '_tier_text', None)
        if tt is not None:
            tt.delete('1.0', 'end')
            if any(tiers):
                tt.insert('1.0', '\n'.join(tiers) + '\n')
        # เติมช่องเรทสุ่มให้ตรงบรรทัด (โชว์/ซ่อนตาม Type ที่ตั้งด้านบนแล้ว)
        rt = getattr(self, '_rate_text', None)
        if rt is not None:
            rt.delete('1.0', 'end')
            if any(rates):
                rt.insert('1.0', '\n'.join(rates) + '\n')
        rewards = b.get('rewards', [])
        self._current_rewards = list(rewards)
        self._current_rewards_name = b['name']
        if rewards:
            summ = '  |  '.join('%s = %s x%s' % (r['type'], r['value'], r['qty']) for r in rewards)
            self.reward_lbl.config(text='Reward (%d): %s' % (len(rewards), summ))
            self.log('  bundle นี้มี reward %d รายการ: %s' % (len(rewards), summ), 'INFO')
        else:
            self.reward_lbl.config(text='')
        self._sync_name_col()      # กำกับชื่อไอเท็มตาม ID ที่เพิ่งโหลด
        self._refresh_detail()

    # ---------------- login / nav ----------------
    def _open_login(self):
        if not PW_OK:
            messagebox.showerror('ผิดพลาด', 'ยังไม่ได้ติดตั้ง playwright')
            return
        url = game_url(self.vgame.get(), 'bundles')
        self.log('กำลังเปิดหน้า login: ' + url, 'INFO')

        def _run():
            async def _open():
                async with async_playwright() as pw:
                    browser = await pw.chromium.launch_persistent_context(**self._launch_kwargs())
                    page = browser.pages[0] if browser.pages else await browser.new_page()
                    await page.goto(url, wait_until='domcontentloaded', timeout=30000)
                    self.log('  Login แล้วปิด browser ได้เลย', 'SUCCESS')
                    await self._idle_until_closed(browser, page)
                    self.log('  Profile saved', 'SUCCESS')
            try:
                asyncio.run(_open())
            except Exception:
                self.log(traceback.format_exc(), 'ERROR')

        threading.Thread(target=_run, daemon=True).start()

    def _clear_profile(self):
        import shutil
        if os.path.isdir(CHROME_PROFILE):
            try:
                shutil.rmtree(CHROME_PROFILE, ignore_errors=True)
                self.log('ล้าง session แล้ว', 'WARNING')
            except Exception:
                self.log('ลบไม่สำเร็จ (อาจมี browser เปิดค้างอยู่)', 'ERROR')
        else:
            self.log('ยังไม่มี session ที่เก็บไว้', 'INFO')

    def _test_nav(self):
        if not PW_OK:
            messagebox.showerror('ผิดพลาด', 'ยังไม่ได้ติดตั้ง playwright')
            return
        if self._busy:
            return
        game = self.vgame.get()
        url = game_url(game, 'bundles')
        self.log('-- ทดสอบ [' + game + '] -> ' + url, 'STEP')
        self._cancel = False
        self._set_busy(True)

        def _run():
            async def _t():
                async with async_playwright() as pw:
                    browser = await pw.chromium.launch_persistent_context(**self._launch_kwargs())
                    page = browser.pages[0] if browser.pages else await browser.new_page()
                    try:
                        await page.goto(url, wait_until='domcontentloaded', timeout=30000)
                    except Exception as e:
                        self.log('FAIL x โหลดหน้าไม่ได้: ' + str(e), 'ERROR')
                        await browser.close()
                        return
                    await page.wait_for_timeout(2500)
                    final = page.url
                    low = final.lower()
                    if '/shop/bundles' in low:
                        self.log('PASS เข้าหน้า Bundles ได้  | url=' + final, 'SUCCESS')
                    elif any(k in low for k in ('login', 'signin', 'sign-in', 'auth')):
                        self.log('FAIL x ถูกเด้งไปหน้า login - กด "เปิดหน้า Login" ก่อน | url=' + final, 'ERROR')
                    else:
                        self.log('WARNING ! ไม่ได้อยู่ที่ /shop/bundles | url=' + final, 'WARNING')
                    await page.wait_for_timeout(4000)
                    await browser.close()
            try:
                asyncio.run(_t())
            except Exception:
                self.log(traceback.format_exc(), 'ERROR')
            finally:
                self._set_busy(False)

        threading.Thread(target=_run, daemon=True).start()

    def _fetch_options(self, on_done=None):
        """ดึงรายการสกุลเงิน/ยศ จากฟอร์มสร้าง bundle ของเกมที่เลือก แล้วเก็บไว้ใช้ต่อ
        on_done = callback (main thread) เรียกเมื่อดึงเสร็จ ไว้ให้หน้าต่างอื่นรีเฟรช dropdown"""
        if not PW_OK:
            messagebox.showerror('ผิดพลาด', 'ยังไม่ได้ติดตั้ง playwright')
            return
        if self._busy:
            return
        game = self.vgame.get()
        url = game_url(game, 'bundles')
        self.log('-- Fetch ตัวเลือก (สกุลเงิน/ยศ) แยกตาม section [%s]' % game, 'STEP')
        self._cancel = False
        self._set_busy(True)

        def _run():
            async def _f():
                async with async_playwright() as pw:
                    browser = await pw.chromium.launch_persistent_context(**self._launch_kwargs())
                    page = browser.pages[0] if browser.pages else await browser.new_page()
                    try:
                        await page.goto(url, wait_until='domcontentloaded', timeout=30000)
                    except Exception as e:
                        self.log('FAIL x โหลดหน้าไม่ได้: ' + str(e), 'ERROR')
                        await browser.close()
                        return
                    await page.wait_for_timeout(2000)
                    # เช็ค/ซ่อม session ให้เอง (เว็บไม่ได้เด้งไป url login เสมอ — บางทีเป็นหน้า
                    # 'No Permission' + ลิงก์ Login เฉย ๆ เช็คจาก url อย่างเดียวเลยหลุด)
                    if not await core.ensure_logged_in(page, log=self.log):
                        self.log('FAIL x ยังไม่ได้ login — กดปุ่ม "เข้าสู่ระบบ" ก่อน', 'ERROR')
                        await browser.close()
                        return
                    btn = await first_visible(page, [
                        "button:has-text('%s')" % SEL['add_bundle_btn'],
                        "text=%s" % SEL['add_bundle_btn'],
                    ])
                    if btn:
                        await btn.click()
                        await page.wait_for_timeout(1500)
                    else:
                        self.log('  หาปุ่ม "เพิ่ม Bundle" ไม่เจอ - ลองอ่าน select เท่าที่มี', 'WARNING')
                    result, total = {}, 0
                    for _col, kind, label in REWARD_COLS:
                        acc = await self._reward_section(page, label)
                        opts = []
                        if acc is not None:
                            try:
                                sel = acc.locator('select').first
                                if await sel.count():
                                    opts = await sel.evaluate(
                                        "el => [...el.options].map(o => (o.textContent||'').trim())")
                            except Exception:
                                opts = []
                        opts = [o for o in opts if o and 'เลือก' not in o]
                        opts = list(dict.fromkeys(opts))
                        result[kind] = opts
                        total += len(opts)
                        self.log('  %s (%d): %s' % (kind, len(opts), ', '.join(opts) or '-'),
                                 'SUCCESS' if opts else 'WARNING')
                    if total == 0:
                        self.log('  ไม่พบ option เลย (ดูว่าฟอร์มเปิด/section แสดงถูกไหม)', 'WARNING')
                    else:
                        try:
                            save_options(game, result)      # เก็บแยกตามเกม/เซิร์ฟ
                            self.log('  เก็บตัวเลือกของ [%s] แล้ว — ใช้ได้ทั้งใน "เพิ่ม Bundle เข้าคิว" '
                                     'และ template' % game, 'INFO')
                        except Exception as e:
                            self.log('  บันทึก options ไม่สำเร็จ: ' + str(e), 'WARNING')
                    await browser.close()
            try:
                asyncio.run(_f())
            except Exception:
                self.log(traceback.format_exc(), 'ERROR')
            finally:
                self._set_busy(False)
                if on_done:                                 # ให้หน้าต่างที่เรียก รีเฟรช dropdown
                    try:
                        self.root.after(0, on_done)
                    except Exception:
                        pass

        threading.Thread(target=_run, daemon=True).start()

    # ---------------- สร้าง bundle (เดี่ยว / ทุกอัน) ----------------
    def _create_bundle_fields(self):
        if not PW_OK:
            messagebox.showerror('ผิดพลาด', 'ยังไม่ได้ติดตั้ง playwright')
            return
        if self._busy:
            return
        name = self.vname.get().strip()
        if not name:
            messagebox.showwarning('กรอกชื่อก่อน', 'กรุณาใส่ชื่อบันเดิล')
            return
        game = self.vgame.get()
        btype = self.vtype.get()
        deliver = self.vdeliver.get()
        do_save = self.vsave.get()
        items = self._get_items()
        rewards = (list(self._current_rewards)
                   if name == self._current_rewards_name else [])
        dup = self._dupes(self._raw_ids())
        if dup:
            if not messagebox.askyesno('พบ item id ซ้ำ',
                                       'มี item id ซ้ำ: %s\nจะตัดให้เหลือตัวเดียวแล้วสร้างต่อไหม?'
                                       % ', '.join(dup)):
                return
        if btype.upper() == 'RANDOM':
            missing = [it['id'] for it in items if not it.get('rate')]
            if missing and not messagebox.askyesno(
                    'เรทสุ่มไม่ครบ',
                    'Type = RANDOM แต่ไอเท็มเหล่านี้ยังไม่มีเรทสุ่ม:\n%s\n\n'
                    'ใส่ค่าในช่อง "เรทสุ่ม" (คอลัมน์สีเหลืองทางขวาของรายการไอเทม) '
                    'ให้ครบก่อน แล้วค่อยสร้าง\n\nจะสร้างต่อโดยข้ามเรทสุ่มเลยไหม?'
                    % ', '.join(missing[:15])):
                return
        self._cancel = False
        self._set_busy(True)
        self._current_group = ''      # สร้างเดี่ยว: ไม่ผูกกลุ่ม (กัน id หลุดไปกลุ่มเก่าจาก batch ก่อนหน้า)

        def _run():
            async def _go():
                async with async_playwright() as pw:
                    browser = await pw.chromium.launch_persistent_context(**self._launch_kwargs())
                    page = browser.pages[0] if browser.pages else await browser.new_page()
                    await self._process_bundle(page, game, name, btype, deliver, items, rewards, do_save)
                    await self._idle_until_closed(browser, page)
            try:
                asyncio.run(_go())
            except Exception:
                self.log(traceback.format_exc(), 'ERROR')
            finally:
                self._set_busy(False)

        threading.Thread(target=_run, daemon=True).start()

    def _create_all(self):
        if not PW_OK:
            messagebox.showerror('ผิดพลาด', 'ยังไม่ได้ติดตั้ง playwright')
            return
        if self._busy:
            return
        if not self._bundles:
            messagebox.showinfo('ว่าง', 'ยังไม่ได้ import ไฟล์ (ไม่มี bundle)')
            return
        self._commit_fields_to_queue()      # ดึงค่าที่เพิ่งแก้ในฟอร์มของตัวที่เลือกอยู่เข้าคิวก่อนสร้าง
        if not messagebox.askyesno('ยืนยัน',
                                   'สร้างทุก bundle อัตโนมัติ %d อัน และ "กด Save จริง" ทุกอัน?\n'
                                   'ดำเนินการต่อไหม?' % len(self._bundles)):
            return
        game = self.vgame.get()
        bundles = [dict(b) for b in self._bundles]
        dup_report = []
        for b in bundles:
            d = self._dupes([it['id'] for it in b['items']])
            if d:
                dup_report.append('%s: %s' % (b['name'], ', '.join(d)))
        if dup_report:
            if not messagebox.askyesno('พบ item id ซ้ำ',
                                       'พบ item id ซ้ำในบางบันเดิล:\n%s\n\nจะตัดซ้ำแล้วสร้างต่อไหม?'
                                       % '\n'.join(dup_report)):
                return
        rate_report = []
        for b in bundles:
            if b.get('type', '').upper() == 'RANDOM':
                miss = [it['id'] for it in b['items'] if not it.get('rate')]
                if miss:
                    rate_report.append('%s: %s' % (b['name'], ', '.join(miss[:10])))
        if rate_report:
            if not messagebox.askyesno('เรทสุ่มไม่ครบ',
                                       'บันเดิล RANDOM เหล่านี้มีไอเท็มที่ยังไม่มีเรทสุ่ม:\n%s\n\n'
                                       'จะสร้างต่อโดยข้ามเรทสุ่มเลยไหม?' % '\n'.join(rate_report)):
                return
        self.log('==== สร้างทุก bundle อัตโนมัติ %d อัน ====' % len(bundles), 'STEP')
        self._cancel = False
        self._set_busy(True)

        def _run():
            async def _go():
                async with async_playwright() as pw:
                    browser = await pw.chromium.launch_persistent_context(**self._launch_kwargs())
                    page = browser.pages[0] if browser.pages else await browser.new_page()
                    ok = 0
                    for n, b in enumerate(bundles, 1):
                        if self._cancel:
                            self.log('หยุดกลางคัน (ทำไป %d/%d)' % (n - 1, len(bundles)), 'WARNING')
                            break
                        self.log('----- [%d/%d] %s -----' % (n, len(bundles), b['name']), 'STEP')
                        self._current_group = b.get('group', '')   # ผูก id ที่จะสร้างเข้ากลุ่มนี้
                        try:
                            saved = await self._process_bundle(
                                page, game, b['name'], b['type'], b['deliver'],
                                b['items'], b.get('rewards', []), True)
                            if saved:
                                ok += 1
                        except Exception:
                            self.log(traceback.format_exc(), 'ERROR')
                        await page.wait_for_timeout(1000)
                    self.log('==== เสร็จ: สร้างสำเร็จ %d/%d อัน ====' % (ok, len(bundles)),
                             'SUCCESS' if ok == len(bundles) else 'WARNING')
                    await self._idle_until_closed(browser, page)
            try:
                asyncio.run(_go())
            except Exception:
                self.log(traceback.format_exc(), 'ERROR')
            finally:
                self._set_busy(False)

        threading.Thread(target=_run, daemon=True).start()

    async def _process_bundle(self, page, game, name, btype, deliver, items, rewards, do_save):
        """สร้าง 1 bundle: กรอกฟอร์ม + items + rewards + tier + (save) + ดึงเลข + รีเชค. คืน saved(bool)"""
        url = game_url(game, 'bundles')
        self.log('== สร้าง Bundle [%s] ชื่อ="%s" type=%s ส่งทันที=%s ไอเทม=%d reward=%d'
                 % (game, name, btype, 'on' if deliver else 'off', len(items), len(rewards)), 'STEP')
        if items:
            seen, uni = set(), []
            for it in items:
                if it['id'] in seen:
                    continue
                seen.add(it['id'])
                uni.append(it)
            if len(uni) != len(items):
                self.log('  (ตัด item id ซ้ำออก %d ตัว)' % (len(items) - len(uni)), 'WARNING')
            items = uni
        try:
            await page.goto(url, wait_until='domcontentloaded', timeout=30000)
        except Exception as e:
            self.log('FAIL x โหลดหน้าไม่ได้: ' + str(e), 'ERROR')
            return False
        await page.wait_for_timeout(2000)
        if any(k in page.url.lower() for k in ('login', 'signin', 'auth')):
            self.log('FAIL x ยังไม่ได้ login (โดนเด้งไป login)', 'ERROR')
            return False

        # 1) คลิก "เพิ่ม Bundle"
        btn = await first_visible(page, [
            "button:has-text('%s')" % SEL['add_bundle_btn'],
            "text=%s" % SEL['add_bundle_btn'],
            "a:has-text('%s')" % SEL['add_bundle_btn'],
        ])
        if not btn:
            self.log('FAIL x หาปุ่ม "เพิ่ม Bundle" ไม่เจอ', 'ERROR')
            await page.wait_for_timeout(2000)
            return False
        await btn.click()
        self.log('  [1] คลิก "เพิ่ม Bundle" แล้ว', 'INFO')
        await page.wait_for_timeout(1500)

        # 2) ชื่อบันเดิล
        name_box = await first_visible(page, [
            "xpath=//*[contains(normalize-space(.),'%s')]/following::input[1]" % SEL['name_label'],
            "input[placeholder*='ชื่อ']",
            "form input[type='text']",
            "input[type='text']",
        ])
        if name_box:
            try:
                await name_box.fill(name)
                self.log('  [2] ใส่ชื่อบันเดิล: ' + name, 'INFO')
            except Exception as e:
                self.log('  [2] กรอกชื่อไม่สำเร็จ: ' + str(e), 'WARNING')
        else:
            self.log('  [2] หาช่องชื่อบันเดิลไม่เจอ', 'WARNING')

        # 3) Bundle Type
        type_done = False
        sel_el = await first_visible(page, [
            "xpath=//*[contains(text(),'Bundle Type')]/following::select[1]",
            "xpath=//label[contains(.,'Bundle Type')]/following::select[1]",
        ])
        if sel_el:
            for how in ('label', 'value'):
                try:
                    if how == 'label':
                        await sel_el.select_option(label=btype)
                    else:
                        await sel_el.select_option(value=btype)
                    type_done = True
                    self.log('  [3] เลือก Bundle Type: ' + btype, 'INFO')
                    break
                except Exception:
                    continue
        if not type_done:
            try:
                trigger = await first_visible(page, [
                    "text=%s" % SEL['type_placeholder'],
                    "xpath=//*[contains(text(),'Bundle Type')]/following::*[contains(normalize-space(.),'%s')][1]" % SEL['type_placeholder'],
                ])
                if trigger:
                    await trigger.click()
                    await page.wait_for_timeout(600)
                    opt = await first_visible(page, [
                        "xpath=//li[normalize-space(.)='%s']" % btype,
                        "xpath=//*[@role='option'][normalize-space(.)='%s']" % btype,
                        "text=%s" % btype,
                    ])
                    if opt:
                        await opt.click()
                        type_done = True
                        self.log('  [3] เลือก Bundle Type (dropdown): ' + btype, 'INFO')
            except Exception as e:
                self.log('  [3] dropdown error: ' + str(e), 'WARNING')
        if not type_done:
            self.log('  [3] เลือก Bundle Type ไม่สำเร็จ (%s)' % btype, 'WARNING')

        # 4) checkbox ส่งของทันที
        try:
            cb = await first_visible(page, [
                "xpath=//*[contains(text(),'%s')]/following::input[@type='checkbox'][1]" % SEL['deliver_label'],
                "xpath=//*[contains(text(),'%s')]/preceding::input[@type='checkbox'][1]" % SEL['deliver_label'],
                "input[type='checkbox']",
            ])
            if cb:
                cur = await cb.is_checked()
                if cur != deliver:
                    await cb.click()
                    self.log('  [4] ปรับ ส่งของทันที -> ' + ('on' if deliver else 'off'), 'INFO')
                else:
                    self.log('  [4] ส่งของทันที ตรงอยู่แล้ว', 'INFO')
            else:
                self.log('  [4] หา checkbox ส่งของทันทีไม่เจอ', 'WARNING')
        except Exception as e:
            self.log('  [4] เซ็ต checkbox ไม่สำเร็จ: ' + str(e), 'WARNING')

        # 5) เพิ่มไอเทม
        if items:
            self.log('  [5] เพิ่มไอเทม %d ตัว' % len(items), 'INFO')
            ok_cnt = 0
            for it in items:
                if self._cancel:
                    self.log('  [5] หยุดกลางคัน', 'WARNING')
                    break
                try:
                    if await self._add_one_item(page, it['id']):
                        ok_cnt += 1
                except Exception as e:
                    self.log('    x %s error: %s' % (it['id'], str(e)), 'WARNING')
            self.log('  [5] เพิ่มไอเทมสำเร็จ %d/%d' % (ok_cnt, len(items)),
                     'SUCCESS' if ok_cnt == len(items) else 'WARNING')

        # 5b) เพิ่ม reward
        if rewards:
            self.log('  [5b] เพิ่ม reward %d รายการ' % len(rewards), 'INFO')
            rok = 0
            for rw in rewards:
                if self._cancel:
                    self.log('  [5b] หยุดกลางคัน', 'WARNING')
                    break
                try:
                    if await self._add_one_reward(page, rw['type'], rw['value'], rw['qty']):
                        rok += 1
                except Exception as e:
                    self.log('    x reward %s error: %s' % (rw.get('type'), str(e)), 'WARNING')
            self.log('  [5b] reward สำเร็จ %d/%d' % (rok, len(rewards)),
                     'SUCCESS' if rok == len(rewards) else 'WARNING')

        # 6) Expand All + ตั้งจำนวน/Tier (item) + Tier reward
        if items or rewards:
            eb = await first_visible(page, [
                "button:has-text('%s')" % SEL['expand_all'],
                "text=%s" % SEL['expand_all'],
            ], timeout_each=2500)
            if eb:
                await eb.click()
                self.log('  [6] กด Expand All', 'INFO')
                await page.wait_for_timeout(1200)
            else:
                self.log('  [6] หาปุ่ม Expand All ไม่เจอ', 'WARNING')
            is_random = btype.upper() == 'RANDOM'
            if items:
                qok = 0
                for it in items:
                    if self._cancel:
                        break
                    try:
                        if await self._set_item_qty_tier(page, it['id'], it['qty'], it['tier']):
                            qok += 1
                    except Exception as e:
                        self.log('    x %s qty/tier error: %s' % (it['id'], str(e)), 'WARNING')
                    # เรทสุ่ม (เฉพาะ bundle type = RANDOM)
                    if is_random:
                        rate = it.get('rate', '')
                        if rate not in ('', None):
                            try:
                                await self._set_item_rate(page, it['id'], rate)
                            except Exception as e:
                                self.log('    x %s rate error: %s' % (it['id'], str(e)), 'WARNING')
                        else:
                            self.log('    ! %s: RANDOM แต่ไม่มีเรทสุ่ม (ข้าม)' % it['id'], 'WARNING')
                self.log('  [6] ตั้งจำนวน/Tier (item) สำเร็จ %d/%d' % (qok, len(items)),
                         'SUCCESS' if qok == len(items) else 'WARNING')
            try:
                swept = await self._fill_unset_tiers(page, DEFAULT_TIER)
                if swept:
                    self.log('  [6] ตั้ง Tier=%s ให้การ์ดที่ยังว่าง %d อัน' % (DEFAULT_TIER, swept), 'INFO')
            except Exception as e:
                self.log('  [6] ตั้ง Tier reward error: %s' % str(e), 'WARNING')

        # 7) Save + 8) ดึงเลข
        saved = False
        created_id = None
        if do_save and not self._cancel:
            sb = await first_visible(page, [
                "button:has-text('%s')" % SEL['save_btn'],
                "text=%s" % SEL['save_btn'],
            ], timeout_each=3000)
            if not sb:
                self.log('  [7] หาปุ่มยืนยันการสร้างบันเดิลไม่เจอ', 'WARNING')
            else:
                resp = None
                try:
                    async with page.expect_response(
                        lambda r: r.request.method in ('POST', 'PUT', 'PATCH') and 'bundle' in r.url.lower(),
                        timeout=15000,
                    ) as resp_info:
                        await sb.click()
                    resp = await resp_info.value
                except Exception:
                    resp = None
                saved = True
                self.log('  [7] กดยืนยันการสร้างบันเดิล (Save) แล้ว', 'SUCCESS')
                await page.wait_for_timeout(1200)
                bid = None
                # ดึงเลข bundle เฉพาะเมื่อ response สำเร็จ (2xx) เท่านั้น — กันดึงเลขจาก error body
                if resp is not None and resp.ok:
                    self.log('  [7] HTTP %d %s' % (resp.status, resp.url), 'INFO')
                    try:
                        data = await resp.json()
                        bid = extract_bundle_id(data)
                        self.log('  [8] response: ' + json.dumps(data, ensure_ascii=False)[:600], 'INFO')
                    except Exception:
                        try:
                            txt = await resp.text()
                            self.log('  [8] response(text): ' + txt[:400], 'INFO')
                            # จับเฉพาะเลขที่อยู่ใกล้ key 'id/bundle' (ไม่ใช่เลขตัวแรกมั่ว ๆ ใน body)
                            m = re.search(r'(?:bundle[_ ]?(?:id|no)|"id")\D{0,6}(\d{3,})', txt, re.I)
                            bid = m.group(1) if m else None
                        except Exception:
                            pass
                elif resp is not None:
                    self.log('  [7] HTTP %d (ไม่สำเร็จ) — ไม่ดึงเลข bundle' % resp.status, 'WARNING')
                if not bid:
                    m = re.search(r'/bundles?/(\d+)', page.url)   # redirect URL = แหล่งที่เชื่อถือได้
                    if m:
                        bid = m.group(1)
                if bid:
                    self.log('  [8] >> เลข Bundle ที่สร้าง: %s' % bid, 'SUCCESS')
                    self._last_bundle_id = bid
                    created_id = bid
                    if self._on_bundle_created:     # ผูก id เข้าคิว Event (Item Code) ผ่าน launcher + กลุ่มต้นทาง
                        try:
                            self._on_bundle_created(bid, self._current_group)
                        except TypeError:
                            self._on_bundle_created(bid)   # เผื่อ callback รุ่นเก่ารับแค่ bid
                        except Exception:
                            pass
                else:
                    self.log('  [8] ดึงเลข bundle ไม่ได้ | url=' + page.url, 'WARNING')
        elif do_save and self._cancel:
            self.log('  [7] ข้ามบันทึก (ถูกยกเลิก)', 'WARNING')
        else:
            self.log('  [7] ข้ามบันทึก (ไม่ได้ติ๊ก Save)', 'INFO')

        # 9) รีเชค
        recheck_id = None
        if saved and not self._cancel:
            try:
                recheck_id = await self._recheck_bundle(page, game, name, created_id)
            except Exception as e:
                self.log('  [9] recheck error: ' + str(e), 'WARNING')
        if saved:
            self.add_result(name, created_id, recheck_id)

        self.log('เสร็จ - %s' % ('บันทึกแล้ว' if saved else 'ยังไม่กดบันทึก'), 'SUCCESS')
        return saved

    # ---------------- helpers: item / reward / tier / recheck ----------------
    async def _add_one_item(self, page, iid):
        box = await first_visible(page, [
            "xpath=//*[contains(text(),'%s')]/following::input[1]" % SEL['item_label'],
            "input[placeholder*='ค้นหา']",
        ], timeout_each=1500)
        if not box:
            try:
                hdr = await first_visible(page, ["xpath=//*[normalize-space(text())='Item']"], timeout_each=1500)
                if hdr:
                    await hdr.click()
                    await page.wait_for_timeout(500)
            except Exception:
                pass
            box = await first_visible(page, [
                "xpath=//*[contains(text(),'%s')]/following::input[1]" % SEL['item_label'],
                "input[placeholder*='ค้นหา']",
            ], timeout_each=1500)
        if not box:
            self.log('    x %s: หาช่องค้นหาไอเท็มไม่เจอ' % iid, 'ERROR')
            return False
        try:
            await box.fill('')
            await box.fill(iid)
        except Exception as e:
            self.log('    x %s: กรอก ID ไม่ได้: %s' % (iid, str(e)), 'WARNING')
            return False
        sbtn = await first_visible(page, [
            "button:has-text('%s')" % SEL['search_btn'],
            "xpath=//*[contains(text(),'%s')]/following::button[normalize-space(.)='%s'][1]" % (SEL['item_label'], SEL['search_btn']),
        ], timeout_each=1500)
        if sbtn:
            await sbtn.click()
        else:
            await box.press('Enter')
        await page.wait_for_timeout(400)
        row_btn = page.locator(
            "xpath=//tr[.//*[normalize-space(text())='%s']]//button[contains(normalize-space(.),'%s')]"
            % (iid, SEL['add_to_bundle'])).first
        try:
            await row_btn.wait_for(state='visible', timeout=8000)
        except Exception:
            row_btn = page.locator(
                "xpath=//tr[contains(.,'%s')]//button[contains(normalize-space(.),'%s')]"
                % (iid, SEL['add_to_bundle'])).first
            try:
                await row_btn.wait_for(state='visible', timeout=3000)
            except Exception:
                self.log('    ! %s: ค้นไม่เจอ/ไม่มีปุ่มเพิ่ม - ข้าม' % iid, 'WARNING')
                return False
        await row_btn.click()
        self.log('    + เพิ่ม %s OK' % iid, 'SUCCESS')
        await page.wait_for_timeout(700)
        return True

    async def _reward_section(self, page, label):
        """หา accordion ของ reward ตามหัวข้อ (Credit/Debit/Mileage/Player Experience) แล้วกางถ้ายังปิด"""
        acc = page.locator(
            "xpath=//h2[@data-testid='flowbite-accordion-heading' and normalize-space(.)='%s']"
            "/ancestor::div[@data-testid='flowbite-accordion'][1]" % label).first
        try:
            await acc.wait_for(state='attached', timeout=4000)
        except Exception:
            acc = page.locator(
                "xpath=//*[normalize-space(text())='%s']/ancestor::div[.//select or .//button][1]" % label).first
            try:
                await acc.wait_for(state='attached', timeout=3000)
            except Exception:
                return None
        # กางถ้ายังไม่มี select ข้างใน
        try:
            if await acc.locator('select').count() == 0:
                await acc.locator('button').first.click()
                await page.wait_for_timeout(600)
        except Exception:
            pass
        return acc

    async def _add_one_reward(self, page, kind, value, qty):
        label = REWARD_LABELS.get(kind, kind)
        acc = await self._reward_section(page, label)
        if acc is None:
            self.log('    ! reward %s: หา section "%s" ไม่เจอ' % (kind, label), 'WARNING')
            return False
        # เลือก dropdown = value (ลอง label ก่อน แล้ว value)
        try:
            sel = acc.locator('select').first
            if await sel.count() == 0:
                self.log('    ! reward %s: ไม่มี select ใน section (กางไม่ได้?)' % kind, 'WARNING')
                return False
            done = False
            for how in ('label', 'value'):
                try:
                    if how == 'label':
                        await sel.select_option(label=value)
                    else:
                        await sel.select_option(value=value)
                    done = True
                    break
                except Exception:
                    continue
            if not done:
                self.log('    ! reward %s: เลือก "%s" ไม่ได้' % (kind, value), 'WARNING')
                return False
        except Exception as e:
            self.log('    ! reward %s select error: %s' % (kind, str(e)), 'WARNING')
            return False
        # จำนวน = input text ใน section (id ลงท้าย _quantity)
        try:
            qbox = acc.locator("css=input[id$='_quantity'], input[name$='_quantity']").first
            if await qbox.count() == 0:
                qbox = acc.locator("css=input:not([type=checkbox]):not([type=radio]):not([type=hidden])").first
            await qbox.fill(str(qty))
        except Exception as e:
            self.log('    ! reward %s qty error: %s' % (kind, str(e)), 'WARNING')
            return False
        # กดปุ่มเพิ่มเข้าบันเดิลใน section
        try:
            btn = acc.locator("xpath=.//button[contains(normalize-space(.),'%s')]" % SEL['add_to_bundle']).first
            await btn.click()
            self.log('    + reward %s = %s x%s OK' % (kind, value, qty), 'SUCCESS')
            await page.wait_for_timeout(700)
            return True
        except Exception as e:
            self.log('    ! reward %s กดเพิ่มไม่ได้: %s' % (kind, str(e)), 'WARNING')
            return False

    async def _locate_card(self, page, iid):
        """หา Locator ของการ์ดไอเทมตาม Item ID
        เว็บใช้ React controlled input -> ค่า Item ID อยู่ที่ property (i.value)
        ไม่ใช่ attribute จึงหาด้วย xpath @value ไม่เจอ -> อ่านผ่าน JS แล้วอิง
        data-rfd-draggable-id ของการ์ด (item-0, item-1, ...)"""
        drag_id = None
        try:
            drag_id = await page.evaluate(
                """(iid) => {
                    const cards = document.querySelectorAll('[data-rfd-draggable-id]');
                    for (const c of cards) {
                        for (const i of c.querySelectorAll('input')) {
                            if ((i.value || '').trim() === String(iid)) {
                                return c.getAttribute('data-rfd-draggable-id');
                            }
                        }
                    }
                    return null;
                }""", str(iid))
        except Exception:
            drag_id = None
        if drag_id:
            loc = page.locator("[data-rfd-draggable-id=\"%s\"]" % drag_id).first
            try:
                await loc.wait_for(state='visible', timeout=4000)
                return loc
            except Exception:
                pass
        # fallback (เผื่อ attribute ถูก set จริง): xpath @value
        for xp in (
            "xpath=//input[@value='%s']/ancestor::*[@data-rfd-draggable-id][1]" % iid,
            "xpath=//input[@value='%s']/ancestor::*[.//select][1]" % iid,
        ):
            loc = page.locator(xp).first
            try:
                if await loc.count():
                    await loc.wait_for(state='visible', timeout=2500)
                    return loc
            except Exception:
                continue
        return None

    async def _set_item_qty_tier(self, page, iid, qty, tier):
        card = await self._locate_card(page, iid)
        if card is None:
            self.log('    ! %s: หาการ์ดไอเทมไม่เจอ' % iid, 'WARNING')
            return False
        ok = True
        try:
            qbox = card.locator(
                "css=input:enabled:not([readonly]):not([type=checkbox]):not([type=radio]):not([type=hidden])").first
            if await qbox.count() == 0:
                qbox = card.locator(
                    "xpath=.//*[contains(text(),'%s')]/following::input[not(@type) or @type='number' or @type='text'][1]" % SEL['qty_label']).first
            await qbox.fill(str(qty))
            self.log('    . %s จำนวน=%s' % (iid, qty), 'INFO')
        except Exception as e:
            self.log('    ! %s ตั้งจำนวนไม่ได้: %s' % (iid, str(e)), 'WARNING')
            ok = False
        try:
            tsel = card.locator('select').first
            done = False
            for how in ('label', 'value'):
                try:
                    if how == 'label':
                        await tsel.select_option(label=tier)
                    else:
                        await tsel.select_option(value=tier)
                    done = True
                    break
                except Exception:
                    continue
            if done:
                self.log('    . %s Tier=%s' % (iid, tier), 'INFO')
            else:
                self.log('    ! %s ตั้ง Tier ไม่ได้ (%s)' % (iid, tier), 'WARNING')
                ok = False
        except Exception as e:
            self.log('    ! %s tier error: %s' % (iid, str(e)), 'WARNING')
            ok = False
        return ok

    async def _set_item_rate(self, page, iid, rate):
        """ตั้งค่า "เรทสุ่ม" ของไอเท็ม (การ์ดใน bundle แบบ RANDOM)
        DOM จริง: เรทสุ่ม = <input type=number required placeholder="0.000 - 100.000">
                  เรทโชว์ = <input type=number> (ไม่มี required) -> ต้องเลี่ยง"""
        card = await self._locate_card(page, iid)
        if card is None:
            self.log('    ! %s: หาการ์ดไอเทมไม่เจอ (เรทสุ่ม)' % iid, 'WARNING')
            return False

        async def _try(loc):
            try:
                return loc if await loc.count() else None
            except Exception:
                return None

        # 1) แถวที่ label มีคำว่า "เรทสุ่ม" -> input ในแถวนั้น (แม่นสุด, ไม่โดนเรทโชว์)
        rbox = await _try(card.locator(
            "xpath=.//div[contains(@class,'flex') and .//div[contains(normalize-space(.),'%s')]]"
            "//input" % SEL['rate_label']).first)
        # 2) input number ที่ required (เรทสุ่ม required, เรทโชว์ไม่ required)
        if rbox is None:
            rbox = await _try(card.locator("css=input[type='number'][required]").first)
        # 3) input number ตัวแรก (เรทสุ่มมาก่อนเรทโชว์ใน DOM)
        if rbox is None:
            rbox = await _try(card.locator("css=input[type='number']").first)
        if rbox is None:
            self.log('    ! %s: หาช่องเรทสุ่มไม่เจอ' % iid, 'WARNING')
            return False
        try:
            await rbox.fill('')
            await rbox.fill(str(rate))
            self.log('    . %s เรทสุ่ม=%s' % (iid, rate), 'INFO')
            return True
        except Exception as e:
            self.log('    ! %s ตั้งเรทสุ่มไม่ได้: %s' % (iid, str(e)), 'WARNING')
            return False

    async def _fill_unset_tiers(self, page, tier='Common'):
        selects = page.locator("xpath=//select[option[normalize-space(.)='%s']]" % tier)
        try:
            n = await selects.count()
        except Exception:
            return 0
        done = 0
        for i in range(n):
            s = selects.nth(i)
            try:
                cur = await s.evaluate(
                    "el => ((el.options[el.selectedIndex] && el.options[el.selectedIndex].textContent) || '').trim()")
            except Exception:
                cur = ''
            if (not cur) or ('เลือก' in cur):
                try:
                    await s.select_option(label=tier)
                    done += 1
                except Exception:
                    pass
        return done

    async def _recheck_bundle(self, page, game, name, created_id):
        self.log('  [9] รีเชคด้วยชื่อ bundle: ' + name, 'INFO')
        try:
            await page.goto(game_url(game, 'bundles'), wait_until='domcontentloaded', timeout=30000)
        except Exception as e:
            self.log('  [9] โหลดหน้า bundles ไม่ได้: ' + str(e), 'WARNING')
            return None
        await page.wait_for_timeout(1500)
        box = await first_visible(page, [
            "xpath=//*[contains(text(),'%s')]/following::input[1]" % SEL['bundle_search'],
            "input[placeholder*='Aztek']",
            "input[placeholder*='Bundle']",
        ])
        if not box:
            self.log('  [9] หาช่องค้นหา bundle ไม่เจอ', 'WARNING')
            return None
        try:
            await box.fill('')
            await box.fill(name)
        except Exception as e:
            self.log('  [9] กรอกชื่อค้นหาไม่ได้: ' + str(e), 'WARNING')
            return None
        resp = None
        try:
            async with page.expect_response(
                lambda r: r.request.method in ('GET', 'POST') and 'bundle' in r.url.lower(),
                timeout=12000,
            ) as ri:
                sb = await first_visible(page, ["button:has-text('%s')" % SEL['search_btn']], timeout_each=1500)
                if sb:
                    await sb.click()
                else:
                    await box.press('Enter')
            resp = await ri.value
        except Exception:
            resp = None
        await page.wait_for_timeout(800)
        rid = None
        if resp is not None:
            try:
                data = await resp.json()
                ids = extract_ids_by_name(data, name)
                rid = pick_recheck_id(ids, created_id) or extract_bundle_id(data)
                if len(ids) > 1:
                    self.log('  [9] เจอ bundle ชื่อซ้ำ %d ตัว (%s) -> เลือกเลข %s'
                             % (len(ids), ', '.join(ids), rid), 'WARNING')
                self.log('  [9] search response: ' + json.dumps(data, ensure_ascii=False)[:500], 'INFO')
            except Exception:
                pass
        if rid is None:
            try:
                row = page.locator("xpath=//tr[contains(., %s)]" % json.dumps(name, ensure_ascii=False)).first
                await row.wait_for(state='visible', timeout=4000)
                txt = await row.inner_text()
                nums = re.findall(r'\d+', txt)
                rid = nums[0] if nums else None
            except Exception:
                rid = None
        if rid:
            ok = (created_id is None) or (str(rid) == str(created_id))
            self.log('  [9] รีเชคได้เลข: %s (%s)'
                     % (rid, 'ตรงกับตอนสร้าง' if ok else 'ไม่ตรงกับ ' + str(created_id)),
                     'SUCCESS' if ok else 'WARNING')
        else:
            self.log('  [9] รีเชคไม่ได้เลข (ดู response/หน้าเว็บ)', 'WARNING')
        return rid

    # ---------------- ตารางผลลัพธ์ ----------------
    def add_result(self, name, created_id, recheck_id):
        self.root.after(0, self._add_result_main, name, created_id, recheck_id)

    def _add_result_main(self, name, created_id, recheck_id):
        if created_id and recheck_id:
            status = 'OK' if str(created_id) == str(recheck_id) else 'MISMATCH'
        elif created_id or recheck_id:
            status = 'partial'
        else:
            status = 'no-id'
        self._results.append({'name': name, 'created': created_id or '',
                              'recheck': recheck_id or '', 'status': status})
        # ชื่อเต็ม (ไม่ตัด) — เลื่อนแนวนอนดูจนจบได้ · %-28s = แพดชื่อสั้นให้คอลัมน์ตรงกัน
        line = '%-28s | สร้าง=%-8s | รีเชค=%-8s | %s' % (
            name, created_id or '-', recheck_id or '-', status)
        col = C['accent2'] if status == 'OK' else (C['danger'] if status == 'MISMATCH' else C['warn'])
        self.result_list.insert('end', line)
        self.result_list.itemconfig(self.result_list.size() - 1, foreground=col)
        self.result_list.see('end')
        try:
            self.nb.select(self._tab_result)
        except Exception:
            pass

    def _copy_ids(self):
        if not self._results:
            messagebox.showinfo('ว่าง', 'ยังไม่มีเลข bundle')
            return
        sel = self.result_list.curselection()
        idxs = list(sel) if sel else list(range(len(self._results)))
        ids = []
        for i in idxs:
            if 0 <= i < len(self._results):
                r = self._results[i]
                v = r.get('created') or r.get('recheck')
                if v:
                    ids.append(str(v))
        if not ids:
            messagebox.showinfo('ว่าง', 'แถวที่เลือกไม่มีเลข')
            return
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append('\n'.join(ids))
            self.root.update()
            self.log('คัดลอกเลข bundle %d ตัว (%s)' % (len(ids), 'ที่เลือก' if sel else 'ทั้งหมด'), 'SUCCESS')
        except Exception as e:
            messagebox.showerror('คัดลอกไม่สำเร็จ', str(e))

    def _clear_results(self):
        self._results = []
        self.result_list.delete(0, tk.END)

    def _export_results(self):
        if not self._results:
            messagebox.showinfo('ว่าง', 'ยังไม่มีข้อมูล')
            return
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(defaultextension='.csv', initialfile='bundle_ids.csv',
                                            filetypes=[('CSV', '*.csv')])
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f)
                w.writerow(['bundle_name', 'created_id', 'recheck_id', 'status'])
                for r in self._results:
                    w.writerow([r['name'], r['created'], r['recheck'], r['status']])
            self.log('บันทึกผล: ' + os.path.basename(path), 'SUCCESS')
        except Exception as e:
            messagebox.showerror('บันทึกไม่สำเร็จ', str(e))


def main():
    try:
        root = tk.Tk()
        root.geometry('620x720')
        App(root)
        root.mainloop()
    except Exception:
        err = traceback.format_exc()
        try:
            messagebox.showerror('Startup Error', err)
        except Exception:
            print(err)


# ทะเบียนสำหรับ All for Cabal launcher
from tool_registry import ToolSpec
TOOL = ToolSpec(
    key='bundle', icon='📦', title='Create Bundle', nav='📦  Create Bundle',
    desc='สร้าง Bundle อัตโนมัติบนเว็บ aztek-tools', boot='success',
    make=lambda lc, fr: App(lc.root, container=fr, game_var=lc.game_var,
                            on_bundle_created=lc._on_bundle_created,
                            on_go_next=lc._go_next))


if __name__ == '__main__':
    main()
