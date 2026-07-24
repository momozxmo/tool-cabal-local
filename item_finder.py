"""item_finder.py — Aztek Item Finder
Reconstructed (clone) ของ "Cabal Item ID Finder.exe"

- ดึงรหัสไอเทม (Aztek ID) จากเว็บ combo-interactive / aztek-tools
- รองรับ CabalM / CabalPC (TH/SEA)
- ค้นชิ้นเดียว หรือหลายชิ้นจาก template (.xlsx/.csv)
- Deep Check: web / img / qty / trade / socket(drill) / crit
- Export เป็น .xlsx / .csv

ปรับปรุงจากต้นฉบับ:
  * web=No short-circuit — ถ้า criterion ของคอลัมน์ web เป็น "no"
    เมื่อ item ผ่านการเช็ค web แล้ว จะหยุดทันที ไม่อ่าน field ที่เหลือ
    (img/qty/trade/socket/crit) เพื่อความเร็ว
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog, font as tkfont
import threading
import asyncio
import traceback
import os
import sys
import json
import csv
from datetime import datetime

try:
    from playwright.async_api import async_playwright
    PW_OK = True
except ImportError:
    PW_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# ธีม/ฟอนต์/ไอคอน จากของกลาง (re-export ชื่อเดิมไว้ให้ launcher/ส่วนอื่นใช้ต่อได้)
from ui_common import C, FM, FB, F9, _find_icon, _set_window_icon
import aztek_core as core
import finder_core

# ---------------------------------------------------------------------------
# เกม/เซิร์ฟ + Chrome + prefs ใช้จาก aztek_core (แหล่งเดียว) — GAMES = URL หน้า items
GAMES = {g: core.game_url(g, 'items') for g in core.GAME_NAMES}
GAME_NAMES = core.GAME_NAMES

TMPL_HEADERS = ['ItemKind', 'itemOption', 'durationIndex', 'ItemName',
                'web', 'img', 'qty', 'trade', 'socket', 'crit']


def _norm_name(s):
    """normalize ชื่อไอเทมเพื่อเทียบแบบยืดหยุ่น (ไม่สนตัวพิมพ์/ช่องว่าง)"""
    return ''.join((s or '').lower().split())


_get_app_dir = core._get_app_dir
find_chrome_exe = core.find_chrome_exe
# login ร่วมกับ tool อื่น: Chrome profile + prefs 'game' ใช้ไฟล์เดียวกับ core
CHROME_PROFILE = core.CHROME_PROFILE
load_prefs = core.load_prefs
save_prefs = core.save_prefs


# ---------------------------------------------------------------------------
def download_template(path):
    """สร้าง xlsx template สำหรับค้นหาหลายชิ้น พร้อมคอลัมน์ Deep Check + Dropdown"""
    search_cols = {'ItemKind', 'itemOption', 'durationIndex', 'ItemName'}
    bool_cols = {'img', 'web', 'trade', 'socket'}
    col_widths = [14, 14, 16, 28, 9, 9, 10, 9, 10, 12]
    NCOLS = len(TMPL_HEADERS)
    last_col_letter = chr(64 + NCOLS)

    if XLSX_OK:
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Items'
        blue_fill = PatternFill('solid', fgColor='1F6FEB')
        orange_fill = PatternFill('solid', fgColor='D29922')
        wht_font = Font(color='FFFFFF', bold=True)
        blk_font = Font(color='0D1117', bold=True)
        note_gray = PatternFill('solid', fgColor='21262D')
        note_lines = [
            'คำแนะนำการกรอกข้อมูล  |  web / img / trade / socket : Any, Yes, No  |  qty / crit : เว้นว่าง=Any  หรือใส่ค่า exact match',
            'สีน้ำเงิน = ตัวกรองค้นหา (ItemKind / itemOption / durationIndex)     สีส้ม = Deep Check (web / img / qty / trade / socket / crit)',
            'ItemName : ไม่ได้ใช้ค้นหา — ใส่ไว้เทียบกับชื่อในเว็บเฉย ๆ (ไม่ตรงจะขึ้น ≠ ให้ตรวจ)  |  กรอก Item ตั้งแต่ Row 5 ลงมา (อย่าลบ Row 1-4)',
        ]
        for r, line in enumerate(note_lines, 1):
            c = ws.cell(row=r, column=1, value=line)
            c.font = Font(italic=True, color='8B949E', size=9)
            c.fill = note_gray
            ws.merge_cells(f'A{r}:{last_col_letter}{r}')

        bool_col_letters = []
        for col, h in enumerate(TMPL_HEADERS, 1):
            cell = ws.cell(row=4, column=col, value=h)
            is_search = h in search_cols
            cell.fill = blue_fill if is_search else orange_fill
            cell.font = wht_font if is_search else blk_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = (
                col_widths[col - 1] if col <= len(col_widths) else 12)
            if h in bool_cols:
                bool_col_letters.append(cell.column_letter)

        ws.append(['4101', '100', '0', 'HP Potion', 'Yes', 'Any', '100', 'Yes', 'Any', ''])
        ws.append(['4104', '100', '0', '', 'Yes', 'Any', '', 'Yes', 'Any', ''])

        for letter in bool_col_letters:
            ws.add_data_validation(DataValidation(
                type='list', formula1='"Any,Yes,No"', allow_blank=True,
                showDropDown=False, sqref=f'{letter}5:{letter}1000'))
        wb.save(path)
        return

    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(TMPL_HEADERS)
        w.writerow(['4101', '100', '0', 'HP Potion', 'Yes', 'Any', '100', 'Yes', 'Any', ''])
        w.writerow(['4104', '100', '0', '', 'Yes', 'Any', '', 'Yes', 'Any', ''])


def read_template(path):
    """อ่าน xlsx หรือ csv template คืน list of dicts (search + deep check fields)"""
    ext = os.path.splitext(path)[1].lower()
    rows = []

    def parse_row(g):
        """g(col_name) -> str value"""
        def norm_bool(v):
            v = v.strip().lower()
            if v in ('yes', 'y', '1', 'true'):
                return 'yes'
            if v in ('no', 'n', '0', 'false'):
                return 'no'
            return 'any'
        return {
            'kind': g('ItemKind'),
            'opt': g('itemOption'),
            'dur': g('durationIndex'),
            'name': g('ItemName'),
            'web': norm_bool(g('web')),
            'img': norm_bool(g('img')),
            'qty_val': g('qty'),
            'trade': norm_bool(g('trade')),
            'drill': norm_bool(g('socket') or g('drill')),
            'crit_val': g('crit'),
        }

    if ext in ('.xlsx', '.xlsm') and XLSX_OK:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row = 1
        for r in ws.iter_rows():
            vals = [str(c.value or '').strip() for c in r]
            if 'ItemKind' in vals:
                header_row = r[0].row
                break
        headers = [str(c.value or '').strip() for c in ws[header_row]]
        idx = {h: i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):
                continue
            raw = [str(v) if v is not None else '' for v in row]

            def g(k, r=raw, ix=idx):
                if k in ix and ix[k] < len(r):
                    return r[ix[k]].strip()
                return ''

            d = parse_row(g)
            if d['kind'] or d['opt'] or d['dur'] or d['name']:
                rows.append(d)
        wb.close()
        return rows

    with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
        reader = csv.DictReader(f)
        for row in reader:
            def g(k, r=row):
                return str(r.get(k, '') or '').strip()
            d = parse_row(g)
            if d['kind'] or d['opt'] or d['dur'] or d['name']:
                rows.append(d)
    return rows


# ---------------------------------------------------------------------------
# Import ตาราง "Event / Prize" (ITEM CODE .xlsx) — header อยู่กลาง sheet, มีได้หลายตาราง/sheet
# ชื่อไอเทมมีได้ 2 ชื่อคอลัมน์: "Item Name" (ให้ priority) หรือ "Display Name"
# บางไฟล์ "Display Name" เป็นชื่อข่าว ไม่ใช่ชื่อไอเทม -> ใช้ "Item Name" ก่อน
_EVENT_HDR = {
    'itemkind': 'kind', 'itemoption': 'opt', 'durationindex': 'dur',
    'itemname': 'name', 'displayname': 'name', 'amt': 'amt', 'amount': 'amt',
    'itemindex': 'index',
}


def _event_num(v):
    """'33559707.0' -> '33559707' ; เว้นว่างถ้า None ; datetime -> Excel serial
    (ItemKind บางตัวถูก Excel format เป็นวันที่)"""
    if v is None:
        return ''
    import datetime as _dt
    if isinstance(v, _dt.date):                    # datetime เป็น subclass ของ date
        try:
            d = v.date() if isinstance(v, _dt.datetime) else v
            return str((d - _dt.date(1899, 12, 30)).days)
        except Exception:
            return ''
    s = str(v).strip()
    if not s:
        return ''
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return ('%f' % f).rstrip('0').rstrip('.')
    except Exception:
        return s


def _event_norm(x):
    return str(x).strip().lower().replace(' ', '').replace('\n', '') if x is not None else ''


def _event_isnum(x):
    try:
        float(str(x).strip())
        return True
    except Exception:
        return False


# ข้อความ generic ที่ไม่เอามาเป็นชื่อกลุ่ม — ข้ามแล้วไล่ขึ้นไปหาชื่อจริง
# (เช่น "Participation" / "Fastest Guild" / "Lucky Winner" ที่อยู่บนสุดของ section)
_EVENT_SKIP_LABEL = set(_EVENT_HDR.keys()) | {'stackable', 'itemmove'}
_EVENT_GENERIC_PREFIX = ('prize', 'conditions', 'codeexpiredate', 'additionalconditions',
                         'number', 'uniquecode', 'mastercode', 'deliverystatus',
                         'activitydetails', 'total', 'applicantfor', 'linkfile', 'no.')


def _event_is_generic(v):
    n = _event_norm(v)
    if not n or n in _EVENT_SKIP_LABEL:
        return True
    return any(n.startswith(p) for p in _EVENT_GENERIC_PREFIX)


_EVENT_LEGEND_COL = 17   # คอลัมน์ >= นี้ = ข้อความอธิบาย template (มี Unique/Master Code ปนกัน) -> ข้าม


def _event_extract_meta(buf):
    """สแกนแถว "บล็อกเงื่อนไข" ที่อยู่เหนือ header ของตาราง Prize ดึงค่า (เฉพาะคอลัมน์ข้อมูลจริง
    < %d ข้ามคอลัมน์ legend ที่มีคำอธิบายทั้ง Unique/Master Code):
    - activity      : ชื่อกิจกรรม (แบนเนอร์ซ้ายบนสุด คอลัมน์ 0 เช่น 'VENI')
    - reward        : ชื่อรางวัล (ค่าใต้หัวคอลัมน์ 'Reward')
    - expire        : วันหมดอายุ (CODE EXPIRE DATE เช่น '2026-08-31 00:00:00' หรือ '31 Aug')
    - codes_per_set : Number [codes per set] (เช่น '750')
    - cannot_repeat : เจอ 'Cannot be repeated' / 'เติมซ้ำไม่ได้'
    - unique_code   : เจอ 'Unique Code' ในบล็อก Prize -> ใช้ Server Generate
    - conditions    : ข้อความ Additional conditions (ไทย/อังกฤษ) ไว้อ้างอิง""" % _EVENT_LEGEND_COL
    meta = {'activity': '', 'reward': '', 'code_no': '', 'expire': '',
            'codes_per_set': '', 'set_count': '', 'total': '',
            'cannot_repeat': False, 'unique_code': False, 'conditions': [],
            'cross_cannot': False, 'cross_can': False,
            'once_per_set': False, 'can_repeat': False}
    n = len(buf)
    for ri, row in enumerate(buf):
        for ci, cell in enumerate(row):
            if ci >= _EVENT_LEGEND_COL:
                continue
            key = _event_norm(cell)
            if not key:
                continue
            raw = str(cell).strip()

            def right():
                for cj in range(ci + 1, min(len(row), _EVENT_LEGEND_COL)):
                    v = row[cj]
                    if v is not None and str(v).strip():
                        return str(v).strip()
                return ''

            def below():
                if ri + 1 < n and ci < len(buf[ri + 1]):
                    v = buf[ri + 1][ci]
                    if v is not None and str(v).strip():
                        return str(v).strip()
                return ''

            # เลข Code จากแบนเนอร์ CONDITIONS/Prize เช่น '... "Code No. 1" ...' -> 'Code No. 1'
            # (ใช้เป็นชื่อรางวัล fallback เมื่อไม่มีแบนเนอร์ section เช่นไฟล์ VENI)
            if not meta['code_no'] and (key.startswith('conditions') or key.startswith('prize')):
                import re as _re
                mm = (_re.search(r'code\s*no\.?\s*(\d+)', raw, _re.I)
                      or _re.search(r'code\s*(\d+)', raw, _re.I))
                if mm:
                    meta['code_no'] = 'Code No. %s' % mm.group(1)
            if key.startswith('codeexpiredate') or 'วันหมดอายุ' in key:
                meta['expire'] = right() or below() or meta['expire']
            # จำนวนโค้ด/ชุด — EN 'Number [codes per set]' / TH 'จำนวน [โค้ด/ชุด]'
            if 'codesperset' in key or ('โค้ด' in key and 'ชุด' in key):
                for cand in (right(), below()):     # เลือกค่าที่เป็น "ตัวเลข" (ข้ามหัวคอลัมน์อื่น)
                    if _event_isnum(cand):
                        meta['codes_per_set'] = _event_num(cand)
                        break
            # จำนวนชุด — EN 'Number [set]' / TH 'จำนวน [ชุด]' (มี 'ชุด' แต่ไม่มี 'โค้ด')
            elif key == 'number[set]' or ('จำนวน' in key and 'ชุด' in key and 'โค้ด' not in key):
                for cand in (right(), below()):
                    if _event_isnum(cand):
                        meta['set_count'] = _event_num(cand)
                        break
            if key == 'total' or 'รวม' in key:      # จำนวนโค้ดทั้งหมด — EN 'Total' / TH 'รวม'
                for cand in (right(), below()):
                    if _event_isnum(cand):
                        meta['total'] = _event_num(cand)
                        break
            if key.startswith('uniquecode'):
                meta['unique_code'] = True
            low = raw.lower()
            if 'cannotberepeat' in key or 'เติมซ้ำไม่ได้' in raw or 'เติมซํ้าไม่ได้' in raw:
                meta['cannot_repeat'] = True
            # เงื่อนไขตัดสินว่า "เจนรวม" หรือ "หลายชุด" (EN + TH)
            if 'even different' in low or 'ข้ามเซ็ตไม่ได้' in raw or 'ข้ามไม่ได้' in raw:
                meta['cross_cannot'] = True         # ไม่ซ้ำข้ามเซ็ต -> เจนรวม
            if 'ข้ามได้' in raw:
                meta['cross_can'] = True            # เติมข้ามได้ -> หลายชุด
            if 'once per set' in low or 'ครั้งเดียวต่อชุด' in raw:
                meta['once_per_set'] = True         # ใช้ครั้งเดียวต่อชุด -> หลายชุด
            if 'can be repeated' in low or 'เติมซ้ำได้' in raw:
                meta['can_repeat'] = True           # เติมซ้ำได้ -> เจนรวม
            if key.startswith('additionalconditions') or 'เงื่อนไข' in key:
                v = right()
                if v:
                    meta['conditions'].append(v)
    return meta


def _parse_event_rows(rows, skipped=None):
    """rows = iterable ของ tuple (values_only) จาก 1 sheet
    หา header row ('Item Kind'/'ItemKind' ...) แล้วดึงแถวข้อมูล รองรับหลายตาราง/sheet
    + จับชื่อกลุ่ม/ตาราง คืน list ของ finder-criteria dict (แต่ละตัวมี key 'group')
    แถวที่มีชื่อไอเทมแต่ ItemKind อ่านไม่ได้ (#VALUE! จาก date-format) -> เก็บชื่อลง skipped"""
    items = []
    col = None
    group = ''
    cur_meta = {}
    tbl = 0
    buf = []
    sheet_activity = ''                            # ชื่อกิจกรรมระดับชีต (เช่น VICI/VENI) เก็บครั้งเดียว
    for row in rows:
        cn = [_event_norm(c) for c in row]
        # ชื่อกิจกรรม = ข้อความคอลัมน์ 0 ตัวแรกที่ไม่ใช่ generic (แบนเนอร์บนสุดของชีต) เก็บครั้งเดียว
        if not sheet_activity and row and row[0] is not None:
            v0 = str(row[0]).strip()
            if v0 and not _event_isnum(v0) and not _event_is_generic(v0):
                sheet_activity = v0.replace('\n', ' ')
        if 'itemkind' in cn:                       # เจอ header -> เริ่มตารางใหม่
            col = {}
            for i, c in enumerate(cn):
                if c == 'itemname':                # ชื่อจริง — ทับ Display Name เสมอ
                    col['name'] = i
                elif c in _EVENT_HDR and _EVENT_HDR[c] != 'name':
                    col[_EVENT_HDR[c]] = i
                elif c == 'displayname' and 'name' not in col:
                    col['name'] = i                # fallback ถ้าไม่มี Item Name
            tbl += 1
            banner = ''                            # แบนเนอร์ section ในคอลัมน์ Item Kind (เช่น 'GM Slayer')
            kc = col.get('kind')
            if kc is not None:
                # ไล่จากใกล้ header ขึ้นไป ข้ามข้อความ generic (Prize/Conditions/...)
                # จนเจอชื่อ section จริง (แบนเนอร์บนสุดของ section เช่น GM Slayer)
                for prev in reversed(buf):
                    if kc < len(prev):
                        v = prev[kc]
                        if v is not None and str(v).strip() and not _event_isnum(v) \
                                and not _event_is_generic(v):
                            banner = str(v).strip().replace('\n', ' ')
                            break
            # ดึงเงื่อนไข (วันหมดอายุ/codes per set/เติมซ้ำไม่ได้/unique code/เลข Code)
            cur_meta = _event_extract_meta(buf)
            cur_meta['activity'] = sheet_activity
            # ชื่อรางวัล = แบนเนอร์ section ถ้ามี ไม่งั้น fallback เป็นเลข Code จาก CONDITIONS/Prize
            reward = banner or cur_meta.get('code_no', '')
            cur_meta['reward'] = reward
            # ชื่อกลุ่ม (แสดง/จัดกลุ่มผล) = 'ชื่อกิจกรรม ชื่อรางวัล'
            disp = (sheet_activity + ' ' + reward).strip() if reward else sheet_activity
            group = disp or banner or ('ตาราง %d' % tbl)
            cur_meta['event_name'] = group
            buf.append(row)
            buf[:] = buf[-16:]
            continue
        buf.append(row)
        buf[:] = buf[-16:]
        if not col or 'kind' not in col:
            continue
        k = col['kind']
        kraw = row[k] if k < len(row) else None
        kind = _event_num(kraw)

        def get(key, r=row, cc=col):
            j = cc.get(key)
            return r[j] if (j is not None and j < len(r)) else None
        nm = get('name')
        name_filled = nm is not None and str(nm).strip() != ''

        if kind and kind.isdigit():
            items.append({
                'kind': kind,
                'opt': _event_num(get('opt')),
                'dur': _event_num(get('dur')),
                'name': str(nm).strip() if name_filled else '',
                'amt': _event_num(get('amt')),
                'group': group,
                'group_meta': cur_meta,
                'web': 'any', 'img': 'any', 'qty_val': '',
                'trade': 'any', 'drill': 'any', 'crit_val': '',
            })
        elif name_filled and kraw is not None and str(kraw).strip() != '':
            # แถวมีชื่อไอเทมแต่ ItemKind อ่านไม่ได้ (#VALUE! จาก date-format) -> ข้าม+เตือน (ไม่จบตาราง)
            if skipped is not None:
                skipped.append(str(nm).strip())
        else:
            col = None                             # แถวว่าง/หมดตาราง -> รอ header ถัดไป
    return items


_SHOP_HDR = {
    'itemkind': 'kind', 'itemoption': 'opt', 'durationindex': 'dur',
    'itemname': 'name', 'itemindex': 'index', 'amt': 'amt', 'amount': 'amt',
    'itemmove': 'move', 'stackable': 'stack',
}


def _shop_sheet_items(rows, sheet_title, skipped=None):
    """template Shop (Cash Shop / Promotion / In Game) 1 sheet -> finder-format items
    เลย์เอาต์ต่างกันแต่ละชีต (Cash Shop: Itemmove อยู่คอลัมน์ J, Promotion: G)
    -> จับจากชื่อหัวคอลัมน์ ไม่ล็อกตำแหน่ง
    Itemmove = Yes -> ตั้ง deep param trade='yes' = ช่อง 'แลกเปลี่ยนได้' (is_tradable) บนเว็บ"""
    items = []
    col = None
    group = ''
    tbl = 0
    buf = []
    for row in rows:
        cn = [_event_norm(c) for c in row]
        if 'itemkind' in cn:                       # หัวตาราง -> เริ่มตารางใหม่
            col = {}
            for i, c in enumerate(cn):
                if c in _SHOP_HDR:
                    key = _SHOP_HDR[c]
                    if key == 'name' and 'name' in col:
                        continue
                    col[key] = i
            tbl += 1
            # ชื่อสินค้า = ป้าย 'Product Name' เหนือหัวตาราง (ค่าอยู่ช่องถัดไปในแถวเดียวกัน)
            group = ''
            for prev in reversed(buf):
                for i, v in enumerate(prev):
                    if _event_norm(v) == 'productname':
                        for j in range(i + 1, len(prev)):
                            if prev[j] is not None and str(prev[j]).strip():
                                group = str(prev[j]).strip().replace('\n', ' ')
                                break
                        break
                if group:
                    break
            if not group:
                group = '%s · ตาราง %d' % (sheet_title, tbl)
            buf.append(row)
            buf[:] = buf[-16:]
            continue
        buf.append(row)
        buf[:] = buf[-16:]
        if not col or 'kind' not in col:
            continue

        def get(key, r=row, cc=col):
            j = cc.get(key)
            return r[j] if (j is not None and j < len(r)) else None
        kraw = get('kind')
        kind = _event_num(kraw)
        nm = get('name')
        name_filled = nm is not None and str(nm).strip() != ''
        if kind and kind.isdigit():
            mv = _event_norm(get('move'))
            tradable = mv in ('yes', 'y', 'true', '1')
            # Itemmove = Yes -> ต้องติ๊ก 'แลกเปลี่ยนได้' (is_tradable) อย่างเดียว
            #   ช่องที่เหลือในกล่อง 'พารามิเตอร์สำหรับแสดงบนเว็บ' ต้องว่าง/ไม่ติ๊ก
            #   (ไม่ติ๊กแลกเปลี่ยนได้ = ขึ้นเป็น 'ผูกมัดไอดี')
            # Itemmove = No/ว่าง -> ไม่ตรวจอะไร (ช่องพวกนี้ไม่โผล่เพราะปิดแสดงผลบนเว็บอยู่)
            items.append({
                'kind': kind,
                'opt': _event_num(get('opt')),
                'dur': _event_num(get('dur')),
                'name': str(nm).strip() if name_filled else '',
                'amt': _event_num(get('amt')),
                'group': group,
                'group_meta': {'is_shop': True, 'shop_sheet': sheet_title, 'product': group},
                # โหมด Shop: ทุกตัวต้องมีรูปภาพไอเท็ม (ช่องรูปอยู่นอกกล่อง 'พารามิเตอร์แสดงบนเว็บ'
                # -> อ่านได้ไม่ว่าจะเปิด/ปิดการแสดงผลบนเว็บ)
                'web': 'any', 'img': 'yes',
                'trade': 'yes' if tradable else 'any',
                'drill': 'no' if tradable else 'any',
                'qty_val': MUST_EMPTY if tradable else '',
                'crit_val': MUST_EMPTY if tradable else '',
            })
        elif name_filled and kraw is not None and str(kraw).strip() != '':
            if skipped is not None:
                skipped.append(str(nm).strip())
        else:
            col = None
    return items


def parse_shop_workbook(path):
    """อ่านไฟล์ Monthly Plan (โหมด Shop) -> (out, skipped) รูปแบบเดียวกับ parse_event_workbook
    อ่านทุกชีตที่มีตารางไอเทม (Cash Shop / Promotion / In Game / ...)"""
    import warnings
    import openpyxl
    _WSP = _orig_init = _patched_init = None
    try:
        import openpyxl.worksheet._reader as _wsr
        _WSP = _wsr.WorkSheetParser
        _orig_init = _WSP.__init__

        def _patched_init(self, *a, **k):        # ItemKind ที่ถูก format เป็นวันที่ -> เอาเลขดิบ
            _orig_init(self, *a, **k)
            try:
                self.date_formats = set()
                self.timedelta_formats = set()
            except Exception:
                pass
    except Exception:
        _WSP = None
    out, skipped = [], []
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        if _WSP is not None:
            _WSP.__init__ = _patched_init
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for sh in wb.sheetnames:
                    rows = list(wb[sh].iter_rows(values_only=True))
                    items = _shop_sheet_items(rows, sh, skipped)
                    for it in items:
                        grp = it.pop('group', '')
                        it['sources'] = [grp] if grp else ['(ไม่มีชื่อสินค้า)']
                    if items:
                        out.append((sh, items))
            finally:
                wb.close()
        finally:
            if _WSP is not None:
                _WSP.__init__ = _orig_init
    return out, skipped


def parse_event_workbook(path):
    """อ่าน .xlsx แล้วคืน (out, skipped)
    out = list ของ (sheet_name, [items]) เฉพาะ sheet ที่มีไอเทม
          แต่ละ item มี 'sources' = ['<group>'] (เช่น Participation, Fastest Guild)
    skipped = list ชื่อไอเทมที่ ItemKind อ่านไม่ได้ (Excel format เป็นวันที่)"""
    import warnings
    import openpyxl

    # template 'Pride Code Request' โครงสร้างคนละแบบ (หลายบล็อกโค้ดต่อชีท)
    # -> ใช้ตัวอ่านเฉพาะ เหมือนโหมด Event เพื่อให้ Import ไฟล์เดียวจบทั้ง 2 โหมด
    try:
        from event_tool import _pride_workbook_items
        pride = _pride_workbook_items(path)
    except Exception:
        pride = None
    if pride is not None:
        return pride, []

    # patch openpyxl: อย่าแปลงเลขที่ cell format เป็น "วันที่" ให้เป็น datetime
    # (ItemKind บาง cell ถูก format เป็นวันที่ -> เดิมกลายเป็น #VALUE! ค่าหาย)
    # ทำให้ date_formats ว่าง -> คืนเลขดิบ (เช่น 50337837) กู้ค่าคืนได้
    _WSP = _orig_init = _patched_init = None
    try:
        import openpyxl.worksheet._reader as _wsr
        _WSP = _wsr.WorkSheetParser
        _orig_init = _WSP.__init__

        def _patched_init(self, *a, **k):
            _orig_init(self, *a, **k)
            try:
                self.date_formats = set()
                self.timedelta_formats = set()
            except Exception:
                pass
    except Exception:
        _WSP = None

    out = []
    skipped = []
    with warnings.catch_warnings():                # กัน UserWarning เรื่อง date serial
        warnings.simplefilter('ignore')            # (read_only อ่านตอน iterate)
        if _WSP is not None:
            _WSP.__init__ = _patched_init
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for sh in wb.sheetnames:
                    items = _parse_event_rows(wb[sh].iter_rows(values_only=True), skipped)
                    if items:
                        for it in items:
                            grp = it.pop('group', '')
                            it['sources'] = [grp] if grp else ['(ไม่มีชื่อกลุ่ม)']
                        out.append((sh, items))
            finally:
                wb.close()
        finally:
            if _WSP is not None:
                _WSP.__init__ = _orig_init          # คืน patch เสมอ
    return out, skipped


# ---------------------------------------------------------------------------
# Theme (C/FM/FB/F9 import จาก ui_common ข้างบน)
def _E(p, **k):
    k.setdefault('bg', C['bg_inp'])
    k.setdefault('fg', C['text'])
    k.setdefault('insertbackground', C['text'])
    k.setdefault('relief', 'flat')
    k.setdefault('font', FM)
    return tk.Entry(p, **k)


def _L(p, **k):
    k.setdefault('bg', C['bg_med'])
    k.setdefault('fg', C['text'])
    k.setdefault('font', FM)
    return tk.Label(p, **k)


def _LF(parent, title):
    return tk.LabelFrame(parent, text=title, bg=C['bg_med'], fg=C['muted'],
                         font=FM, bd=1, relief='solid')


DEEP_OPTS = ['any', 'yes', 'no']

# ค่าพิเศษของช่อง จำนวน/คริติคอล: "ต้องไม่มีค่าในช่องนั้น"
# (ช่องว่างธรรมดา = ไม่ตรวจ ซึ่งคนละความหมายกัน)
MUST_EMPTY = '(ต้องว่าง)'

# ป้ายพารามิเตอร์ Deep Check — ใช้คำสั้น ๆ จะได้เหลือที่ให้ชื่อไอเทม (มีคำอธิบายกำกับข้างบน)
_DEEP_BOOL = (('web', 'เว็บ'), ('img', 'รูป'), ('trade', 'เทรด'), ('drill', 'เจาะ'))
_DEEP_VAL = (('qty_val', 'จำนวน'), ('crit_val', 'คริต'))
DEEP_LEGEND = ('✓=ต้องใช่  ✗=ต้องไม่  ∅=ต้องว่าง   |   '
               'เว็บ=แสดงผลบนเว็บ  เทรด=แลกเปลี่ยนได้  เจาะ=Socket  คริต=คริติคอล')


def deep_summary(r):
    """สรุปว่าไอเทมนี้จะถูก 'เช็คพารามิเตอร์อะไรบ้าง' แบบย่อ ('-' = ไม่เช็คอะไรเลย)
    เช่น Shop ไอเทมเทรดได้ -> 'เว็บ✓ เทรด✓ เจาะ✗ จำนวน∅ คริต∅'  (ดูคำอธิบายที่ DEEP_LEGEND)"""
    parts = []
    for k, lab in _DEEP_BOOL:
        v = r.get(k, 'any')
        if v in ('yes', 'no'):
            parts.append(lab + ('✓' if v == 'yes' else '✗'))
    for k, lab in _DEEP_VAL:
        v = str(r.get(k, '') or '').strip()
        if v:
            parts.append(lab + ('∅' if v == MUST_EMPTY else '=' + v))
    return ' '.join(parts) if parts else '-'


def _fmt_web_val(v):
    """แปลงค่าจริงที่อ่านจากเว็บให้อ่านง่าย: None/ว่าง = '-', bool = 'Y'/'N', อื่น ๆ = ค่าเดิม"""
    if v is None:
        return '-'
    if isinstance(v, bool):
        return 'Y' if v else 'N'
    s = str(v).strip()
    return s if s else '-'


def web_values_summary(item_data, detail):
    """โหมด 'มี': โชว์ค่าจริงบนเว็บของช่องที่ 'ไม่ได้เช็ค' (Excel ไม่ได้สั่ง) เพื่อให้เห็นครบทุกตัว
    ข้ามช่องที่เช็คอยู่แล้ว (deep_summary โชว์ให้แล้ว) กันซ้ำ -> คืน '' ถ้าไม่มีอะไรต้องโชว์"""
    extra = []
    if item_data.get('trade', 'any') == 'any':
        extra.append('เทรด=' + _fmt_web_val(detail.get('tradeable')))
    if item_data.get('drill', 'any') == 'any':
        extra.append('เจาะ=' + _fmt_web_val(detail.get('drillable')))
    if not str(item_data.get('qty_val', '') or '').strip():
        extra.append('จำนวน=' + _fmt_web_val(detail.get('qty')))
    if not str(item_data.get('crit_val', '') or '').strip():
        extra.append('คริต=' + _fmt_web_val(detail.get('critVal')))
    return ' '.join(extra)


class App:

    def __init__(self, root, container=None, on_send_to_bundle=None, game_var=None,
                 on_queue_bundles=None):
        self.root = root
        # container = frame ที่จะฝัง UI ลงไป (ถ้า None = สร้างลง root โดยตรง = โหมดรันเดี่ยว)
        self.container = container if container is not None else root
        self._embedded = container is not None    # เปิดผ่าน launcher -> ซ่อน เกม/Login ของตัวเอง (ใช้แถบบน)
        # callback รับ list[{'id','name'}] เพื่อส่งไอเทมไปเครื่องมือ Create Bundle
        # (มีเฉพาะตอนรันในโปรแกรมรวม — โหมดเดี่ยวเป็น None ปุ่มจะไม่โผล่)
        self._on_send_to_bundle = on_send_to_bundle
        # callback รับ list[{'name','group','items':[{'id','name'}]}] -> รวมผลค้นหาเป็นบันเดิลรายกลุ่ม
        # แล้วเก็บเข้าคิว Create Bundle โดยตรง (ผ่านหน้ารีวิว: แก้ชื่อ/ตัดไอเทมก่อนส่ง)
        self._on_queue_bundles = on_queue_bundles
        self._event_plans = []          # คิว Event ที่ parse จาก Monthly Plan (โหมด Event) ให้ Create Event ดึงไปเติม
        # game_var = StringVar เกม/เซิร์ฟที่แชร์ร่วมกับ tool อื่น (เลือกที่ไหนก็เปลี่ยนพร้อมกัน)
        self._game_var = game_var
        self._shared_game = game_var is not None
        if container is None:
            self.root.title('Aztek Item Finder')
            self.root.configure(bg=C['bg_dark'])
            self.root.minsize(680, 720)
            _set_window_icon(self.root)
        self._running = False
        self._cancel = False
        self._results = []
        self._imported = []
        self._occurrences = []          # ทุก occurrence (group+ลำดับเอกสาร) ก่อน dedup — ใช้เรียงผลตามชุด
        self._event_group_meta = {}     # group(ชื่อกิจกรรม/Code) -> เงื่อนไข (expire/codes/…) สำหรับ Item Code
        self._not_found = []
        self._build_ui()
        self._load_prefs()

    # ---- UI scaffolding -----------------------------------------------------
    def _build_ui(self):
        hdr = tk.Frame(self.container, bg=C['bg_med'], height=48)
        hdr.pack(fill='x')
        tk.Label(hdr, text='Aztek Item Finder', bg=C['bg_med'], fg=C['text'],
                 font=('Segoe UI', 13, 'bold')).pack(side='left', padx=14, pady=10)
        tk.Label(hdr, text='combo-interactive', bg=C['bg_med'], fg=C['teal'],
                 font=FM).pack(side='left', pady=10)

        s = ttk.Style()
        s.theme_use('clam')
        s.configure('TNotebook', background=C['bg_dark'], borderwidth=0)
        s.configure('TNotebook.Tab', background=C['bg_card'], foreground=C['muted'],
                    padding=[12, 5], font=FM)
        s.map('TNotebook.Tab',
              background=[('selected', C['bg_med'])],
              foreground=[('selected', C['text'])])
        s.configure('green.Horizontal.TProgressbar', troughcolor=C['bg_card'],
                    background=C['accent2'], thickness=8)
        s.configure('TCombobox', fieldbackground=C['bg_inp'], background=C['bg_card'],
                    foreground=C['text'], arrowcolor=C['text'])

        self.nb = ttk.Notebook(self.container)
        self.nb.pack(fill='both', expand=True, padx=8, pady=8)
        self.tm = tk.Frame(self.nb, bg=C['bg_dark'])
        self.tr = tk.Frame(self.nb, bg=C['bg_dark'])
        self.tl = tk.Frame(self.nb, bg=C['bg_dark'])
        self.nb.add(self.tm, text='  ค้นหา  ')
        self.nb.add(self.tr, text='  ผลลัพธ์  ')
        self.nb.add(self.tl, text='  Log  ')
        self._init_search_vars()   # ตัวแปรค้นหา/Deep Check (เดิมอยู่แท็บ 'ค้นหาชิ้นเดียว' ที่เอา UI ออกแล้ว)
        self._multi_tab()
        self._results_tab()
        self._log_tab()

    def _init_search_vars(self):
        """ตัวแปรค้นหา/Deep Check (เดิมอยู่แท็บ 'ค้นหาชิ้นเดียว' ที่เอา UI ออกแล้ว)
        โหมดหลายชิ้นใช้ค่าพวกนี้: vgame (เกม), vbatch/vhdl (มี UI ในแท็บหลายชิ้น),
        และ global deep (vweb/vimg/…) เป็นค่า default กลาง — ค่าจริงของแต่ละไอเทมมาจาก template แล้ว override ทับ"""
        self.vgame = self._game_var if self._game_var is not None else tk.StringVar(value='CabalPC SEA')
        # global Deep Check — คงไว้เป็น default (หลายชิ้น override รายไอเทมจาก template)
        self.vdeep = tk.BooleanVar(value=False)
        self.vweb = tk.StringVar(value='any')
        self.vimg = tk.StringVar(value='any')
        self.vqty_val = tk.StringVar()
        self.vtrade = tk.StringVar(value='any')
        self.vdrill = tk.StringVar(value='any')
        self.vcrit_val = tk.StringVar()
        # ตั้งค่าประมวลผล (มี UI ในแท็บหลายชิ้น)
        self.vbatch = tk.IntVar(value=10)
        self.vhdl = tk.BooleanVar(value=False)

    def _multi_tab(self):
        wrap = tk.Frame(self.tm, bg=C['bg_dark'])
        wrap.pack(fill='both', expand=True, padx=12, pady=10)

        top = _LF(wrap, ' หลายชิ้น (จาก Template) ')
        top.pack(fill='x', pady=(0, 8))
        _L(top, text='เซิร์ฟเวอร์ :').grid(row=0, column=0, sticky='w', padx=8, pady=8)
        # ตัวแปรเกมเดียวกับ launcher/prefs -> เลือกที่ไหนก็ sync กันหมด
        self.vmulti_game = self.vgame
        if not self._embedded:
            # รันเดี่ยว -> มี dropdown เกม + ปุ่ม Login (เดิมอยู่แท็บค้นหาเดี่ยว)
            ttk.Combobox(top, textvariable=self.vmulti_game, values=GAME_NAMES,
                         state='readonly', width=20).grid(row=0, column=1, sticky='w', padx=4)
            tk.Button(top, text='🔓  Open for Login', command=self._open_login,
                      bg=C['bg_card'], fg=C['text'], font=FM, relief='flat',
                      padx=10, pady=4, cursor='hand2').grid(row=0, column=2, sticky='w', padx=8)
        else:
            tk.Label(top, textvariable=self.vgame, bg=C['bg_med'], fg=C['teal'], font=FB).grid(
                row=0, column=1, sticky='w', padx=4)
            _L(top, text='(เลือก/Login บนแถบบน)', fg=C['muted'], font=F9).grid(
                row=0, column=2, sticky='w', padx=8)

        # โหมดงาน: Event (เดิม จบที่ Create Bundle) / Item Code (ไป Create Item Code)
        _L(top, text='โหมด :').grid(row=1, column=0, sticky='w', padx=8, pady=(0, 8))
        self.vmode = tk.StringVar(value='event')
        mrow = tk.Frame(top, bg=C['bg_med'])
        mrow.grid(row=1, column=1, columnspan=3, sticky='w', padx=4)
        for txt, val in (('Event', 'event'), ('Item Code', 'itemcode'), ('Shop', 'shop')):
            tk.Radiobutton(mrow, text=txt, variable=self.vmode, value=val,
                           command=self._on_mode_change, bg=C['bg_med'], fg=C['text'],
                           selectcolor=C['bg_inp'], activebackground=C['bg_med'],
                           activeforeground=C['text'], font=FM).pack(side='left', padx=(0, 10))

        # โหมด "เปิดใช้งานการแสดงผลบนเว็บ" — เข้าไปเช็คในหน้ารายละเอียดแต่ละไอเทม (Deep Check)
        _L(top, text='แสดงผลบนเว็บ :').grid(row=2, column=0, sticky='w', padx=8, pady=(0, 8))
        self.vmulti_web = tk.StringVar(value='any')
        self._web_before_itemcode = None      # ค่าก่อนโดนโหมดบังคับ (ไว้คืนค่า)
        self._web_forced_value = None         # ค่าที่เราตั้งให้ล่าสุด (ไว้ดูว่าผู้ใช้เปลี่ยนเองไหม)
        wrow = tk.Frame(top, bg=C['bg_med'])
        wrow.grid(row=2, column=1, columnspan=3, sticky='w', padx=4)
        self._web_radios = []
        for txt, val in (('ทั้งหมด', 'any'), ('มี', 'yes'), ('ไม่มี', 'no')):
            rb = tk.Radiobutton(wrow, text=txt, variable=self.vmulti_web, value=val,
                                bg=C['bg_med'], fg=C['text'], selectcolor=C['bg_inp'],
                                activebackground=C['bg_med'], activeforeground=C['text'],
                                font=FM)
            rb.pack(side='left', padx=(0, 10))
            self._web_radios.append(rb)
        self.multi_hint = _L(top, text='(มี/ไม่มี = เข้าไปเช็คทีละไอเทม จะช้ากว่า)',
                             fg=C['muted'], font=F9)
        self.multi_hint.grid(row=3, column=1, columnspan=3, sticky='w', padx=4, pady=(0, 6))

        # ตั้งค่าประมวลผล (เดิมอยู่แท็บค้นหาเดี่ยว) — Batch size / Headless
        _L(top, text='ตั้งค่า :').grid(row=4, column=0, sticky='w', padx=8, pady=(0, 8))
        srow = tk.Frame(top, bg=C['bg_med'])
        srow.grid(row=4, column=1, columnspan=3, sticky='w', padx=4, pady=(0, 8))
        _L(srow, text='Batch size').pack(side='left')
        _E(srow, textvariable=self.vbatch, width=5).pack(side='left', padx=(4, 14), ipady=2)
        tk.Checkbutton(srow, text='Headless (ซ่อน browser)', variable=self.vhdl,
                       bg=C['bg_med'], fg=C['text'], selectcolor=C['bg_inp'],
                       activebackground=C['bg_med'], activeforeground=C['text'],
                       font=FM).pack(side='left')

        ext = 'xlsx' if XLSX_OK else 'csv'
        bf = tk.Frame(wrap, bg=C['bg_dark'])
        bf.pack(fill='x', pady=(0, 8))
        tk.Button(bf, text=f'⬇  โหลด Template (.{ext})', command=self._download_template,
                  bg=C['bg_card'], fg=C['text'], font=FM, relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='left')
        tk.Button(bf, text='📂  Import Template', command=self._import_template,
                  bg=C['accent'], fg='#fff', font=FM, relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='left', padx=6)
        tk.Button(bf, text='🎁  Import Event/Prize', command=self._import_event_file,
                  bg=C['accent2'], fg='#fff', font=FM, relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='left', padx=(0, 6))
        tk.Button(bf, text='ล้างรายการ', command=self._clear_imported,
                  bg=C['bg_card'], fg=C['muted'], font=FM, relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='left')

        self.multi_info = tk.Label(wrap, text='ยังไม่ได้ import', bg=C['bg_dark'],
                                   fg=C['muted'], font=F9)
        self.multi_info.pack(anchor='w', pady=2)
        # คำอธิบายตัวย่อของคอลัมน์ "พารามิเตอร์ที่เช็ค"
        tk.Label(wrap, text=DEEP_LEGEND, bg=C['bg_dark'], fg=C['teal'], font=F9,
                 anchor='w', justify='left').pack(fill='x')

        lf = tk.Frame(wrap, bg=C['bg_dark'])
        lf.pack(fill='both', expand=True)
        self.multi_list = tk.Listbox(lf, bg=C['bg_med'], fg=C['text'],
                                     font=('Consolas', 10), relief='flat',
                                     selectbackground=C['accent'], activestyle='none')
        vsb = ttk.Scrollbar(lf, command=self.multi_list.yview)
        hsb = ttk.Scrollbar(lf, orient='horizontal', command=self.multi_list.xview)
        self.multi_list.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')          # ชื่อ/กลุ่มยาว -> เลื่อนดูได้จนจบบรรทัด
        self.multi_list.pack(side='left', fill='both', expand=True)

        runbar = tk.Frame(wrap, bg=C['bg_dark'])
        runbar.pack(anchor='w', pady=8)
        self.multi_run_btn = tk.Button(runbar, text='🔍  ค้นหาทั้งหมด',
                                       command=lambda: self._start('multi'),
                                       bg=C['accent'], fg='#fff', font=FB, relief='flat',
                                       padx=18, pady=7, cursor='hand2')
        self.multi_run_btn.pack(side='left')
        self.cancel_btn = tk.Button(runbar, text='⏹  หยุด', command=self._cancel_run,
                                    bg=C['danger'], fg='#fff', font=FM, relief='flat',
                                    padx=14, pady=7, cursor='hand2', state='disabled')
        self.cancel_btn.pack(side='left', padx=8)

    def _download_template(self):
        ext = 'xlsx' if XLSX_OK else 'csv'
        ftypes = [('Excel', '*.xlsx'), ('CSV', '*.csv')] if XLSX_OK else [('CSV', '*.csv')]
        path = filedialog.asksaveasfilename(defaultextension=f'.{ext}', filetypes=ftypes,
                                            initialfile=f'item_finder_template.{ext}')
        if not path:
            return
        try:
            download_template(path)
            messagebox.showinfo('Template', f'บันทึกแล้ว:\n{path}')
        except Exception as e:
            messagebox.showerror('Error', str(e))

    def _import_template(self):
        ftypes = [('Excel/CSV', '*.xlsx *.xlsm *.csv'), ('All', '*.*')]
        path = filedialog.askopenfilename(filetypes=ftypes)
        if not path:
            return
        try:
            rows = read_template(path)
            self._imported = rows
            self._refresh_multi_list()
            self.multi_info.config(
                text=f'Import สำเร็จ — {len(rows)} รายการ จาก {os.path.basename(path)}',
                fg=C['accent2'])
            self.log(f'Imported {len(rows)} items from {os.path.basename(path)}', 'SUCCESS')
        except Exception as e:
            messagebox.showerror('Import Error', str(e))

    # ---- Import Event/Prize (.xlsx ตารางรางวัล หลาย sheet) -------------------
    def _import_event_file(self):
        if not XLSX_OK:
            messagebox.showerror('ต้องมี openpyxl', 'ฟีเจอร์นี้ต้องติดตั้ง openpyxl ก่อน')
            return
        path = filedialog.askopenfilename(
            title='เลือกไฟล์ Event / Prize (.xlsx)',
            filetypes=[('Excel', '*.xlsx *.xlsm'), ('All', '*.*')])
        if not path:
            return
        # โหมด Event ใช้ template คนละแบบ (Monthly Plan) -> parse แล้วส่งไป Create Event เลย
        if self.vmode.get() == 'event':
            self._import_event_plan(path)
            return
        if self.vmode.get() == 'shop':
            self._import_shop_plan(path)
            return
        self.multi_info.config(text='กำลังอ่านไฟล์... (มีหลาย sheet อาจใช้เวลาสักครู่)', fg=C['warn'])
        self.log(f'อ่าน event file: {os.path.basename(path)}', 'STEP')

        def _work():
            try:
                data, skipped = parse_event_workbook(path)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda: (
                    self.multi_info.config(text='อ่านไฟล์ไม่สำเร็จ', fg=C['danger']),
                    messagebox.showerror('อ่านไฟล์ไม่สำเร็จ', msg)))
                return
            self.root.after(0, lambda: self._event_loaded(path, data, skipped))

        threading.Thread(target=_work, daemon=True).start()

    def _import_shop_plan(self, path):
        """โหมด Shop: อ่าน Monthly Plan (Cash Shop / Promotion / In Game) เข้า pipeline ค้นหาเดิม
        ไอเทมที่ Itemmove=Yes จะติดพารามิเตอร์ 'แลกเปลี่ยนได้'=yes ไว้ให้ Deep Check ตรวจกับเว็บ"""
        self.multi_info.config(text='กำลังอ่าน Shop plan...', fg=C['warn'])
        self.log('อ่าน Shop plan: %s' % os.path.basename(path), 'STEP')

        def _work():
            try:
                data, skipped = parse_shop_workbook(path)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda: (
                    self.multi_info.config(text='อ่านไฟล์ไม่สำเร็จ', fg=C['danger']),
                    messagebox.showerror('อ่านไฟล์ไม่สำเร็จ', msg)))
                return
            n_move = sum(1 for _sh, items in data for it in items if it.get('trade') == 'yes')
            self.root.after(0, lambda: (
                self.log('โหมด Shop: เช็ค "มีรูปภาพ" ทุกตัว | Itemmove=Yes %d ตัว -> '
                         'เช็ค "แลกเปลี่ยนได้" เพิ่ม (ช่องอื่นต้องว่าง)' % n_move,
                         'SUCCESS' if n_move else 'INFO'),
                self._event_loaded(path, data, skipped)))

        threading.Thread(target=_work, daemon=True).start()

    def _import_event_plan(self, path):
        """โหมด Event: อ่าน Monthly Plan (template คนละแบบ) -> แปลงเป็นไอเทมแล้วเข้า
        pipeline ค้นหาเดิม (เหมือน Item Code ทุกอย่าง แค่ template ต่างกัน)"""
        self.multi_info.config(text='กำลังอ่าน Event Plan...', fg=C['warn'])
        self.log('อ่าน Event Plan: %s' % os.path.basename(path), 'STEP')

        def _work():
            try:
                import event_tool
                data, skipped = event_tool.parse_event_plan_workbook(path)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda: (
                    self.multi_info.config(text='อ่านไฟล์ไม่สำเร็จ', fg=C['danger']),
                    messagebox.showerror('อ่านไฟล์ไม่สำเร็จ', msg)))
                return
            self.root.after(0, lambda: self._event_loaded(path, data, skipped))

        threading.Thread(target=_work, daemon=True).start()

    def _event_loaded(self, path, data, skipped=None):
        if skipped:
            self.log('อ่าน ItemKind ไม่ได้ %d ตัว (Excel format เป็นวันที่ ค่าหาย): %s'
                     % (len(skipped), ', '.join(skipped[:15])
                        + (' ...' if len(skipped) > 15 else '')), 'WARNING')
        if not data:
            self.multi_info.config(text='ไม่พบตารางไอเทมในไฟล์นี้', fg=C['danger'])
            messagebox.showinfo('ไม่พบข้อมูล',
                                'ไม่พบตารางที่มีหัวคอลัมน์ Item Kind / ItemOption / '
                                'DurationIndex / Display Name ในไฟล์นี้')
            return
        total = sum(len(items) for _, items in data)
        self.multi_info.config(
            text=f'อ่านสำเร็จ — {len(data)} sheet มีของ, รวม {total} ไอเทม', fg=C['accent2'])
        if len(data) == 1:
            self._apply_event_items(data[0][1], src=os.path.basename(path))
        else:
            self._sheet_picker(path, data)

    def _sheet_picker(self, path, data):
        win = tk.Toplevel(self.root)
        win.title('เลือก Sheet ที่จะ import')
        win.configure(bg=C['bg_dark'])
        win.geometry('580x580')
        win.minsize(480, 440)
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass

        tk.Label(win, text=os.path.basename(path), bg=C['bg_dark'], fg=C['teal'],
                 font=FB, anchor='w').pack(fill='x', padx=12, pady=(10, 2))
        tk.Label(win, text='คลิกเลือก sheet ที่ต้องการ (เลือกได้หลายอัน ไม่ต้องกด Ctrl · คลิกซ้ำ = ยกเลิก):',
                 bg=C['bg_dark'], fg=C['muted'], font=F9, anchor='w').pack(fill='x', padx=12)

        top = tk.Frame(win, bg=C['bg_dark'])
        top.pack(fill='x', padx=12, pady=4)
        tk.Button(top, text='เลือกทั้งหมด', command=lambda: lb.selection_set(0, tk.END),
                  bg=C['bg_card'], fg=C['text'], font=F9, relief='flat', padx=8, pady=2).pack(side='left')
        tk.Button(top, text='ล้าง', command=lambda: lb.selection_clear(0, tk.END),
                  bg=C['bg_card'], fg=C['muted'], font=F9, relief='flat', padx=8, pady=2).pack(side='left', padx=6)

        lf = tk.Frame(win, bg=C['bg_dark'])
        lf.pack(fill='both', expand=True, padx=12, pady=4)
        lb = tk.Listbox(lf, bg=C['bg_med'], fg=C['text'], font=('Consolas', 10), relief='flat',
                        selectmode='multiple', selectbackground=C['accent'],
                        selectforeground='#fff', activestyle='none')
        sb = ttk.Scrollbar(lf, command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        lb.pack(side='left', fill='both', expand=True)
        for sh, items in data:
            lb.insert(tk.END, '%-44s (%d)' % (sh[:44], len(items)))

        bar = tk.Frame(win, bg=C['bg_dark'])
        bar.pack(fill='x', padx=12, pady=(4, 10))

        def _do():
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo('เลือกก่อน', 'ยังไม่ได้เลือก sheet', parent=win)
                return
            items = []
            for i in sel:
                items.extend(data[i][1])
            win.destroy()
            self._apply_event_items(items, src=os.path.basename(path))

        tk.Button(bar, text='Import ที่เลือก', command=_do, bg=C['accent2'], fg='#fff',
                  font=FB, relief='flat', padx=16, pady=6, cursor='hand2').pack(side='left')
        tk.Button(bar, text='ยกเลิก', command=win.destroy, bg=C['bg_card'], fg=C['muted'],
                  font=FM, relief='flat', padx=14, pady=6, cursor='hand2').pack(side='left', padx=8)
        # handle สำหรับทดสอบ
        win.listbox, win.do_import = lb, _do
        return win

    def _apply_event_items(self, items, src=''):
        """เพิ่มไอเทมจาก event เข้า self._imported — ตัดซ้ำ kind/opt/dur/name เพื่อค้นครั้งเดียว
        แต่ "รวมรายชื่อกลุ่ม (sources)" ของตัวซ้ำไว้ เพื่อให้ผลลัพธ์รู้ว่าไอเทมอยู่ตารางไหนบ้าง"""
        def key(r):
            return (r.get('kind', ''), r.get('opt', ''), r.get('dur', ''),
                    (r.get('name', '') or '').strip())
        index = {}
        for r in self._imported:
            r.setdefault('sources', [])
            index[key(r)] = r
        added = merged = 0
        for it in items:
            it.setdefault('sources', [])
            # เก็บเงื่อนไขของแต่ละกลุ่ม (สำหรับ auto-fill Item Code) แล้วตัด group_meta ออกจากไอเทม
            gm = it.pop('group_meta', None)
            if gm:
                for s in it.get('sources', []):
                    # overwrite (last import wins) — import เดือนใหม่ทับกลุ่มชื่อซ้ำ ไม่ให้ค้าง expire เดือนเก่า
                    self._event_group_meta[s] = gm
            # เก็บทุก occurrence (ตามลำดับเอกสาร/กลุ่ม) ไว้เรียงผลลัพธ์ทีหลัง
            self._occurrences.append(dict(it))
            k = key(it)
            if k in index:                         # ซ้ำ -> รวมกลุ่มเข้าตัวเดิม
                tgt = index[k]
                for s in it.get('sources', []):
                    if s not in tgt['sources']:
                        tgt['sources'].append(s)
                        merged += 1
            else:
                self._imported.append(it)
                index[k] = it
                added += 1
        self._refresh_multi_list()
        self.multi_info.config(
            text='Import Event สำเร็จ — เพิ่ม %d ไอเทมใหม่%s (รวมทั้งหมด %d รายการ)%s'
                 % (added, (' · รวมกลุ่มซ้ำ %d' % merged) if merged else '',
                    len(self._imported), (' จาก ' + src) if src else ''),
            fg=C['accent2'])
        self.log('Import event: +%d new, merged %d groups (total %d)'
                 % (added, merged, len(self._imported)), 'SUCCESS')

    def _refresh_multi_list(self):
        self.multi_list.delete(0, tk.END)
        # ชื่อ/พารามิเตอร์ยาวสุดเท่าไหร่ ก็จัดคอลัมน์ให้เท่านั้น (ไม่ตัดทิ้ง — เลื่อนแนวนอนดูได้)
        # ห้ามใส่เพดานความกว้าง: ไม่ได้ตัดข้อความ ถ้าเพดานต่ำกว่าของจริงคอลัมน์จะเหลื่อมกัน
        name_w = max([len((r.get('name', '') or '')) for r in self._imported] + [20])
        par = [deep_summary(r) for r in self._imported]
        par_w = max([len(p) for p in par] + [16])
        for i, r in enumerate(self._imported):
            nm = (r.get('name', '') or '')
            srcs = r.get('sources')
            if srcs:                               # ไอเทมจาก event/shop -> โชว์กลุ่มที่สังกัด
                tail = (' , '.join(srcs) if len(srcs) <= 2
                        else '[%d กลุ่ม] %s , ...' % (len(srcs), srcs[0]))
            else:
                tail = '-'
            line = (f"{i+1:<4} {r.get('kind',''):<10} {r.get('opt',''):<6} "
                    f"{r.get('dur',''):<5} {nm:<{name_w}} | {par[i]:<{par_w}} | {tail}")
            self.multi_list.insert(tk.END, line)

    def _clear_imported(self):
        self._imported.clear()
        self._occurrences.clear()
        self._event_group_meta.clear()
        self.multi_list.delete(0, tk.END)
        self.multi_info.config(text='ล้างรายการแล้ว', fg=C['muted'])

    def _regroup_results(self):
        """เรียงผลลัพธ์ใหม่ให้ 'แยกตามชุด (group) ตามลำดับในเอกสาร'
        กระจายไอเทมที่พบกลับตามทุก occurrence (ไอเทมที่อยู่หลายชุด = โผล่หลายแถว แถวละชุด)"""
        if not self._occurrences or not self._results:
            return
        # index ผลที่พบ ด้วย (kind, opt, dur) -> เก็บ "ทุกตัวที่ค้นเจอ" ไม่ตัดทิ้ง
        # (1 kind/opt/dur อาจเจอหลาย id เพราะไม่ได้กรองด้วยชื่อแล้ว — ต้องโชว์ครบให้ตรวจ)
        found = {}
        for r in self._results:
            k = (str(r.get('item_kind', '')).strip(),
                 str(r.get('item_option', '')).strip(),
                 str(r.get('duration_index', '')).strip())
            lst = found.setdefault(k, [])
            if not any(x.get('aztek_id') == r.get('aztek_id') for x in lst):
                lst.append(r)                    # กันซ้ำเฉพาะ "แถวเดียวกัน" ที่อ่านมาหลายรอบ
        new_results = []
        for occ in self._occurrences:            # เรียงตามลำดับในเอกสาร
            k = (str(occ.get('kind', '')).strip(),
                 str(occ.get('opt', '')).strip(),
                 str(occ.get('dur', '')).strip())
            for r in found.get(k, []):           # occurrence เดียวเจอหลายตัว -> โชว์ทุกตัว
                row = dict(r)
                row['sources'] = list(occ.get('sources', []))   # ชุดเดียวของ occurrence นี้
                row['file_name'] = occ.get('name', '')          # ชื่อของ occurrence นี้ (ตามเอกสาร)
                new_results.append(row)
        if not new_results:
            return                                   # กันเคส match พลาดแล้วผลหาย
        self._results = new_results
        self.result_tree.delete(*self.result_tree.get_children())
        for it in new_results:
            self.add_result_row(it)
        self._update_count()
        self.log('เรียงผลลัพธ์ตามชุด (เอกสาร) — %d แถว' % len(new_results), 'SUCCESS')

    def _results_tab(self):
        pf = tk.Frame(self.tr, bg=C['bg_dark'])
        pf.pack(fill='x', padx=10, pady=(8, 4))
        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(pf, variable=self.progress_var, maximum=100,
                        style='green.Horizontal.TProgressbar').pack(fill='x')
        self.progress_lbl = tk.Label(pf, text='', bg=C['bg_dark'], fg=C['muted'], font=F9)
        self.progress_lbl.pack(anchor='e', pady=2)

        # แถบเลือก — คลิกเลือกได้หลายชิ้นโดยไม่ต้องกด Ctrl (คลิกซ้ำ = ยกเลิก)
        selbar = tk.Frame(self.tr, bg=C['bg_dark'])
        selbar.pack(fill='x', padx=10, pady=(0, 2))
        tk.Label(selbar, text='คลิกเลือกได้หลายชิ้น (คลิกซ้ำเพื่อยกเลิก)', bg=C['bg_dark'],
                 fg=C['muted'], font=F9).pack(side='left')
        tk.Label(self.tr, text=DEEP_LEGEND, bg=C['bg_dark'], fg=C['teal'], font=F9,
                 anchor='w', justify='left').pack(fill='x', padx=10)
        tk.Button(selbar, text='ล้างที่เลือก', command=self._clear_result_selection,
                  bg=C['bg_card'], fg=C['muted'], font=F9, relief='flat',
                  padx=8, pady=2, cursor='hand2').pack(side='right')
        tk.Button(selbar, text='เลือกทั้งหมด', command=self._select_all_results,
                  bg=C['bg_card'], fg=C['text'], font=F9, relief='flat',
                  padx=8, pady=2, cursor='hand2').pack(side='right', padx=(0, 6))

        lf2 = tk.Frame(self.tr, bg=C['bg_dark'])
        lf2.pack(fill='both', expand=True, padx=8)
        st = ttk.Style()
        st.configure('Result.Treeview', background=C['bg_med'], foreground=C['text'],
                     fieldbackground=C['bg_med'], font=('Segoe UI', 10), rowheight=26,
                     borderwidth=0)
        st.configure('Result.Treeview.Heading', background=C['bg_card'], foreground=C['text'],
                     font=FB, relief='flat')
        st.map('Result.Treeview', background=[('selected', C['accent'])],
               foreground=[('selected', '#ffffff')])
        cols = ('id', 'name', 'fname', 'params', 'groups', 'desc')
        self.result_tree = ttk.Treeview(lf2, columns=cols, show='headings',
                                        selectmode='extended', style='Result.Treeview')
        self.result_tree.heading('id', text='Aztek ID')
        self.result_tree.heading('name', text='ชื่อในเว็บ')
        self.result_tree.heading('fname', text='ชื่อในไฟล์ (ไม่ได้ใช้ค้นหา)')
        self.result_tree.heading('desc', text='คำอธิบายไอเทม')
        self.result_tree.heading('params', text='พารามิเตอร์ที่เช็ค')
        self.result_tree.heading('groups', text='กลุ่ม / ตาราง')
        # 'desc' โผล่เฉพาะโหมด Shop (คุมด้วย displaycolumns ใน _set_desc_col)
        self.result_tree.configure(displaycolumns=('id', 'name', 'fname', 'params', 'groups'))
        # ทุกคอลัมน์ stretch=False: ถ้าตั้งให้ยืด Tk จะบีบทุกคอลัมน์ให้พอดีจอเสมอ
        # -> รวมกันไม่เคยเกินความกว้างจอ = แถบเลื่อนแนวนอนเลื่อนไม่ได้
        # ความกว้างจะโตตามเนื้อหาจริงใน _fit_result_cols() แล้วค่อยเลื่อนดูส่วนที่เกิน
        self._tree_font = tkfont.Font(family='Segoe UI', size=10)
        self._col_px = dict(self._COL_MIN)
        for col, wpx in self._COL_MIN.items():
            self.result_tree.column(col, width=wpx, minwidth=min(wpx, 70),
                                    anchor='w', stretch=False)
        # ชื่อไม่ตรงกัน = ไฮไลต์ให้ตรวจ (ไม่ได้แปลว่าผิด — ค้นด้วย kind/opt/dur แล้ว)
        self.result_tree.tag_configure('diff', foreground=C['warn'])
        vsb2 = ttk.Scrollbar(lf2, command=self.result_tree.yview)
        hsb2 = ttk.Scrollbar(lf2, orient='horizontal', command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
        vsb2.pack(side='right', fill='y')
        hsb2.pack(side='bottom', fill='x')     # ชื่อยาวเกินจอ -> เลื่อนดูจนจบได้
        self.result_tree.pack(side='left', fill='both', expand=True)

        # คลิกเลือกได้หลายแถวโดยไม่ต้องกด Ctrl (คลิกซ้ำ = ยกเลิก)
        def _tree_toggle(e):
            row = self.result_tree.identify_row(e.y)
            if not row:
                return None
            if row in self.result_tree.selection():
                self.result_tree.selection_remove(row)
            else:
                self.result_tree.selection_add(row)
            return 'break'
        self.result_tree.bind('<Button-1>', _tree_toggle)

        btmf = tk.Frame(self.tr, bg=C['bg_dark'])
        btmf.pack(fill='x', padx=8, pady=8)
        self.count_lbl = tk.Label(btmf, text='พบ 0 items', bg=C['bg_dark'], fg=C['muted'], font=FM)
        self.count_lbl.pack(side='left')
        # โผล่เฉพาะตอนรันในโปรแกรมรวม: รวมผลค้นหาเป็นบันเดิลรายกลุ่ม -> เข้าคิว Create Bundle ตรง ๆ
        # (แก้ชื่อ/ตัดไอเทม/ลบบันเดิล/ตั้ง Type+reward ทำต่อที่หน้า Create Bundle — โฟลว์เดียว)
        if self._on_queue_bundles:
            tk.Button(btmf, text='📦  รวมเป็นบันเดิล (รีวิว)', command=self._review_bundles,
                      bg=C['accent2'], fg='#fff', font=FB, relief='flat',
                      padx=12, pady=5, cursor='hand2').pack(side='left', padx=(12, 0))
        tk.Button(btmf, text='Copy Selected IDs', command=self._copy_selected,
                  bg=C['bg_card'], fg=C['text'], font=FM, relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='right')
        tk.Button(btmf, text='Copy All IDs', command=self._copy_all,
                  bg=C['accent'], fg='#fff', font=FM, relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='right', padx=(0, 6))
        tk.Button(btmf, text='⬇  Download Excel', command=self._download_results,
                  bg=C['warn'], fg=C['bg_dark'], font=FB, relief='flat',
                  padx=12, pady=5, cursor='hand2').pack(side='right', padx=(0, 6))

    def _log_tab(self):
        top = tk.Frame(self.tl, bg=C['bg_dark'])
        top.pack(fill='x')
        tk.Button(top, text='Clear Log', command=self._clear_log, bg=C['bg_card'],
                  fg=C['muted'], font=FM, relief='flat', padx=8, pady=3).pack(
            side='right', padx=8, pady=4)
        self.log_area = scrolledtext.ScrolledText(self.tl, bg=C['bg_dark'], fg=C['text'],
                                                  font=('Consolas', 9), relief='flat',
                                                  state='disabled', wrap='word')
        self.log_area.pack(fill='both', expand=True, padx=4, pady=(0, 4))
        for lvl, col in (('STEP', C['accent']), ('SUCCESS', C['accent2']),
                         ('WARNING', C['warn']), ('ERROR', C['danger']), ('INFO', C['text'])):
            self.log_area.tag_config(lvl, foreground=col)

    # ---- prefs --------------------------------------------------------------
    def _load_prefs(self):
        # โหมดแชร์เกม (อยู่ในโปรแกรมรวม) ปล่อยให้ launcher คุมค่าเกม ไม่ override จาก prefs ตัวเอง
        if self._shared_game:
            return
        p = load_prefs()
        if p.get('game') in GAMES:
            self.vgame.set(p['game'])

    def _save_prefs(self):
        save_prefs({'game': self.vgame.get()})

    # โหมดที่ตั้ง "แสดงผลบนเว็บ" ให้เองตอนเข้าโหมด (Item Code ล็อกด้วย, Shop เปลี่ยนเองได้)
    _MODE_WEB = {'itemcode': 'no', 'shop': 'no'}

    def _on_mode_change(self):
        """ตั้งค่า 'แสดงผลบนเว็บ' ตามโหมด แล้วคืนค่าเดิมเมื่อออกจากโหมดนั้น
        ต้องคืนค่าด้วย — ไม่งั้นค่า 'ไม่มี' ค้างข้ามโหมด (เคยทำให้ผลลัพธ์ตกหมด)"""
        mode = self.vmode.get()
        forced = self._MODE_WEB.get(mode)
        cur = self.vmulti_web.get()
        if forced:
            # ถ้าผู้ใช้เปลี่ยนเอง (ค่าปัจจุบัน != ค่าที่เราตั้งไว้ล่าสุด) ให้จำค่าใหม่แทน
            if self._web_before_itemcode is None or cur != self._web_forced_value:
                self._web_before_itemcode = cur
            self.vmulti_web.set(forced)
            self._web_forced_value = forced
        elif self._web_before_itemcode is not None:
            # คืนค่าเดิมเฉพาะตอนที่ค่ายังเป็นของที่เราตั้งให้ (ผู้ใช้เปลี่ยนเอง = เคารพค่าเขา)
            if cur == self._web_forced_value:
                self.vmulti_web.set(self._web_before_itemcode)
            self._web_before_itemcode = self._web_forced_value = None
        # ล็อกเฉพาะ Item Code — โหมด Shop ตั้งให้เป็นค่าเริ่มต้น แต่ผู้ใช้เปลี่ยนเองได้
        for rb in getattr(self, '_web_radios', []):
            rb.config(state='disabled' if mode == 'itemcode' else 'normal')

    # ---- log / progress (thread-safe via root.after) ------------------------
    def log(self, msg, lvl='INFO'):
        ts = datetime.now().strftime('%H:%M:%S')
        line = f'[{ts}] {msg}\n'

        def _do():
            self.log_area.config(state='normal')
            self.log_area.insert('end', line, lvl)
            self.log_area.see('end')
            self.log_area.config(state='disabled')

        self.root.after(0, _do)

    def _clear_log(self):
        self.log_area.config(state='normal')
        self.log_area.delete('1.0', tk.END)
        self.log_area.config(state='disabled')

    def set_progress(self, cur, total, name=''):
        pct = (cur / total * 100) if total else 0

        def _do():
            self.progress_var.set(pct)
            self.progress_lbl.config(text=f'{cur} / {total}  {name}')

        self.root.after(0, _do)

    # ความกว้างเริ่มต้นของคอลัมน์ผลลัพธ์ (px) — จะโตตามเนื้อหาจริงทีหลัง
    _COL_MIN = {'id': 80, 'name': 260, 'fname': 260, 'desc': 240, 'params': 165, 'groups': 200}
    _COL_MAX = 640          # เพดานต่อคอลัมน์ ที่เกินจากนี้ใช้เลื่อนแนวนอนเอา

    def _set_desc_col(self, show):
        """โชว์/ซ่อนคอลัมน์ 'คำอธิบายไอเทม' — โผล่เฉพาะโหมด Shop"""
        cols = (('id', 'name', 'fname', 'params', 'groups', 'desc') if show
                else ('id', 'name', 'fname', 'params', 'groups'))
        try:
            self.result_tree.configure(displaycolumns=cols)
        except Exception:
            pass

    def _fit_result_cols(self, values):
        """ขยายคอลัมน์ให้พอดีเนื้อหาที่ยาวสุดเท่าที่เคยเจอ (เรียกตอนเพิ่มแถว)
        โตอย่างเดียว ไม่หด -> คอลัมน์ไม่กระตุกตอนผลลัพธ์ทยอยเข้ามา"""
        try:
            for col, val in zip(self.result_tree['columns'], values):
                need = min(self._tree_font.measure(str(val)) + 26, self._COL_MAX)
                if need > self._col_px.get(col, 0):
                    self._col_px[col] = need
                    self.result_tree.column(col, width=need)
        except Exception:
            pass

    def _reset_result_cols(self):
        self._col_px = dict(self._COL_MIN)
        for col, wpx in self._COL_MIN.items():
            try:
                self.result_tree.column(col, width=wpx)
            except Exception:
                pass

    def add_result_row(self, item):
        aid = item.get('aztek_id', '?')
        name = item.get('item_name', '')
        fname = item.get('file_name', '') or ''
        groups = ' , '.join(item.get('sources', []))
        params = deep_summary(item)          # เช็คพารามิเตอร์อะไรไปบ้างกับตัวนี้
        wv = item.get('_web_vals')           # โหมด "มี": ต่อท้ายด้วยค่าจริงบนเว็บของช่องที่ไม่ได้เช็ค
        if wv:
            params = (params + ' | ' + wv) if params and params != '-' else wv
        # เทียบชื่อไฟล์กับชื่อในเว็บ -> ไม่ตรงติดธง (ค้นด้วย kind/opt/dur ไม่ได้ใช้ชื่อ)
        tags = ()
        if fname and _norm_name(fname) not in _norm_name(name):
            fname, tags = '≠ ' + fname, ('diff',)
        desc = item.get('_desc', '') or ''   # โหมด Shop: คำอธิบายไอเทม (โหมดอื่นว่าง+คอลัมน์ถูกซ่อน)
        vals = (aid, name, fname, params, groups, desc)
        self.root.after(0, lambda: (self.result_tree.insert('', 'end', values=vals, tags=tags),
                                    self._fit_result_cols(vals)))

    def _selected_result_indexes(self):
        children = self.result_tree.get_children()
        return [children.index(iid) for iid in self.result_tree.selection()]

    def _update_count(self):
        n = len(self._results)
        self.root.after(0, lambda: self.count_lbl.config(text=f'พบ {n} items'))

    # ---- copy / export ------------------------------------------------------
    def _copy_all(self):
        if not self._results:
            messagebox.showinfo('Copy', 'ยังไม่มีผลลัพธ์')
            return
        ids = [r.get('aztek_id', '') for r in self._results]
        self.root.clipboard_clear()
        self.root.clipboard_append(', '.join(ids))
        self.log(f'Copied {len(ids)} IDs', 'SUCCESS')

    def _copy_selected(self):
        idxs = self._selected_result_indexes()
        if not idxs:
            messagebox.showinfo('Copy', 'เลือก item ก่อน')
            return
        ids = [self._results[i].get('aztek_id', '') for i in idxs if i < len(self._results)]
        self.root.clipboard_clear()
        self.root.clipboard_append(', '.join(ids))
        self.log(f'Copied {len(ids)} selected IDs', 'SUCCESS')

    def _select_all_results(self):
        self.result_tree.selection_set(self.result_tree.get_children())

    def _clear_result_selection(self):
        self.result_tree.selection_remove(self.result_tree.selection())

    def _send_to_bundle(self):
        """ส่งไอเทมที่เลือก (ไม่เลือก = ทั้งหมด) ไปยังเครื่องมือ Create Bundle"""
        if not self._on_send_to_bundle:
            return
        if not self._results:
            messagebox.showinfo('ส่งไป Bundle', 'ยังไม่มีผลลัพธ์ให้ส่ง')
            return
        sel_idxs = self._selected_result_indexes()
        idxs = sel_idxs if sel_idxs else list(range(len(self._results)))
        items = []
        grp_count = {}
        sent = set()          # ผลลัพธ์โชว์ซ้ำได้ (ไอเทมเดียวหลายชุด) แต่ส่งเข้า bundle ครั้งเดียวพอ
        for i in idxs:
            if 0 <= i < len(self._results):
                r = self._results[i]
                aid = r.get('aztek_id', '')
                if aid and aid not in sent:
                    sent.add(aid)
                    items.append({'id': aid, 'name': r.get('item_name', '')})
                for s in r.get('sources', []):     # นับกลุ่ม (Code) เพื่อผูก bundle id เข้าคิว Event
                    grp_count[s] = grp_count.get(s, 0) + 1
        if not items:
            messagebox.showinfo('ส่งไป Bundle', 'ไม่มีไอเทมให้ส่ง')
            return
        group = max(grp_count, key=grp_count.get) if grp_count else ''
        try:
            self._on_send_to_bundle(items, group)
        except Exception as e:
            messagebox.showerror('ส่งไป Bundle ไม่สำเร็จ', str(e))
            return
        how = 'ที่เลือก' if sel_idxs else 'ทั้งหมด'
        self.log(f'ส่ง {len(items)} ไอเทม ({how}) ไปหน้าต่างสร้าง Bundle', 'SUCCESS')

    def _bundle_name_for_group(self, g):
        """ชื่อบันเดิล = 'ชื่อแท็บ - ชื่อตาราง' (เอาชื่อชีต/แท็บที่เลือกมานำหน้าชื่อกลุ่ม)
        ชีตดึงจาก group_meta ของกลุ่มนั้น (Shop=shop_sheet, Event=activity)
        ถ้าไม่มีชีต หรือชื่อกลุ่มขึ้นต้นด้วยชื่อชีตอยู่แล้ว -> ใช้ชื่อกลุ่มเดิม (กันซ้ำซ้อน)"""
        meta = self._event_group_meta.get(g) or {}
        sheet = str(meta.get('shop_sheet') or meta.get('activity') or '').strip()
        if sheet and not g.startswith(sheet):
            return '%s - %s' % (sheet, g)
        return g

    def _review_bundles(self):
        """รวมผลค้นหาเป็นบันเดิลรายกลุ่ม (1 กลุ่มในคอลัมน์ "กลุ่ม" = 1 บันเดิล)
        ชื่อบันเดิล = 'ชื่อแท็บ - ชื่อตาราง' แล้วเปิดฟอร์ม Bundle ทีละกลุ่มให้เติม/แก้ก่อนเก็บเข้าคิว
        เลือกแถวไว้ = เฉพาะที่เลือก, ไม่เลือก = ทั้งหมด
        ไอเทมที่ใช้ร่วมหลายบันเดิล (sources มีมากกว่า 1 กลุ่ม) จะถูกดันไป 'ล่างสุด' เสมอในทุกบันเดิล
        ที่มันอยู่ — ไม่งั้นตำแหน่งจะขึ้นกับลำดับดิบในเอกสาร ซึ่งบางกลุ่มอาจดันมันขึ้นไปอยู่บนได้"""
        if not self._on_queue_bundles:
            return
        if not self._results:
            messagebox.showinfo('รวมเป็นบันเดิล', 'ยังไม่มีผลลัพธ์ให้รวม')
            return
        sel = self._selected_result_indexes()
        # เรียง index ตามลำดับในตาราง (=ลำดับเอกสาร) เสมอ — selection คืนมาไม่เรียงก็ได้
        # ลำดับไอเทมในบันเดิลสำคัญ ต้องตรงกับลำดับที่เห็นในผลค้นหา
        idxs = sorted(sel) if sel else list(range(len(self._results)))

        # รอบแรก: หาว่าไอเทมตัวไหน (aid) ถูกใช้ร่วมกันมากกว่า 1 กลุ่ม (รวม sources ของทุก occurrence)
        aid_groups = {}
        for i in idxs:
            if not (0 <= i < len(self._results)):
                continue
            r = self._results[i]
            aid = str(r.get('aztek_id', '') or '').strip()
            if not aid:
                continue
            gs = aid_groups.setdefault(aid, set())
            for g in (r.get('sources') or ['(ไม่มีกลุ่ม)']):
                gs.add(str(g or '').strip() or '(ไม่มีกลุ่ม)')

        order, groups = [], {}
        for i in idxs:
            if not (0 <= i < len(self._results)):
                continue
            r = self._results[i]
            aid = str(r.get('aztek_id', '') or '').strip()
            if not aid:
                continue
            nm = r.get('item_name', '') or ''
            shared = len(aid_groups.get(aid, ())) > 1     # ใช้ร่วมหลายกลุ่ม -> ดันไปท้ายทุกบันเดิล
            for g in (r.get('sources') or ['(ไม่มีกลุ่ม)']):
                g = str(g or '').strip() or '(ไม่มีกลุ่ม)'
                if g not in groups:
                    groups[g] = {'name': self._bundle_name_for_group(g), 'group': g,
                                 'seen': set(), 'items': [], 'shared_items': []}
                    order.append(g)
                gd = groups[g]
                if aid in gd['seen']:            # ไอเทมซ้ำในกลุ่มเดียว -> ใส่ครั้งเดียว
                    continue
                gd['seen'].add(aid)
                (gd['shared_items'] if shared else gd['items']).append({'id': aid, 'name': nm})
        bundles = [{'name': groups[g]['name'], 'group': groups[g]['group'],
                    'items': groups[g]['items'] + groups[g]['shared_items']} for g in order
                   if groups[g]['items'] or groups[g]['shared_items']]
        if not bundles:
            messagebox.showinfo('รวมเป็นบันเดิล', 'ไม่มีไอเทมให้รวม')
            return
        try:
            self._on_queue_bundles(bundles)         # -> เปิดฟอร์ม "เพิ่ม Bundle เข้าคิว" ทีละกลุ่ม
        except Exception as e:
            messagebox.showerror('รวมเป็นบันเดิลไม่สำเร็จ', str(e))
            return
        total = sum(len(b['items']) for b in bundles)
        self.log('รวมเป็น %d กลุ่ม (%d ไอเทม) — เปิดฟอร์ม Bundle ทีละกลุ่มให้เติม/แก้ก่อนเก็บเข้าคิว'
                 % (len(bundles), total), 'SUCCESS')

    def _download_results(self):
        if not self._results:
            messagebox.showinfo('Download', 'ยังไม่มีผลลัพธ์')
            return
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f'item_finder_results_{ts}'
        if XLSX_OK:
            path = filedialog.asksaveasfilename(defaultextension='.xlsx',
                                                filetypes=[('Excel', '*.xlsx'), ('CSV', '*.csv')],
                                                initialfile=default_name + '.xlsx')
        else:
            path = filedialog.asksaveasfilename(defaultextension='.csv',
                                                filetypes=[('CSV', '*.csv')],
                                                initialfile=default_name + '.csv')
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.xlsx' and XLSX_OK:
                self._export_xlsx(path)
            else:
                self._export_csv(path)
            self.log(f'บันทึกผลลัพธ์: {os.path.basename(path)}', 'SUCCESS')
            messagebox.showinfo('Download', 'บันทึกสำเร็จ:\n' + path)
        except Exception as e:
            messagebox.showerror('Export Error', str(e))
            self.log(str(e), 'ERROR')

    def _export_xlsx(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Results'
        hdr_fill = PatternFill('solid', fgColor='1F6FEB')
        pass_fill = PatternFill('solid', fgColor='1A3A1A')
        wht_font = Font(color='FFFFFF', bold=True)
        grn_font = Font(color='3FB950')
        ctr = Alignment(horizontal='center')
        gray_fill = PatternFill('solid', fgColor='21262D')
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        game = self.vgame.get()
        summary_lines = [
            f'Aztek Item Finder — Export  |  เกม: {game}  |  วันที่: {ts}',
            f'พบทั้งหมด {len(self._results)} items  |  สร้างโดย item_finder.py',
            '',
        ]
        COLS = 11
        for r, line in enumerate(summary_lines, 1):
            c = ws.cell(row=r, column=1, value=line)
            c.font = Font(italic=True, color='8B949E', size=9)
            c.fill = gray_fill
            ws.merge_cells(f'A{r}:{chr(64 + COLS)}{r}')
        headers = ['#', 'Aztek ID', 'Item Name', 'ItemKind', 'itemOption',
                   'durationIndex', 'Game', 'Notes', 'Criteria #', 'Groups (อยู่ตารางไหนบ้าง)',
                   'คำอธิบายไอเทม']
        col_w = [5, 12, 30, 12, 14, 16, 14, 24, 10, 42, 40]
        for col, (h, w) in enumerate(zip(headers, col_w), 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = wht_font
            cell.alignment = ctr
            ws.column_dimensions[cell.column_letter].width = w
        for i, item in enumerate(self._results, 1):
            row = [i, item.get('aztek_id', ''), item.get('item_name', ''),
                   item.get('item_kind', ''), item.get('item_option', ''),
                   item.get('duration_index', ''), item.get('game', ''),
                   item.get('notes', 'passed'), item.get('_ci', ''),
                   ' | '.join(item.get('sources', [])), item.get('_desc', '')]
            ws.append(row)
            r = 4 + i
            for col in range(1, COLS + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = pass_fill
                cell.font = grn_font
                cell.alignment = Alignment(
                    horizontal='center' if col in (1, 2, 4, 5, 6, 7, 9) else 'left')
        ws.auto_filter.ref = f'A4:{chr(64 + COLS)}4'
        ws.freeze_panes = 'A5'
        wb.save(path)

    def _export_csv(self, path):
        headers = ['#', 'aztek_id', 'item_name', 'item_kind', 'item_option',
                   'duration_index', 'game', 'notes', 'criteria_no', 'groups', 'description']
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(headers)
            for i, item in enumerate(self._results, 1):
                w.writerow([i, item.get('aztek_id', ''), item.get('item_name', ''),
                            item.get('item_kind', ''), item.get('item_option', ''),
                            item.get('duration_index', ''), item.get('game', ''),
                            item.get('notes', 'passed'), item.get('_ci', ''),
                            ' | '.join(item.get('sources', [])), item.get('_desc', '')])

    # ---- login --------------------------------------------------------------
    def _open_login(self):
        if not PW_OK:
            messagebox.showerror('No Playwright',
                                 'pip install playwright\npython -m playwright install chromium')
            return
        game = self.vgame.get()
        url = GAMES.get(game, list(GAMES.values())[0])
        self.nb.select(2)              # แท็บ Log (index 2 หลังเอาแท็บค้นหาเดี่ยวออก)
        self.log(f'Opening Chrome for login ({game})...', 'STEP')

        def _run():
            async def _open():
                chrome_exe = find_chrome_exe()
                kw = finder_core.build_launch_kwargs(headless=False,
                                                     user_data_dir=CHROME_PROFILE, chrome_exe=chrome_exe)
                async with async_playwright() as pw:
                    browser = await pw.chromium.launch_persistent_context(**kw)
                    page = browser.pages[0] if browser.pages else await browser.new_page()
                    await page.goto(url, wait_until='domcontentloaded', timeout=30000)
                    self.log('  Login แล้วปิด browser ได้เลย', 'SUCCESS')
                    try:
                        await page.wait_for_event('close', timeout=0)
                    except Exception:
                        pass
                    self.log('  Profile saved', 'SUCCESS')
            # ใช้ profile ร่วมกับ tool อื่น -> จอง browser ก่อนเปิด login แล้วคืนเสมอ
            try:
                core.acquire_browser('Item Finder')
            except core.BrowserBusy as ex:
                self.log('✋ Browser กำลังถูกใช้โดย tool อื่น — ปิด browser เดิมก่อน', 'ERROR')
                self.log(str(ex), 'WARNING')
                return
            try:
                asyncio.run(_open())
            finally:
                core.release_browser()

        threading.Thread(target=_run, daemon=True).start()

    # ---- run control --------------------------------------------------------
    def _start(self, mode='multi'):
        if not PW_OK:
            messagebox.showerror('No Playwright',
                                 'pip install playwright\npython -m playwright install chromium')
            return
        if self._running:
            return
        game = self.vgame.get()

        if mode == 'multi':
            if not self._imported:
                messagebox.showwarning('ค้นหา', 'ยัง import template ไม่ได้ หรือ template ว่างเปล่า')
                return
            multi_game = self.vmulti_game.get()
            multi = [dict(r) for r in self._imported]

            # โหมด "แสดงผลบนเว็บ" รวม — override web ของทุกไอเทม (มี=yes / ไม่มี=no)
            web_mode = self.vmulti_web.get()
            if web_mode in ('yes', 'no'):
                # ข้อจำกัดของเว็บ: ช่อง แลกเปลี่ยนได้/เจาะรูได้/จำนวน/คริติคอล จะโผล่ให้ตรวจ
                # ก็ต่อเมื่อติ๊ก "เปิดใช้งานการแสดงผลบนเว็บ" ไว้เท่านั้น
                # -> แถวที่ต้องเช็คของพวกนี้ (เช่น Shop: Itemmove=Yes) ต้องเป็น "มี" เสมอ
                def _needs_web(r):
                    return (r.get('trade', 'any') != 'any' or r.get('drill', 'any') != 'any'
                            or str(r.get('qty_val', '') or '').strip()
                            or str(r.get('crit_val', '') or '').strip())
                n_exc = 0
                for r in multi:
                    if web_mode == 'no' and _needs_web(r):
                        r['web'] = 'yes'
                        n_exc += 1
                    else:
                        r['web'] = web_mode
                self.log('⚠ ตัวเลือก "แสดงผลบนเว็บ" = %s -> ทับทุกไอเทม (%d ตัว)'
                         % ('มี' if web_mode == 'yes' else 'ไม่มี', len(multi)), 'WARNING')
                if n_exc:
                    self.log('   ยกเว้น %d ตัวที่ต้องเช็ค แลกเปลี่ยนได้/เจาะรูได้/จำนวน/คริติคอล '
                             '-> ใช้ "มี" (ถ้าปิดแสดงผลบนเว็บ จะไม่มีช่องพวกนี้ให้ตรวจ)' % n_exc,
                             'INFO')
                # เลือก "มี" -> เว็บเปิดทุกตัว ช่อง เทรด/เจาะ/จำนวน/คริต โผล่ให้อ่านได้
                # -> ดึงค่าจริงมาโชว์เพิ่มในคอลัมน์ผลลัพธ์ (ไม่ใช้ตัดผล แค่ให้เห็นครบ)
                if web_mode == 'yes':
                    for r in multi:
                        r['_show_web_vals'] = True

            def _row_has_deep(r):
                return (r.get('web', 'any') != 'any' or r.get('img', 'any') != 'any'
                        or r.get('qty_val', '') != '' or r.get('trade', 'any') != 'any'
                        or r.get('drill', 'any') != 'any' or r.get('crit_val', '') != '')

            auto_deep = any(_row_has_deep(r) for r in multi)
        else:
            return   # โหมดค้นหาชิ้นเดียวถูกถอดออกแล้ว (เหลือเฉพาะค้นหาจาก template)

        use_game = multi_game
        # โหมด Shop -> อ่าน "คำอธิบายไอเทม" จากหน้ารายละเอียดมาโชว์เพิ่ม (โหมดอื่นไม่อ่าน/ไม่โชว์)
        read_desc = (getattr(self, 'vmode', None) is not None and self.vmode.get() == 'shop')
        data = {
            'game': use_game,
            'url': GAMES[use_game],
            'multi': multi,
            'deep': self.vdeep.get() or auto_deep,
            'web': self.vweb.get(),
            'img': self.vimg.get(),
            'qty_val': self.vqty_val.get().strip(),
            'trade': self.vtrade.get(),
            'drill': self.vdrill.get(),
            'crit_val': self.vcrit_val.get().strip(),
            'batch': self.vbatch.get(),
            'headless': self.vhdl.get(),
            'read_desc': read_desc,
        }

        self._results.clear()
        self.result_tree.delete(*self.result_tree.get_children())
        self._reset_result_cols()      # เริ่มรอบใหม่ -> คอลัมน์กลับไปขนาดเริ่มต้น
        self._set_desc_col(read_desc)  # โชว์คอลัมน์ "คำอธิบายไอเทม" เฉพาะโหมด Shop
        self.progress_var.set(0)
        self.progress_lbl.config(text='')
        self.count_lbl.config(text='พบ 0 items')
        self._cancel = False
        self._running = True
        self.multi_run_btn.config(state='disabled', text='Running...')
        self.cancel_btn.config(state='normal')
        self._save_prefs()
        self.nb.select(2)              # แท็บ Log (index 2 หลังเอาแท็บค้นหาเดี่ยวออก)
        self.log('=' * 52)
        self.log(f'Items : {len(multi)}  Server : {multi_game}', 'STEP')

        if data['deep']:
            if not self.vdeep.get():
                self.log('  Deep Check เปิดอัตโนมัติ (พบค่า deep check ใน template)', 'INFO')
            self.log(f"  Global deep — web:{data['web']} img:{data['img']} "
                     f"qty:{data['qty_val']!r} trade:{data['trade']} "
                     f"drill:{data['drill']} crit:{data['crit_val']!r}", 'INFO')
            self.log('  (แต่ละแถวใช้ค่า deep check ของตัวเองจาก template)', 'INFO')

        threading.Thread(target=self._run_thread, args=(data,), daemon=True).start()

    def _cancel_run(self):
        self._cancel = True
        self.log('Cancelling...', 'WARNING')

    def _run_thread(self, data):
        def _rst():
            self.multi_run_btn.config(state='normal', text='🔍  ค้นหาทั้งหมด')
            self.cancel_btn.config(state='disabled')
        try:
            asyncio.run(self._auto(data))
        except core.BrowserBusy as ex:
            self.log('✋ Browser กำลังถูกใช้โดย tool อื่น — ปิด browser เดิมก่อนแล้วค้นใหม่', 'ERROR')
            self.log(str(ex), 'WARNING')
        except Exception as ex:
            self.log(str(ex), 'ERROR')
            self.log(traceback.format_exc(), 'ERROR')
        finally:
            self._running = False
            self.root.after(0, _rst)

    # ---- async engine -------------------------------------------------------
    async def _auto(self, data):
        chrome_exe = find_chrome_exe()
        kw = finder_core.build_launch_kwargs(headless=data['headless'],
                                             user_data_dir=CHROME_PROFILE, chrome_exe=chrome_exe)
        if not chrome_exe:
            self.log('Chrome not found — using Playwright Chromium', 'WARNING')
        # ใช้ profile ร่วมกับ tool อื่น -> จอง browser ก่อน (อาจ raise BrowserBusy) แล้วคืนเสมอ
        core.acquire_browser('Item Finder')
        try:
            async with async_playwright() as pw:
                browser = await pw.chromium.launch_persistent_context(**kw)
                page = browser.pages[0] if browser.pages else await browser.new_page()
                try:
                    await self._search_all(page, data)
                finally:
                    await browser.close()
        finally:
            core.release_browser()

    async def _search_all(self, page, data):
        all_items = []
        total_passed = 0
        not_found = []
        multi = data['multi']
        self.nb.select(1)              # แท็บ ผลลัพธ์ (index 1 หลังเอาแท็บค้นหาเดี่ยวออก)
        base_url = data['url']

        def _clabel(i, cr):
            parts = [f"Kind={cr.get('kind', '')!r}"]
            if cr.get('opt'):
                parts.append(f"Opt={cr['opt']!r}")
            if cr.get('dur'):
                parts.append(f"Dur={cr['dur']!r}")
            if cr.get('name'):
                parts.append(f"Name={cr['name']!r}")
            return f"#{i + 1} " + ' '.join(parts)

        for idx, criteria in enumerate(multi):
            if self._cancel:
                break
            self.log(f"── {idx+1}/{len(multi)}: Kind={criteria['kind']!r} "
                     f"Opt={criteria['opt']!r} Dur={criteria['dur']!r} "
                     f"Name={criteria.get('name', '')!r}", 'STEP')
            # ค้นหา + retry ถ้าตารางว่าง (กัน timing miss: บางทีเจอ บางทีไม่เจอ)
            rows = []
            for attempt in range(2):
                if self._cancel:
                    break
                await page.goto(base_url, wait_until='domcontentloaded', timeout=30000)
                await page.wait_for_timeout(2000)          # ให้ SPA เรนเดอร์ (เหมือนเดิม)
                await self._apply_filters(page, criteria)
                # นับแถวดิบหลังค้นหา (ก่อนกรอง opt/dur/name)
                try:
                    raw_count = await page.evaluate(
                        "() => document.querySelectorAll('table tbody tr').length")
                except Exception:
                    raw_count = 0
                rows = await self._read_all_pages(page, criteria)
                if rows or raw_count > 0:
                    break  # เจอผล หรือมีแถวดิบ (กรองแล้วไม่ตรงก็ไม่ retry)
                if attempt < 1:
                    self.log('  (ตารางว่าง — ลองใหม่อีกครั้ง)', 'WARNING')
                    await page.wait_for_timeout(1500)

            new_items = []
            for item in rows:
                aid = item.get('aztek_id', '')
                if not aid:
                    continue
                item['_ci'] = idx + 1
                item['game'] = data['game']
                item['file_name'] = criteria.get('name', '')   # ชื่อจากไฟล์ — ไว้เทียบกับชื่อในเว็บ
                for k in ('web', 'img', 'qty_val', 'trade', 'drill', 'crit_val', 'sources', 'amt',
                          '_show_web_vals'):
                    if k in criteria:
                        item[k] = criteria[k]
                new_items.append(item)

            self.log(f'  → {len(new_items)} items', 'INFO')
            if not new_items:
                not_found.append((_clabel(idx, criteria), 'ไม่พบ row ที่ตรงเงื่อนไข'))
                continue

            if data['deep']:
                passed_this = 0
                self.log(f'  Deep check {len(new_items)} items (criteria {idx+1})', 'STEP')
                filtered_list_url = page.url
                for item in new_items:
                    if self._cancel:
                        break
                    item_data = dict(data)
                    item_data['list_url'] = filtered_list_url
                    for k in ('web', 'img', 'qty_val', 'trade', 'drill', 'crit_val'):
                        if k in item:
                            item_data[k] = item[k]
                    self.set_progress(len(self._results) + 1,
                                      len(new_items) * (idx + 1),
                                      item.get('item_name', ''))
                    ok, notes = await self._check_item_detail(page, item, item_data)
                    if ok:
                        item['notes'] = notes
                        total_passed += 1
                        passed_this += 1
                        self._results.append(item)
                        self.add_result_row(item)
                        self._update_count()
                        self.log(f"  ✓ {item['aztek_id']}  {item['item_name']}  [{notes}]", 'SUCCESS')
                    else:
                        self.log(f"  ✗ {item['aztek_id']}  {item['item_name']}", 'INFO')
                    all_items.append(item)
                if passed_this == 0 and not self._cancel:
                    not_found.append((_clabel(idx, criteria),
                                      f'เจอ {len(new_items)} แต่ไม่ผ่าน deep check'))
            else:
                # ไม่ตัดของซ้ำข้ามเงื่อนไข — เจอมาเท่าไหร่โชว์หมด (ไอเทมเดียวอยู่ได้หลายชุด)
                for item in new_items:
                    all_items.append(item)
                    self._results.append(item)
                    self.add_result_row(item)
                self._update_count()

        if self._cancel:
            return
        self._not_found = not_found
        total_items = len(all_items)
        self.log(f'Total items processed: {total_items}', 'INFO')
        ids = ', '.join(r.get('aztek_id') for r in self._results)
        if data['deep']:
            self.set_progress(total_items, total_items, 'เสร็จสิ้น')
            self.log(f'Done! {total_passed}/{total_items} passed deep check', 'SUCCESS')
        elif all_items:
            self.log(f'Done! {len(self._results)} items', 'SUCCESS')
        if ids:
            self.log(f'IDs: {ids}', 'SUCCESS')

        # สรุปรายการที่หาไม่เจอ
        if not_found:
            self.log(f'──  หาไม่เจอ {len(not_found)}/{len(multi)} รายการ  ──', 'WARNING')
            for label, reason in not_found:
                self.log(f'  ✗ {label}   ({reason})', 'WARNING')
        elif multi:
            self.log('✓ เจอครบทุกรายการ', 'SUCCESS')
        if not all_items:
            self.log('ไม่พบ item ที่ตรงเงื่อนไขเลย', 'WARNING')

        # เรียงผลลัพธ์ตามชุด (เอกสาร) หลังค้นเสร็จ — ทำบน main thread
        if self._occurrences and self._results:
            self.root.after(0, self._regroup_results)

    async def _wait_table_ready(self, page, tries=10, interval=250):
        """รอให้ตาราง 'มีข้อมูลและนิ่ง' กัน race อ่านตอนตารางว่างชั่วขณะ (ระหว่างโหลด)
        - ถ้ามีแถว: คืนเมื่อ signature เดิม 2 ครั้งติด (นิ่ง)
        - ถ้ายังว่าง: ไม่คืน รอต่อจนมีแถว หรือครบเวลา (~tries*interval)"""
        last, stable = None, 0
        for _ in range(tries):
            if self._cancel:
                return
            try:
                r = await page.evaluate("""() => {
                    const rows = document.querySelectorAll('table tbody tr');
                    const f = rows[0] && rows[0].querySelector('td')
                              ? (rows[0].querySelector('td').innerText || '').trim() : '';
                    return {n: rows.length, sig: rows.length + '|' + f};
                }""")
            except Exception:
                r = None
            if r and r['n'] > 0:                    # มีแถวแล้ว -> เช็คนิ่ง
                if r['sig'] == last:
                    stable += 1
                    if stable >= 2:
                        return
                else:
                    stable, last = 0, r['sig']
            else:
                stable, last = 0, None              # ยังว่าง (กำลังโหลด) -> รอต่อ ไม่คืน
            await page.wait_for_timeout(interval)

    async def _wait_detail_ready(self, page, tries=18, interval=300):
        """รอจนฟอร์มรายละเอียด 'โหลดข้อมูลไอเทมจริงและนิ่ง' ก่อนอ่าน (Deep Check)
        คืน True ถ้าโหลดสำเร็จ, False ถ้าหมดเวลา (ฟอร์มยังเปล่า)
        - ตอนโหลดไม่เสร็จ ทุกช่องเปล่า+checkbox ไม่ติ๊ก -> อ่านได้ค่าผิด (web=False หลอก)
        - สัญญาณว่าโหลดแล้ว = มี input ที่ค่าเป็นเลข >=4 หลัก (ItemKind/Option/Id) + ค่านิ่ง"""
        last, stable = None, 0
        for _ in range(tries):
            if self._cancel:
                return False
            try:
                r = await page.evaluate("""() => {
                    const inputs = [...document.querySelectorAll('input')];
                    let hasData = false;
                    for (const i of inputs) {
                        if (/^\\d{4,}$/.test((i.value || '').trim())) { hasData = true; break; }
                    }
                    const cbs = document.querySelectorAll(
                        'input[type=checkbox],[role=checkbox]').length;
                    const vals = inputs.map(i => (i.value || '') + (i.checked ? '1' : '0')).join('|');
                    return {hasData, cbs, sig: cbs + '::' + vals};
                }""")
            except Exception:
                r = None
            if r and r['hasData'] and r['cbs'] > 0:     # ข้อมูลโหลดแล้ว -> เช็คนิ่ง
                if r['sig'] == last:
                    stable += 1
                    if stable >= 2:
                        return True
                else:
                    stable, last = 0, r['sig']
            else:
                stable, last = 0, None                  # ยังไม่มีข้อมูล (โหลดอยู่) -> รอต่อ
            await page.wait_for_timeout(interval)
        return False

    async def _apply_filters(self, page, criteria):
        """ใช้แค่ ItemKind (top search) เท่านั้น
        ไม่ใส่ opt/dur column filter เพราะเว็บทำ client-side filtering
        ซึ่งทำให้ปุ่ม Next หายไปและอ่านได้แค่หน้าแรก
        opt/dur จะ filter ด้วย Python หลังอ่านทุกหน้าแล้ว"""
        await page.wait_for_timeout(300)
        kind = criteria.get('kind', '')
        if not kind:
            return

        box = None
        # 1. Try exact input[name="searchBox"] or placeholders with visibility wait
        for sel in (
            'input[name="searchBox"]',
            'input[name="searchbox"]',
            'input[placeholder*="Aztek Item Id"]',
            'input[placeholder*="Item Name"]',
            'input[placeholder*="ItemKind"]',
            'input[placeholder*="ค้นหา"]',
        ):
            loc = page.locator(sel).first
            try:
                if await loc.count() > 0:
                    await loc.wait_for(state='visible', timeout=5000)
                    box = loc
                    break
            except Exception:
                continue

        # 2. Fallback: find any visible text/search input
        if box is None:
            for sel in (
                'input[name*="search"]:visible',
                'input[placeholder*="Search"]:visible',
                'input[type="text"]:visible',
                'input:not([type="hidden"]):visible',
            ):
                loc = page.locator(sel).first
                try:
                    if await loc.count() > 0 and await loc.is_visible():
                        box = loc
                        break
                except Exception:
                    continue

        # 3. Check frames as fallback
        if box is None:
            for frame in page.frames:
                for sel in ('input[name="searchBox"]', 'input[placeholder*="Aztek"]', 'input[type="text"]:visible'):
                    try:
                        loc = frame.locator(sel).first
                        if await loc.count() > 0 and await loc.is_visible():
                            box = loc
                            break
                    except Exception:
                        continue
                if box is not None:
                    break

        if box is None:
            self.log(f'  หาช่องค้นหาไม่เจอ (หน้าปัจจุบัน: {page.url})', 'WARNING')
            return
        try:
            await box.fill(kind)
            self.log(f'  Top search: {kind}', 'INFO')
        except Exception as e:
            self.log(f'  กรอกช่องค้นหาไม่ได้: {e}', 'WARNING')
            return
        await page.wait_for_timeout(200)
        clicked = False
        # v2 puts several elements labelled "ค้นหา" on the page (a date filter,
        # column headers); only the type=submit one actually runs the search, so
        # try the submit button first. The generic label selectors stay as a
        # fallback for the v1 layout.
        for btn_sel in (
            'button[type="submit"]:has-text("ค้นหา")',
            'button[type="submit"]:has-text("Search")',
            'button:has-text("Search")',
            'button:has-text("ค้นหา")',
        ):
            try:
                loc = page.locator(btn_sel).first
                if await loc.count() > 0 and await loc.is_visible():
                    await loc.click()
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            await box.press('Enter')
        # base wait (เหมือนเดิมที่เคยเวิร์ก) + รอตารางมีข้อมูลนิ่ง (ไม่คืนตอนว่าง)
        await page.wait_for_timeout(1500)
        await self._wait_table_ready(page)

    async def _read_all_pages(self, page, criteria=None, mode='param'):
        all_items = []
        seen_ids = set()
        page_num = 1
        await self._dump_pager(page)   # log วินิจฉัย pagination (ครั้งแรกของรอบ)
        while True:
            if self._cancel:
                break
            rows = await self._read_table_page(page)
            new = [r for r in rows if r.get('aztek_id') and r.get('aztek_id') not in seen_ids]
            for r in new:
                seen_ids.add(r.get('aztek_id'))
            all_items.extend(new)
            self.log(f'  Page {page_num}: {len(rows)} rows ({len(new)} ใหม่)', 'INFO')
            # ถ้าหน้าถัดมาไม่มีแถวใหม่เลย = วน/จบแล้ว ออกกันลูปค้าง
            if page_num > 1 and not new:
                break
            first_id = rows[0].get('aztek_id') if rows else None
            if not await self._go_next_page(page, first_id):
                break
            page_num += 1
            await page.wait_for_timeout(800)

        if criteria:
            before = len(all_items)
            opt = criteria.get('opt', '').strip()
            dur = criteria.get('dur', '').strip()
            name = criteria.get('name', '').strip()
            # ค้นด้วย Kind/Option/Duration เท่านั้น — ไม่กรองด้วยชื่อ
            # (ชื่อในไฟล์กับชื่อในเว็บไม่ตรงกันบ่อย ถ้ากรองจะทำให้ของที่ถูกต้องหายไป)
            # ชื่อยังเก็บไว้แสดง/เทียบให้ตรวจเอง | โหมด all = ไม่กรองอะไรเลย
            if mode in ('all', 'noparam'):
                opt = dur = ''
            if opt or dur:
                pairs = sorted({(r.get('item_option', ''), r.get('duration_index', ''))
                                for r in all_items})
                self.log(f'  opt/dur ที่พบในผลค้นหา ({len(all_items)} แถว): {pairs}', 'INFO')
            if opt:
                all_items = [r for r in all_items if r.get('item_option', '') == opt]
            if dur:
                all_items = [r for r in all_items if r.get('duration_index', '') == dur]
            if opt or dur:
                self.log(f'  Python filter opt={opt} dur={dur}: {before} → {len(all_items)}', 'INFO')
            if name:      # ไม่ได้กรอง แค่บอกให้เทียบ — ชื่อไม่ตรงไม่ได้แปลว่าผิด
                diff = sorted({r.get('item_name', '') for r in all_items
                               if _norm_name(name) not in _norm_name(r.get('item_name', ''))})
                if diff:
                    self.log(f'  ⚠ ชื่อไม่ตรงกับไฟล์ ({name!r}) -> ในเว็บ: {diff} (ไม่ได้กรองออก ตรวจเอง)',
                             'WARNING')
        return all_items

    async def _dump_pager(self, page):
        """log โครงสร้าง pagination ครั้งเดียวต่อรอบ เพื่อช่วยวินิจฉัยว่าอ่านได้กี่หน้า"""
        if getattr(self, '_pager_dumped', False):
            return
        self._pager_dumped = True
        try:
            info = await page.evaluate("""() => {
                const out=[];
                const sels=['.pagination','.ant-pagination','.MuiPagination-root',
                            '[role="navigation"]','nav','ul.page'];
                for(const s of sels){
                    const el=document.querySelector(s);
                    if(el) out.push(s+' => '+(el.innerText||'').replace(/\\s+/g,' ').trim().slice(0,140));
                }
                const selects=[...document.querySelectorAll('select')]
                    .map(s=>[...s.options].map(o=>(o.text||'').trim()).join('/')).filter(Boolean);
                return {pager:out, selects:selects};
            }""")
            if info.get('pager'):
                for line in info['pager']:
                    self.log('  [pager] ' + line, 'INFO')
            else:
                self.log('  [pager] ไม่พบ pagination control มาตรฐาน (อาจอ่านได้แค่หน้าเดียว)', 'WARNING')
            if info.get('selects'):
                self.log('  [selects] ' + ' | '.join(info['selects'])[:200], 'INFO')
        except Exception:
            pass

    async def _read_table_page(self, page):
        return await page.evaluate("""() => {
            const results=[];
            const tbody=document.querySelector('table tbody');
            if(!tbody) return results;
            tbody.querySelectorAll('tr').forEach(tr=>{
                const tds=[...tr.querySelectorAll('td')];
                if(tds.length<5) return;
                const aztek_id=(tds[0].innerText||'').trim();
                if(!aztek_id||!/^\\d+$/.test(aztek_id)) return;
                results.push({
                    aztek_id, item_name:(tds[1].innerText||'').trim(),
                    item_kind:(tds[2].innerText||'').trim(),
                    item_option:(tds[3].innerText||'').trim(),
                    duration_index:(tds[4].innerText||'').trim(),
                });
            });
            return results;
        }""")

    async def _go_next_page(self, page, prev_first_id=None):
        kind = await page.evaluate("""() => {
            const norm = t => (t||'').trim().toLowerCase();
            const nextWords = ['next','ถัดไป','>','next >','›','»','>>'];
            // Words that mean "previous" — never treat these as a Next button,
            // even though "‹ Previous" also contains a chevron.
            const prevWords = ['prev','previous','ก่อนหน้า','‹','«','<'];
            const isDisabled = el => el.disabled || el.getAttribute('disabled')!==null ||
                el.classList.contains('disabled') ||
                el.getAttribute('aria-disabled')==='true' ||
                (el.closest && el.closest('.disabled,[disabled],[aria-disabled="true"]')!==null);
            const cands=[...document.querySelectorAll('button,a,[role="button"],li')];
            // 1) exact-match a Next label (original v1 behaviour)
            for(const el of cands){
                const t=(el.innerText||el.textContent||'').trim();
                if(!nextWords.includes(norm(t))) continue;
                if(isDisabled(el)) continue;
                const target=(el.tagName==='LI')?(el.querySelector('a,button')||el):el;
                target.click(); return 'text';
            }
            // 1b) v2 labels like "Next ›" (word + chevron): match if the text
            // CONTAINS a next word and no previous word.
            for(const el of cands){
                const t=norm(el.innerText||el.textContent||'');
                if(!t) continue;
                if(prevWords.some(w=>t.includes(w))) continue;
                if(!(t.includes('next')||t.includes('ถัดไป')||t.includes('›')||t.includes('»'))) continue;
                if(isDisabled(el)) continue;
                const target=(el.tagName==='LI')?(el.querySelector('a,button')||el):el;
                target.click(); return 'text2';
            }
            // 2) aria-label / rel=next / title
            const aria=document.querySelector('[rel="next"],[aria-label="Next"],[aria-label*="ถัดไป"],button[title="Next"]');
            if(aria && !isDisabled(aria)){ aria.click(); return 'aria'; }
            // 3) pagination แบบเลขหน้า -> คลิกตัวถัดจากหน้าที่ active
            const active=document.querySelector('.pagination .active,.pagination [aria-current="page"],li.active,.page-item.active,.ant-pagination-item-active,.Mui-selected');
            if(active){
                let nx=active.nextElementSibling;
                while(nx && !(nx.querySelector && nx.querySelector('a,button'))) nx=nx.nextElementSibling;
                if(nx){ const a=nx.querySelector('a,button')||nx; a.click(); return 'num'; }
            }
            return false;
        }""")
        if not kind:
            return False
        await page.wait_for_timeout(1200)
        # ยืนยันว่าเนื้อหาเปลี่ยนจริง (กันกรณีกดปุ่มได้แต่อยู่หน้าเดิม = หน้าสุดท้าย)
        new_first = await page.evaluate("""() => {
            const td=document.querySelector('table tbody tr td');
            return td ? (td.innerText||'').trim() : null;
        }""")
        if prev_first_id is not None and new_first == prev_first_id:
            return False
        return True

    async def _run_deep_check(self, page, items, data):
        total = len(items)
        passed = 0
        batch_n = data['batch']
        self.log(f'Deep check: {total} items  batch={batch_n}', 'STEP')
        for i, item in enumerate(items):
            if self._cancel:
                self.log('Cancelled', 'WARNING')
                break
            self.set_progress(i + 1, total, item.get('item_name', ''))
            item_data = dict(data)
            for k in ('web', 'img', 'qty_val', 'trade', 'drill', 'crit_val'):
                if k in item:
                    item_data[k] = item[k]
            ok, notes = await self._check_item_detail(page, item, item_data)
            if ok:
                item['notes'] = notes
                passed += 1
                self._results.append(item)
                self.add_result_row(item)
                self._update_count()
                self.log(f"  ✓ {item['aztek_id']}  {item['item_name']}  [{notes}]", 'SUCCESS')
            else:
                self.log(f"  ✗ {item['aztek_id']}  {item['item_name']}", 'INFO')
            if (i + 1) % batch_n == 0 and (i + 1) < total:
                self.log(f'  — batch {(i + 1) // batch_n} done, pause 0.5s —', 'INFO')
                await page.wait_for_timeout(500)
        self.set_progress(total, total, 'เสร็จสิ้น')
        ids = ', '.join(r.get('aztek_id', '') for r in self._results)
        self.log(f'Done! {passed}/{total} passed', 'SUCCESS')
        if ids:
            self.log(f'IDs: {ids}', 'SUCCESS')

    async def _check_item_detail(self, page, item, data):
        aztek_id = item['aztek_id']
        back_url = data.get('list_url', data['url'])
        try:
            clicked = await page.evaluate("""([aid])=>{
                for(const tr of document.querySelectorAll('table tbody tr')){
                    const tds=tr.querySelectorAll('td');
                    if(tds.length&&(tds[0].innerText||'').trim()===aid){
                        const btn=tr.querySelector('a,button');
                        if(btn){btn.click();return true;}
                    }
                }
                return false;
            }""", [aztek_id])
            if not clicked:
                for url_try in (f"{data['url']}/{aztek_id}", f"{data['url']}/{aztek_id}/edit"):
                    try:
                        await page.goto(url_try, wait_until='domcontentloaded', timeout=12000)
                        if aztek_id in page.url:
                            break
                    except Exception:
                        continue
            else:
                await page.wait_for_timeout(1500)

            self.log(f'    URL: {page.url}', 'INFO')
            try:
                await page.wait_for_selector('input[type="checkbox"],[role="checkbox"]', timeout=8000)
            except Exception:
                pass
            # รอจนฟอร์มโหลด "ข้อมูลจริง" และนิ่ง (กันอ่านตอนฟอร์มยังเปล่า -> ค่าผิด)
            loaded = await self._wait_detail_ready(page)
            if not loaded and not self._cancel:
                self.log(f'    ⚠ {aztek_id}: ข้อมูลยังไม่มา — รีโหลดแล้วลองใหม่', 'WARNING')
                try:
                    await page.reload(wait_until='domcontentloaded', timeout=15000)
                    await page.wait_for_selector('input[type="checkbox"],[role="checkbox"]', timeout=8000)
                except Exception:
                    pass
                loaded = await self._wait_detail_ready(page)
            if not loaded:
                # อ่านไม่ได้จริง -> ไม่นับผ่าน (กันผ่านผิดเพราะฟอร์มเปล่า)
                self.log(f'    ✗ {aztek_id}: โหลดรายละเอียดไม่สำเร็จ ข้าม (ไม่นับผ่าน)', 'WARNING')
                await self._go_back(page, back_url)
                return (False, '')

            detail = await page.evaluate("""()=>{
                let webEnabled=null,qty='',tradeable=null,drillable=null,critVal='';
                let hasImage=false,desc='';

                // คำอธิบายไอเทม (textarea name="detail") — โหมด Shop เอามาโชว์เพิ่มในตาราง
                const dta=document.querySelector('textarea[name="detail"]');
                if(dta) desc=(dta.value||dta.textContent||'').trim();

                // Image check — look for item image that actually loaded
                const imgs=[...document.querySelectorAll('img')];
                for(const img of imgs){
                    if(img.naturalWidth>0&&img.naturalHeight>0&&
                       img.src&&!img.src.startsWith('data:')&&
                       img.getBoundingClientRect().width>30){
                        hasImage=true; break;
                    }
                }

                // Helper: get container text
                function parentText(el,levels=6){
                    let cur=el;
                    for(let i=0;i<levels;i++){
                        if(!cur||!cur.parentElement) break;
                        cur=cur.parentElement;
                        const t=(cur.innerText||cur.textContent||'').trim();
                        if(t.length>2&&t.length<200) return t;
                    }
                    return '';
                }

                // Scan checkboxes
                const cbs=[
                    ...document.querySelectorAll('input[type="checkbox"]'),
                    ...document.querySelectorAll('[role="checkbox"]'),
                ];
                const cbInfo=[];
                for(const cb of cbs){
                    const checked=cb.checked!==undefined?cb.checked
                                  :cb.getAttribute('aria-checked')==='true'
                                  ||cb.classList.contains('checked');
                    const txt=parentText(cb);
                    const nm=(cb.getAttribute('name')||'').toLowerCase();
                    cbInfo.push({txt:txt.substring(0,60),checked,name:nm});
                    // ItemMove ในไฟล์ shop = ช่องนี้ (ไม่ติ๊ก = 'ผูกมัดไอดี')
                    // จับจาก name ก่อน แม่นกว่าอ่านข้อความรอบ ๆ
                    if(nm==='is_tradable'){ tradeable=checked; continue; }
                    if(webEnabled===null&&(txt.includes('เปิดใช้งาน')||txt.includes('แสดงผล')))
                        webEnabled=checked;
                    if(tradeable===null&&txt.includes('แลกเปลี่ยน')) tradeable=checked;
                    if(drillable===null&&txt.includes('เจาะรู')) drillable=checked;
                }

                // Scan text inputs
                const inputs=[...document.querySelectorAll(
                    'input[type="text"],input[type="number"],input:not([type])')];
                for(const inp of inputs){
                    const txt=parentText(inp);
                    if(!qty&&txt.includes('จำนวน')) qty=(inp.value||'').trim();
                    if(!critVal&&txt.includes('คริติคอล')) critVal=(inp.value||'').trim();
                }

                return {webEnabled,qty,tradeable,drillable,critVal,hasImage,desc,cbInfo};
            }""")

            self.log(f"    web:{detail.get('webEnabled')} img:{detail.get('hasImage')} "
                     f"qty:{detail.get('qty')!r} trade:{detail.get('tradeable')} "
                     f"drill:{detail.get('drillable')} crit:{detail.get('critVal')!r}", 'INFO')

            # โหมด Shop: เก็บ "คำอธิบายไอเทม" ไว้โชว์เพิ่มในตาราง (ไม่กระทบผ่าน/ตก)
            # ยุบช่องว่าง/ขึ้นบรรทัดใหม่ให้เป็นช่องเดียว -> โชว์ในเซลล์แถวเดียวได้สวย
            if data.get('read_desc'):
                item['_desc'] = ' '.join((detail.get('desc') or '').split())
                if item['_desc']:
                    self.log(f"    คำอธิบาย: {item['_desc']}", 'INFO')

            notes_parts = []

            def chk_bool(key, field, label):
                want = data[key] == 'yes'
                val = detail.get(field)
                if val is None:
                    # ช่อง แลกเปลี่ยนได้/เจาะรูได้/จำนวน/คริติคอล จะโผล่ก็ต่อเมื่อ
                    # ติ๊ก "เปิดใช้งานการแสดงผลบนเว็บ" ไว้ -> ปิดอยู่ = ไม่มีอะไรให้ตรวจ
                    if detail.get('webEnabled') is False:
                        self.log(f'    ✗ {label}: ไอเทมนี้ปิด "แสดงผลบนเว็บ" อยู่ '
                                 f'-> ไม่มีช่อง "{label}" ให้ตรวจ (ต้องเปิดแสดงผลบนเว็บก่อน)',
                                 'WARNING')
                    else:
                        self.log(f'    ✗ {label} = None (อ่านค่าไม่ได้)', 'WARNING')
                    return False
                if val != want:
                    self.log(f'    ✗ {label}: got {val}, want {want}', 'INFO')
                    return False
                notes_parts.append(f"{label}={'Y' if val else 'N'}")
                return True

            def chk_val(key, field, label):
                filt = data[key]
                if not filt:
                    return True
                actual = detail.get(field, '').strip()
                if filt == MUST_EMPTY:          # ต้องไม่มีค่าในช่องนี้
                    if actual:
                        self.log(f'    ✗ {label}: ต้องเว้นว่าง แต่มีค่า {actual!r}', 'INFO')
                        return False
                    notes_parts.append(f'{label}=ว่าง')
                    return True
                if actual != filt:
                    self.log(f'    ✗ {label}: got {actual!r}, want {filt!r}', 'INFO')
                    return False
                notes_parts.append(f'{label}={actual}')
                return True

            # (key, active, fn)
            checks = [
                ('web',   data['web'] != 'any',    lambda: chk_bool('web', 'webEnabled', 'web')),
                ('img',   data['img'] != 'any',    lambda: chk_bool('img', 'hasImage', 'img')),
                ('qty',   bool(data['qty_val']),   lambda: chk_val('qty_val', 'qty', 'qty')),
                ('trade', data['trade'] != 'any',  lambda: chk_bool('trade', 'tradeable', 'trade')),
                ('drill', data['drill'] != 'any',  lambda: chk_bool('drill', 'drillable', 'drill')),
                ('crit',  bool(data['crit_val']),  lambda: chk_val('crit_val', 'critVal', 'crit')),
            ]
            active_keys = [k for k, active, _ in checks if active]
            for key, active, fn in checks:
                if not active:
                    continue
                if not fn():
                    await self._go_back(page, back_url)
                    return (False, '')
                # web=No short-circuit — ใช้ได้เฉพาะตอนไม่มีพารามิเตอร์อื่นให้เช็คเท่านั้น
                # ถ้ามีอย่างอื่นด้วย (เช่น 'แลกเปลี่ยนได้' ของโหมด Shop) ต้องเช็คให้ครบ
                # ไม่งั้นจะ "ผ่านทั้งที่ไม่ได้ตรวจ" แต่ผลลัพธ์ดันโชว์ว่าตรวจแล้ว
                if key == 'web' and data['web'] == 'no' and active_keys == ['web']:
                    self.log('    web=No → ผ่านแล้ว (ไม่มีพารามิเตอร์อื่น) ข้ามการอ่านที่เหลือ', 'INFO')
                    break

            # โหมด "มี": เก็บค่าจริงบนเว็บของช่องที่ไม่ได้เช็ค ไว้โชว์ในตาราง (ไม่กระทบผ่าน/ตก)
            if item.get('_show_web_vals'):
                item['_web_vals'] = web_values_summary(data, detail)
            await self._go_back(page, back_url)
            return (True, ', '.join(notes_parts) if notes_parts else 'passed')
        except Exception as ex:
            self.log(f'    ⚠ {aztek_id}: {ex}', 'WARNING')
            try:
                await self._go_back(page, back_url)
            except Exception:
                pass
            return (False, '')

    async def _go_back(self, page, back_url):
        """Navigate back to the list page (filtered URL preferred over base URL)."""
        try:
            await page.go_back(wait_until='domcontentloaded', timeout=8000)
            await page.wait_for_timeout(600)
        except Exception:
            try:
                await page.goto(back_url, wait_until='domcontentloaded', timeout=15000)
                await page.wait_for_timeout(800)
            except Exception:
                pass


def main():
    try:
        root = tk.Tk()
        root.geometry('680x740')
        root.withdraw()
        App(root)
        root.deiconify()
        root.lift()
        root.focus_force()
        root.mainloop()
    except Exception:
        err = traceback.format_exc()
        try:
            messagebox.showerror('Startup Error', err)
        except Exception:
            print(err)


# ทะเบียนสำหรับ All for Cabal launcher (เพิ่ม tool = ใส่ TOOL นี้ใน all_for_cabal.TOOLS)
from tool_registry import ToolSpec
TOOL = ToolSpec(
    key='finder', icon='🔍', title='Cabal Item ID Finder', nav='🔍  Item Finder',
    desc='ค้นหา / ตรวจรหัสไอเทม (Aztek ID) จากเว็บ', boot='info',
    make=lambda lc, fr: App(lc.root, container=fr, game_var=lc.game_var,
                            on_send_to_bundle=lc._handoff_to_bundle,
                            on_queue_bundles=lc._review_from_finder))


if __name__ == '__main__':
    main()
