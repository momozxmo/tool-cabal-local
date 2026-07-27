# -*- coding: utf-8 -*-
"""Reading a plan file into a draft Item Code.

The numbers here decide how many codes get generated on the live site, so the
rules the desktop worked out against real plans are pinned rather than
described: buffers, which conditions mean one batch and which mean one per
set, and where a date that cannot be read is reported instead of guessed.
"""
from __future__ import annotations

import os
import sys
import tempfile
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import pytest  # noqa: E402

import item_finder  # noqa: E402
from web.itemcode_plan import build_itemcodes, code_name  # noqa: E402

NOW = datetime(2026, 7, 26, 15, 40)
SEA, TH = 'CabalM SEA', 'CabalPC TH'


def _event(**extra):
    """A prize table's conditions, as Item Finder reads them."""
    meta = {'activity': 'Storm Chaser', 'reward': 'WINNER REWARDS',
            'expire': '2026-08-31 00:00:00', 'codes_per_set': '35',
            'set_count': '1', 'total': '35', 'unique_code': True,
            'cannot_repeat': False, 'can_repeat': True, 'once_per_set': True,
            'cross_can': False, 'cross_cannot': False}
    meta.update(extra)
    return meta


def _pride(**extra):
    block = {'title': 'CBM Master Code (PH)', 'event_name': 'Live Stream',
             'expire': datetime(2026, 12, 9), 'fix_code': 'CBPA7X9Q2MW',
             'unique_code': False, 'refill_limit': '3000', 'code_count': '',
             'cannot_repeat': True, 'note': ''}
    block.update(extra)
    return {'is_pride': True, 'pride': block}


def _one(meta, game=SEA, group='g'):
    drafts = build_itemcodes({group: meta}, game, now=NOW)
    assert len(drafts) == 1
    return drafts[0]


# --------------------------------- naming ---------------------------------

def test_the_name_is_the_activity_and_the_table_it_came_from():
    """A sheet holds several prize tables and each is its own Item Code, so the
    activity alone would name them all the same thing."""
    assert code_name(_event()) == 'Storm Chaser - WINNER REWARDS'
    assert code_name(_pride()) == 'Live Stream - CBM Master Code (PH)'


def test_the_slug_carries_the_server_it_is_for():
    assert _one(_event())['slug'] == 'storm-chaser-winner-rewards-msea'
    assert _one(_event(), TH)['slug'] == 'storm-chaser-winner-rewards-pcth'


def test_a_thai_name_cannot_become_a_slug_and_says_so():
    """Aztek takes a-z0-9- only; an empty required field with no explanation is
    worse than being told to type one."""
    draft = _one(_event(activity='กิจกรรมทดสอบ', reward=''))
    assert draft['slug'] == ''
    assert any('slug' in note for note in draft['notes'])


# --------------------------------- counts ---------------------------------

def test_a_single_set_is_generated_as_one_batch_with_the_buffer():
    """35 codes asked for, +5 on SEA — the buffer covers ones that fail."""
    reward = _one(_event())['rewards'][0]
    assert reward['num_codes'] == '40'
    assert (reward['quantity'], reward['remaining']) == ('40', '40')
    assert reward['limited'] is True


def test_the_item_code_itself_carries_no_counts():
    """The counts belong to the reward set — the code-wide "จำกัดจำนวน" is not
    part of how these are written, so nothing here may set it."""
    draft = _one(_event())
    assert not {'limited', 'quantity', 'remaining', 'kind',
                'desc_th', 'desc_en'} & set(draft)


def test_the_buffer_is_bigger_on_pc_th():
    assert _one(_event(), TH)['rewards'][0]['num_codes'] == '45'


def test_once_per_set_makes_one_reward_set_each_and_buffers_only_the_first_two():
    draft = _one(_event(set_count='4', codes_per_set='100', total='400',
                        once_per_set=True, can_repeat=False))
    assert [r['num_codes'] for r in draft['rewards']] == ['105', '105', '100', '100']
    assert [r['name_th'] for r in draft['rewards']][:2] == [
        'Storm Chaser - WINNER REWARDS 1', 'Storm Chaser - WINNER REWARDS 2']


def test_a_code_that_may_not_be_reused_is_one_use_per_player():
    draft = _one(_event(cannot_repeat=True, can_repeat=False))
    assert draft['uses_per_user'] == '1'
    assert all(r['uses_per_user'] == '1' for r in draft['rewards'])


def test_a_repeatable_code_lets_a_player_use_every_one_of_them():
    assert _one(_event(cannot_repeat=False, can_repeat=True))['uses_per_user'] == '40'


def test_an_absurd_number_of_sets_is_capped_and_reported():
    draft = _one(_event(set_count='5000', codes_per_set='1', total='5000',
                        once_per_set=True, can_repeat=False))
    assert len(draft['rewards']) == 300
    assert any('5000' in note for note in draft['notes'])


# ---------------------------------- codes ----------------------------------

def test_a_unique_code_is_generated_by_the_server():
    reward = _one(_event(unique_code=True))['rewards'][0]
    assert reward['code_type'] == '2'
    assert reward['code_list'] == ''


def test_a_master_code_is_the_one_written_in_the_file():
    reward = _one(_pride())['rewards'][0]
    assert reward['code_type'] == '1'
    assert reward['code_list'] == 'CBPA7X9Q2MW'
    # A fixed code is capped by how many times it may be redeemed.
    assert (reward['quantity'], reward['remaining']) == ('3000', '3000')


def test_a_pride_block_with_no_code_at_all_is_flagged():
    draft = _one(_pride(fix_code='', unique_code=False))
    assert any('โค้ด' in note for note in draft['notes'])


def test_a_generated_pride_code_uses_the_count_it_asked_for():
    reward = _one(_pride(unique_code=True, code_count='500'))['rewards'][0]
    assert (reward['code_type'], reward['num_codes']) == ('2', '500')


def test_a_note_in_the_file_is_carried_over():
    draft = _one(_pride(note='คนละวันกับ ID (วันแรก)'))
    assert 'คนละวันกับ ID (วันแรก)' in ' '.join(draft['notes'])


# ---------------------------------- dates ----------------------------------

def test_the_run_starts_today_and_ends_when_the_file_says():
    draft = _one(_event())
    assert draft['start_time'] == '2026-07-26T00:00:00'
    # SEA closes an hour earlier than TH — a regional rule, not a rounding.
    assert draft['end_time'] == '2026-08-31T22:59:59'
    assert _one(_event(), TH)['end_time'] == '2026-08-31T23:59:59'


def test_an_expiry_that_arrived_as_an_excel_serial_is_still_a_date():
    """Reading the plan turns date formats off to recover ItemKind, which
    leaves the expiry as a serial number."""
    assert _one(_event(expire='46265.0'))['end_time'] == '2026-08-31T22:59:59'


def test_an_unreadable_expiry_is_reported_not_guessed():
    draft = _one(_event(expire='2 เดือน'))
    assert draft['end_time'] == ''
    assert any('2 เดือน' in note for note in draft['notes'])


def test_an_expiry_that_has_already_passed_is_flagged():
    """Plan files get re-used for the next run of the same activity."""
    draft = _one(_event(expire='2026-07-01 00:00:00'))
    assert any('ผ่านมาแล้ว' in note for note in draft['notes'])


def test_a_missing_expiry_stays_blank_and_says_to_fill_it():
    draft = _one(_event(expire=''))
    assert draft['end_time'] == ''
    assert any('วันหมดอายุ' in note for note in draft['notes'])


# --------------------------------- selection ---------------------------------

def test_only_the_groups_the_plan_describes_become_drafts():
    drafts = build_itemcodes({'a': _event()}, SEA, groups=['a', 'never-imported'],
                             now=NOW)
    assert [d['group'] for d in drafts] == ['a']


def test_the_order_asked_for_is_the_order_returned():
    meta = {'b': _event(reward='B'), 'a': _event(reward='A')}
    assert [d['group'] for d in build_itemcodes(meta, SEA, groups=['a', 'b'],
                                                now=NOW)] == ['a', 'b']


# ---------------------------- reading the workbook ----------------------------

def _sheet(rows):
    """Write one sheet and read it back through the real import path."""
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'plan'
    for row in rows:
        sheet.append(list(row))
    handle, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(handle)
    book.save(path)
    try:
        return item_finder.parse_event_workbook(path)[0]
    finally:
        os.unlink(path)


def _block(code_no, banner, per_set, kinds):
    """One prize table the way the ITEM CODE plan writes it.

    ``code_no`` of 0 is a sheet that numbers nothing — real ones exist, and
    they are the case where two tables have nothing to tell them apart.
    """
    head = ('CONDITIONS (Only for "Code No. %d")' % code_no) if code_no \
        else 'CONDITIONS'
    return [
        ['', '', '', '', '', '', '', head],
        ['', '', '', '', '', '', '', 'CODE EXPIRE DATE', '', '2026-08-31 00:00:00'],
        ['', '', '', '', '', '', '', 'Additional conditions A', '', 'Cannot be repeated'],
        [],
        ['', '', '', '', '', '', '', 'Prize'],
        ['', '', '', '', '', '', '', banner, 'Unique Code', '', '',
         '', '', 'Number [code/set]', 'Number [set]', '', 'Total'],
        ['', '', '', '', '', '', '', '', '', '', '', '', '', per_set, 1, '', per_set],
        [],
        ['', '', '', '', '', '', '', 'Item Kind', 'Item Index', 'ItemOption',
         'DurationIndex', 'Stackable', 'Display Name'],
    ] + [['', '', '', '', '', '', '', kind, '1', '0', '10', 'No', 'Thing %s' % kind]
         for kind in kinds] + [[]]


def _headerless_block():
    """A repeated live-plan prize block which omits its Item Kind header."""
    return [
        ['', '', '', '', '', '', '', 'CONDITIONS (Only for "Code No. 2")'],
        ['', '', '', '', '', '', '', 'CODE EXPIRE DATE', '',
         '2026-08-31 00:00:00'],
        ['', '', '', '', '', '', '', 'Additional conditions A', '',
         'Cannot be repeated'],
        [],
        ['', '', '', '', '', '', '', 'Prize'],
        ['', '', '', '', '', '', '', 'GIVEAWAY', 'Unique Code', '', '',
         '', '', 'Number [code/set]', 'Number [set]', '', 'Total'],
        ['', '', '', '', '', '', '', '', '', '', '', '', '', 20, 16, '', 320],
        [],
        ['Prize'],
        ['', '', '', '', '', '', '', '333', '1', '0', '10', 'No',
         'Giveaway Thing'],
    ]


def test_a_plan_sheet_becomes_one_group_per_prize_table():
    sheets = _sheet([['Storm Chaser']]
                    + _block(1, 'WINNER REWARDS', 35, ['111', '222'])
                    + _block(2, 'Participation Rewards', 350, ['333']))
    assert len(sheets) == 1
    groups = [row['sources'][0] for _name, rows in sheets for row in rows]
    assert groups == ['Storm Chaser WINNER REWARDS'] * 2 \
        + ['Storm Chaser Participation Rewards']


def test_stacked_tables_stay_apart_even_with_nothing_to_tell_them_apart():
    """Two codes with no section banner and no code number used to be read as
    one group, which merged two Item Codes into a single form."""
    sheets = _sheet([['August E-Card']]
                    + _block(0, '', 50, ['111'])
                    + _block(0, '', 60, ['222']))
    groups = [row['sources'][0] for _name, rows in sheets for row in rows]
    assert groups == ['August E-Card', 'August E-Card (2)']


def test_a_later_prize_block_can_reuse_the_first_tables_missing_header():
    """Some live plan tabs omit the repeated Item Kind header on their second
    prize block.  Its rows still use the same columns and must become another
    Item Code instead of disappearing from the handoff."""
    sheets = _sheet([['Apple, Orange, Ensaymada?']]
                    + _block(1, 'WINNER REWARDS', 20, ['111'])
                    + _headerless_block())

    rows = sheets[0][1]
    groups = [row['sources'][0] for row in rows]
    assert groups == [
        'Apple, Orange, Ensaymada? WINNER REWARDS',
        'Apple, Orange, Ensaymada? GIVEAWAY',
    ]
    assert rows[1]['group_meta']['codes_per_set'] == '20'
    assert rows[1]['group_meta']['set_count'] == '16'


def test_the_singular_spelling_of_the_codes_column_is_read_too():
    """Older sheets write 'Number [code/set]' rather than 'codes per set'."""
    sheets = _sheet([['Quiz Night']] + _block(1, 'WINNER REWARDS', 50, ['111']))
    meta = sheets[0][1][0]['group_meta']
    assert meta['codes_per_set'] == '50'
    assert meta['set_count'] == '1'


def _workbook_bytes(rows, title='plan'):
    import io

    import openpyxl

    book = openpyxl.Workbook()
    book.active.title = title
    for row in rows:
        book.active.append(list(row))
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_finder_handoff_keeps_a_later_headerless_item_code(client):
    payload = _workbook_bytes(
        [['Apple, Orange, Ensaymada?']]
        + _block(1, 'WINNER REWARDS', 20, ['111'])
        + _headerless_block(),
        title='PH July- STM - Apple, Orange, E',
    )
    started = client.post(
        '/api/import-plan', data={'mode': 'itemcode'},
        files={'file': ('plan.xlsx', payload,
                        'application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.sheet')},
    ).json()
    assert started['sheets'] == [{
        'name': 'PH July- STM - Apple, Orange, E',
        'count': 2,
    }]
    applied = client.post('/api/import-plan/apply', json={
        'pending_id': started['pending_id'],
        'selected_sheets': ['PH July- STM - Apple, Orange, E'],
    })
    assert applied.status_code == 200, applied.text

    drafts = client.get(
        '/api/workspaces/%s/itemcodes' % started['workspace_id']
    ).json()['itemcodes']
    assert [draft['name_th'] for draft in drafts] == [
        'Apple, Orange, Ensaymada? - WINNER REWARDS',
        'Apple, Orange, Ensaymada? - GIVEAWAY',
    ]
    assert len(drafts[1]['rewards']) == 16


def test_importing_a_plan_on_this_page_needs_a_session(anonymous_client):
    assert anonymous_client.post('/api/itemcodes/import').status_code == 401


def test_import_reports_the_tabs_and_stamps_each_draft_with_its_own(client):
    """The page offers the tabs to pick from, so a draft has to know which one
    it came from — a plan holds dozens of activities and a run is about a few."""
    payload = _workbook_bytes(
        [['Quiz Night']] + _block(1, 'WINNER REWARDS', 50, ['111'])
        + _block(2, 'Participation', 60, ['222']), title='ID COM Quiz')
    response = client.post(
        '/api/itemcodes/import', data={'game': SEA},
        files={'file': ('plan.xlsx', payload,
                        'application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.sheet')})
    assert response.status_code == 200, response.text
    body = response.json()
    assert body['sheets'] == [{'name': 'ID COM Quiz', 'count': 2}]
    assert [d['sheet'] for d in body['itemcodes']] == ['ID COM Quiz'] * 2
    assert [d['name_th'] for d in body['itemcodes']] == [
        'Quiz Night - WINNER REWARDS', 'Quiz Night - Participation']
    assert body['itemcodes'][0]['rewards'][0]['num_codes'] == '55'


def test_only_the_tabs_chosen_on_the_finder_page_are_kept(client):
    """A plan holds dozens of activities. What the operator did not tick must
    not turn up later as an Item Code nobody asked for."""
    import io

    import openpyxl

    book = openpyxl.Workbook()
    for index, (title, activity) in enumerate(
            [('Wanted', 'Quiz Night'), ('Skipped', 'Other Thing')]):
        sheet = book.active if index == 0 else book.create_sheet()
        sheet.title = title
        for row in [[activity]] + _block(1, 'WINNER REWARDS', 50, ['111']):
            sheet.append(list(row))
    buffer = io.BytesIO()
    book.save(buffer)

    started = client.post(
        '/api/import-plan', data={'mode': 'itemcode'},
        files={'file': ('plan.xlsx', buffer.getvalue(),
                        'application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.sheet')}).json()
    assert {sheet['name'] for sheet in started['sheets']} == {'Wanted', 'Skipped'}
    applied = client.post('/api/import-plan/apply', json={
        'pending_id': started['pending_id'], 'selected_sheets': ['Wanted']})
    assert applied.status_code == 200, applied.text

    drafts = client.get('/api/workspaces/%s/itemcodes'
                        % started['workspace_id']).json()['itemcodes']
    assert [d['name_th'] for d in drafts] == ['Quiz Night - WINNER REWARDS']


def test_a_file_with_no_prize_table_imports_as_nothing_rather_than_failing(client):
    response = client.post(
        '/api/itemcodes/import', data={'game': SEA},
        files={'file': ('empty.xlsx', _workbook_bytes([['just a note']]),
                        'application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.sheet')})
    assert response.status_code == 200
    assert response.json()['itemcodes'] == []


@pytest.mark.parametrize('game, expected', [(SEA, '55'), (TH, '60')])
def test_the_whole_way_through_from_workbook_to_draft(game, expected):
    sheets = _sheet([['Quiz Night']] + _block(1, 'WINNER REWARDS', 50, ['111']))
    row = sheets[0][1][0]
    drafts = build_itemcodes({row['sources'][0]: row['group_meta']}, game, now=NOW)
    assert drafts[0]['name_th'] == 'Quiz Night - WINNER REWARDS'
    assert drafts[0]['rewards'][0]['num_codes'] == expected
    assert drafts[0]['rewards'][0]['code_type'] == '2'
    assert drafts[0]['uses_per_user'] == '1'          # Cannot be repeated
