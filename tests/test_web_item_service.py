# -*- coding: utf-8 -*-
"""Parity tests for the browser Item Finder service (GUI-free)."""
import csv
import io
import os
import sys
import threading
import time

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from web import item_service as svc  # noqa: E402
import event_tool  # noqa: E402
import item_finder  # noqa: E402


def test_mode_policy_matches_desktop():
    assert svc.mode_policy('event') == {'web_mode': 'any', 'web_locked': False,
                                        'read_desc': False}
    assert svc.mode_policy('itemcode') == {'web_mode': 'no', 'web_locked': True,
                                           'read_desc': False}
    assert svc.mode_policy('shop') == {'web_mode': 'no', 'web_locked': False,
                                           'read_desc': True}


def test_parser_for_mode_matches_desktop_import_paths():
    assert svc.parser_for_mode('event') is event_tool.parse_event_plan_workbook
    assert svc.parser_for_mode('itemcode') is item_finder.parse_event_workbook
    assert svc.parser_for_mode('shop') is item_finder.parse_shop_workbook


def test_merge_imported_deduplicates_search_but_preserves_occurrences_and_metadata():
    rows = [
        {'kind': '1', 'opt': '2', 'dur': '0', 'name': 'A', 'sources': ['G1'],
         'group_meta': {'activity': 'Sheet 1'}},
        {'kind': '1', 'opt': '2', 'dur': '0', 'name': 'A', 'sources': ['G2'],
         'group_meta': {'activity': 'Sheet 2'}},
        {'kind': '3', 'opt': '', 'dur': '', 'name': 'B', 'sources': ['G1']},
    ]
    merged = svc.merge_imported([], [], {}, rows)
    assert len(merged.criteria) == 2
    assert merged.criteria[0]['sources'] == ['G1', 'G2']
    assert len(merged.occurrences) == 3
    assert merged.group_meta['G1']['activity'] == 'Sheet 1'
    assert merged.group_meta['G2']['activity'] == 'Sheet 2'
    assert all('group_meta' not in row for row in merged.criteria)


def test_regroup_results_follows_document_occurrences():
    found = [
        {'aztek_id': '10', 'item_kind': '1', 'item_option': '2',
         'duration_index': '0', 'item_name': 'A-web', 'sources': ['G1', 'G2']},
        {'aztek_id': '30', 'item_kind': '3', 'item_option': '',
         'duration_index': '', 'item_name': 'B-web', 'sources': ['G1']},
    ]
    occurrences = [
        {'kind': '1', 'opt': '2', 'dur': '0', 'name': 'A-file', 'sources': ['G1']},
        {'kind': '3', 'opt': '', 'dur': '', 'name': 'B-file', 'sources': ['G1']},
        {'kind': '1', 'opt': '2', 'dur': '0', 'name': 'A-file', 'sources': ['G2']},
    ]
    rows = svc.regroup_results(found, occurrences)
    assert [r['aztek_id'] for r in rows] == ['10', '30', '10']
    assert [r['sources'] for r in rows] == [['G1'], ['G1'], ['G2']]
    assert rows[0]['file_name'] == 'A-file'


def test_regroup_and_bundles_carry_amt_as_qty():
    # The imported 'Amt' column must flow occurrence -> row -> bundle qty.
    found = [{'aztek_id': '10', 'item_kind': '1', 'item_option': '2',
              'duration_index': '0', 'item_name': 'A-web', 'sources': ['G1']}]
    occurrences = [{'kind': '1', 'opt': '2', 'dur': '0', 'name': 'A',
                    'amt': '5', 'sources': ['G1']}]
    rows = svc.regroup_results(found, occurrences)
    assert rows[0]['amt'] == '5'
    bundles = svc.build_bundles(
        [{'aztek_id': '10', 'item_name': 'A-web', 'amt': '5', 'sources': ['G1']}],
        {})
    assert bundles[0]['items'][0]['qty'] == '5'
    # Missing/blank Amt falls back to '1'.
    blank = svc.build_bundles(
        [{'aztek_id': '11', 'item_name': 'B', 'sources': ['G1']}], {})
    assert blank[0]['items'][0]['qty'] == '1'


def test_build_bundles_keeps_the_order_the_plan_file_listed():
    """The plan file's order is the order the bundle should read in.

    Shared items used to be swept to the bottom of every group, which quietly
    rearranged bundles the operator had already put in sequence. They keep their
    place now and are found by their highlight instead.
    """
    rows = [
        {'aztek_id': '99', 'item_name': 'Shared', 'sources': ['G1']},
        {'aztek_id': '11', 'item_name': 'Only G1', 'sources': ['G1']},
        {'aztek_id': '99', 'item_name': 'Shared', 'sources': ['G2']},
        {'aztek_id': '22', 'item_name': 'Only G2', 'sources': ['G2']},
    ]
    meta = {'G1': {'activity': 'Event'}, 'G2': {'shop_sheet': 'Cash Shop'}}
    bundles = svc.build_bundles(rows, meta)
    assert [b['name'] for b in bundles] == ['Event - G1', 'Cash Shop - G2']
    assert [[it['id'] for it in b['items']] for b in bundles] == [
        ['99', '11'], ['99', '22']]
    # Still flagged, just not moved.
    assert [it['shared'] for it in bundles[0]['items']] == [True, False]


def test_missing_criteria_are_the_rows_nothing_came_back_for():
    """Searching again should cost only the rows that failed, not the whole plan."""
    criteria = [
        {'kind': '1', 'opt': '', 'dur': '', 'name': 'found'},
        {'kind': '2', 'opt': '5', 'dur': '', 'name': 'missing'},
        {'kind': '3', 'opt': '', 'dur': '7', 'name': 'also missing'},
    ]
    results = [{'aztek_id': '10', 'item_kind': '1', 'item_option': '',
                'duration_index': ''}]
    missing = svc.missing_criteria(criteria, results)
    assert [row['name'] for row in missing] == ['missing', 'also missing']
    # Each keeps the number it had in the plan, not its place in the retry.
    assert missing[0]['_label'].startswith('#2 ')
    assert missing[1]['_label'].startswith('#3 ')


def test_a_retry_adds_to_what_was_already_found():
    """The first pass's results have to survive a run that only redoes the
    misses — and the newcomers belong at their place in the document, not
    stacked at the bottom."""
    occurrences = [
        {'kind': '1', 'opt': '', 'dur': '', 'name': 'first', 'sources': ['G1']},
        {'kind': '2', 'opt': '', 'dur': '', 'name': 'second', 'sources': ['G1']},
        {'kind': '3', 'opt': '', 'dur': '', 'name': 'third', 'sources': ['G2']},
    ]
    previous = [
        {'aztek_id': '10', 'item_kind': '1', 'item_option': '', 'duration_index': ''},
        {'aztek_id': '30', 'item_kind': '3', 'item_option': '', 'duration_index': ''},
    ]
    fresh = [
        {'aztek_id': '20', 'item_kind': '2', 'item_option': '', 'duration_index': ''}]
    merged = svc.merge_found(previous, fresh, occurrences)
    assert [row['aztek_id'] for row in merged] == ['10', '20', '30']
    assert [row['groups'] for row in merged] == ['G1', 'G1', 'G2']


def test_a_retry_does_not_duplicate_what_it_finds_again():
    occurrences = [{'kind': '1', 'opt': '', 'dur': '', 'sources': ['G1']}]
    row = {'aztek_id': '10', 'item_kind': '1', 'item_option': '',
           'duration_index': ''}
    assert len(svc.merge_found([row], [dict(row)], occurrences)) == 1


def test_bundle_items_carry_what_the_document_said():
    """A bundle is checked against the plan file, so the plan file's own words
    travel with it — otherwise the check means going back to a results table
    holding every group at once."""
    rows = [{'aztek_id': '11', 'item_name': 'Leaf Gem [30d]',
             'file_name': 'Leaf Gem 30 วัน', 'name_mismatch': True,
             'params': 'เว็บ✓ เทรด✓ จำนวน∅', 'desc': 'ใช้แล้วได้ Gem',
             'amt': '5', 'sources': ['G1']}]
    item = svc.build_bundles(rows, {})[0]['items'][0]
    assert item['file_name'] == 'Leaf Gem 30 วัน'
    assert item['name_mismatch'] is True
    assert item['params'] == 'เว็บ✓ เทรด✓ จำนวน∅'
    assert item['desc'] == 'ใช้แล้วได้ Gem'
    # The count starts at what the document asked for, and says so separately so
    # the page can flag a value typed over it.
    assert (item['qty'], item['doc_qty']) == ('5', '5')


def test_a_random_box_arrives_as_a_random_bundle_with_its_odds():
    """The plan's rate column is the whole reason that product is a random box;
    re-typing the odds by hand is exactly what the import is there to avoid."""
    rows = [{'aztek_id': '11', 'item_name': 'A', 'rate': '40', 'sources': ['G1']},
            {'aztek_id': '22', 'item_name': 'B', 'rate': '60', 'sources': ['G1']}]
    bundle = svc.build_bundles(rows, {'G1': {'is_random': True}})[0]
    assert bundle['is_random'] is True
    assert [it['rate'] for it in bundle['items']] == ['40', '60']


def test_a_product_with_no_odds_is_not_a_random_bundle():
    rows = [{'aztek_id': '11', 'item_name': 'A', 'sources': ['G1']}]
    bundle = svc.build_bundles(rows, {'G1': {'is_shop': True}})[0]
    assert bundle['is_random'] is False
    assert bundle['items'][0]['rate'] == ''


def test_bundle_name_uses_event_name_from_import():
    # Event imports carry 'event_name'; it should title the bundle.
    rows = [{'aztek_id': '1', 'item_name': 'X', 'sources': ['รางวัลที่ 1']}]
    meta = {'รางวัลที่ 1': {'event_name': 'กิจกรรมปีใหม่', 'is_event': True}}
    bundles = svc.build_bundles(rows, meta)
    assert bundles[0]['name'] == 'กิจกรรมปีใหม่ - รางวัลที่ 1'
    assert bundles[0]['group'] == 'รางวัลที่ 1'


def test_exports_include_full_result_columns():
    rows = [{'aztek_id': '10', 'item_name': 'A', 'item_kind': '1',
             'item_option': '2', 'duration_index': '0', 'game': 'CabalM SEA',
             'notes': 'passed', '_ci': 1, 'sources': ['G1'], '_desc': 'คำอธิบาย'}]
    csv_bytes = svc.export_csv_bytes(rows)
    parsed = list(csv.reader(io.StringIO(csv_bytes.decode('utf-8-sig'))))
    assert parsed[0][-1] == 'description'
    assert parsed[1][1] == '10' and parsed[1][-1] == 'คำอธิบาย'

    xlsx_bytes = svc.export_xlsx_bytes(rows, 'CabalM SEA')
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    assert ws.cell(4, 11).value == 'คำอธิบายไอเทม'
    assert ws.cell(5, 2).value == '10'
    assert ws.cell(5, 11).value == 'คำอธิบาย'


def test_exports_escape_spreadsheet_formulas():
    rows = [{'aztek_id': '10', 'item_name': '=HYPERLINK("bad")',
             'sources': ['+CMD'], '_desc': '@payload'}]
    parsed = list(csv.reader(io.StringIO(
        svc.export_csv_bytes(rows).decode('utf-8-sig'))))
    assert parsed[1][2].startswith("'=")
    assert parsed[1][9].startswith("'+")
    assert parsed[1][10].startswith("'@")
    workbook = openpyxl.load_workbook(
        io.BytesIO(svc.export_xlsx_bytes(rows, 'CabalM SEA')), data_only=False)
    assert workbook.active.cell(5, 3).value.startswith("'=")


def test_workspace_store_applies_selected_sheets_and_preserves_state():
    store = svc.WorkspaceStore()
    workspace = store.create('shop', 'monthly.xlsx')
    upload = store.add_pending(
        workspace.id,
        [('Cash Shop', [{'kind': '1', 'opt': '', 'dur': '', 'name': 'A',
                         'sources': ['P1'], 'group_meta': {'shop_sheet': 'Cash Shop'}}]),
         ('Ignore Me', [{'kind': '2', 'opt': '', 'dur': '', 'name': 'B',
                         'sources': ['P2']}])],
        ['Bad item'],
    )
    applied = store.apply_pending(upload.id, ['Cash Shop'])
    assert applied.id == workspace.id
    assert [row['kind'] for row in applied.criteria] == ['1']
    assert applied.occurrences[0]['sources'] == ['P1']
    assert applied.group_meta['P1']['shop_sheet'] == 'Cash Shop'
    assert applied.skipped == ['Bad item']
    assert store.get(workspace.id) is applied


def test_workspace_store_template_replaces_items_and_clear_removes_workspace():
    store = svc.WorkspaceStore()
    first = store.create('event', 'one.xlsx', [{'kind': '1'}])
    replaced = store.replace_template(first.id, 'two.xlsx', [{'kind': '2'}, {'kind': '3'}])
    assert replaced.filename == 'two.xlsx'
    assert [row['kind'] for row in replaced.criteria] == ['2', '3']
    assert replaced.occurrences == []
    store.delete(first.id)
    try:
        store.get(first.id)
        assert False, 'workspace should be gone'
    except KeyError:
        pass


def test_workspace_store_concurrent_apply_does_not_lose_an_import():
    store = svc.WorkspaceStore()
    workspace = store.create('event')
    first = store.add_pending(workspace.id, [('One', [{'kind': '1'}])])
    second = store.add_pending(workspace.id, [('Two', [{'kind': '2'}])])
    original = svc.merge_imported
    calls = 0
    calls_lock = threading.Lock()

    def slow_first(*args, **kwargs):
        nonlocal calls
        with calls_lock:
            calls += 1
            call_number = calls
        if call_number == 1:
            time.sleep(0.08)
        return original(*args, **kwargs)

    svc.merge_imported = slow_first
    try:
        threads = [
            threading.Thread(target=store.apply_pending, args=(first.id, ['One'])),
            threading.Thread(target=store.apply_pending, args=(second.id, ['Two'])),
        ]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join(timeout=2)
        assert sorted(row['kind'] for row in store.get(workspace.id).criteria) == ['1', '2']
    finally:
        svc.merge_imported = original


def test_parse_workbook_locked_serializes_global_openpyxl_patch_users():
    active = 0
    max_active = 0
    guard = threading.Lock()

    def parser(path):
        nonlocal active, max_active
        with guard:
            active += 1
            max_active = max(max_active, active)
        time.sleep(0.05)
        with guard:
            active -= 1
        return path

    threads = [threading.Thread(target=svc.parse_workbook_locked,
                                args=(parser, str(index))) for index in range(3)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=2)
    assert max_active == 1


if __name__ == '__main__':
    tests = [v for k, v in sorted(globals().items()) if k.startswith('test_')]
    for test in tests:
        test()
        print('PASS', test.__name__)
    print('ALL PASS (%d)' % len(tests))
