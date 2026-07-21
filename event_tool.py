# -*- coding: utf-8 -*-
"""event_tool.py — เครื่องมือ "สร้าง Event" บนเว็บ aztek-tools (events/create)

หน้าตาเหมือน Create Item Code แต่:
  - ประเภท (Type) = WINNER เสมอ
  - จำนวนครั้งที่ใช้ได้ / จำนวนคงเหลือ = 0
  - ชุดรางวัลมีแค่: ชื่อ (ไทย/อังกฤษ) + Limited + ใช้ต่อ 1 User + Bundles
  - bundle_id ของแต่ละชุดรางวัล "อยู่ในไฟล์ template แล้ว" (คอลัมน์ Bundle ID) จึงไม่ต้องหา/สร้างใหม่

Template (Monthly Plan): ชื่อ Event อยู่หลังคำว่า "Event Name" ; แต่ละชุดรางวัลมีแบนเนอร์
"Reward:  <ชื่อ>: N รางวัล" (ตัด ": N รางวัล" ออก) + ตาราง item ที่มี Bundle ID อยู่ในแถวหัวตาราง

รัน:  python event_tool.py   (หรือเปิดผ่าน All for Cabal)
"""

import os
import re
import asyncio
import threading
import traceback
from datetime import datetime, timedelta

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import aztek_core as core
import config as cfg
# web helper กลาง (form driver aztek-tools)
from aztek_core import _find, _fill, _set_datetime, _add_bundle
# UI kit จากของกลาง (ไม่พึ่ง itemcode_tool อีก)
from ui_common import (C, F, FB_, MONO, SM, make_button, card, labeled_entry,
                       DateTimePicker, _find_icon, _set_window_icon)
# parser ที่ยังใช้ร่วมกับ Create Item Code
from itemcode_tool import slugify, parse_expire

LIST_PATH = "events"
CREATE_PATH = "events/create"


# --------------------------------------------------------------------------- parser
def _norm(x):
    return str(x).strip().lower().replace(' ', '').replace('\n', '') if x is not None else ''


def _isnum(x):
    try:
        float(str(x).strip())
        return True
    except Exception:
        return False


def _parse_thai_time(s):
    """'12.00 น.' -> (12,0) ; '23:59 น.' -> (23,59) ; อ่านไม่ได้คืน None"""
    m = re.search(r'(\d{1,2})[.:](\d{2})', str(s or ''))
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if 0 <= h < 24 and 0 <= mi < 60:
            return h, mi
    return None


def _clean_reward_name(s):
    """'Reward:  ตอบถูก! ตอบไว!: 10 รางวัล' -> 'ตอบถูก! ตอบไว!' (ตัด 'Reward:' และ ': N รางวัล')"""
    s = str(s or '').strip()
    s = re.sub(r'^\s*reward\s*:\s*', '', s, flags=re.I)
    s = re.sub(r'\s*:\s*\d+\s*ราง?วัล.*$', '', s).strip()   # ตัด ": 10 รางวัล" ท้ายสุด
    return s


def _parse_plan_sheet(rows):
    """คืน dict {name, rewards:[{name, bundle_id, items:[...]}]} จาก 1 sheet (Event template)"""
    event_name = ''
    for row in rows[:40]:
        for ci, c in enumerate(row):
            if _norm(c) == 'eventname':
                for cj in range(ci + 1, len(row)):
                    if row[cj] is not None and str(row[cj]).strip():
                        event_name = str(row[cj]).strip().replace('\n', ' ')
                        break
                break
        if event_name:
            break

    # ระยะเวลากิจกรรม: แถว 'Start' (วันที่ col3 + เวลา 'น.') / 'End' (เวลา 'น.' อาจไม่มีวัน)
    start_dt = end_date = end_time = None
    for row in rows[:40]:
        c2 = _norm(row[2]) if len(row) > 2 else ''
        if c2 in ('start', 'end'):
            d = t = None
            for c in row:
                if isinstance(c, datetime):
                    d = c
                elif isinstance(c, (int, float)) and not isinstance(c, bool) and 20000 < float(c) < 90000:
                    d = datetime(1899, 12, 30) + timedelta(days=int(c))   # Excel serial (patch ปิด date_formats)
                elif isinstance(c, str) and 'น.' in c:
                    t = _parse_thai_time(c)
            if c2 == 'start' and d:
                hh, mm = t or (0, 0)
                start_dt = d.replace(hour=hh, minute=mm, second=0, microsecond=0)
            elif c2 == 'end':
                end_date, end_time = d, t

    rewards = []
    banner = ''
    col = None
    for row in rows:
        # แบนเนอร์ชื่อรางวัล 'Reward: ...' (ไม่สนใจ Bundle ID / คอลัมน์นอกกรอบ)
        for c in row:
            s = str(c).strip() if c is not None else ''
            if s.lower().startswith('reward:'):
                banner = _clean_reward_name(s)
        cn = [_norm(c) for c in row]
        if 'itemkind' in cn:                       # แถวหัวตาราง item ของชุดรางวัลนี้
            col = {}
            for i, k in enumerate(cn):
                if k == 'itemkind':
                    col['kind'] = i
                elif k == 'itemindex':
                    col['index'] = i
                elif k in ('itemname', 'displayname'):
                    col.setdefault('name', i)
                elif k in ('amount', 'amt'):
                    col['amt'] = i
            rewards.append({'name': banner or ('รางวัลที่ %d' % (len(rewards) + 1)),
                            'items': [], '_col': col})
            banner = ''
            continue
        if col is not None and rewards:
            kc = col.get('kind')
            kv = row[kc] if (kc is not None and kc < len(row)) else None
            if kv is not None and _isnum(kv):
                nm = row[col['name']] if (col.get('name') is not None and col['name'] < len(row)) else ''
                rewards[-1]['items'].append({
                    'kind': str(int(float(str(kv).strip()))),
                    'name': str(nm).strip() if nm else '',
                })
            elif kv is None or str(kv).strip() == '':
                col = None                         # จบตารางชุดนี้

    for rw in rewards:
        rw.pop('_col', None)
    return {'name': event_name, 'rewards': rewards,
            'start_dt': start_dt, 'end_date': end_date, 'end_time': end_time}


def _num(v):
    """แปลงค่าเป็นสตริงเลขจำนวนเต็ม/ทศนิยม ; datetime/serial ไม่แตะ (ItemKind เป็นเลขอยู่แล้ว)"""
    if v is None:
        return ''
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        return str(int(f)) if f == int(f) else ('%f' % f).rstrip('0').rstrip('.')
    s = str(v).strip()
    if not s:
        return ''
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except Exception:
        return s


def _plan_period(rows):
    """ระยะเวลากิจกรรม: Start (วันที่+เวลา 'น.') / End (เวลา 'น.') -> (start_dt, end_date, end_time)"""
    start_dt = end_date = end_time = None
    for row in rows[:40]:
        c2 = _norm(row[2]) if len(row) > 2 else ''
        if c2 in ('start', 'end'):
            d = t = None
            for c in row:
                if isinstance(c, datetime):
                    d = c
                elif isinstance(c, (int, float)) and not isinstance(c, bool) and 20000 < float(c) < 90000:
                    d = datetime(1899, 12, 30) + timedelta(days=int(c))
                elif isinstance(c, str) and 'น.' in c:
                    t = _parse_thai_time(c)
            if c2 == 'start' and d:
                hh, mm = t or (0, 0)
                start_dt = d.replace(hour=hh, minute=mm, second=0, microsecond=0)
            elif c2 == 'end':
                end_date, end_time = d, t
    return start_dt, end_date, end_time


def _plan_event_name(rows):
    for row in rows[:40]:
        for ci, c in enumerate(row):
            if _norm(c) == 'eventname':
                for cj in range(ci + 1, len(row)):
                    if row[cj] is not None and str(row[cj]).strip():
                        return str(row[cj]).strip().replace('\n', ' ')
    return ''


def parse_event_plan_workbook(path):
    """อ่าน Event template -> (out, skipped) รูปแบบเดียวกับ item_finder.parse_event_workbook
    เพื่อเข้า pipeline "ค้นหาไอเทม" เดิมได้เลย (แค่ template ต่าง)
      out = [(sheet, items)] ; แต่ละ item มี kind/opt/dur/name + group(ชื่อรางวัล) + group_meta(ข้อมูล Event)"""
    import warnings
    import openpyxl
    _WSP = _orig = None
    try:
        import openpyxl.worksheet._reader as _wsr
        _WSP = _wsr.WorkSheetParser
        _orig = _WSP.__init__

        def _patched(self, *a, **k):
            _orig(self, *a, **k)
            try:
                self.date_formats = set()
                self.timedelta_formats = set()
            except Exception:
                pass
    except Exception:
        _WSP = None
    # template 'Pride Code Request' โครงสร้างคนละแบบ (หลายบล็อกโค้ดต่อชีท) -> เช็คก่อน
    pride = _pride_workbook_items(path)
    if pride is not None:
        return pride, []
    out, skipped = [], []
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        if _WSP is not None:
            _WSP.__init__ = _patched
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for sh in wb.sheetnames:
                    rows = list(wb[sh].iter_rows(values_only=True))
                    if any(_norm(c) == 'webreward' for row in rows for c in row):
                        items = _topspender_sheet_items(rows, sh, skipped)   # template Top Spender
                    else:
                        items = _plan_sheet_items(rows, skipped)
                    for it in items:               # group -> sources (ให้เข้า pipeline เดิม)
                        grp = it.pop('group', '')
                        it['sources'] = [grp] if grp else ['(ไม่มีชื่อรางวัล)']
                    if items:
                        out.append((sh, items))
            finally:
                wb.close()
        finally:
            if _WSP is not None:
                _WSP.__init__ = _orig
    return out, skipped


def _pride_workbook_items(path):
    """template 'Cabal Pride Code Request' -> [(sheet, items)] | ไม่ใช่ template นี้ -> None
    1 บล็อกโค้ด = 1 กลุ่ม -> ค้นหาไอเทม -> ส่งสร้าง bundle -> Item Code ดึงไปเติมได้เลย
    (import ครั้งเดียวจบที่ Item Finder ไม่ต้องเปิดไฟล์ซ้ำที่ Item Code)"""
    try:
        from itemcode_tool import read_pride_codes
        blocks = read_pride_codes(path)
    except Exception:
        return None
    # ต้องเจอป้าย MASTER/UNIQUE CODE จริง ๆ ไม่งั้นเป็นแค่ตาราง Item Kind ของ template อื่น
    blocks = [b for b in (blocks or []) if b.get('has_code_label') and b.get('items')]
    if not blocks:
        return None
    out, order, used = {}, [], {}
    for b in blocks:
        base = (b.get('title') or 'Code').strip() or 'Code'
        used[base] = used.get(base, 0) + 1
        # ชื่อกลุ่มต้องไม่ซ้ำ — meta/bundle ผูกด้วยชื่อกลุ่ม (ซ้ำ = ทับกัน)
        grp = base if used[base] == 1 else '%s (%d)' % (base, used[base])
        meta = {'event_name': b.get('event_name') or '', 'reward': grp,
                'is_pride': True, 'pride': b}
        sh = b.get('sheet') or 'Pride'
        if sh not in out:
            out[sh] = []
            order.append(sh)
        for it in b['items']:
            out[sh].append({
                'kind': it.get('kind', ''), 'opt': it.get('opt', ''), 'dur': it.get('dur', ''),
                'name': it.get('name', ''), 'amt': it.get('amt', ''),
                'sources': [grp], 'group_meta': meta,
                'web': 'any', 'img': 'any', 'qty_val': '',
                'trade': 'any', 'drill': 'any', 'crit_val': '',
            })
    return [(sh, out[sh]) for sh in order]


def _plan_sheet_items(rows, skipped):
    """ดึงไอเทมจาก Event template 1 sheet -> list finder-format (group=ชื่อรางวัล, group_meta=ข้อมูล Event)"""
    event_name = _plan_event_name(rows)
    start_dt, end_date, end_time = _plan_period(rows)
    items = []
    banner = ''
    col = None
    cur_group = ''
    cur_meta = {}
    ntab = 0
    for row in rows:
        for c in row:
            s = str(c).strip() if c is not None else ''
            if s.lower().startswith('reward:'):
                banner = _clean_reward_name(s)
        cn = [_norm(c) for c in row]
        if 'itemkind' in cn:
            col = {}
            for i, k in enumerate(cn):
                if k == 'itemkind':
                    col['kind'] = i
                elif k == 'itemoption':
                    col['opt'] = i
                elif k == 'durationindex':
                    col['dur'] = i
                elif k in ('itemname', 'displayname'):
                    col.setdefault('name', i)
                elif k in ('amount', 'amt'):
                    col['amt'] = i
            ntab += 1
            cur_group = banner or ('รางวัลที่ %d' % ntab)
            cur_meta = {'event_name': event_name, 'reward': cur_group, 'is_event': True,
                        'start_dt': start_dt, 'end_date': end_date, 'end_time': end_time}
            banner = ''
            continue
        if col is None or 'kind' not in col:
            continue

        def get(key, r=row, cc=col):
            j = cc.get(key)
            return r[j] if (j is not None and j < len(r)) else None
        kv = get('kind')
        kind = _num(kv)
        nm = get('name')
        name_filled = nm is not None and str(nm).strip() != ''
        if kind and kind.isdigit():
            items.append({
                'kind': kind, 'opt': _num(get('opt')), 'dur': _num(get('dur')),
                'name': str(nm).strip() if name_filled else '', 'amt': _num(get('amt')),
                'group': cur_group, 'group_meta': cur_meta,
                'web': 'any', 'img': 'any', 'qty_val': '',
                'trade': 'any', 'drill': 'any', 'crit_val': '',
            })
        elif name_filled and kv is not None and str(kv).strip() != '':
            if skipped is not None:
                skipped.append(str(nm).strip())
        else:
            col = None
    return items


# --------------------------------------------------------------------------- template "Monthly Top Spender"
def _cell_date(c):
    """คืน datetime (วันที่) จาก cell ที่เป็น datetime หรือ Excel serial (ตอนปิด date_formats) ; None ถ้าไม่ใช่"""
    if isinstance(c, bool):
        return None
    if isinstance(c, datetime):
        return c
    if isinstance(c, (int, float)) and 20000 < float(c) < 90000:
        return datetime(1899, 12, 30) + timedelta(days=int(float(c)))
    return None


def _cell_time(c):
    """คืน (ชม.,นาที) จาก cell เวลา: datetime.time / datetime / เศษของวัน 0-1 (ตอนปิด date_formats) / 'HH:MM'"""
    from datetime import time as _time
    if isinstance(c, bool):
        return None
    if isinstance(c, _time):
        return c.hour, c.minute
    if isinstance(c, datetime):
        return c.hour, c.minute
    if isinstance(c, (int, float)):
        frac = float(c) - int(float(c))
        if 0 < frac < 1:
            total = round(frac * 24 * 60)
            return (total // 60) % 24, total % 60
        return None
    if isinstance(c, str):
        m = re.search(r'(\d{1,2})[.:](\d{2})', c)
        if m:
            h, mi = int(m.group(1)), int(m.group(2))
            if 0 <= h < 24 and 0 <= mi < 60:
                return h, mi
    return None


def _clean_sheet_title(t):
    """ชื่อชีท -> ชื่อ Event (ตัด '(Done)' ท้าย, ยุบช่องว่าง)"""
    t = re.sub(r'\s*\(\s*done\s*\)\s*$', '', str(t or '').strip(), flags=re.I)
    return re.sub(r'\s+', ' ', t).strip()


def _parse_topspender_sheet(rows, sheet_title=''):
    """parser สำหรับ template 'Monthly Item Top Spender':
      - แต่ละ tier ('Web Reward' + ชื่อ tier) = 1 ชุดรางวัล (Black/Prestige/Top Spender)
      - ชื่อ Event = ชื่อชีท (ไฟล์นี้ไม่มีเซลล์ 'Event Name' และไม่มี Bundle ID)
      - Start/End: ป้ายกำกับ + ค่าวันที่/เวลา อยู่คอลัมน์ไหนก็ได้ (บางชีทเลื่อน A->B, C->D, E->F)
        จึง "สแกนหา" ไม่ล็อกคอลัมน์ -> วันจบดึงถูกทุกชีท
      คืน dict shape เดียวกับ _parse_plan_sheet + start_event/end_event (datetime จริงจากไฟล์)"""
    rewards = []
    cur = None
    col = None
    ev_start = ev_end = None
    for row in rows:
        cn = [_norm(c) for c in row]
        if 'webreward' in cn:                     # เริ่ม tier ใหม่
            wi = cn.index('webreward')
            tier = ''
            for cj in range(wi + 1, len(row)):
                if row[cj] is not None and str(row[cj]).strip():
                    tier = str(row[cj]).strip().replace('\n', ' ')
                    break
            cur = {'name': tier or ('รางวัลที่ %d' % (len(rewards) + 1)), 'bundle_id': '', 'items': []}
            rewards.append(cur)
            col = None
            continue
        if cur is None:
            continue
        # Start/End — ป้ายอยู่คอลัมน์ A หรือ B ก็ได้ แล้วสแกนหา date/time ในเซลล์ถัด ๆ ไป
        lab = lab_i = None
        for i, k in enumerate(cn[:3]):
            if k in ('start', 'end'):
                lab, lab_i = k, i
                break
        if lab:
            d = t = None
            for c in row[lab_i + 1:]:
                if d is None:
                    d = _cell_date(c)
                if t is None:
                    t = _cell_time(c)
            if d is not None:
                hh, mm = t or (0, 0)
                dt = d.replace(hour=hh, minute=mm, second=0, microsecond=0)
                if lab == 'start':
                    ev_start = dt if (ev_start is None or dt < ev_start) else ev_start
                else:
                    ev_end = dt if (ev_end is None or dt > ev_end) else ev_end
            continue
        if 'itemkind' in cn:                      # หัวตาราง item ของ tier นี้
            col = {'kind': cn.index('itemkind')}
            for i, k in enumerate(cn):
                if k in ('itemname', 'displayname'):
                    col['name'] = i
                    break
            continue
        if col is not None:                       # แถวไอเทม
            kc = col['kind']
            kv = row[kc] if kc < len(row) else None
            if kv is not None and _isnum(kv):
                nm = row[col['name']] if (col.get('name') is not None and col['name'] < len(row)) else ''
                cur['items'].append({'kind': str(int(float(str(kv).strip()))),
                                     'name': str(nm).strip() if nm else ''})
            elif kv is None or str(kv).strip() == '':
                col = None                        # จบตารางของ tier นี้
    return {'name': _clean_sheet_title(sheet_title), 'rewards': rewards,
            'start_dt': ev_start, 'end_date': ev_end, 'end_time': None,
            'start_event': ev_start, 'end_event': ev_end}


def _parse_any_sheet(rows, sheet_title=''):
    """เลือก parser อัตโนมัติ: ชีทที่มี 'Web Reward' = template Top Spender, ไม่งั้นใช้ template เดิม"""
    if any(_norm(c) == 'webreward' for row in rows for c in row):
        return _parse_topspender_sheet(rows, sheet_title)
    return _parse_plan_sheet(rows)


def _topspender_sheet_items(rows, sheet_title, skipped):
    """Top Spender 1 sheet -> finder-format items (group=tier, group_meta=Event+วันเวลา)
    โครงสร้างเดียวกับ _plan_sheet_items เพื่อไหลเข้า pipeline Item Finder/provider ได้เลย
    -> ทำให้ 'Import Event/Prize' ใน Item Finder + 'เติมจาก Plan' ใน Event ใช้ไฟล์นี้ได้"""
    event_name = _clean_sheet_title(sheet_title)
    # วันเริ่ม/จบของทั้ง sheet (ทุก tier ใช้ร่วมกัน) — สแกนไม่ล็อกคอลัมน์
    ev_start = ev_end = None
    for row in rows:
        cn = [_norm(c) for c in row]
        lab = lab_i = None
        for i, k in enumerate(cn[:3]):
            if k in ('start', 'end'):
                lab, lab_i = k, i
                break
        if not lab:
            continue
        d = t = None
        for c in row[lab_i + 1:]:
            if d is None:
                d = _cell_date(c)
            if t is None:
                t = _cell_time(c)
        if d is not None:
            hh, mm = t or (0, 0)
            dt = d.replace(hour=hh, minute=mm, second=0, microsecond=0)
            if lab == 'start':
                ev_start = dt if (ev_start is None or dt < ev_start) else ev_start
            else:
                ev_end = dt if (ev_end is None or dt > ev_end) else ev_end
    end_time = (ev_end.hour, ev_end.minute) if ev_end else None

    items = []
    tier_i = 0
    cur_group = ''
    cur_meta = {}
    col = None
    for row in rows:
        cn = [_norm(c) for c in row]
        if 'webreward' in cn:
            wi = cn.index('webreward')
            tier = ''
            for cj in range(wi + 1, len(row)):
                if row[cj] is not None and str(row[cj]).strip():
                    tier = str(row[cj]).strip().replace('\n', ' ')
                    break
            tier_i += 1
            cur_group = tier or ('รางวัลที่ %d' % tier_i)
            cur_meta = {'event_name': event_name, 'reward': cur_group, 'is_event': True,
                        'start_dt': ev_start, 'end_date': ev_end, 'end_time': end_time}
            col = None
            continue
        if 'itemkind' in cn:
            col = {}
            for i, k in enumerate(cn):
                if k == 'itemkind':
                    col['kind'] = i
                elif k == 'itemoption':
                    col['opt'] = i
                elif k == 'durationindex':
                    col['dur'] = i
                elif k in ('itemname', 'displayname'):
                    col.setdefault('name', i)
                elif k in ('amount', 'amt', 'จำนวน'):
                    col['amt'] = i
            continue
        if col is None or 'kind' not in col or not cur_group:
            continue

        def get(key, r=row, cc=col):
            j = cc.get(key)
            return r[j] if (j is not None and j < len(r)) else None
        kv = get('kind')
        kind = _num(kv)
        nm = get('name')
        name_filled = nm is not None and str(nm).strip() != ''
        if kind and kind.isdigit():
            items.append({
                'kind': kind, 'opt': _num(get('opt')), 'dur': _num(get('dur')),
                'name': str(nm).strip() if name_filled else '', 'amt': _num(get('amt')),
                'group': cur_group, 'group_meta': cur_meta,
                'web': 'any', 'img': 'any', 'qty_val': '',
                'trade': 'any', 'drill': 'any', 'crit_val': '',
            })
        elif kv is None or str(kv).strip() == '':
            col = None
    return items


def parse_event_plan(path):
    """อ่าน .xlsx -> list ของ (sheet_name, event_dict) เฉพาะ sheet ที่มี Event + ชุดรางวัล"""
    import warnings
    import openpyxl
    out = []
    # ปิด date_formats กู้ ItemKind ที่ถูก format เป็นวันที่ (เหมือน parse_event_workbook)
    _WSP = _orig = None
    try:
        import openpyxl.worksheet._reader as _wsr
        _WSP = _wsr.WorkSheetParser
        _orig = _WSP.__init__

        def _patched(self, *a, **k):
            _orig(self, *a, **k)
            try:
                self.date_formats = set()
                self.timedelta_formats = set()
            except Exception:
                pass
    except Exception:
        _WSP = None
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        if _WSP is not None:
            _WSP.__init__ = _patched
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for sh in wb.sheetnames:
                    rows = list(wb[sh].iter_rows(values_only=True))
                    ev = _parse_any_sheet(rows, sh)
                    if ev['rewards']:
                        out.append((sh, ev))
            finally:
                wb.close()
        finally:
            if _WSP is not None:
                _WSP.__init__ = _orig
    return out


# --------------------------------------------------------------------------- web automation
EVENT_START_LABELS = ["เวลาเริ่มกิจกรรม", "เวลาจบกิจกรรม",
                      "เวลาเริ่มรับของรางวัล", "เวลาสิ้นสุดรับของรางวัล"]


async def _select_type_winner(page, log=None):
    sel = await _find(page, ['xpath=//*[normalize-space(text())="ประเภท"]/following::select[1]', 'select'], timeout=1500)
    if sel:
        for how in (dict(label="WINNER"), dict(value="WINNER"), dict(label="Winner")):
            try:
                await sel.select_option(**how)
                return True
            except Exception:
                continue
    if log:
        log("  ! ตั้งประเภท WINNER ไม่สำเร็จ - ตรวจอีกที", "WARNING")
    return False


async def _add_event_reward(page, log, rw, index):
    """เพิ่มชุดรางวัล Event 1 ชุด (ชื่อ + Limited + ใช้ต่อ user + bundle)
    คืน True ถ้ากรอกชื่อรางวัล (บังคับ) สำเร็จ"""
    ok = True
    btn = await _find(page, ['button:has-text("เพิ่มชุดรางวัล")', 'xpath=//button[contains(.,"เพิ่มชุดรางวัล")]'], timeout=2500)
    if btn:
        await btn.click()
        await page.wait_for_timeout(500)
    else:
        log("  ! หาปุ่ม 'เพิ่มชุดรางวัล' ไม่เจอ", "WARNING")
        ok = False
    i0 = index - 1
    log("  ชุดรางวัลที่ %d: %s (bundle %s)" % (index, rw.get("name"), rw.get("bundle_id") or "-"), "STEP")
    if not await _fill(page, rw.get("name"),
                       ['input[name="rewards.%d.name_th"]' % i0, 'input[placeholder="ชื่อรางวัล (ไทย)"]'],
                       last=True, log=log, label="ชื่อรางวัลไทย"):
        ok = False
    if not await _fill(page, rw.get("name"),
                       ['input[name="rewards.%d.name_en"]' % i0, 'input[placeholder="ชื่อรางวัล (อังกฤษ)"]'],
                       last=True, log=log, label="ชื่อรางวัลอังกฤษ"):
        ok = False
    await _fill(page, rw.get("uses_per_user") or "1",
                ['input[name="rewards.%d.per_player_limit"]' % i0,
                 'input[placeholder="จำนวนการใช้งานต่อ 1 User"]',
                 'xpath=(//*[contains(text(),"จำนวนการใช้งานต่อ 1 User")]/following::input)[last()]'],
                log=log, label="ใช้ต่อ user")
    bid = str(rw.get("bundle_id") or "").strip()
    if bid:
        await _add_bundle(page, i0, bid, log)
    else:
        log("  ! ชุดรางวัลที่ %d ไม่มี bundle_id ในไฟล์ (ตรวจ/ใส่เอง)" % index, "WARNING")
    return ok


async def create_event(page, log, ev, submit=True):
    """สร้าง Event 1 อัน (ev มี name_th/name_en/slug/dates/rewards)"""
    log("=== Event: %s (slug=%s) ===" % (ev.get("name_th"), ev.get("slug")), "STEP")
    fails = []
    def chk(ok, label):
        if not ok:
            fails.append(label)
    chk(await _fill(page, ev.get("name_th"), ['input[placeholder="ชื่อกิจกรรม (ไทย)"]'], log=log, label="ชื่อไทย"), "ชื่อไทย")
    chk(await _fill(page, ev.get("name_en"), ['input[placeholder="ชื่อกิจกรรม (อังกฤษ)"]'], log=log, label="ชื่ออังกฤษ"), "ชื่ออังกฤษ")
    chk(await _fill(page, ev.get("slug"), ['input[placeholder="Slug"]'], log=log, label="slug"), "slug")
    await _select_type_winner(page, log)
    await _fill(page, ev.get("uses_per_user") or "1",
                ['xpath=//*[contains(text(),"จำนวนการใช้งานต่อ 1 User")]/following::input[1]'], log=log, label="ใช้ต่อ user")
    await _fill(page, "0", ['input[placeholder="จำนวนครั้งที่สามารถใช้งานได้"]'], log=log, label="จำนวนครั้ง=0")
    await _fill(page, "0", ['input[placeholder="จำนวนคงเหลือ"]'], log=log, label="คงเหลือ=0")
    for label in EVENT_START_LABELS:
        chk(await _set_datetime(page, ev.get("dates", {}).get(label), label, log), label)

    rewards = ev.get("rewards") or []
    if not rewards:
        log("  ! ไม่มีชุดรางวัล", "WARNING")
        fails.append("ชุดรางวัล")
    for i, rw in enumerate(rewards, 1):
        if not await _add_event_reward(page, log, rw, i):
            fails.append("ชุดรางวัลที่ %d" % i)

    if not submit:
        log("  กรอกครบแล้ว (โหมดพรีวิว - ยังไม่กดบันทึก)", "SUCCESS")
        return True
    if fails:   # กรอกไม่ครบ -> ไม่กดบันทึก (กัน Event ข้อมูลหายถูกสร้างบนเว็บจริง)
        log("  ✋ กรอกไม่ครบ: %s — ไม่กดบันทึกอัตโนมัติ ตรวจ/แก้แล้วกดบันทึกเองในหน้าเว็บ"
            % ", ".join(fails), "WARNING")
        return False
    btn = await _find(page, ['button:has-text("บันทึก")', 'button:has-text("สร้าง")',
                             'button:has-text("Save")', 'button[type="submit"]'], timeout=2500)
    if btn:
        await btn.click()
        await page.wait_for_timeout(1200)
        log("  บันทึกแล้ว", "SUCCESS")
        return True
    log("  ! หาปุ่มบันทึกไม่เจอ - กรอกให้แล้ว กดเองได้", "WARNING")
    return False


# --------------------------------------------------------------------------- reward block (UI)
class EventRewardBlock:
    def __init__(self, app, parent, index):
        self.app = app
        self.frame = card(parent, " ของรางวัลชุดที่ %d " % index)
        self.frame.pack(fill="x", pady=4)
        bg = C["bg_card"]
        top = tk.Frame(self.frame, bg=bg); top.pack(fill="x")
        make_button(top, "ลบชุดนี้", self.remove, "danger").pack(side="right")
        self.name_th = labeled_entry(self.frame, "ชื่อรางวัล (ไทย) *")
        self.name_en = labeled_entry(self.frame, "ชื่อรางวัล (อังกฤษ) *")
        lf = tk.Frame(self.frame, bg=bg); lf.pack(fill="x", pady=2)
        tk.Label(lf, text="จำกัดการใช้งาน", bg=bg, fg=C["text"], font=F, width=20, anchor="w").pack(side="left")
        self.limited = tk.BooleanVar(value=False)
        tk.Checkbutton(lf, text="Limited", variable=self.limited, bg=bg, fg=C["text"],
                       selectcolor=C["bg_inp"], activebackground=bg, activeforeground=C["text"], font=F).pack(side="left")
        self.uses_per_user = labeled_entry(self.frame, "ใช้งานต่อ 1 User", "1")
        self.bundle_id = labeled_entry(self.frame, "Bundle ID (ค้น/ใส่เอง)")

    def remove(self):
        self.frame.destroy()
        self.app.reward_blocks = [b for b in self.app.reward_blocks if b is not self]
        self.app._renumber_rewards()

    def collect(self):
        return {
            "name": self.name_th.get().strip() or self.name_en.get().strip(),
            "name_th": self.name_th.get().strip(),
            "name_en": self.name_en.get().strip(),
            "limited": "TRUE" if self.limited.get() else "FALSE",
            "uses_per_user": self.uses_per_user.get().strip() or "1",
            "bundle_id": self.bundle_id.get().strip(),
        }


# --------------------------------------------------------------------------- App
class App:
    def __init__(self, root, container=None, game_var=None, event_plan_provider=None):
        self.root = root
        self.container = container if container is not None else root
        self._embedded = container is not None    # เปิดผ่าน launcher -> ซ่อน เกม/Login ของตัวเอง (ใช้แถบบน)
        self._game_var = game_var
        self._shared_game = game_var is not None
        # callable() -> list[(sheet, event_dict)] จากคิว Item Finder (โหมด Event)
        self._event_plan_provider = event_plan_provider
        self._cancel = False
        self._plans = []          # [(sheet, event_dict)] จาก import ในตัว (fallback)
        self.reward_blocks = []
        if container is None:
            root.title("Create Event")
            root.configure(bg=C["bg_dark"])
            root.minsize(820, 720)
            root.geometry("880x820")
        self.vgame = game_var if game_var is not None else \
            tk.StringVar(value=core.load_prefs().get("game", core.GAME_NAMES[0]))
        self.vgame.trace_add("write", lambda *a: self._refresh_url())   # เกมเปลี่ยน (แถบบน) -> อัปเดต URL
        self.item_vars = {}
        self._build_ui()
        self._refresh_url()

    def _build_ui(self):
        hdr = tk.Frame(self.container, bg=C["bg_med"], height=46); hdr.pack(fill="x")
        tk.Label(hdr, text="Create Event", bg=C["bg_med"], fg="#ffffff",
                 font=("Segoe UI", 15, "bold")).pack(side="left", padx=14, pady=8)
        tk.Label(hdr, text="combo-interactive", bg=C["bg_med"], fg=C["teal"], font=FB_).pack(side="left", pady=12)

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        for seq in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
            self.root.bind_class("TCombobox", seq, lambda e: "break")

        bar = card(self.container, " ข้อมูล " if self._embedded else " เลือกเกม / เซิร์ฟเวอร์ + เข้าสู่ระบบ ")
        bar.pack(fill="x", padx=10, pady=(8, 4))
        row = tk.Frame(bar, bg=C["bg_card"]); row.pack(fill="x")
        if not self._embedded:
            tk.Label(row, text="เกม:", bg=C["bg_card"], fg=C["text"], font=F).pack(side="left")
            self.game_cb = ttk.Combobox(row, textvariable=self.vgame, values=core.GAME_NAMES, state="readonly", width=16, font=F)
            self.game_cb.pack(side="left", padx=8)
            self.game_cb.bind("<<ComboboxSelected>>", lambda e: (self._save_prefs(), self._refresh_url()))
            make_button(row, "เปิดหน้า Login", self._open_login, "accent").pack(side="left", padx=(6, 4))
        else:
            tk.Label(row, text="เกม:", bg=C["bg_card"], fg=C["muted"], font=F).pack(side="left")
            tk.Label(row, textvariable=self.vgame, bg=C["bg_card"], fg=C["teal"], font=FB_).pack(side="left", padx=(4, 8))
            tk.Label(row, text="(เลือก/Login บนแถบบน)", bg=C["bg_card"], fg=C["muted"], font=SM).pack(side="left")
        make_button(row, "เคลียข้อมูลที่กรอก", self._clear_form, "muted").pack(side="left", padx=4)
        self.url_lbl = tk.Label(bar, text="", bg=C["bg_card"], fg=C["teal"], font=MONO, anchor="w")
        self.url_lbl.pack(fill="x", pady=(6, 0))

        nb = ttk.Notebook(self.container); nb.pack(fill="both", expand=True, padx=8, pady=6)
        self.tab_form = tk.Frame(nb, bg=C["bg_dark"])
        self.tab_bulk = tk.Frame(nb, bg=C["bg_dark"])
        self.tab_log = tk.Frame(nb, bg=C["bg_dark"])
        nb.add(self.tab_form, text="  สร้าง Event  ")
        nb.add(self.tab_bulk, text="  นำเข้า Plan  ")
        nb.add(self.tab_log, text="  Log  ")
        self._build_form(self.tab_form)
        self._build_bulk(self.tab_bulk)
        self._build_log(self.tab_log)

    # ---- form ----
    def _wheel_scroll(self, event):
        cv = getattr(self, "_form_canvas", None)
        if cv is None:
            return
        num = getattr(event, "num", None)
        if num == 4:
            cv.yview_scroll(-1, "units")
        elif num == 5:
            cv.yview_scroll(1, "units")
        else:
            cv.yview_scroll(int(-event.delta / 120), "units")
        return "break"

    def _bind_wheel(self, w):
        for seq in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
            w.bind(seq, self._wheel_scroll)
        for c in w.winfo_children():
            self._bind_wheel(c)

    def _build_form(self, parent):
        canvas = tk.Canvas(parent, bg=C["bg_dark"], highlightthickness=0)
        self._form_canvas = canvas
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=C["bg_dark"])
        canvas.create_window((0, 0), window=inner, anchor="nw", width=840)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        item = card(inner, " ข้อมูลกิจกรรม ")
        item.pack(fill="x", padx=8, pady=6)
        self.item_vars["slug"] = labeled_entry(item, "Slug *")
        tk.Label(item, text="slug: a-z 0-9 - เท่านั้น", bg=C["bg_card"], fg=C["muted"], font=SM).pack(anchor="w", padx=2)
        self.item_vars["name_th"] = labeled_entry(item, "ชื่อกิจกรรม (ไทย) *")
        self.item_vars["name_en"] = labeled_entry(item, "ชื่อกิจกรรม (อังกฤษ) *")
        tf = tk.Frame(item, bg=C["bg_card"]); tf.pack(fill="x", pady=2)
        tk.Label(tf, text="ประเภท", bg=C["bg_card"], fg=C["text"], font=F, width=20, anchor="w").pack(side="left")
        self.item_vars["type"] = tk.StringVar(value="WINNER")
        ttk.Combobox(tf, textvariable=self.item_vars["type"], values=["WINNER"], state="readonly", width=24, font=F).pack(side="left", padx=4)
        self.item_vars["uses_per_user"] = labeled_entry(item, "จำนวนใช้งานต่อ 1 User *", "1")
        self.item_vars["total_uses"] = labeled_entry(item, "จำนวนครั้งที่ใช้ได้", "0")
        self.item_vars["remaining"] = labeled_entry(item, "จำนวนคงเหลือ", "0")
        now = datetime.now()
        self.item_vars["start_event"] = DateTimePicker(item, "เวลาเริ่มกิจกรรม *", now)
        self.item_vars["end_event"] = DateTimePicker(item, "เวลาจบกิจกรรม *", now + timedelta(days=30))
        self.item_vars["start_reward"] = DateTimePicker(item, "เวลาเริ่มรับของรางวัล *", now)
        self.item_vars["end_reward"] = DateTimePicker(item, "เวลาสิ้นสุดรับของรางวัล *", now + timedelta(days=30))

        rhead = tk.Frame(inner, bg=C["bg_dark"]); rhead.pack(fill="x", padx=8, pady=(4, 0))
        tk.Label(rhead, text="ชุดรางวัล (มีได้หลายชุด)", bg=C["bg_dark"], fg=C["text"], font=FB_).pack(side="left")
        make_button(rhead, "เพิ่มชุดรางวัล", self._add_reward, "accent").pack(side="left", padx=8)
        if self._event_provider_ok():
            make_button(rhead, "📥 เติมจาก Plan", self._plan_picker, "green").pack(side="right", padx=8)

        self.reward_host = tk.Frame(inner, bg=C["bg_dark"]); self.reward_host.pack(fill="x", padx=8)
        btns = tk.Frame(inner, bg=C["bg_dark"]); btns.pack(pady=10)
        make_button(btns, "กรอกให้ดูก่อน (ไม่กดบันทึก)", self._preview_one, "warn").pack(side="left", padx=6)
        make_button(btns, "สร้าง + กดบันทึก", self._create_one, "green").pack(side="left", padx=6)
        self._add_reward()
        self._bind_wheel(inner)

    def _event_provider_ok(self):
        return True

    def _add_reward(self):
        idx = len(self.reward_blocks) + 1
        block = EventRewardBlock(self, self.reward_host, idx)
        self.reward_blocks.append(block)
        if getattr(self, "_form_canvas", None) is not None:
            self._bind_wheel(block.frame)

    def _renumber_rewards(self):
        for i, b in enumerate(self.reward_blocks, 1):
            b.frame.config(text=" ของรางวัลชุดที่ %d " % i)

    def _clear_form(self):
        if not messagebox.askyesno("เคลียข้อมูล", "ล้างข้อมูลที่กรอกทั้งหมดในหน้า 'สร้าง Event'?"):
            return
        for k in ("slug", "name_th", "name_en"):
            self.item_vars[k].set("")
        self.item_vars["uses_per_user"].set("1")
        self.item_vars["total_uses"].set("0")
        self.item_vars["remaining"].set("0")
        self.item_vars["type"].set("WINNER")
        now = datetime.now()
        self.item_vars["start_event"].set(now)
        self.item_vars["end_event"].set(now + timedelta(days=30))
        self.item_vars["start_reward"].set(now)
        self.item_vars["end_reward"].set(now + timedelta(days=30))
        while self.reward_blocks:
            self.reward_blocks[-1].remove()
        self._add_reward()
        self.log("เคลียข้อมูลในฟอร์มแล้ว", "INFO")

    # ---- prefill จาก Plan ----
    def prefill_from_plan(self, ev):
        """เติมฟอร์ม Event จาก event_dict ที่ parse มา (name + rewards[{name,bundle_id}])"""
        ev = ev or {}
        name = str(ev.get("name") or "").strip()
        try:
            server = (self.vgame.get() or "").strip()
        except Exception:
            server = ""
        scode = cfg.server_code(server)
        if name:
            self.item_vars["name_th"].set(name)
            self.item_vars["name_en"].set(name)
            base = slugify(name)                # ชื่ออังกฤษ -> slug ; ชื่อไทยล้วนจะว่าง
            if base:
                self.item_vars["slug"].set(base + ("-" + scode if scode else ""))
            else:
                self.log("  ! ชื่อ Event เป็นไทยล้วน สร้าง slug อัตโนมัติไม่ได้ กรอก slug เอง", "WARNING")
        self.item_vars["type"].set("WINNER")
        self.item_vars["uses_per_user"].set("1")
        self.item_vars["total_uses"].set("0")
        self.item_vars["remaining"].set("0")
        now = datetime.now()
        # เริ่ม = วันนี้เที่ยงคืนเสมอ (fix ไว้ ไม่ดึงจากไฟล์ ตามที่ผู้ใช้ต้องการ)
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = cfg.region_end_of_day(server, now + timedelta(days=7))   # ค่าเริ่มต้นเผื่อไฟล์ไม่มีวันจบ
        # ดึงเฉพาะ 'วันจบ' จากไฟล์ถ้ามี — รองรับ end_event (Top Spender) หรือ end_date(+end_time) (provider/ทั่วไป)
        fe = ev.get("end_event")
        if not isinstance(fe, datetime):
            ed = ev.get("end_date")
            if isinstance(ed, datetime):
                et = ev.get("end_time")
                if isinstance(et, (tuple, list)) and len(et) == 2:
                    fe = ed.replace(hour=int(et[0]), minute=int(et[1]), second=0, microsecond=0)
                else:
                    fe = ed.replace(hour=23, minute=59, second=0, microsecond=0)
        if isinstance(fe, datetime):
            end = fe
        self.item_vars["start_event"].set(start)
        self.item_vars["end_event"].set(end)
        self.item_vars["start_reward"].set(start)
        self.item_vars["end_reward"].set(end)

        rewards = ev.get("rewards") or []
        n = max(1, len(rewards))
        while len(self.reward_blocks) > n:
            self.reward_blocks[-1].remove()
        while len(self.reward_blocks) < n:
            self._add_reward()
        for rb, rw in zip(self.reward_blocks, rewards):
            rwn = rw.get("name", "")
            # ชื่อชุดรางวัล = 'ชื่อกิจกรรม - ชื่อรางวัลของแต่ละอัน'
            rname = ("%s - %s" % (name, rwn)) if (name and rwn) else (rwn or name)
            rb.name_th.set(rname)
            rb.name_en.set(rname)
            rb.uses_per_user.set("1")
            rb.bundle_id.set(str(rw.get("bundle_id") or ""))   # bundle id ที่สร้างจากค้นหา->บันเดิล
        self.log("เติมจาก Plan: %s | %d ชุดรางวัล | ระยะเวลา %s -> %s" % (
            name or "-", len(rewards),
            start.strftime("%Y-%m-%d %H:%M"), end.strftime("%Y-%m-%d %H:%M")), "SUCCESS")

    def _get_plans(self):
        """คิว Event: ถ้า import ไฟล์ไว้ในแท็บนี้ = ใช้ไฟล์นั้นก่อน (เจตนาชัดว่าจะใช้ไฟล์)
        ไม่งั้นค่อยใช้คิวจาก Item Finder (pipeline) -> ทำให้ 'เติมจาก Plan' ใช้ไฟล์ที่เพิ่ง import ได้"""
        if self._plans:
            return self._plans
        if self._event_plan_provider:
            try:
                q = list(self._event_plan_provider() or [])
                if q:
                    return q
            except Exception:
                pass
        return self._plans

    def _plan_picker(self, plans=None):
        # plans=None -> ใช้คิวปกติ (provider ก่อน แล้ว self._plans) ; ส่ง list มาเอง = บังคับใช้ตัวนั้น
        if plans is None:
            plans = self._get_plans()
        if not plans:
            messagebox.showinfo("ยังไม่มีข้อมูล",
                                "ไป Item Finder → โหมด Event → Import Event/Prize เลือก sheet ก่อน\n"
                                "(หรือใช้แท็บ 'นำเข้า Plan' ในหน้านี้)")
            return
        win = tk.Toplevel(self.root)
        win.title("เลือก Event มาเติมฟอร์ม")
        win.configure(bg=C["bg_dark"]); win.geometry("560x460"); win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass
        tk.Label(win, text="เลือก 1 Event → เติมฟอร์มให้ (แก้ต่อได้)", bg=C["bg_dark"],
                 fg=C["muted"], font=F).pack(anchor="w", padx=12, pady=(10, 4))
        lf = tk.Frame(win, bg=C["bg_dark"]); lf.pack(fill="both", expand=True, padx=12, pady=4)
        lb = tk.Listbox(lf, bg=C["bg_med"], fg=C["text"], font=("Consolas", 10), relief="flat",
                        selectbackground=C["accent"], selectforeground="#fff", activestyle="none")
        sb = ttk.Scrollbar(lf, command=lb.yview); lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); lb.pack(side="left", fill="both", expand=True)
        for sh, ev in plans:
            lb.insert(tk.END, "%-30s · %d รางวัล" % ((ev.get("name") or sh)[:30], len(ev.get("rewards", []))))

        def _do():
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo("เลือกก่อน", "ยังไม่ได้เลือก Event", parent=win)
                return
            ev = plans[sel[0]][1]
            win.destroy()
            self.prefill_from_plan(ev)
        bar = tk.Frame(win, bg=C["bg_dark"]); bar.pack(fill="x", padx=12, pady=(4, 10))
        make_button(bar, "เติมฟอร์ม", _do, "green").pack(side="left")
        make_button(bar, "ปิด", win.destroy, "muted").pack(side="left", padx=8)
        win.listbox, win.do_pick = lb, _do
        return win

    # ---- bulk import ----
    def _build_bulk(self, parent):
        top = tk.Frame(parent, bg=C["bg_dark"]); top.pack(fill="x", padx=10, pady=10)
        make_button(top, "เลือกไฟล์ Event (.xlsx)", self._pick_excel, "accent").pack(side="left")
        make_button(top, "เลือกเดือน → เติมฟอร์ม", lambda: self._plan_picker(self._plans), "green").pack(side="left", padx=6)
        make_button(top, "เติม Event แรก", self._prefill_first, "warn").pack(side="left", padx=6)
        tk.Label(parent, text="รองรับ 2 แบบ: (1) template เดิม (Event Name + Reward: + Bundle ID)  "
                              "(2) Monthly Top Spender (ชื่อชีท + Web Reward tiers + วันเวลาในไฟล์)",
                 bg=C["bg_dark"], fg=C["muted"], font=SM).pack(anchor="w", padx=12)
        bwrap = tk.Frame(parent, bg=C["bg_dark"]); bwrap.pack(fill="both", expand=True, padx=12, pady=(6, 12))
        bvsb = ttk.Scrollbar(bwrap, orient="vertical")
        bhsb = ttk.Scrollbar(bwrap, orient="horizontal")
        bvsb.pack(side="right", fill="y"); bhsb.pack(side="bottom", fill="x")
        self.bulk_view = tk.Text(bwrap, bg=C["bg_inp"], fg=C["text"], font=MONO, relief="flat",
                                 wrap="none", insertbackground=C["text"],
                                 yscrollcommand=bvsb.set, xscrollcommand=bhsb.set)
        self.bulk_view.pack(side="left", fill="both", expand=True)
        bvsb.config(command=self.bulk_view.yview); bhsb.config(command=self.bulk_view.xview)
        for tag, cl in (("h", C["teal"]), ("k", C["accent"]), ("w", C["warn"])):
            self.bulk_view.tag_config(tag, foreground=cl)
        self.bulk_view.insert("end", "(ยังไม่ได้เลือกไฟล์)")

    def _pick_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
        if not path:
            return
        try:
            self._plans = parse_event_plan(path)
        except ImportError:
            self.log("ต้องติดตั้ง openpyxl: pip install openpyxl", "ERROR"); return
        except Exception:
            self.log(traceback.format_exc(), "ERROR"); return
        self._render_plans(os.path.basename(path))
        self.log("อ่าน %d Event จาก %s" % (len(self._plans), os.path.basename(path)), "INFO")

    def _render_plans(self, src=""):
        v = self.bulk_view
        v.delete("1.0", "end")
        if not self._plans:
            v.insert("end", "(ไม่พบ Event ในไฟล์นี้)"); return
        for i, (sh, ev) in enumerate(self._plans, 1):
            v.insert("end", "#%d  %s\n" % (i, ev.get("name") or sh), "h")
            v.insert("end", "    sheet: %s | ชุดรางวัล: %d\n" % (sh, len(ev.get("rewards", []))), "k")
            for j, rw in enumerate(ev.get("rewards", []), 1):
                v.insert("end", "      [%d] %s  (%d ไอเทม)\n" % (
                    j, rw.get("name"), len(rw.get("items", []))), "w")
            v.insert("end", "\n")

    def _prefill_first(self):
        if not self._plans:
            messagebox.showinfo("ยังไม่มีข้อมูล", "เลือกไฟล์ Event ก่อน"); return
        self.prefill_from_plan(self._plans[0][1])

    # ---- log ----
    def _build_log(self, parent):
        top = tk.Frame(parent, bg=C["bg_dark"]); top.pack(fill="x", padx=8, pady=(8, 0))
        make_button(top, "ล้าง Log", lambda: self.log_area.delete("1.0", "end"), "muted").pack(side="right")
        self.log_area = tk.Text(parent, bg=C["bg_med"], fg=C["text"], font=MONO, relief="flat", wrap="word", insertbackground=C["text"])
        self.log_area.pack(fill="both", expand=True, padx=8, pady=8)
        for lvl, cl in (("SUCCESS", C["accent2"]), ("WARNING", C["warn"]), ("ERROR", C["danger"]),
                        ("INFO", C["text"]), ("STEP", C["accent"])):
            self.log_area.tag_config(lvl, foreground=cl)
        self.log("มี session เก็บไว้แล้ว" if core.has_session() else "ยังไม่มี session ที่เก็บไว้", "INFO")

    def log(self, msg, level="INFO"):
        # thread-safe: Playwright coroutine เรียกจาก worker thread -> marshal เข้า Tk main thread
        try:
            self.root.after(0, self._log_main, str(msg), level)
        except Exception:
            pass

    def _log_main(self, msg, level):
        self.log_area.insert("end", msg + "\n", level); self.log_area.see("end")

    # ---- helpers ----
    def _refresh_url(self):
        self.url_lbl.config(text=core.build_url(self.vgame.get(), CREATE_PATH))

    def _save_prefs(self):
        if not self._shared_game:
            p = core.load_prefs(); p["game"] = self.vgame.get(); core.save_prefs(p)

    def _run_async(self, coro_factory):
        def worker():
            try:
                asyncio.run(coro_factory())
            except Exception:
                self.log(traceback.format_exc(), "ERROR")
        threading.Thread(target=worker, daemon=True).start()

    def _collect_event(self):
        return {
            "name_th": self.item_vars["name_th"].get().strip(),
            "name_en": self.item_vars["name_en"].get().strip(),
            "slug": self.item_vars["slug"].get().strip(),
            "type": "WINNER",
            "uses_per_user": self.item_vars["uses_per_user"].get().strip() or "1",
            "total_uses": "0", "remaining": "0",
            "dates": {
                "เวลาเริ่มกิจกรรม": self.item_vars["start_event"].get(),
                "เวลาจบกิจกรรม": self.item_vars["end_event"].get(),
                "เวลาเริ่มรับของรางวัล": self.item_vars["start_reward"].get(),
                "เวลาสิ้นสุดรับของรางวัล": self.item_vars["end_reward"].get(),
            },
            "rewards": [b.collect() for b in self.reward_blocks],
        }

    def _validate_event(self, ev):
        if not ev.get("name_th") or not ev.get("name_en"):
            return False, "กรุณากรอกชื่อกิจกรรม (ไทย/อังกฤษ)"
        if not re.fullmatch(r"[a-z0-9\-]+", ev.get("slug", "")):
            return False, "slug ต้องเป็น a-z 0-9 - เท่านั้น"
        if not ev.get("rewards"):
            return False, "ต้องมีชุดรางวัลอย่างน้อย 1 ชุด"
        for i, rw in enumerate(ev["rewards"], 1):
            if not rw.get("name_th") or not rw.get("name_en"):
                return False, "ชุดรางวัลที่ %d: ต้องมีชื่อไทย/อังกฤษ" % i
        return True, ""   # Bundle ID ไม่บังคับ (ค้น/ใส่เองบนเว็บ)

    # ---- login / run ----
    def _open_login(self):
        self._save_prefs()
        s = core.AztekSession(log=self.log, should_cancel=lambda: self._cancel)
        self._run_async(lambda: s.open_login(core.build_url(self.vgame.get(), LIST_PATH)))

    def _preview_one(self):
        self._run_form(submit=False)

    def _create_one(self):
        self._run_form(submit=True)

    def _run_form(self, submit):
        self._save_prefs()
        ev = self._collect_event()
        ok, err = self._validate_event(ev)
        if not ok:
            messagebox.showwarning("ข้อมูลไม่ครบ/ไม่ถูกต้อง", err); return
        s = core.AztekSession(log=self.log, should_cancel=lambda: self._cancel)
        url = core.build_url(self.vgame.get(), CREATE_PATH)

        async def run():
            await s.run_page(url, lambda page: create_event(page, self.log, ev, submit=submit), hold=True)
        self._run_async(run)


def main():
    root = tk.Tk()
    _set_window_icon(root)
    App(root)
    root.mainloop()


# ทะเบียนสำหรับ All for Cabal launcher
from tool_registry import ToolSpec
TOOL = ToolSpec(
    key='event', icon='🎉', title='Create Event', nav='🎉  Event',
    desc='สร้าง Event (WINNER) จาก Monthly Plan บนเว็บ aztek-tools', boot='danger',
    make=lambda lc, fr: App(lc.root, container=fr, game_var=lc.game_var,
                            event_plan_provider=lc._event_plans_provider))


if __name__ == "__main__":
    main()
