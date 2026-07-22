# -*- coding: utf-8 -*-
"""
itemcode_tool.py  -  tool "สร้าง Item Code" บนเว็บ aztek-tools (ธีมเดียวกับ Create bundle)

โครงสร้าง: 1 Item Code มีได้หลาย "ชุดรางวัล" (reward set)
Excel = 2 ชีตเชื่อมด้วย slug:
  - ชีต 'itemcodes' : 1 แถว = 1 Item Code
  - ชีต 'rewards'   : 1 แถว = 1 ชุดรางวัล (ผูกด้วย itemcode_slug ; หลายแถว = หลายชุด)
หน้า create จริง: {base}/itemcodes/create

ต้องมี:  pip install playwright openpyxl   แล้ว   playwright install chromium
รัน:    python itemcode_tool.py
"""

import os
import re
import asyncio
import threading
import traceback
from datetime import datetime, timedelta

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import aztek_core as core
import config as cfg
# web helper กลาง (form driver aztek-tools) — ใช้ร่วมทุก tool
from aztek_core import _find, _fill, _set_datetime, _add_bundle
# UI kit กลาง (palette/ฟอนต์/วิดเจ็ต/ไอคอน) — event_tool re-import ต่อจากไฟล์นี้ได้เหมือนเดิม
from ui_common import (C, F, FB_, FM, FB, MONO, SM,
                       make_button, card, labeled_entry, DateTimePicker,
                       _find_icon, _set_window_icon)

LIST_PATH = "itemcodes"
CREATE_PATH = "itemcodes/create"
CODE_TYPES = ["Fix Codes", "Server Generate Codes"]
REVEAL_DELAY = 2500  # ms รอ animation หลังกด Limited / เลือก Server ก่อนกรอกช่องที่เพิ่งโผล่

# ---- ฟิลด์ระดับ Item Code (ชีต itemcodes) ----
# key, ป้าย, default, บังคับ
ITEM_FIELDS = [
    ("slug",          "Slug",                          "",    True),
    ("name_th",       "ชื่อ Item Code (ไทย)",          "",    True),
    ("name_en",       "ชื่อ Item Code (อังกฤษ)",        "",    True),
    ("type",          "ประเภท",                        "ALL", True),
    ("uses_per_user", "จำนวนการใช้งานต่อ 1 User",       "1",   True),
    ("total_uses",    "จำนวนครั้งที่สามารถใช้งานได้",     "",    False),
    ("remaining",     "จำนวนคงเหลือ",                  "",    False),
    ("start_time",    "เวลาเริ่มใช้งาน (YYYY-MM-DD HH:MM:SS)", "", True),
    ("end_time",      "เวลาสิ้นสุด (YYYY-MM-DD HH:MM:SS)",     "", True),
]
ITEM_KEYS = [f[0] for f in ITEM_FIELDS]
ITEM_HEADERS = [f[1] for f in ITEM_FIELDS]

# ---- ฟิลด์ระดับชุดรางวัล (ชีต rewards) ----
REWARD_FIELDS = [
    ("itemcode_slug",  "itemcode_slug (ตรงกับ slug ในชีต itemcodes)", "", True),
    ("reward_name_th", "ชื่อรางวัล (ไทย)",              "", True),
    ("reward_name_en", "ชื่อรางวัล (อังกฤษ)",            "", True),
    ("limited",        "Limited (TRUE/FALSE)",          "FALSE", False),
    ("uses_per_user",  "จำนวนการใช้งานต่อ 1 User",       "1", False),
    ("total_uses",     "จำนวนครั้งที่สามารถใช้งานได้ (เมื่อ Limited)", "", False),
    ("remaining",      "จำนวนคงเหลือ (เมื่อ Limited)",   "", False),
    ("code_type",      "ประเภทของ Code (Fix Codes / Server Generate Codes)", "Fix Codes", False),
    ("code_list",      "รายการ Code (เมื่อ Fix ; หลายอันขึ้นบรรทัดใหม่)", "", False),
    ("prefix",         "prefix ของแต่ละโค้ด (เมื่อ Server)", "", False),
    ("num_codes",      "จำนวนโค้ดที่ต้องการ (เมื่อ Server)", "", False),
    ("bundles",        "bundles (id คั่นด้วย , เช่น 208106,207938)", "", True),
]
REWARD_KEYS = [f[0] for f in REWARD_FIELDS]
REWARD_HEADERS = [f[1] for f in REWARD_FIELDS]


# --------------------------------------------------------------------------- utils
def slugify(s):
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-z0-9\-]", "", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def valid_slug(s):
    return bool(s) and re.fullmatch(r"[a-z0-9\-]+", s) is not None


def as_bool(v):
    return str(v).strip().lower() in ("1", "true", "yes", "y", "t", "จริง", "เปิด")


def now_str(offset_days=0):
    return (datetime.now() + timedelta(days=offset_days)).strftime("%Y-%m-%d %H:%M:%S")


def split_ids(s):
    return [x.strip() for x in re.split(r"[,\s]+", str(s or "").strip()) if x.strip()]


def is_server(code_type):
    return "server" in str(code_type).lower()


def parse_expire(text, now=None):
    """แปลงวันหมดอายุจากไฟล์ Event -> datetime
    รองรับ:
    - datetime เต็ม '2026-08-31 00:00:00' (คงเวลาเดิม)
    - Excel serial number (openpyxl ปิด date_formats เพื่อกู้ ItemKind -> วันหมดอายุกลายเป็นเลข)
    - '31 Aug', '2026-08-31', '31/08/2026' ฯลฯ (ไม่มีปี -> ปีปัจจุบัน, ตั้งเวลา 00:00:00)"""
    if text is None or str(text).strip() == "":
        return None
    now = now or datetime.now()
    t = str(text).strip()
    # Excel serial (เช่น 46265 = 2026-08-31) — จำกัดช่วง ~2009-2064 กันเลขสุ่ม (จำนวน/id) ถูกอ่านเป็นวัน
    try:
        f = float(t)
        if 40000 < f < 60000:
            return datetime(1899, 12, 30) + timedelta(days=int(f))
    except Exception:
        pass
    dt = None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",   # ISO มี 'T' (Export Code Simulator)
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S"):
        try:
            dt = datetime.strptime(t, fmt); break         # มีเวลาในไฟล์ -> ใช้ตามนั้น (prefill แทนเวลาตามเซิร์ฟ)
        except Exception:
            pass
    if dt is None:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
                    "%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y",
                    "%d %b, %Y", "%d %B, %Y"):
            try:
                dt = datetime.strptime(t, fmt); break     # ทั้งวัน -> เวลา 00:00:00
            except Exception:
                pass
    if dt is None:
        for fmt in ("%d %b", "%d %B", "%b %d", "%B %d"):   # ไม่มีปี -> เติมปีปัจจุบัน
            try:
                dt = datetime.strptime(t, fmt).replace(year=now.year)
                if dt < now:                              # วันผ่านมาแล้ว -> เลื่อนเป็นปีหน้า
                    dt = dt.replace(year=now.year + 1)
                break
            except Exception:
                pass
    if dt is None:
        return None
    if dt.year > 2400:                                    # พ.ศ. (เช่น 2569) -> ค.ศ.
        dt = dt.replace(year=dt.year - 543)
    return dt


# --------------------------------------------------------------------------- Excel
def build_template_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    hf = PatternFill("solid", fgColor="21262D")
    hfont = Font(bold=True, color="FFFFFF")

    def head(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[cell.column_letter].width = max(16, min(40, len(h) + 2))
        ws.freeze_panes = "A2"

    ws1 = wb.active; ws1.title = "itemcodes"; head(ws1, ITEM_HEADERS)
    for r, row in enumerate([
        ["event-newyear-2026", "โค้ดปีใหม่ 2026", "New Year 2026", "ALL", 1, 1000, 1000, now_str(), now_str(30)],
    ], 2):
        for c, v in enumerate(row, 1):
            ws1.cell(row=r, column=c, value=v)

    ws2 = wb.create_sheet("rewards"); head(ws2, REWARD_HEADERS)
    for r, row in enumerate([
        # ชุดที่ 1: Fix Codes (ไม่ limited) -> ใส่ code_list, ผูก bundle 208106
        ["event-newyear-2026", "รางวัลที่ 1", "Prize 1", "FALSE", 1, "", "",
         "Fix Codes", "CODE-A1\nCODE-A2\nCODE-A3", "", "", "208106"],
        # ชุดที่ 2: Server Generate (limited) -> ใส่ prefix + num_codes + total/remaining, ผูก 2 bundle
        ["event-newyear-2026", "รางวัลที่ 2", "Prize 2", "TRUE", 1, 500, 500,
         "Server Generate Codes", "", "NY-", 500, "208106,207938"],
    ], 2):
        for c, v in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=v)

    # ---- dropdown (data validation) ในไฟล์ Excel ----
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    def add_dv(ws, col_idx, options, allow_blank=True):
        col = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(options),
                            allow_blank=allow_blank, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add("%s2:%s500" % (col, col))

    add_dv(ws1, ITEM_KEYS.index("type") + 1, ["ALL"], allow_blank=False)
    add_dv(ws2, REWARD_KEYS.index("limited") + 1, ["TRUE", "FALSE"])
    add_dv(ws2, REWARD_KEYS.index("code_type") + 1, ["Fix Codes", "Server Generate Codes"])
    # itemcode_slug -> เลือกจาก slug ที่กรอกไว้ในชีต itemcodes (คอลัมน์ A)
    scol = get_column_letter(REWARD_KEYS.index("itemcode_slug") + 1)
    dv_slug = DataValidation(type="list", formula1="=itemcodes!$A$2:$A$500",
                             allow_blank=True, showErrorMessage=False)  # เป็นตัวช่วยเลือก ไม่บังคับ (พิมพ์ชื่อแทนก็ได้)
    ws2.add_data_validation(dv_slug)
    dv_slug.add("%s2:%s500" % (scol, scol))

    note = wb.create_sheet("วิธีใช้")
    tips = [
        "วิธีใช้ template (2 ชีตเชื่อมกันด้วย slug)",
        "",
        "ชีต 'itemcodes' = ข้อมูลหลักของ Item Code (1 แถว = 1 โค้ด)",
        "  - slug เว้นว่างได้ (โปรแกรมจะสร้างจากชื่ออังกฤษให้อัตโนมัติ) ถ้ากรอกเองใช้ได้เฉพาะ a-z 0-9 - และห้ามซ้ำ",
        "  - ประเภท (type) = ALL เสมอ | เวลา YYYY-MM-DD HH:MM:SS",
        "",
        "ชีต 'rewards' = ชุดรางวัลของแต่ละ Item Code (1 แถว = 1 ชุดรางวัล)",
        "  - itemcode_slug ใส่ slug หรือ 'ชื่อ' (ไทย/อังกฤษ) ของ Item Code ก็ได้ (โปรแกรมจับคู่ให้)",
        "  - หลายชุดรางวัลในโค้ดเดียว = ใส่หลายแถวที่ itemcode_slug เดียวกัน (เรียงตามลำดับแถว)",
        "  - Limited = FALSE -> ประเภท Code บังคับเป็น Fix Codes (ไม่ต้องใส่ total_uses/remaining)",
        "  - Limited = TRUE  -> ใส่ total_uses/remaining ได้ และเลือก code_type เป็น Server Generate Codes ได้",
        "  - code_type Fix Codes -> กรอก 'รายการ Code' (code_list) หลายอันขึ้นบรรทัดใหม่ในเซลล์",
        "  - code_type Server Generate Codes -> กรอก prefix + num_codes (จำนวนโค้ดที่ต้องการ)",
        "  - bundles = ใส่ bundle id ที่ตั้งไว้แล้ว คั่นด้วย , เช่น 208106,207938 (โปรแกรมจะค้นหาแล้วกดเพิ่มให้)",
        "",
        "เสร็จแล้วกลับไปโปรแกรม -> แท็บ 'นำเข้า Excel' -> เลือกไฟล์นี้ -> กด 'ตรวจไฟล์' เช็กก่อน -> สร้างทั้งหมด",
    ]
    for i, t in enumerate(tips, 1):
        cell = note.cell(row=i, column=1, value=t)
        if i == 1:
            cell.font = Font(bold=True, size=13)
    note.column_dimensions["A"].width = 95
    wb.save(path)
    return path


def _read_sheet(ws, keys, headers):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {}
    for k, h in zip(keys, headers):
        if h in header:
            idx[k] = header.index(h)
    out = []
    for r in rows[1:]:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue
        d = {}
        for i, k in enumerate(keys):
            pos = idx.get(k, i)
            d[k] = r[pos] if pos < len(r) and r[pos] is not None else ""
        out.append(d)
    return out


def read_template_xlsx(path):
    """อ่าน 2 ชีต -> list ของ item dict (แต่ละ item มี key 'rewards' = list ของชุดรางวัล)
    - ถ้า slug ว่าง จะสร้างจากชื่ออังกฤษ/ไทยให้อัตโนมัติ (it['_slug_auto']=True)
    - ชีต rewards อ้างอิง itemcode ด้วย slug หรือชื่อ (ไทย/อังกฤษ) ก็ได้ (จับคู่ยืดหยุ่น)
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws_items = wb["itemcodes"] if "itemcodes" in wb.sheetnames else wb.worksheets[0]
    items = _read_sheet(ws_items, ITEM_KEYS, ITEM_HEADERS)
    rewards = []
    if "rewards" in wb.sheetnames:
        rewards = _read_sheet(wb["rewards"], REWARD_KEYS, REWARD_HEADERS)

    lookup = {}  # key(ตัวเล็ก) -> item

    def addkey(k, it):
        k = str(k or "").strip().lower()
        if k:
            lookup.setdefault(k, it)

    for it in items:
        raw = str(it.get("slug") or "").strip()
        if raw:
            it["slug"] = slugify(raw)
            it["_slug_auto"] = False
        else:
            it["slug"] = slugify(it.get("name_en") or it.get("name_th"))
            it["_slug_auto"] = True
        it["rewards"] = []
        addkey(it["slug"], it)
        addkey(slugify(it.get("name_en")), it)
        addkey(slugify(it.get("name_th")), it)
        addkey(it.get("name_en"), it)
        addkey(it.get("name_th"), it)

    orphan = []
    for rw in rewards:
        raw = str(rw.get("itemcode_slug") or "").strip()
        it = lookup.get(raw.lower()) or lookup.get(slugify(raw))
        if it is not None:
            it["rewards"].append(rw)
        else:
            orphan.append(rw)
    return items, orphan


# --------------------------------------------------------------------------- template "Cabal Pride Code Request"
def _pn(x):
    """normalize: strip lower ตัดช่องว่าง/ขึ้นบรรทัด"""
    return str(x).strip().lower().replace(' ', '').replace('\n', '') if x is not None else ''


def _num_str(v):
    """float-ish -> สตริงเลขสะอาด (33644216.0 -> '33644216') ; อ่านไม่ได้คืนสตริงเดิม"""
    if v is None:
        return ''
    s = str(v).strip()
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else str(f)
    except Exception:
        return s


def _int_str(v):
    """ดึงจำนวนเต็มตัวแรกออกจากข้อความ ('3000 ครั้ง' -> '3000', ' 2700 โค้ด' -> '2700')"""
    m = re.search(r'\d[\d,]*', str(v or ''))
    return m.group(0).replace(',', '') if m else ''


def read_pride_codes(path):
    """อ่าน template 'Cabal Pride Code Request' -> list ของบล็อกโค้ด (แต่ละบล็อก = 1 Item Code)
    เลย์เอาต์: แต่ละชีทมีหลายบล็อกวางเรียงเป็นคอลัมน์ (เยื้องกันไปเรื่อย ๆ)
    จับบล็อกจากหัวตาราง 'Item Kind' — คอลัมน์ที่เจอ = คอลัมน์ฐานของบล็อกนั้น
    field อื่น (title/expire/code/limit) อยู่เหนือหัวตาราง คอลัมน์เดียวกัน, ค่าเยื้องขวา +2"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    blocks = []
    try:
        for ws in wb.worksheets:
            grid = list(ws.iter_rows(values_only=True))
            nrows = len(grid)

            def cell(r, c, _g=grid, _n=nrows):
                if 1 <= r <= _n:
                    row = _g[r - 1]
                    if 1 <= c <= len(row):
                        return row[c - 1]
                return None

            # Event Name (ต่อชีท) — ป้าย 'Event Name' แล้วค่าถัดไปในแถวเดียวกัน
            event_name = ''
            for r in range(1, min(nrows, 12) + 1):
                row = grid[r - 1]
                for ci, v in enumerate(row):
                    if _pn(v) == 'eventname':
                        for cj in range(ci + 1, len(row)):
                            if row[cj] is not None and str(row[cj]).strip():
                                event_name = str(row[cj]).strip().replace('\n', ' ')
                                break
                        break
                if event_name:
                    break

            # ทุกจุดที่เจอ 'Item Kind' = 1 บล็อก
            for r in range(1, nrows + 1):
                row = grid[r - 1]
                for ci, v in enumerate(row):
                    if _pn(v) == 'itemkind':
                        blk = _pride_block(cell, nrows, r, ci + 1, event_name, ws.title)
                        if blk:
                            blocks.append(blk)
    finally:
        wb.close()
    return blocks


def _pride_block(cell, nrows, r_head, c_ik, event_name, sheet):
    """ดึงข้อมูล 1 บล็อกโค้ด: c_ik = คอลัมน์ Item Kind (ฐาน), r_head = แถวหัวตาราง"""
    title = code_label = note = ''
    code_val = expire = expire_time = refill = code_count = None
    cannot_repeat = False
    for row in range(max(1, r_head - 13), r_head):          # meta อยู่เหนือหัวตาราง
        base = _pn(cell(row, c_ik))
        if base == 'conditions':
            t = cell(row - 1, c_ik)                          # ชื่อบล็อกอยู่เหนือ CONDITIONS
            if t:
                title = str(t).strip().replace('\n', ' ')
                # โน้ตไทยต่อท้ายชื่อ (เช่น 'คนละวันกับ ID (วันแรก)') = คำอธิบาย ไม่ใช่ชื่อโค้ด -> แยกออก
                mth = re.search(r'[฀-๿]', title)
                if mth and title[:mth.start()].strip():
                    note = title[mth.start():].strip()
                    title = title[:mth.start()].strip()
        elif 'codeexpiredate' in base:
            expire = cell(row, c_ik + 2)
            expire_time = cell(row, c_ik + 3)
        elif base in ('mastercode', 'uniquecode'):
            code_label = base
            code_val = cell(row, c_ik + 2)
        elif 'จำนวนการเติมจำกัด' in base:
            refill = cell(row, c_ik + 2)
        elif base.startswith('จำนวนcode'):                  # 'จำนวน Code ที่สร้าง'
            code_count = cell(row, c_ik + 2)
        for c in range(c_ik, c_ik + 9):                      # เติมซ้ำไม่ได้ (ที่ไหนก็ได้ในบล็อก)
            if 'เติมซ้ำไม่ได้' in _pn(cell(row, c)):
                cannot_repeat = True

    items = []
    rr = r_head + 1
    while rr <= nrows:
        kv = cell(rr, c_ik)
        if kv is None or str(kv).strip() == '':
            break
        try:
            kind = str(int(float(str(kv).strip())))
        except Exception:
            break
        nm = cell(rr, c_ik + 5)
        items.append({'kind': kind, 'opt': _num_str(cell(rr, c_ik + 2)),
                      'dur': _num_str(cell(rr, c_ik + 3)),
                      'name': str(nm).strip() if nm else '', 'amt': _num_str(cell(rr, c_ik + 7))})
        rr += 1

    if not title and not (code_val and str(code_val).strip()) and not items:
        return None
    tl = title.lower()
    game_hint = 'CBPC' if 'cbpc' in tl else ('CBM' if 'cbm' in tl else '')
    is_master = (code_label == 'mastercode') and bool(str(code_val or '').strip())
    return {
        'sheet': sheet, 'title': title or 'Code', 'note': note,
        'event_name': event_name, 'game_hint': game_hint,
        'expire': expire, 'expire_time': expire_time,
        'fix_code': str(code_val).strip() if is_master else '',
        'unique_code': not is_master,                        # ไม่มีโค้ดตายตัว = Server Generate
        'has_code_label': bool(code_label),                  # เจอป้าย MASTER/UNIQUE CODE จริง = บล็อกของ template นี้
        'refill_limit': _int_str(refill), 'code_count': _int_str(code_count),
        'cannot_repeat': cannot_repeat, 'items': items,
    }


# --------------------------------------------------------------------------- Playwright helpers
# _find / _fill / _set_datetime / _add_bundle ย้ายไป aztek_core (import ข้างบน) — ใช้ร่วมทุก tool


async def _select_item_type_all(page, log=None):
    sel = await _find(page, ['xpath=//*[normalize-space(text())="ประเภท"]/following::select[1]', 'select'], timeout=1500)
    if sel:
        for how in (dict(label="ALL"), dict(value="ALL"), dict(label="All")):
            try:
                await sel.select_option(**how); return True
            except Exception:
                continue
    if log:
        log("  ! ตั้งประเภท ALL (ระดับ Item) ไม่สำเร็จ - ตรวจอีกที", "WARNING")
    return False


async def _set_checkbox_last(page, label_th, want, log=None):
    loc = await _find(page, [
        'xpath=(//label[contains(.,"%s")]//input[@type="checkbox"])[last()]' % label_th,
        'xpath=(//*[contains(normalize-space(text()),"%s")]/preceding::input[@type="checkbox"])[last()]' % label_th,
        'xpath=(//*[contains(normalize-space(text()),"%s")]/following::input[@type="checkbox"])[1]' % label_th,
        'xpath=(//input[@type="checkbox"])[last()]'], last=False, timeout=1500)
    if not loc:
        return None
    try:
        checked = await loc.is_checked()
        if checked != bool(want):
            await loc.click()
        return True
    except Exception:
        return None


async def _select_code_type_last(page, code_type, log=None):
    sel = await _find(page, ['xpath=(//*[contains(text(),"ประเภทของ Code")]/following::select)[last()]'], last=False, timeout=1500)
    if not sel:
        sel = await _find(page, ['select'], last=True, timeout=1500)
    if sel:
        for how in (dict(label=code_type), dict(value=code_type)):
            try:
                await sel.select_option(**how); return True
            except Exception:
                continue
    if log:
        log("  ! เลือก code_type ไม่สำเร็จ: " + str(code_type), "WARNING")
    return False


async def _check_by_name(page, name, want, log=None):
    """ติ๊ก/เอาติ๊กออก checkbox ตาม name attribute (แม่นสุด)"""
    loc = await _find(page, ['input[type="checkbox"][name="%s"]' % name, 'input[name="%s"]' % name], timeout=2000)
    if not loc:
        if log:
            log("  ! หา checkbox name=%s ไม่เจอ ลองแบบ label" % name, "WARNING")
        return await _set_checkbox_last(page, "Limited", want, log)
    try:
        if want:
            await loc.check()
        else:
            await loc.uncheck()
        return True
    except Exception:
        try:
            await loc.click(); return True
        except Exception:
            return None


async def _select_code_type(page, i0, code_type, log=None):
    """เลือกประเภทของ Code ตรงจาก select[name=rewards.{i0}.is_generate_code] (value 1=Fix, 2=Server)"""
    val = "2" if is_server(code_type) else "1"
    sel = await _find(page, ['select[name="rewards.%d.is_generate_code"]' % i0], timeout=2000)
    if sel:
        for how in (dict(value=val), dict(label=code_type)):
            try:
                await sel.select_option(**how); return True
            except Exception:
                continue
    if log:
        log("  ! เลือกประเภท Code ไม่สำเร็จ: " + str(code_type), "WARNING")
    return False


async def _add_reward_set(page, log, rw, index):
    """เพิ่มชุดรางวัล 1 ชุด (กดปุ่ม 'เพิ่มชุดรางวัล' แล้วกรอกฟิลด์ในบล็อกล่าสุด)
    คืน True ถ้ากรอกช่องบังคับ (ชื่อรางวัล) สำเร็จ, False ถ้ามีปัญหา"""
    ok = True
    btn = await _find(page, ['button:has-text("เพิ่มชุดรางวัล")', 'xpath=//button[contains(.,"เพิ่มชุดรางวัล")]'], timeout=2500)
    if btn:
        await btn.click()
        await page.wait_for_timeout(500)
    else:
        log("  ! หาปุ่ม 'เพิ่มชุดรางวัล' ไม่เจอ", "WARNING")
        ok = False

    i0 = index - 1  # ฟอร์มใช้ index 0-based: name="rewards.{i0}.<field>"
    log("  ชุดรางวัลที่ %d: %s / %s" % (index, rw.get("reward_name_th"), rw.get("reward_name_en")), "STEP")
    if not await _fill(page, rw.get("reward_name_th"),
                       ['input[name="rewards.%d.name_th"]' % i0, 'input[placeholder="ชื่อรางวัล (ไทย)"]'],
                       last=True, log=log, label="ชื่อรางวัลไทย"):
        ok = False
    if not await _fill(page, rw.get("reward_name_en"),
                       ['input[name="rewards.%d.name_en"]' % i0, 'input[placeholder="ชื่อรางวัล (อังกฤษ)"]'],
                       last=True, log=log, label="ชื่อรางวัลอังกฤษ"):
        ok = False

    limited = as_bool(rw.get("limited"))
    code_type = rw.get("code_type") or "Fix Codes"
    if limited:
        await _check_by_name(page, "rewards.%d.is_limited" % i0, True, log)
        await page.wait_for_timeout(REVEAL_DELAY)  # รอช่องที่เพิ่งโผล่ (animation)
        await _fill(page, rw.get("uses_per_user") or "1",
                    ['input[name="rewards.%d.per_player_limit"]' % i0, 'input[placeholder="จำนวนการใช้งานต่อ 1 User"]',
                     'xpath=(//*[contains(text(),"จำนวนการใช้งานต่อ 1 User")]/following::input)[last()]'],
                    log=log, label="ใช้ต่อ user")
        await _fill(page, rw.get("total_uses"),
                    ['input[name="rewards.%d.quantity"]' % i0, 'input[placeholder="จำนวนครั้งที่สามารถใช้งานได้"]',
                     'xpath=(//*[contains(text(),"จำนวนครั้งที่สามารถใช้งานได้")]/following::input)[last()]'],
                    log=log, label="จำนวนครั้ง")
        await _fill(page, rw.get("remaining"),
                    ['input[name="rewards.%d.remaining"]' % i0, 'input[placeholder="จำนวนคงเหลือ"]',
                     'xpath=(//*[contains(text(),"จำนวนคงเหลือ")]/following::input)[last()]'],
                    log=log, label="คงเหลือ")
        await _select_code_type(page, i0, code_type, log)
        if is_server(code_type):
            await page.wait_for_timeout(REVEAL_DELAY)  # รอ prefix/จำนวนโค้ด ที่เพิ่งโผล่
            await _fill(page, rw.get("prefix"),
                        ['input[name="rewards.%d.prefix"]' % i0, 'input[placeholder="prefix ของแต่ละโค้ด"]'],
                        last=True, log=log, label="prefix")
            await _fill(page, rw.get("num_codes"),
                        ['input[name="rewards.%d.amount"]' % i0, 'input[placeholder="จำนวนโค้ดที่ต้องการ"]',
                         'xpath=(//*[contains(text(),"จำนวนโค้ดที่ต้องการ")]/following::input)[last()]'],
                        log=log, label="จำนวนโค้ด")
        else:
            await _fill(page, rw.get("code_list"),
                        ['textarea[name="rewards.%d.serial"]' % i0, 'textarea[placeholder="รายการ Code"]'],
                        last=True, log=log, label="รายการ Code")
    else:
        # ไม่ limited -> Fix Codes บังคับ
        await _fill(page, rw.get("uses_per_user") or "1",
                    ['input[name="rewards.%d.per_player_limit"]' % i0, 'input[placeholder="จำนวนการใช้งานต่อ 1 User"]',
                     'xpath=(//*[contains(text(),"จำนวนการใช้งานต่อ 1 User")]/following::input)[last()]'],
                    log=log, label="ใช้ต่อ user")
        await _fill(page, rw.get("code_list"),
                    ['textarea[name="rewards.%d.serial"]' % i0, 'textarea[placeholder="รายการ Code"]'],
                    last=True, log=log, label="รายการ Code")

    for bid in split_ids(rw.get("bundles")):
        await _add_bundle(page, i0, bid, log)
    return ok


async def create_itemcode(page, log, item, submit=True):
    """สร้าง Item Code 1 อัน (item ต้องมี key 'rewards' = list ของชุดรางวัล)"""
    log("=== Item Code: %s (slug=%s) ===" % (item.get("name_th"), item.get("slug")), "STEP")
    fails = []   # เก็บ label ของช่อง "บังคับ" ที่กรอกไม่สำเร็จ -> ถ้ามี จะไม่กดบันทึกอัตโนมัติ
    def chk(ok, label):
        if not ok:
            fails.append(label)
    # ระดับ Item
    chk(await _fill(page, item.get("name_th"), ['input[placeholder="ชื่อ Item Code (ไทย)"]'], log=log, label="ชื่อไทย"), "ชื่อไทย")
    chk(await _fill(page, item.get("name_en"), ['input[placeholder="ชื่อ Item Code (อังกฤษ)"]'], log=log, label="ชื่ออังกฤษ"), "ชื่ออังกฤษ")
    chk(await _fill(page, item.get("slug"), ['input[placeholder="Slug"]'], log=log, label="slug"), "slug")
    await _select_item_type_all(page, log)
    await _fill(page, item.get("uses_per_user"),
                ['xpath=//*[contains(text(),"จำนวนการใช้งานต่อ 1 User")]/following::input[1]'], log=log, label="ใช้ต่อ user")
    await _fill(page, item.get("total_uses"),
                ['input[placeholder="จำนวนครั้งที่สามารถใช้งานได้"]'], log=log, label="จำนวนครั้ง")
    await _fill(page, item.get("remaining"),
                ['input[placeholder="จำนวนคงเหลือ"]'], log=log, label="คงเหลือ")
    chk(await _set_datetime(page, item.get("start_time"), "เวลาเริ่มใช้งาน", log), "เวลาเริ่ม")
    chk(await _set_datetime(page, item.get("end_time"), "เวลาสิ้นสุด", log), "เวลาสิ้นสุด")

    rewards = item.get("rewards") or []
    if not rewards:
        log("  ! ไม่มีชุดรางวัล - ต้องมีอย่างน้อย 1 ชุด", "WARNING")
        fails.append("ชุดรางวัล")
    for i, rw in enumerate(rewards, 1):
        if not await _add_reward_set(page, log, rw, i):
            fails.append("ชุดรางวัลที่ %d" % i)

    if not submit:
        log("  กรอกครบแล้ว (โหมดพรีวิว - ยังไม่กดบันทึก ตรวจแล้วกดเองได้)", "SUCCESS")
        return True
    if fails:   # กรอกไม่ครบ -> ไม่กดบันทึก (กัน Item Code ข้อมูลหายถูกสร้างบนเว็บจริง)
        log("  ✋ กรอกไม่ครบ: %s — ไม่กดบันทึกอัตโนมัติ ตรวจ/แก้แล้วกดบันทึกเองในหน้าเว็บ"
            % ", ".join(fails), "WARNING")
        return False
    btn = await _find(page, [
        'button:has-text("บันทึก")', 'button:has-text("สร้าง Item")', 'button:has-text("Save")',
        'button[type="submit"]'], timeout=2500)
    if btn:
        await btn.click(); await page.wait_for_timeout(1200)
        log("  บันทึกแล้ว", "SUCCESS"); return True
    log("  ! หาปุ่มบันทึกไม่เจอ - กรอกให้แล้ว กดบันทึกเองได้", "WARNING")
    return False


# --------------------------------------------------------------------------- reward block (dynamic)
class RewardBlock:
    def __init__(self, app, parent, index):
        self.app = app
        self.frame = card(parent, " ของรางวัลชุดที่ %d " % index)
        self.frame.pack(fill="x", pady=4)
        bg = C["bg_card"]

        top = tk.Frame(self.frame, bg=bg); top.pack(fill="x")
        make_button(top, "ลบชุดนี้", self.remove, "danger").pack(side="right")

        self.name_th = labeled_entry(self.frame, "ชื่อรางวัล (ไทย) *")
        self.name_en = labeled_entry(self.frame, "ชื่อรางวัล (อังกฤษ) *")

        lf = tk.Frame(self.frame, bg=bg); lf.pack(fill="x", pady=2)
        tk.Label(lf, text="จำกัดการใช้งาน", bg=bg, fg=C["text"], font=F, width=20, anchor="w").pack(side="left")
        self.limited = tk.BooleanVar(value=False)
        tk.Checkbutton(lf, text="Limited", variable=self.limited, command=self._toggle, bg=bg, fg=C["text"],
                       selectcolor=C["bg_inp"], activebackground=bg, activeforeground=C["text"], font=F).pack(side="left")

        self.uses_per_user = labeled_entry(self.frame, "ใช้งานต่อ 1 User", "1")
        self.total_uses = labeled_entry(self.frame, "จำนวนครั้งที่ใช้ได้ *")
        self.remaining = labeled_entry(self.frame, "จำนวนคงเหลือ")

        cf = tk.Frame(self.frame, bg=bg); cf.pack(fill="x", pady=2)
        tk.Label(cf, text="ประเภทของ Code", bg=bg, fg=C["text"], font=F, width=20, anchor="w").pack(side="left")
        self.code_type = tk.StringVar(value="Fix Codes")
        self.code_cb = ttk.Combobox(cf, textvariable=self.code_type, values=CODE_TYPES, state="disabled",
                                    width=24, font=F)
        self.code_cb.pack(side="left", padx=4)
        self.code_cb.bind("<<ComboboxSelected>>", lambda e: self._toggle())

        # Fix -> code_list ; Server -> prefix + num_codes
        clf = tk.Frame(self.frame, bg=bg); clf.pack(fill="x", pady=2)
        tk.Label(clf, text="รายการ Code (Fix)", bg=bg, fg=C["text"], font=F, width=20, anchor="nw").pack(side="left")
        self.code_list = tk.Text(clf, bg=C["bg_inp"], fg=C["text"], insertbackground=C["text"],
                                 relief="flat", font=F, width=26, height=3)
        self.code_list.pack(side="left", padx=4)

        self.prefix = labeled_entry(self.frame, "prefix (Server)")
        self.num_codes = labeled_entry(self.frame, "จำนวนโค้ดที่ต้องการ (Server)")
        self.bundles = labeled_entry(self.frame, "bundles (id คั่น ,) *")

        tk.Label(self.frame, text="Limited ปิด = Fix เท่านั้น | Limited เปิด = เลือก Server ได้ (ใส่ prefix+จำนวนโค้ด)",
                 bg=bg, fg=C["muted"], font=SM).pack(anchor="w", padx=2)
        self._toggle()

    def _toggle(self):
        lim = self.limited.get()
        self.code_cb.config(state="readonly" if lim else "disabled")
        if not lim:
            self.code_type.set("Fix Codes")
        for w in (self.total_uses, self.remaining):
            pass  # ค่ายังกรอกได้ (ไม่ล็อก) - ใช้เฉพาะตอน limited
        server = lim and is_server(self.code_type.get())
        # เปิด/ปิด code_list vs prefix/num_codes
        try:
            self.code_list.config(state="disabled" if server else "normal",
                                  bg=C["bg_med"] if server else C["bg_inp"])
        except Exception:
            pass

    def remove(self):
        self.frame.destroy()
        self.app.reward_blocks = [b for b in self.app.reward_blocks if b is not self]
        self.app._renumber_rewards()

    def collect(self):
        return {
            "reward_name_th": self.name_th.get().strip(),
            "reward_name_en": self.name_en.get().strip(),
            "limited": "TRUE" if self.limited.get() else "FALSE",
            "uses_per_user": self.uses_per_user.get().strip() or "1",
            "total_uses": self.total_uses.get().strip(),
            "remaining": self.remaining.get().strip(),
            "code_type": self.code_type.get(),
            "code_list": self.code_list.get("1.0", "end").strip(),
            "prefix": self.prefix.get().strip(),
            "num_codes": self.num_codes.get().strip(),
            "bundles": self.bundles.get().strip(),
        }


# --------------------------------------------------------------------------- App
class App:
    def __init__(self, root, container=None, game_var=None, event_queue_provider=None):
        self.root = root
        # container = frame ที่จะฝัง UI ลงไป (ถ้า None = สร้างลง root โดยตรง = โหมดรันเดี่ยว)
        self.container = container if container is not None else root
        self._embedded = container is not None    # เปิดผ่าน launcher -> ซ่อน เกม/Login ของตัวเอง (ใช้แถบบน)
        # game_var = StringVar เกม/เซิร์ฟที่แชร์ร่วมกับ tool อื่น (เลือกที่ไหนก็เปลี่ยนพร้อมกัน)
        self._game_var = game_var
        self._shared_game = game_var is not None
        # callable() -> dict {group: {'meta':..., 'bundle_id':...}} จากคิว Event ของ launcher
        self._event_queue_provider = event_queue_provider
        self._cancel = False
        self._items = []          # จาก Excel
        self.reward_blocks = []
        if container is None:
            root.title("Create itemcode")
            root.configure(bg=C["bg_dark"])
            root.minsize(820, 720)
            root.geometry("880x820")
        self.vgame = game_var if game_var is not None else \
            tk.StringVar(value=core.load_prefs().get("game", core.GAME_NAMES[0]))
        self.vgame.trace_add("write", lambda *a: self._refresh_url())   # เกมเปลี่ยน (แถบบน) -> อัปเดต URL
        self.item_vars = {}
        self._build_ui()
        self._refresh_url()

    def _build_ui(self):
        hdr = tk.Frame(self.container, bg=C["bg_med"], height=46); hdr.pack(fill="x")
        tk.Label(hdr, text="Create itemcode", bg=C["bg_med"], fg="#ffffff",
                 font=("Segoe UI", 15, "bold")).pack(side="left", padx=14, pady=8)
        tk.Label(hdr, text="combo-interactive", bg=C["bg_med"], fg=C["teal"], font=FB_).pack(side="left", pady=12)

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("TNotebook", background=C["bg_dark"], borderwidth=0)
        style.configure("TNotebook.Tab", background=C["bg_med"], foreground=C["muted"], padding=(14, 6), font=F)
        style.map("TNotebook.Tab", background=[("selected", C["bg_card"])], foreground=[("selected", "#ffffff")])
        style.configure("TCombobox", fieldbackground=C["bg_inp"], background=C["bg_card"], foreground=C["text"],
                        arrowcolor=C["text"], selectbackground=C["bg_inp"], selectforeground=C["text"])
        self.root.option_add("*TCombobox*Listbox.background", C["bg_inp"])
        self.root.option_add("*TCombobox*Listbox.foreground", C["text"])
        self.root.option_add("*TCombobox*Listbox.selectBackground", C["accent"])
        # กัน dropdown เปลี่ยนค่าตอนเลื่อนลูกกลิ้งเมาส์ผ่าน (readonly combobox) — ต้องคลิกเลือกเอง
        for seq in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
            self.root.bind_class("TCombobox", seq, lambda e: "break")

        bar = card(self.container, " ข้อมูล " if self._embedded else " เลือกเกม / เซิร์ฟเวอร์ + เข้าสู่ระบบ ")
        bar.pack(fill="x", padx=10, pady=(8, 4))
        row = tk.Frame(bar, bg=C["bg_card"]); row.pack(fill="x")
        if not self._embedded:
            # รันเดี่ยว -> มี เกม + Login ครบ
            tk.Label(row, text="เกม:", bg=C["bg_card"], fg=C["text"], font=F).pack(side="left")
            self.game_cb = ttk.Combobox(row, textvariable=self.vgame, values=core.GAME_NAMES, state="readonly", width=16, font=F)
            self.game_cb.pack(side="left", padx=8)
            self.game_cb.bind("<<ComboboxSelected>>", lambda e: (self._save_prefs(), self._refresh_url()))
            make_button(row, "เปิดหน้า Login", self._open_login, "accent").pack(side="left", padx=(6, 4))
            make_button(row, "ทดสอบเข้าหน้า Item Code", self._test_nav, "warn").pack(side="left", padx=4)
        else:
            # เปิดผ่าน launcher -> เกม/Login อยู่แถบบน แสดงแค่ชื่อเกมปัจจุบัน
            tk.Label(row, text="เกม:", bg=C["bg_card"], fg=C["muted"], font=F).pack(side="left")
            tk.Label(row, textvariable=self.vgame, bg=C["bg_card"], fg=C["teal"], font=FB_).pack(side="left", padx=(4, 8))
            tk.Label(row, text="(เลือก/Login บนแถบบน)", bg=C["bg_card"], fg=C["muted"], font=SM).pack(side="left")
        make_button(row, "เคลียข้อมูลที่กรอก", self._clear_form, "muted").pack(side="left", padx=4)
        self.url_lbl = tk.Label(bar, text="", bg=C["bg_card"], fg=C["teal"], font=MONO, anchor="w")
        self.url_lbl.pack(fill="x", pady=(6, 0))

        nb = ttk.Notebook(self.container); nb.pack(fill="both", expand=True, padx=8, pady=6)
        self.tab_form = tk.Frame(nb, bg=C["bg_dark"])
        self.tab_bulk = tk.Frame(nb, bg=C["bg_dark"])
        self.tab_log = tk.Frame(nb, bg=C["bg_dark"])
        nb.add(self.tab_form, text="  สร้าง Item Code  ")
        nb.add(self.tab_bulk, text="  นำเข้า Excel  ")
        nb.add(self.tab_log, text="  Log  ")
        self._build_form(self.tab_form)
        self._build_bulk(self.tab_bulk)
        self._build_log(self.tab_log)

    def _wheel_scroll(self, event):
        cv = getattr(self, "_form_canvas", None)
        if cv is None:
            return
        num = getattr(event, "num", None)
        if num == 4:
            cv.yview_scroll(-1, "units")
        elif num == 5:
            cv.yview_scroll(1, "units")
        else:
            cv.yview_scroll(int(-event.delta / 120), "units")
        return "break"   # เลื่อนหน้าฟอร์มอย่างเดียว ไม่ให้ widget ใต้เมาส์ (combobox) โดนเปลี่ยนค่า

    def _bind_wheel(self, w):
        w.bind("<MouseWheel>", self._wheel_scroll)   # Windows / macOS
        w.bind("<Button-4>", self._wheel_scroll)     # Linux ขึ้น
        w.bind("<Button-5>", self._wheel_scroll)     # Linux ลง
        for c in w.winfo_children():
            self._bind_wheel(c)

    def _build_form(self, parent):
        # scrollable
        canvas = tk.Canvas(parent, bg=C["bg_dark"], highlightthickness=0)
        self._form_canvas = canvas
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=C["bg_dark"])
        canvas.create_window((0, 0), window=inner, anchor="nw", width=840)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        item = card(inner, " ข้อมูล Item Code ")
        item.pack(fill="x", padx=8, pady=6)
        self.item_vars["slug"] = labeled_entry(item, "Slug *")
        tk.Label(item, text="slug: a-z 0-9 - เท่านั้น", bg=C["bg_card"], fg=C["muted"], font=SM).pack(anchor="w", padx=2)
        self.item_vars["name_th"] = labeled_entry(item, "ชื่อ Item Code (ไทย) *")
        self.item_vars["name_en"] = labeled_entry(item, "ชื่อ Item Code (อังกฤษ) *")
        tf = tk.Frame(item, bg=C["bg_card"]); tf.pack(fill="x", pady=2)
        tk.Label(tf, text="ประเภท", bg=C["bg_card"], fg=C["text"], font=F, width=20, anchor="w").pack(side="left")
        self.item_vars["type"] = tk.StringVar(value="ALL")
        ttk.Combobox(tf, textvariable=self.item_vars["type"], values=["ALL"], state="readonly", width=24, font=F).pack(side="left", padx=4)
        self.item_vars["uses_per_user"] = labeled_entry(item, "จำนวนใช้งานต่อ 1 User *", "1")
        self.item_vars["total_uses"] = labeled_entry(item, "จำนวนครั้งที่ใช้ได้")
        self.item_vars["remaining"] = labeled_entry(item, "จำนวนคงเหลือ")
        self.item_vars["start_time"] = DateTimePicker(item, "เวลาเริ่มใช้งาน *", datetime.now())
        self.item_vars["end_time"] = DateTimePicker(item, "เวลาสิ้นสุด *", datetime.now() + timedelta(days=30))

        rhead = tk.Frame(inner, bg=C["bg_dark"]); rhead.pack(fill="x", padx=8, pady=(4, 0))
        tk.Label(rhead, text="ชุดรางวัล (มีได้หลายชุด)", bg=C["bg_dark"], fg=C["text"], font=FB_).pack(side="left")
        make_button(rhead, "เพิ่มชุดรางวัล", self._add_reward, "accent").pack(side="left", padx=8)
        make_button(rhead, "📥 เติมจาก Event", self._event_prefill_picker, "green").pack(side="right", padx=8)

        self.reward_host = tk.Frame(inner, bg=C["bg_dark"]); self.reward_host.pack(fill="x", padx=8)
        btns = tk.Frame(inner, bg=C["bg_dark"]); btns.pack(pady=10)
        make_button(btns, "กรอกให้ดูก่อน (ไม่กดบันทึก)", self._preview_one, "warn").pack(side="left", padx=6)
        make_button(btns, "สร้าง + กดบันทึก", self._create_one, "green").pack(side="left", padx=6)
        self._add_reward()  # เริ่มด้วย 1 ชุด
        self._bind_wheel(inner)  # ผูกลูกกลิ้งเมาส์กับทุก widget ในฟอร์ม

    def _add_reward(self):
        idx = len(self.reward_blocks) + 1
        block = RewardBlock(self, self.reward_host, idx)
        self.reward_blocks.append(block)
        if getattr(self, "_form_canvas", None) is not None:
            self._bind_wheel(block.frame)

    def _renumber_rewards(self):
        for i, b in enumerate(self.reward_blocks, 1):
            b.frame.config(text=" ของรางวัลชุดที่ %d " % i)

    def _clear_form(self):
        """ล้างข้อมูลที่กรอกทั้งหมดในหน้า 'สร้าง Item Code' -> กลับเป็นฟอร์มเปล่า"""
        if not messagebox.askyesno("เคลียข้อมูล",
                                   "ล้างข้อมูลที่กรอกทั้งหมดในหน้า 'สร้าง Item Code'?"):
            return
        for k in ("slug", "name_th", "name_en", "total_uses", "remaining"):
            self.item_vars[k].set("")
        self.item_vars["uses_per_user"].set("1")
        self.item_vars["type"].set("ALL")
        self.item_vars["start_time"].set(datetime.now())
        self.item_vars["end_time"].set(datetime.now() + timedelta(days=30))
        while self.reward_blocks:                # ลบชุดรางวัลทั้งหมดแล้วเริ่มใหม่ 1 ชุดเปล่า
            self.reward_blocks[-1].remove()
        self._add_reward()
        self.log("เคลียข้อมูลในฟอร์มแล้ว", "INFO")

    # ---- auto-fill จาก Event ----
    def prefill_from_event(self, meta, bundle_id=""):
        """เติมฟอร์ม Item Code จากเงื่อนไขที่อ่านได้จากไฟล์ Event — ทุกช่องแก้เองได้
        กฎ: ชื่อ=ชื่อกิจกรรม | Unique Code -> Server Generate | codes per set -> จำนวนโค้ด
             total/คงเหลือ = จำนวนโค้ด+10 | Cannot be repeated -> ใช้ต่อ 1 User=1
             วันเริ่ม = วันนี้เที่ยง | วันสิ้นสุด = CODE EXPIRE DATE"""
        meta = meta or {}
        if meta.get("is_pride") and meta.get("pride"):       # template Pride -> กติกาคนละชุด
            return self.prefill_from_pride(meta["pride"], bundle_id)
        activity = str(meta.get("activity") or meta.get("event_name") or "").strip()
        reward = str(meta.get("reward") or "").strip()
        name = (activity + " " + reward).strip() if reward else activity   # ชื่อ = 'กิจกรรม รางวัล'
        slug_src = (activity + "-" + reward) if reward else activity        # slug = 'กิจกรรม-รางวัล-เซิร์ฟ' (eng)
        try:
            server = (self.vgame.get() or "").strip()
        except Exception:
            server = ""
        # ตัดคำว่า 'Cabal' ออก เหลือ platform+region ติดกัน: CabalM TH->mth, CabalPC SEA->pcsea
        scode = cfg.server_code(server)
        if scode:
            slug_src = slug_src + "-" + scode
        # ---- จำนวนโค้ด: codes/set, set count, total ----
        cps = str(meta.get("codes_per_set") or "").strip()
        sets = str(meta.get("set_count") or "").strip()
        total = str(meta.get("total") or "").strip()
        n_sets = int(sets) if sets.isdigit() else 1
        n_cps = int(cps) if cps.isdigit() else 0
        n_total = int(total) if total.isdigit() else (n_cps * n_sets if n_cps else 0)
        # เจนรวม (1 ชุด num=Total): ไม่ซ้ำข้ามเซ็ต / เติมซ้ำได้ | หลายชุด: once per set / เติมข้ามได้
        combined_sig = bool(meta.get("cross_cannot") or meta.get("can_repeat"))
        multiple_sig = bool(meta.get("once_per_set") or meta.get("cross_can"))
        if n_sets <= 1:
            combined = True
        elif multiple_sig:
            combined = False
        elif combined_sig:
            combined = True
        else:
            combined = False        # set>1 ไม่มีสัญญาณชัด -> แยกหลายชุดไว้ก่อน
        # เจนรวม: 1 ชุดรางวัล num=Total | once per set: สร้าง n_sets ชุด แต่ละชุด num=codes/set
        if combined:
            n_blocks = 1
            per_base = n_total
        else:
            n_blocks = max(1, min(n_sets, 300))     # กันสร้างมากเกินไป
            per_base = n_cps
        # buffer ตาม config (PCTH +10, เซิร์ฟอื่น +5) | เฉพาะ 2 ชุดแรกพอ — ที่เหลือใส่เลขเป๊ะ
        buf = cfg.code_buffer(server)
        item_total = str(per_base * n_blocks + buf * min(n_blocks, 2)) if per_base else ""

        if name:
            self.item_vars["name_th"].set(name)
            self.item_vars["name_en"].set(name)
            self.item_vars["slug"].set(slugify(slug_src))
        # จำนวนโค้ดรวมใส่ที่ระดับ Item Code ด้วย (ไม่ใช่แค่ชุดรางวัล)
        if item_total:
            self.item_vars["total_uses"].set(item_total)
            self.item_vars["remaining"].set(item_total)
        if meta.get("cannot_repeat"):
            self.item_vars["uses_per_user"].set("1")
        elif meta.get("can_repeat") and item_total:
            self.item_vars["uses_per_user"].set(item_total)   # เติมซ้ำได้ -> ใช้ต่อ 1 User = จำนวนโค้ด+10
        now = datetime.now()
        # เวลาเริ่ม = วันปัจจุบัน 00:00:00 เสมอ
        self.item_vars["start_time"].set(now.replace(hour=0, minute=0, second=0, microsecond=0))
        end = parse_expire(meta.get("expire"), now)
        # ถ้า expire format ผิด (parse ไม่ได้) ใช้วัน default (วันนี้+30) แต่เวลายังตั้งตามภูมิภาค
        if end is None:
            _exp_raw = str(meta.get("expire") or "").strip()
            if _exp_raw:   # มีค่าในไฟล์แต่อ่านไม่ออก -> เตือนดังๆ (อย่าเงียบ) เพราะเสี่ยงวันหมดอายุผิด
                self.log("⚠ อ่านวันหมดอายุไม่ได้: '%s' — ใช้ค่าเริ่มต้น (วันนี้+30) กรุณาตรวจ/แก้เอง" % _exp_raw, "WARNING")
            end = now + timedelta(days=30)
        # เวลาสิ้นสุดตามภูมิภาคเซิร์ฟ: TH = 23:59:59, SEA = 22:59:59 (เสมอ) — กติกาอยู่ใน config
        end = cfg.region_end_of_day(server, end)
        self.item_vars["end_time"].set(end)

        # ปรับจำนวนชุดรางวัลให้เท่ากับ n_blocks แล้วเติมทุกชุด
        while len(self.reward_blocks) > n_blocks:
            self.reward_blocks[-1].remove()
        while len(self.reward_blocks) < n_blocks:
            self._add_reward()
        for idx, rb in enumerate(self.reward_blocks):
            # หลายชุด -> ต่อท้ายชื่อด้วยเลขลำดับชุด (เช่น 'FP Participation 1', ' 2', ...)
            rname = ("%s %d" % (name, idx + 1)) if (name and n_blocks > 1) else name
            if rname:
                rb.name_th.set(rname)
                rb.name_en.set(rname)
            num_i = str(per_base + (buf if idx < 2 else 0)) if per_base else ""   # +buf เฉพาะ 2 ชุดแรก
            if meta.get("unique_code"):
                rb.limited.set(True)
                rb.code_type.set("Server Generate Codes")
                rb._toggle()
                if num_i:
                    rb.num_codes.set(num_i)
                    rb.total_uses.set(num_i)
                    rb.remaining.set(num_i)
            if meta.get("cannot_repeat"):
                rb.uses_per_user.set("1")
            elif meta.get("can_repeat") and num_i:
                rb.uses_per_user.set(num_i)   # เติมซ้ำได้ -> ใช้ต่อ 1 User = จำนวนโค้ด+10
            if bundle_id:
                rb.bundles.set(str(bundle_id))
        try:
            mode = "เจนรวม Total 1 ชุด" if combined else ("แยก %d ชุด (once per set)" % n_blocks)
            if per_base:
                numdesc = ("%d (+%d=%d) ×2, %d ×%d" % (per_base, buf, per_base + buf, per_base, n_blocks - 2)
                           if n_blocks > 2 else "%d (+%d=%d)" % (per_base, buf, per_base + buf))
            else:
                numdesc = "-"
            self.log("เติมจาก Event: %s | expire=%s | codes/set=%s x set=%s = total=%s | %s | ชุดละ num=%s | %s%s" % (
                name or "-", meta.get("expire") or "-", cps or "-", sets or "1", n_total or "-",
                mode, numdesc,
                "Server Generate" if meta.get("unique_code") else "Fix",
                " | bundle=%s" % bundle_id if bundle_id else ""), "SUCCESS")
            if not combined and n_blocks < n_sets:
                self.log("  ⚠ Number[set]=%s เกิน 300 -> สร้างให้ %d ชุด (เพิ่มเองถ้าต้องการครบ)" % (sets, n_blocks), "WARNING")
        except Exception:
            pass

    # ---- Pride Code Request (template คนละแบบ: หลายบล็อกต่อชีท) ----
    def prefill_from_pride(self, block, bundle_id=""):
        """เติมฟอร์ม Item Code จาก 1 บล็อกใน 'Cabal Pride Code Request'
        Master(มีค่าโค้ด)->Fix Codes ใส่โค้ดตายตัว | Unique->Server Generate
        เริ่ม=วันนี้ 00:00 | จบ=CODE EXPIRE DATE (เวลาตามภูมิภาค)
        bundle_id = id ที่สร้างจากกลุ่มนี้ใน Create Bundle (ถ้ามี)"""
        b = block or {}
        try:
            server = (self.vgame.get() or "").strip()
        except Exception:
            server = ""
        scode = cfg.server_code(server)
        ev = str(b.get("event_name") or "").strip()
        title = str(b.get("title") or "").strip()
        name = (ev + " " + title).strip() if ev else title
        if name:
            self.item_vars["name_th"].set(name)
            self.item_vars["name_en"].set(name)
            base = slugify(name)
            self.item_vars["slug"].set((base + ("-" + scode if scode else "")) if base else "")
        self.item_vars["type"].set("ALL")
        now = datetime.now()
        self.item_vars["start_time"].set(now.replace(hour=0, minute=0, second=0, microsecond=0))
        end = parse_expire(str(b.get("expire") or "").strip(), now)
        if end is None:
            raw = str(b.get("expire") or "").strip()
            if raw:
                self.log("⚠ อ่านวันหมดอายุไม่ได้: '%s' — ใช้ค่าเริ่มต้น (วันนี้+30) ตรวจ/แก้เอง" % raw, "WARNING")
            end = now + timedelta(days=30)
        end = cfg.region_end_of_day(server, end)
        self.item_vars["end_time"].set(end)

        unique = bool(b.get("unique_code"))
        limit = str((b.get("code_count") if unique else b.get("refill_limit")) or "").strip()
        if limit:
            self.item_vars["total_uses"].set(limit)
            self.item_vars["remaining"].set(limit)
        self.item_vars["uses_per_user"].set("1")   # เติมซ้ำไม่ได้ = 1 (ค่าเริ่มต้นของ template นี้)

        while len(self.reward_blocks) > 1:
            self.reward_blocks[-1].remove()
        while len(self.reward_blocks) < 1:
            self._add_reward()
        rb = self.reward_blocks[0]
        rb.name_th.set(name or title)
        rb.name_en.set(name or title)
        rb.uses_per_user.set("1")
        # ล้างช่องโค้ดเดิมก่อน (กันค่าค้างเมื่อสลับบล็อก Fix<->Server)
        try:
            rb.code_list.delete("1.0", "end")
        except Exception:
            pass
        rb.num_codes.set("")
        rb.prefix.set("")
        if unique:
            rb.limited.set(True)
            rb.code_type.set("Server Generate Codes")
            rb._toggle()
            if limit:
                rb.num_codes.set(limit)
                rb.total_uses.set(limit)
                rb.remaining.set(limit)
        else:
            rb.limited.set(False)
            rb.code_type.set("Fix Codes")
            rb._toggle()
            fix = str(b.get("fix_code") or "").strip()
            if fix:
                try:
                    rb.code_list.delete("1.0", "end")
                    rb.code_list.insert("1.0", fix)
                except Exception:
                    pass
            if limit:
                rb.total_uses.set(limit)
                rb.remaining.set(limit)
        if bundle_id:
            rb.bundles.set(str(bundle_id))

        items = b.get("items") or []
        self.log("นำเข้า Pride: %s [%s] | หมดอายุ %s | %s | %d ไอเทม%s" % (
            name or "-", b.get("game_hint") or "-", end.strftime("%Y-%m-%d %H:%M"),
            ("Server Generate x%s" % (limit or "?")) if unique else ("Fix code=%s" % (b.get("fix_code") or "-")),
            len(items),
            (" | bundle=%s" % bundle_id) if bundle_id
            else ("  ← ยังไม่มี bundle: ค้นหาไอเทมกลุ่มนี้ที่ Item Finder แล้วส่งสร้าง bundle ก่อน" if items else "")),
            "SUCCESS")
        if b.get("note"):
            self.log("  หมายเหตุจากไฟล์: " + b["note"], "WARNING")
        if not bundle_id and items:
            preview = ", ".join(("%s %s" % (it["kind"], it["name"])).strip() for it in items[:6])
            self.log("  ไอเทม: " + preview + (" ..." if len(items) > 6 else ""), "INFO")

    def _event_queue_rows(self):
        """คิวจาก launcher (Item Finder Import ไฟล์เดียวจบ) -> แถวในตัวเลือก (info, label)"""
        queue = {}
        if self._event_queue_provider is not None:
            try:
                queue = self._event_queue_provider() or {}
            except Exception:
                queue = {}
        rows = []
        for g in queue.keys():
            info = queue[g] or {}
            m = info.get("meta") or {}
            bid = info.get("bundle_id") or ""
            pr = m.get("pride") if m.get("is_pride") else None
            tag = []
            if pr:                                  # template Pride: โชว์เกม + ชนิดโค้ด
                tag.append(pr.get("game_hint") or "?")
                tag.append("Server x%s" % (pr.get("code_count") or "?") if pr.get("unique_code")
                           else "Fix:%s" % (pr.get("fix_code") or "-"))
            else:
                if m.get("total"):
                    tag.append("โค้ดรวม %s" % m["total"])
                elif m.get("codes_per_set"):
                    tag.append("โค้ด %s" % m["codes_per_set"])
                if m.get("unique_code"):
                    tag.append("Server")
            tag.append("bundle %s" % bid if bid else "ยังไม่มี bundle")
            rows.append((info, "%-34s · %s" % (g, " · ".join(tag))))   # ไม่ตัดชื่อ (เลื่อนแนวนอนดูได้)
        return rows

    def _event_prefill_picker(self):
        """เลือกกลุ่มจากคิว (Item Finder) มาเติมฟอร์ม — รองรับทั้ง Event และ Pride Code Request"""
        rows = self._event_queue_rows()
        if not rows:
            messagebox.showinfo("ยังไม่มีคิว",
                                "ยังไม่มีข้อมูล — ไป Item Finder แล้ว Import ไฟล์ Event / Pride Code Request ก่อน")
            return
        win = tk.Toplevel(self.root)
        win.title("เลือกกลุ่มมาเติม Item Code")
        win.configure(bg=C["bg_dark"]); win.geometry("660x500"); win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass
        tk.Label(win, text="เลือก 1 กลุ่ม → เติมฟอร์มให้อัตโนมัติ (แก้ต่อได้) · Pride: ตั้งเกม/เซิร์ฟให้ตรง CBM/CBPC เองก่อน",
                 bg=C["bg_dark"], fg=C["muted"], font=F).pack(anchor="w", padx=12, pady=(10, 4))
        lf = tk.Frame(win, bg=C["bg_dark"]); lf.pack(fill="both", expand=True, padx=12, pady=4)
        lb = tk.Listbox(lf, bg=C["bg_med"], fg=C["text"], font=("Consolas", 10), relief="flat",
                        selectbackground=C["accent"], selectforeground="#fff", activestyle="none")
        sb = ttk.Scrollbar(lf, command=lb.yview); lb.configure(yscrollcommand=sb.set)
        hsb = ttk.Scrollbar(lf, orient="horizontal", command=lb.xview)   # ชื่อยาว -> เลื่อนดูได้
        lb.configure(xscrollcommand=hsb.set)
        sb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        lb.pack(side="left", fill="both", expand=True)
        for _info, label in rows:
            lb.insert(tk.END, label)

        def _do():
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo("เลือกก่อน", "ยังไม่ได้เลือกกลุ่ม", parent=win)
                return
            info = rows[sel[0]][0]
            win.destroy()
            # meta ที่มี is_pride จะถูกส่งต่อให้ prefill_from_pride เองใน prefill_from_event
            self.prefill_from_event(info.get("meta") or {}, info.get("bundle_id") or "")

        bar = tk.Frame(win, bg=C["bg_dark"]); bar.pack(fill="x", padx=12, pady=(4, 10))
        make_button(bar, "เติมฟอร์ม", _do, "green").pack(side="left")
        make_button(bar, "ปิด", win.destroy, "muted").pack(side="left", padx=8)
        win.listbox, win.do_pick, win.rows = lb, _do, rows
        return win

    def _build_bulk(self, parent):
        top = tk.Frame(parent, bg=C["bg_dark"]); top.pack(fill="x", padx=10, pady=10)
        make_button(top, "โหลด Template (.xlsx)", self._save_template, "accent").pack(side="left")
        make_button(top, "เลือกไฟล์ Excel", self._pick_excel, "muted").pack(side="left", padx=6)
        make_button(top, "ตรวจไฟล์", self._validate_file, "muted").pack(side="left", padx=(0, 6))
        make_button(top, "กรอกให้ดูก่อน (อันแรก)", self._preview_bulk, "warn").pack(side="left", padx=(0, 6))
        make_button(top, "สร้างทั้งหมด", self._create_all, "green").pack(side="left")
        tk.Label(parent, text="2 ชีต: itemcodes (โค้ดหลัก) + rewards (ชุดรางวัล ผูกด้วย itemcode_slug)",
                 bg=C["bg_dark"], fg=C["muted"], font=SM).pack(anchor="w", padx=12)
        bwrap = tk.Frame(parent, bg=C["bg_dark"]); bwrap.pack(fill="both", expand=True, padx=12, pady=(6, 12))
        bvsb = ttk.Scrollbar(bwrap, orient="vertical")
        bhsb = ttk.Scrollbar(bwrap, orient="horizontal")   # เลื่อนแนวนอน ดูชื่อยาวๆ ได้
        bvsb.pack(side="right", fill="y")
        bhsb.pack(side="bottom", fill="x")
        self.bulk_view = tk.Text(bwrap, bg=C["bg_inp"], fg=C["text"], font=MONO, relief="flat",
                                 wrap="none", insertbackground=C["text"],
                                 yscrollcommand=bvsb.set, xscrollcommand=bhsb.set)
        self.bulk_view.pack(side="left", fill="both", expand=True)
        bvsb.config(command=self.bulk_view.yview)
        bhsb.config(command=self.bulk_view.xview)
        for tag, col in (("h", C["teal"]), ("k", C["accent"]), ("w", C["warn"]), ("m", C["muted"])):
            self.bulk_view.tag_config(tag, foreground=col)
        self.bulk_view.insert("end", "(ยังไม่ได้เลือกไฟล์)")

    def _build_log(self, parent):
        top = tk.Frame(parent, bg=C["bg_dark"]); top.pack(fill="x", padx=8, pady=(8, 0))
        make_button(top, "ล้าง Log", self._clear_log, "muted").pack(side="right")
        self.log_area = tk.Text(parent, bg=C["bg_med"], fg=C["text"], font=MONO, relief="flat", wrap="word", insertbackground=C["text"])
        self.log_area.pack(fill="both", expand=True, padx=8, pady=8)
        for lvl, col in (("SUCCESS", C["accent2"]), ("WARNING", C["warn"]), ("ERROR", C["danger"]),
                         ("INFO", C["text"]), ("STEP", C["accent"])):
            self.log_area.tag_config(lvl, foreground=col)
        self.log("มี session เก็บไว้แล้ว" if core.has_session() else "ยังไม่มี session ที่เก็บไว้", "INFO")
        if not core.PW_OK:
            self.log("ไม่พบ playwright - รัน: pip install playwright แล้ว playwright install chromium", "WARNING")

    # ---- helpers ----
    def log(self, msg, level="INFO"):
        # thread-safe: Playwright coroutine เรียกจาก worker thread -> marshal เข้า Tk main thread
        try:
            self.root.after(0, self._log_main, str(msg), level)
        except Exception:
            pass

    def _log_main(self, msg, level):
        self.log_area.insert("end", msg + "\n", level); self.log_area.see("end")

    def _clear_log(self):
        self.log_area.delete("1.0", "end")

    def _refresh_url(self):
        self.url_lbl.config(text=core.build_url(self.vgame.get(), CREATE_PATH))

    def _save_prefs(self):
        p = core.load_prefs(); p["game"] = self.vgame.get(); core.save_prefs(p)

    def _run_async(self, coro_factory):
        def worker():
            try:
                asyncio.run(coro_factory())
            except Exception:
                self.log(traceback.format_exc(), "ERROR")
        threading.Thread(target=worker, daemon=True).start()

    def _create_url(self):
        return core.build_url(self.vgame.get(), CREATE_PATH)

    def _collect_item(self):
        d = {k: self.item_vars[k].get().strip() for k in ITEM_KEYS}
        d["rewards"] = [b.collect() for b in self.reward_blocks]
        return d

    def _validate_item(self, it):
        for k, label, _dft, req in ITEM_FIELDS:
            if req and not str(it.get(k, "")).strip():
                return False, "Item: กรุณากรอก " + label
        if not valid_slug(it.get("slug", "")):
            return False, "slug ต้องเป็น a-z 0-9 - เท่านั้น"
        if not it.get("rewards"):
            return False, "ต้องมีชุดรางวัลอย่างน้อย 1 ชุด"
        for i, rw in enumerate(it["rewards"], 1):
            if not rw.get("reward_name_th") or not rw.get("reward_name_en"):
                return False, "ชุดรางวัลที่ %d: ต้องมีชื่อไทย/อังกฤษ" % i
            if not split_ids(rw.get("bundles")):
                return False, "ชุดรางวัลที่ %d: ต้องใส่ bundles อย่างน้อย 1 id" % i
            if as_bool(rw.get("limited")) and is_server(rw.get("code_type")) and not str(rw.get("num_codes")).strip():
                return False, "ชุดรางวัลที่ %d: Server ต้องใส่จำนวนโค้ด" % i
        return True, ""

    # ---- login/nav ----
    def _open_login(self):
        self._save_prefs()
        s = core.AztekSession(log=self.log, should_cancel=lambda: self._cancel)
        url = core.build_url(self.vgame.get(), LIST_PATH)   # อ่าน StringVar บน main thread ก่อน (thread-safe)
        self._run_async(lambda: s.open_login(url))

    def _clear_session(self):
        ok = core.clear_profile()
        self.log("ล้าง session แล้ว" if ok else "ล้าง session ไม่สำเร็จ", "SUCCESS" if ok else "ERROR")

    def _test_nav(self):
        self._save_prefs()
        s = core.AztekSession(log=self.log, should_cancel=lambda: self._cancel)
        url = core.build_url(self.vgame.get(), LIST_PATH)

        async def run():
            async with s.context(url) as (browser, page):
                if await core.is_logged_in(page):
                    self.log("PASS เข้าหน้า Item Code ได้ | url=" + page.url, "SUCCESS")
                else:
                    self.log("FAIL x ยังไม่ได้ login | url=" + page.url, "ERROR")
                await page.wait_for_timeout(1500)
        self._run_async(run)

    def _preview_one(self):
        """กรอกฟอร์มให้แต่ไม่กดบันทึก แล้วเปิด browser ค้างไว้ให้ตรวจ"""
        self._run_form(submit=False)

    def _create_one(self):
        """กรอก + กดบันทึก แล้วเปิด browser ค้างไว้ให้ดูผล"""
        self._run_form(submit=True)

    def _run_form(self, submit):
        self._save_prefs()
        it = self._collect_item()
        ok, err = self._validate_item(it)
        if not ok:
            messagebox.showwarning("ข้อมูลไม่ครบ/ไม่ถูกต้อง", err); return
        s = core.AztekSession(log=self.log, should_cancel=lambda: self._cancel)
        url = self._create_url()

        async def run():
            await s.run_page(url, lambda page: create_itemcode(page, self.log, it, submit=submit), hold=True)
        self._run_async(run)

    # ---- Excel ----
    def _save_template(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="itemcode_template.xlsx",
                                            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            build_template_xlsx(path); self.log("บันทึก template: " + path, "SUCCESS")
        except ImportError:
            self.log("ต้องติดตั้ง openpyxl: pip install openpyxl", "ERROR")
        except Exception:
            self.log(traceback.format_exc(), "ERROR")

    def _pick_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
        if not path:
            return
        try:
            self._items, orphan = read_template_xlsx(path)
        except ImportError:
            self.log("ต้องติดตั้ง openpyxl: pip install openpyxl", "ERROR"); return
        except Exception:
            self.log(traceback.format_exc(), "ERROR"); return
        self._render_items(orphan)
        if orphan:
            self.log("เตือน: มี %d แถวในชีต rewards ที่ itemcode_slug ไม่ตรงกับชีต itemcodes" % len(orphan), "WARNING")
        self.log("อ่าน %d Item Code จาก %s" % (len(self._items), os.path.basename(path)), "INFO")
        self._validate_file(popup=False)

    def _render_items(self, orphan=None):
        v = self.bulk_view
        v.delete("1.0", "end")
        if not self._items:
            v.insert("end", "(ไม่มีข้อมูล)"); return
        for i, it in enumerate(self._items, 1):
            rewards = it.get("rewards", [])
            v.insert("end", "#%d  %s\n" % (i, it.get("name_th") or "(ไม่มีชื่อ)"), "h")
            v.insert("end", "    slug=%s%s | type=%s | ต่อuser=%s | ครั้ง=%s | คงเหลือ=%s\n" %
                     (it.get("slug"), " (auto)" if it.get("_slug_auto") else "",
                      it.get("type"), it.get("uses_per_user"), it.get("total_uses") or "-", it.get("remaining") or "-"))
            v.insert("end", "    ชื่ออังกฤษ=%s\n" % (it.get("name_en") or "-"))
            v.insert("end", "    เวลา: %s  ->  %s\n" % (it.get("start_time"), it.get("end_time")))
            v.insert("end", "    ชุดรางวัล: %d ชุด\n" % len(rewards), "k")
            for j, rw in enumerate(rewards, 1):
                lim = as_bool(rw.get("limited"))
                v.insert("end", "      [%d] %s / %s\n" % (j, rw.get("reward_name_th") or "-", rw.get("reward_name_en") or "-"))
                v.insert("end", "          Limited=%s | ต่อuser=%s | ครั้ง=%s | คงเหลือ=%s | code_type=%s\n" %
                         (rw.get("limited"), rw.get("uses_per_user") or "1",
                          rw.get("total_uses") or "-", rw.get("remaining") or "-", rw.get("code_type")))
                if lim and is_server(rw.get("code_type")):
                    v.insert("end", "          prefix=%s | จำนวนโค้ด=%s\n" % (rw.get("prefix") or "-", rw.get("num_codes") or "-"))
                else:
                    codes = [c for c in str(rw.get("code_list") or "").splitlines() if c.strip()]
                    v.insert("end", "          รายการ Code (%d): %s\n" % (len(codes), ", ".join(codes) if codes else "-"))
                v.insert("end", "          bundles: %s\n" % (", ".join(split_ids(rw.get("bundles"))) or "-"), "w")
            v.insert("end", "\n")
        if orphan:
            v.insert("end", "── แถว rewards ที่ slug ไม่ตรง (orphan) ──\n", "w")
            for rw in orphan:
                v.insert("end", "    itemcode_slug=%s | %s\n" % (rw.get("itemcode_slug"), rw.get("reward_name_th")), "m")

    def _preview_bulk(self):
        """กรอก 'อันแรก' จาก Excel ให้ดู โดยไม่กดบันทึก แล้วค้าง browser ไว้ตรวจ"""
        if not self._items:
            messagebox.showinfo("ยังไม่มีข้อมูล", "เลือกไฟล์ Excel ก่อน"); return
        self._save_prefs()
        it = self._items[0]
        ok, err = self._validate_item(it)
        if not ok:
            messagebox.showwarning("ข้อมูลไม่ครบ/ไม่ถูกต้อง", "อันแรก: " + err); return
        s = core.AztekSession(log=self.log, should_cancel=lambda: self._cancel)
        url = self._create_url()

        async def run():
            self.log("พรีวิวอันแรก: %s (%d ชุดรางวัล)" % (it.get("name_th"), len(it.get("rewards", []))), "STEP")
            await s.run_page(url, lambda page: create_itemcode(page, self.log, it, submit=False), hold=True)
        self._run_async(run)

    def _validate_file(self, popup=True):
        """ตรวจทุกแถวก่อนรัน บอกแถวที่มีปัญหา"""
        if not self._items:
            if popup:
                messagebox.showinfo("ยังไม่มีข้อมูล", "เลือกไฟล์ Excel ก่อน")
            return 0
        problems = 0
        seen = {}
        self.log("===== ตรวจไฟล์ =====", "STEP")
        for i, it in enumerate(self._items, 1):
            errs = []
            ok, err = self._validate_item(it)
            if not ok:
                errs.append(err)
            s = it.get("slug")
            if s in seen:
                errs.append("slug '%s' ซ้ำกับ #%d" % (s, seen[s]))
            else:
                seen[s] = i
            auto = " (auto slug)" if it.get("_slug_auto") else ""
            if errs:
                problems += 1
                self.log("[X] #%d %s%s" % (i, it.get("name_th"), auto), "ERROR")
                for e in errs:
                    self.log("      - " + e, "WARNING")
            else:
                self.log("[OK] #%d %s%s | slug=%s | %d ชุดรางวัล" %
                         (i, it.get("name_th"), auto, s, len(it.get("rewards", []))), "SUCCESS")
        msg = "ตรวจ %d อัน — %s" % (len(self._items),
                                     "ผ่านหมด พร้อมสร้าง" if problems == 0 else "มีปัญหา %d อัน (ดู Log)" % problems)
        self.log(msg, "SUCCESS" if problems == 0 else "WARNING")
        if popup:
            (messagebox.showinfo if problems == 0 else messagebox.showwarning)("ผลตรวจไฟล์", msg)
        return problems

    def _create_all(self):
        if not self._items:
            messagebox.showinfo("ยังไม่มีข้อมูล", "เลือกไฟล์ Excel ก่อน"); return
        self._save_prefs()
        items = list(self._items)
        s = core.AztekSession(log=self.log, should_cancel=lambda: self._cancel)
        url = self._create_url()

        async def run():
            done = 0; failed = 0
            async with s.context(url) as (browser, page):
                for i, it in enumerate(items, 1):
                    if self._cancel:
                        self.log("ยกเลิกกลางคัน", "WARNING"); break
                    ok, err = self._validate_item(it)
                    if not ok:
                        self.log("ข้าม #%d (%s): %s" % (i, it.get("name_th"), err), "WARNING"); failed += 1; continue
                    self.log("========== [%d/%d] %s ==========" % (i, len(items), it.get("name_th")), "STEP")
                    try:
                        ok = await create_itemcode(page, self.log, it)
                        if ok:
                            done += 1
                        else:
                            # คืน False = กรอกไม่ครบ/หาปุ่มบันทึกไม่เจอ -> ยังไม่ได้ถูกสร้างบนเว็บจริง
                            failed += 1
                            self.log("ไม่สำเร็จ #%d: %s (กรอกไม่ครบ/บันทึกไม่ได้ — ไม่ถูกสร้างบนเว็บ)"
                                     % (i, it.get("name_th")), "WARNING")
                    except Exception:
                        failed += 1
                        self.log("พลาด #%d: %s" % (i, it.get("name_th")), "ERROR")
                        self.log(traceback.format_exc(), "ERROR")
                    # กลับไปหน้า create เพื่อทำอันถัดไป (ต่อให้อันนี้พลาดก็ไปต่อ)
                    try:
                        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                        await page.wait_for_timeout(700)
                    except Exception:
                        pass
            self.log("สร้างเสร็จ %d/%d Item Code (พลาด/ข้าม %d)" % (done, len(items), failed),
                     "SUCCESS" if failed == 0 else "WARNING")
        self._run_async(run)


def main():
    root = tk.Tk()
    _set_window_icon(root)
    App(root)
    root.mainloop()


# ทะเบียนสำหรับ All for Cabal launcher
from tool_registry import ToolSpec
TOOL = ToolSpec(
    key='itemcode', icon='🎟️', title='Create Item Code', nav='🎟️  Item Code',
    desc='สร้าง Item Code + ชุดรางวัลบนเว็บ aztek-tools', boot='warning',
    make=lambda lc, fr: App(lc.root, container=fr, game_var=lc.game_var,
                            event_queue_provider=lc._event_queue))


if __name__ == "__main__":
    main()
