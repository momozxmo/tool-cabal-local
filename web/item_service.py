# -*- coding: utf-8 -*-
"""GUI-free Item Finder helpers shared by the web endpoints.

This module contains only deterministic data transforms and byte exporters.
It deliberately does not know about FastAPI, tkinter, or browser sessions.
"""
from dataclasses import dataclass
from datetime import datetime
import csv
import io
import threading
import uuid

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

import event_tool
import item_finder
from web import event_plan


_WORKBOOK_PARSE_LOCK = threading.Lock()


def parse_workbook_locked(parser, path):
    """Serialize parsers that temporarily monkey-patch openpyxl internals."""
    with _WORKBOOK_PARSE_LOCK:
        return parser(path)


@dataclass
class MergeResult:
    criteria: list
    occurrences: list
    group_meta: dict
    added: int
    merged: int


@dataclass
class Workspace:
    id: str
    mode: str
    filename: str
    criteria: list
    occurrences: list
    group_meta: dict
    skipped: list
    results: list
    not_found: list
    game: str = ''


@dataclass
class PendingImport:
    id: str
    workspace_id: str
    sheets: list
    skipped: list


class WorkspaceStore:
    """Small in-memory store for the local MVP; API shape can later use Postgres."""
    def __init__(self):
        self._workspaces = {}
        self._pending = {}
        self._lock = threading.RLock()

    def create(self, mode, filename='', criteria=None):
        with self._lock:
            mode_policy(mode)
            workspace = Workspace(
                id=uuid.uuid4().hex, mode=mode, filename=filename,
                criteria=[dict(row) for row in (criteria or [])], occurrences=[],
                group_meta={}, skipped=[], results=[], not_found=[])
            self._workspaces[workspace.id] = workspace
            return workspace

    def get(self, workspace_id):
        with self._lock:
            return self._workspaces[workspace_id]

    def delete(self, workspace_id):
        with self._lock:
            del self._workspaces[workspace_id]

    def replace_template(self, workspace_id, filename, criteria):
        with self._lock:
            workspace = self.get(workspace_id)
            workspace.filename = filename
            workspace.criteria = [dict(row) for row in criteria]
            workspace.occurrences = []
            workspace.group_meta = {}
            workspace.skipped = []
            workspace.results = []
            workspace.not_found = []
            return workspace

    def add_pending(self, workspace_id, sheets, skipped=None):
        with self._lock:
            self.get(workspace_id)
            pending = PendingImport(uuid.uuid4().hex, workspace_id,
                                    [(name, [dict(row) for row in rows])
                                     for name, rows in sheets], list(skipped or []))
            self._pending[pending.id] = pending
            return pending

    def pending(self, pending_id):
        with self._lock:
            return self._pending[pending_id]

    def apply_pending(self, pending_id, selected_sheets):
        with self._lock:
            pending = self._pending.pop(pending_id)
            workspace = self.get(pending.workspace_id)
            selected = set(selected_sheets)
            items = []
            for sheet_name, rows in pending.sheets:
                if sheet_name in selected:
                    items.extend(
                        stamp_sheet_rows(sheet_name, rows)
                        if workspace.mode == 'event' else rows)
            merged = merge_imported(workspace.criteria, workspace.occurrences,
                                    workspace.group_meta, items)
            workspace.criteria = merged.criteria
            workspace.occurrences = merged.occurrences
            workspace.group_meta = merged.group_meta
            workspace.skipped.extend(pending.skipped)
            workspace.results = []
            workspace.not_found = []
            return workspace


def parser_for_mode(mode):
    parsers = {
        'event': event_tool.parse_event_plan_workbook,
        'itemcode': item_finder.parse_event_workbook,
        'shop': item_finder.parse_shop_workbook,
    }
    try:
        return parsers[mode]
    except KeyError:
        raise ValueError('ไม่รู้จักโหมด: %s' % mode)


def mode_policy(mode):
    policies = {
        'event': {'web_mode': 'any', 'web_locked': False, 'read_desc': False},
        'itemcode': {'web_mode': 'no', 'web_locked': True, 'read_desc': False},
        'shop': {'web_mode': 'no', 'web_locked': False, 'read_desc': True},
    }
    if mode not in policies:
        raise ValueError('ไม่รู้จักโหมด: %s' % mode)
    return dict(policies[mode])


def _criteria_key(row):
    return (str(row.get('kind', '') or ''), str(row.get('opt', '') or ''),
            str(row.get('dur', '') or ''), str(row.get('name', '') or '').strip())


def stamp_sheet_rows(sheet_name, rows):
    """Copy imported rows and attach worksheet-aware reward identities."""
    group_keys = {}
    stamped = []
    for source in rows:
        row = dict(source)
        sources = list(row.get('sources') or [])
        raw_meta = dict(row.get('group_meta') or {})
        reward_index = raw_meta.get('reward_index')
        internal = []
        for group in sources:
            label = str(group or '').strip()
            identity = (label, reward_index) if reward_index else label
            if identity not in group_keys:
                ordinal = reward_index or (len(group_keys) + 1)
                group_keys[identity] = event_plan.make_group_key(
                    sheet_name, label, ordinal)
            internal.append(group_keys[identity])
        row['sources'] = sources
        row['group_keys'] = internal
        if row.get('group_meta'):
            meta = raw_meta
            meta['sheet'] = str(sheet_name or '').strip()
            meta['sheet_key'] = event_plan.make_sheet_key(sheet_name)
            if internal:
                meta['group_key'] = internal[0]
                meta.setdefault('group', str(sources[0] or '').strip())
            row['group_meta'] = meta
        stamped.append(row)
    return stamped


def merge_imported(criteria, occurrences, group_meta, items):
    """Mirror App._apply_event_items without any widget dependencies."""
    criteria = [dict(row) for row in criteria]
    occurrences = [dict(row) for row in occurrences]
    group_meta = dict(group_meta)
    index = {}
    for row in criteria:
        row.setdefault('sources', [])
        if 'group_keys' in row:
            row['group_keys'] = list(row.get('group_keys') or [])
        index[_criteria_key(row)] = row

    added = merged = 0
    for source in items:
        row = dict(source)
        row['sources'] = list(row.get('sources') or [])
        keys = (list(row.get('group_keys') or [])
                if 'group_keys' in row else None)
        if keys is not None:
            row['group_keys'] = keys
        meta = row.pop('group_meta', None)
        if meta:
            for group_key, group in zip(
                    keys or row['sources'], row['sources']):
                saved = dict(meta)
                saved['group_key'] = group_key
                saved.setdefault('group', group)
                group_meta[group_key] = saved
        occurrences.append(dict(row))
        key = _criteria_key(row)
        if key in index:
            target = index[key]
            for position, group in enumerate(row['sources']):
                if group not in target['sources']:
                    target['sources'].append(group)
                    merged += 1
                if keys is not None:
                    group_key = keys[position]
                    target.setdefault('group_keys', [])
                    if group_key not in target['group_keys']:
                        target['group_keys'].append(group_key)
        else:
            criteria.append(row)
            index[key] = row
            added += 1
    return MergeResult(criteria, occurrences, group_meta, added, merged)


def _found_keys(results):
    """The (kind, opt, dur) triples that some result already covers."""
    return {(str(row.get('item_kind', '') or '').strip(),
             str(row.get('item_option', '') or '').strip(),
             str(row.get('duration_index', '') or '').strip())
            for row in results}


def missing_criteria(criteria, results):
    """The plan rows that came back with nothing, ready to search again.

    Each carries the label it had in the full run, so a retry's 'not found'
    list still names rows by their number in the plan rather than by their
    position in the retry.
    """
    found = _found_keys(results)
    missing = []
    for index, row in enumerate(criteria):
        key = (str(row.get('kind', '') or '').strip(),
               str(row.get('opt', '') or '').strip(),
               str(row.get('dur', '') or '').strip())
        if key in found:
            continue
        retry = dict(row)
        retry['_label'] = item_finder.criteria_label(index, row)
        missing.append(retry)
    return missing


def merge_found(previous, fresh, occurrences):
    """Fold a retry's finds into what the first pass already found.

    A retry only runs the rows that came back empty, so the earlier results
    have to survive it. Everything is put back through document order at the
    end, otherwise the newcomers would simply pile up at the bottom.
    """
    seen = set()
    combined = []
    for row in list(previous) + list(fresh):
        key = (str(row.get('aztek_id', '') or ''),
               str(row.get('item_kind', '') or ''),
               str(row.get('item_option', '') or ''),
               str(row.get('duration_index', '') or ''))
        if key in seen:
            continue
        seen.add(key)
        combined.append(dict(row))
    if not occurrences:
        return combined
    ordered = regroup_results(combined, occurrences)
    for row in ordered:
        # regroup rewrites 'sources'; the display column has to follow it.
        row['groups'] = ' , '.join(str(s) for s in (row.get('sources') or []))
    return ordered


def regroup_results(results, occurrences):
    """Expand found rows back into document/group order, matching desktop."""
    if not occurrences or not results:
        return [dict(row) for row in results]
    found = {}
    for row in results:
        key = (str(row.get('item_kind', '') or '').strip(),
               str(row.get('item_option', '') or '').strip(),
               str(row.get('duration_index', '') or '').strip())
        bucket = found.setdefault(key, [])
        if not any(old.get('aztek_id') == row.get('aztek_id') for old in bucket):
            bucket.append(row)
    expanded = []
    for occurrence in occurrences:
        key = (str(occurrence.get('kind', '') or '').strip(),
               str(occurrence.get('opt', '') or '').strip(),
               str(occurrence.get('dur', '') or '').strip())
        for found_row in found.get(key, []):
            row = dict(found_row)
            row['sources'] = list(occurrence.get('sources') or [])
            if 'group_keys' in occurrence:
                row['group_keys'] = list(
                    occurrence.get('group_keys') or [])
            row['file_name'] = occurrence.get('name', '') or ''
            # Carry the quantity ('Amt' column) from the imported row so the
            # bundle dialog can auto-fill qty instead of defaulting to 1.
            row['amt'] = occurrence.get('amt', '') or ''
            # A random box's plan lists a draw rate per item; carrying it means
            # the operator never types the odds back in by hand.
            row['rate'] = occurrence.get('rate', '') or ''
            expanded.append(row)
    return expanded or [dict(row) for row in results]


def _bundle_name(group, group_meta, group_key=None):
    # Prefix the bundle name with the activity/event title so the operator sees
    # which event a bundle belongs to. Event imports carry 'event_name'; shop
    # imports carry 'shop_sheet'/'activity'.
    meta = group_meta.get(group_key or group) or group_meta.get(group) or {}
    title = str(
        meta.get('event_name') or meta.get('shop_sheet') or meta.get('activity')
        or '').strip()
    if title and not group.startswith(title):
        return '%s - %s' % (title, group)
    return group


def build_bundles(results, group_meta):
    """Build one bundle per group, items in the order the plan file listed them."""
    item_groups = {}
    for row in results:
        item_id = str(row.get('aztek_id', '') or '').strip()
        if not item_id:
            continue
        groups = row.get('sources') or ['(ไม่มีกลุ่ม)']
        keys = row.get('group_keys') or groups
        item_groups.setdefault(item_id, set()).update(
            str(key or group or '').strip() or '(ไม่มีกลุ่ม)'
            for group, key in zip(groups, keys))

    order, grouped = [], {}
    for row in results:
        item_id = str(row.get('aztek_id', '') or '').strip()
        if not item_id:
            continue
        shared = len(item_groups.get(item_id, ())) > 1
        groups = row.get('sources') or ['(ไม่มีกลุ่ม)']
        keys = row.get('group_keys') or groups
        for raw_group, raw_key in zip(groups, keys):
            group = str(raw_group or '').strip() or '(ไม่มีกลุ่ม)'
            group_key = str(raw_key or '').strip() or group
            if group_key not in grouped:
                grouped[group_key] = {
                    'group': group, 'seen': set(), 'items': []}
                order.append(group_key)
            bucket = grouped[group_key]
            if item_id in bucket['seen']:
                continue
            bucket['seen'].add(item_id)
            # Document order is the operator's order: the plan file lists items
            # the way the bundle should read, and results are already expanded
            # back into it. Shared items keep their place and are found by their
            # highlight, not by being swept to the bottom.
            # 'shared' marks an item that appears in more than one bundle, so the
            # UI can highlight it and let the user drop the duplicate.
            # qty comes from the imported 'Amt' column when present; else 1.
            qty = str(row.get('amt', '') or '').strip() or '1'
            # Everything the operator checks a bundle against travels with the
            # item. Sending only id/name/qty meant the check had to happen back
            # in the results table, which holds every group at once — far more
            # to read than the handful of rows actually in this bundle.
            bucket['items'].append(
                {'id': item_id, 'name': row.get('item_name', '') or '',
                 'shared': shared, 'qty': qty,
                 # The name as the document wrote it. The search matches on
                 # kind/opt/dur, never on the name, so a name that does not line
                 # up is the one thing only a human can settle.
                 'file_name': row.get('file_name', '') or '',
                 'name_mismatch': bool(row.get('name_mismatch')),
                 # What deep check verified: 'เว็บ✗ รูป✓ เทรด∅ …' — ✓ must be
                 # set, ✗ must not, ∅ must be blank. Passed through as recorded,
                 # because what counts as right differs per mode and per item.
                 'params': row.get('params', '') or '',
                 'desc': row.get('desc', '') or '',
                 'doc_qty': str(row.get('amt', '') or '').strip(),
                 # Odds from the plan's random-box table, already a percentage.
                 'rate': str(row.get('rate', '') or '').strip()})
    return [
        {'name': _bundle_name(grouped[key]['group'], group_meta, key),
         'group': grouped[key]['group'], 'group_key': key,
         # A product whose plan gave draw rates is a random box, so the bundle
         # is created as RANDOM without the operator having to spot it.
         'is_random': bool(
             (group_meta.get(key)
              or group_meta.get(grouped[key]['group'])
              or {}).get('is_random')),
         'items': grouped[key]['items']}
        for key in order
    ]


_EXPORT_HEADERS = ['#', 'aztek_id', 'item_name', 'item_kind', 'item_option',
                   'duration_index', 'game', 'notes', 'criteria_no', 'groups',
                   'description']


def _safe_cell(value):
    """Prevent CSV/XLSX formula execution while preserving displayed text."""
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return "'" + value
    return value


def _export_row(index, item):
    values = [index, item.get('aztek_id', ''), item.get('item_name', ''),
              item.get('item_kind', ''), item.get('item_option', ''),
              item.get('duration_index', ''), item.get('game', ''),
              item.get('notes', 'passed'), item.get('_ci', ''),
              ' | '.join(item.get('sources') or []), item.get('_desc', '')]
    return [_safe_cell(value) for value in values]


def export_csv_bytes(results):
    out = io.StringIO(newline='')
    writer = csv.writer(out)
    writer.writerow(_EXPORT_HEADERS)
    for index, item in enumerate(results, 1):
        writer.writerow(_export_row(index, item))
    return ('\ufeff' + out.getvalue()).encode('utf-8')


def export_xlsx_bytes(results, game):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Results'
    header_fill = PatternFill('solid', fgColor='1F6FEB')
    row_fill = PatternFill('solid', fgColor='1A3A1A')
    white = Font(color='FFFFFF', bold=True)
    green = Font(color='3FB950')
    gray_fill = PatternFill('solid', fgColor='21262D')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    summaries = [
        'Aztek Item Finder — Export  |  เกม: %s  |  วันที่: %s' % (game, timestamp),
        'พบทั้งหมด %d items  |  สร้างโดย All for Cabal Web' % len(results), '',
    ]
    for row_index, text in enumerate(summaries, 1):
        cell = sheet.cell(row=row_index, column=1, value=text)
        cell.font = Font(italic=True, color='8B949E', size=9)
        cell.fill = gray_fill
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=11)

    xlsx_headers = ['#', 'Aztek ID', 'Item Name', 'ItemKind', 'itemOption',
                    'durationIndex', 'Game', 'Notes', 'Criteria #',
                    'Groups (อยู่ตารางไหนบ้าง)', 'คำอธิบายไอเทม']
    widths = [5, 12, 30, 12, 14, 16, 14, 24, 10, 42, 40]
    for column, (header, width) in enumerate(zip(xlsx_headers, widths), 1):
        cell = sheet.cell(row=4, column=column, value=header)
        cell.fill = header_fill
        cell.font = white
        cell.alignment = Alignment(horizontal='center')
        sheet.column_dimensions[cell.column_letter].width = width
    for index, item in enumerate(results, 1):
        sheet.append(_export_row(index, item))
        row_index = 4 + index
        for column in range(1, 12):
            cell = sheet.cell(row=row_index, column=column)
            cell.fill = row_fill
            cell.font = green
            cell.alignment = Alignment(
                horizontal='center' if column in (1, 2, 4, 5, 6, 7, 9) else 'left')
    sheet.auto_filter.ref = 'A4:K4'
    sheet.freeze_panes = 'A5'
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()
